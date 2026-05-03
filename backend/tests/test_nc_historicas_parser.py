from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

import openpyxl
import pytest

from app.services.nc_historicas_parser import (
    NCHistoricasParser,
    ParseError,
    ParseResult,
)

SHEET = NCHistoricasParser.SHEET_NAME


def _build_xlsx(
    rows: list[list],
    headers: list[str] | None = None,
    sheet_name: str = SHEET,
) -> bytes:
    """Build a minimal xlsx with the given headers and data rows."""
    if headers is None:
        headers = ["tipo", "folio", "fecha", "motivo", "neto", "iva", "total"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _default_row(
    *,
    tipo: str = "NC",
    folio: int = 1001,
    fecha: str = "2024-01-15",
    motivo: str = "Corrección precio",
    neto: int = 100000,
    iva: int = 19000,
    total: int = 119000,
    folio_referencia: int | None = None,
    tipo_referencia: int | None = None,
) -> list:
    return [tipo, folio, fecha, motivo, neto, iva, total, folio_referencia, tipo_referencia]


_FULL_HEADERS = [
    "tipo", "folio", "fecha", "motivo", "neto", "iva", "total",
    "folio_referencia", "tipo_referencia",
]


# ---------------------------------------------------------------------------
# Happy-path tests
# ---------------------------------------------------------------------------

def test_valid_nc_row_created():
    data = _build_xlsx([_default_row()], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.a_crear == 1
    assert result.a_omitir == 0
    assert result.invalid_count == 0
    row = result.valid_rows[0]
    assert row.tipo == "NC"
    assert row.folio == 1001
    assert row.fecha == date(2024, 1, 15)
    assert row.motivo == "Corrección precio"
    assert row.neto == Decimal("100000")
    assert row.iva == Decimal("19000")
    assert row.total == Decimal("119000")
    assert row.folio_referencia is None
    assert row.tipo_referencia is None
    assert row.status == "crear"


def test_valid_nd_row_created():
    data = _build_xlsx(
        [_default_row(tipo="ND", folio=2001, motivo="Diferencia precio")],
        headers=_FULL_HEADERS,
    )
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.a_crear == 1
    row = result.valid_rows[0]
    assert row.tipo == "ND"
    assert row.folio == 2001
    assert row.status == "crear"


def test_tipo_case_insensitive():
    data = _build_xlsx([_default_row(tipo="nc")], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.a_crear == 1
    assert result.valid_rows[0].tipo == "NC"


def test_optional_reference_columns_parsed():
    data = _build_xlsx(
        [_default_row(folio_referencia=500, tipo_referencia=33)],
        headers=_FULL_HEADERS,
    )
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    row = result.valid_rows[0]
    assert row.folio_referencia == 500
    assert row.tipo_referencia == 33


def test_date_dd_mm_yyyy_format():
    data = _build_xlsx([_default_row(fecha="15-01-2024")], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.valid_rows[0].fecha == date(2024, 1, 15)


# ---------------------------------------------------------------------------
# Validation error tests
# ---------------------------------------------------------------------------

def test_invalid_tipo_rejected():
    data = _build_xlsx([_default_row(tipo="NF")], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.a_crear == 0
    assert result.invalid_count == 1
    assert "NC o ND" in result.invalid_rows[0].motivo


def test_invalid_folio_not_integer():
    data = _build_xlsx([_default_row(folio="ABC")], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.invalid_count == 1
    assert "folio" in result.invalid_rows[0].motivo.lower()


def test_invalid_folio_zero():
    data = _build_xlsx([_default_row(folio=0)], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.invalid_count == 1


def test_invalid_folio_negative():
    data = _build_xlsx([_default_row(folio=-5)], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.invalid_count == 1


def test_neto_plus_iva_not_equal_total_fails():
    # neto=100000, iva=19000, total=130000 → diff=11000 > 1
    data = _build_xlsx(
        [_default_row(neto=100000, iva=19000, total=130000)],
        headers=_FULL_HEADERS,
    )
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.invalid_count == 1
    assert "no coincide" in result.invalid_rows[0].motivo


def test_neto_plus_iva_within_tolerance_passes():
    # diff = 0.50 ≤ 1 → valid
    data = _build_xlsx(
        [_default_row(neto=100000, iva=19000, total=119000)],
        headers=_FULL_HEADERS,
    )
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.a_crear == 1


def test_motivo_required():
    data = _build_xlsx([_default_row(motivo="")], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.invalid_count == 1
    assert "motivo" in result.invalid_rows[0].motivo.lower()


def test_motivo_too_long():
    long_motivo = "x" * 501
    data = _build_xlsx([_default_row(motivo=long_motivo)], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.invalid_count == 1


def test_invalid_date():
    data = _build_xlsx([_default_row(fecha="not-a-date")], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.invalid_count == 1
    assert "fecha" in result.invalid_rows[0].motivo.lower()


def test_negative_neto_rejected():
    data = _build_xlsx([_default_row(neto=-1)], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.invalid_count == 1


# ---------------------------------------------------------------------------
# Idempotency
# ---------------------------------------------------------------------------

def test_idempotency_existing_folio_becomes_omitir():
    existing = {"NC": {1001}}
    data = _build_xlsx([_default_row(tipo="NC", folio=1001)], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", existing)
    assert result.a_omitir == 1
    assert result.a_crear == 0
    assert result.valid_rows[0].status == "omitir"


def test_idempotency_different_tipo_is_not_omitted():
    # NC folio 1001 exists, but we're importing ND 1001 — different tipo
    existing = {"NC": {1001}}
    data = _build_xlsx([_default_row(tipo="ND", folio=1001)], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", existing)
    assert result.a_crear == 1
    assert result.valid_rows[0].status == "crear"


def test_idempotency_new_folio_is_created():
    existing = {"NC": {999}}
    data = _build_xlsx([_default_row(tipo="NC", folio=1001)], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", existing)
    assert result.a_crear == 1
    assert result.valid_rows[0].status == "crear"


# ---------------------------------------------------------------------------
# Doc-row skipping
# ---------------------------------------------------------------------------

def test_doc_row_is_skipped():
    """Row 2 with documentation keywords should be silently skipped."""
    doc_row = ["requerido: NC o ND", "requerido: número DTE", "YYYY-MM-DD", "descripción", 0, 0, 0, None, None]
    real_row = _default_row()
    data = _build_xlsx([doc_row, real_row], headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.a_crear == 1
    assert result.invalid_count == 0


# ---------------------------------------------------------------------------
# Sheet name validation
# ---------------------------------------------------------------------------

def test_wrong_sheet_name_raises_parse_error():
    data = _build_xlsx([_default_row()], headers=_FULL_HEADERS, sheet_name="Hoja1")
    with pytest.raises(ParseError, match="NC-ND Históricas"):
        NCHistoricasParser.parse(data, "test.xlsx", {})


def test_missing_required_column_raises_parse_error():
    # Omit "total" column
    headers = ["tipo", "folio", "fecha", "motivo", "neto", "iva"]
    data = _build_xlsx([["NC", 1001, "2024-01-15", "Motivo", 100000, 19000]], headers=headers)
    with pytest.raises(ParseError, match="total"):
        NCHistoricasParser.parse(data, "test.xlsx", {})


def test_not_excel_raises_parse_error():
    with pytest.raises(ParseError):
        NCHistoricasParser.parse(b"not an excel file", "test.xlsx", {})


# ---------------------------------------------------------------------------
# Mixed valid/invalid rows
# ---------------------------------------------------------------------------

def test_mixed_rows_counted_correctly():
    rows = [
        _default_row(folio=1001),  # valid
        _default_row(folio=1002, tipo="BAD"),  # invalid tipo
        _default_row(folio=1003),  # valid
    ]
    data = _build_xlsx(rows, headers=_FULL_HEADERS)
    result = NCHistoricasParser.parse(data, "test.xlsx", {})
    assert result.a_crear == 2
    assert result.invalid_count == 1


# ---------------------------------------------------------------------------
# generate_template
# ---------------------------------------------------------------------------

def test_generate_template_returns_bytes():
    result = NCHistoricasParser.generate_template()
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_generate_template_has_correct_sheet():
    template = NCHistoricasParser.generate_template()
    wb = openpyxl.load_workbook(io.BytesIO(template))
    assert SHEET in wb.sheetnames


def test_generate_template_has_all_columns():
    template = NCHistoricasParser.generate_template()
    wb = openpyxl.load_workbook(io.BytesIO(template))
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    expected = ["tipo", "folio", "fecha", "motivo", "neto", "iva", "total",
                "folio_referencia", "tipo_referencia"]
    assert headers == expected


def test_generate_template_example_row_parseable():
    """The example data row (row 3) should pass parsing when doc row (row 2) is skipped."""
    template = NCHistoricasParser.generate_template()
    result = NCHistoricasParser.parse(template, "template.xlsx", {})
    # Template has 1 example row — it must be valid
    assert result.a_crear == 1
    assert result.invalid_count == 0
    row = result.valid_rows[0]
    assert row.tipo == "NC"
    assert row.folio == 1001

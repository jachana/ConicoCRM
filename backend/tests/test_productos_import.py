import io

import openpyxl

from app.services.producto_parser import ALL_COLUMNS


def _xlsx(rows: list[list], header=ALL_COLUMNS) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(header))
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, token, content: bytes, path: str = "/api/productos/import"):
    return client.post(
        path,
        files={"archivo": ("p.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )


def test_template_descarga_admin(client, admin_token):
    r = client.get("/api/productos/import/template", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert len(r.content) > 0


def test_template_requiere_admin(client, vendedor_token):
    r = client.get("/api/productos/import/template", headers={"Authorization": f"Bearer {vendedor_token}"})
    assert r.status_code == 403


def test_template_requiere_auth(client):
    r = client.get("/api/productos/import/template")
    assert r.status_code == 401


def test_preview_valido(client, admin_token):
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([
        ["SKU-001", "Producto A", "100.00", "50.00", "19", "Descripción A", "Electrónica", "kg", "1"],
        ["SKU-002", "Producto B", "200.00", "100.00", "0", "", "Ropa", "", "1"],
    ])
    r = _upload(client, admin_token, content, "/api/productos/import/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["filas_validas"] == 2
    assert body["filas_invalidas"] == 0
    assert body["a_crear"] == 2
    assert body["a_actualizar"] == 0


def test_preview_mixto(client, admin_token):
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([
        ["SKU-001", "Bueno", "100.00", "50.00", "19", "", "", "", ""],
        ["SKU-002", "Precio inválido", "abc", "50.00", "19", "", "", "", ""],
        ["", "Sin SKU", "100.00", "50.00", "19", "", "", "", ""],
    ])
    r = _upload(client, admin_token, content, "/api/productos/import/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["filas_validas"] == 1
    assert body["filas_invalidas"] == 2


def test_preview_sin_columnas_requeridas(client, admin_token):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["sku", "descripcion"])
    ws.append(["SKU-001", "x"])
    buf = io.BytesIO()
    wb.save(buf)
    r = _upload(client, admin_token, buf.getvalue(), "/api/productos/import/preview")
    assert r.status_code == 400
    assert "nombre" in r.json()["detail"]


def test_preview_marca_actualizar_si_existe(client, admin_token):
    # crea uno primero por POST normal
    client.post(
        "/api/productos/",
        json={"nombre": "Antiguo", "sku": "sku-001"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([
        ["SKU-001", "Nuevo nombre", "100.00", "50.00", "19", "", "", "", ""],
        ["SKU-002", "Nuevo", "200.00", "100.00", "0", "", "", "", ""],
    ])
    r = _upload(client, admin_token, content, "/api/productos/import/preview")
    body = r.json()
    assert body["a_crear"] == 1, f"Expected 1 to create, got {body['a_crear']}. Filas: {body['filas']}"
    assert body["a_actualizar"] == 1


def test_import_crea_productos(client, admin_token):
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([
        ["SKU-001", "Producto 1", "100.00", "50.00", "19", "Desc 1", "Electrónica", "kg", "1"],
        ["SKU-002", "Producto 2", "200.00", "100.00", "0", "", "Ropa", "unidad", "0"],
    ])
    r = _upload(client, admin_token, content)
    assert r.status_code == 200
    body = r.json()
    assert body["creadas"] == 2
    assert body["actualizadas"] == 0
    assert body["errores"] == 0
    # verify created
    get_r = client.get(
        "/api/productos/?q=SKU-001",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert get_r.status_code == 200
    prods = get_r.json()
    assert len(prods) > 0
    assert prods[0]["sku"] == "SKU-001"
    assert prods[0]["nombre"] == "Producto 1"
    assert float(prods[0]["precio_venta"]) == 100.0
    assert float(prods[0]["precio_costo"]) == 50.0


def test_import_actualiza_existentes(client, admin_token):
    # create first
    client.post(
        "/api/productos/",
        json={"nombre": "Viejo", "sku": "sku-001", "precio_venta": 50.0, "precio_costo": 25.0},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([
        ["SKU-001", "Nuevo nombre", "150.00", "75.00", "19", "Nueva desc", "Categoría", "L", "1"],
    ])
    r = _upload(client, admin_token, content)
    assert r.status_code == 200
    body = r.json()
    assert body["creadas"] == 0
    assert body["actualizadas"] == 1
    assert body["errores"] == 0
    # verify update
    get_r = client.get(
        "/api/productos/?q=SKU-001",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    prods = get_r.json()
    assert prods[0]["nombre"] == "Nuevo nombre"
    assert prods[0]["descripcion"] == "Nueva desc"
    assert float(prods[0]["precio_venta"]) == 150.0
    assert float(prods[0]["precio_costo"]) == 75.0


def test_import_duplicados_en_archivo(client, admin_token):
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([
        ["SKU-001", "Producto A", "100.00", "50.00", "19", "", "", "", ""],
        ["SKU-001", "Producto A Duplicado", "150.00", "75.00", "19", "", "", "", ""],
    ])
    r = _upload(client, admin_token, content, "/api/productos/import/preview")
    body = r.json()
    # El segundo debería ser válido (keep last), el primero inválido (duplicate)
    assert body["filas_validas"] == 1
    assert body["filas_invalidas"] == 1


def test_import_iva_validacion(client, admin_token):
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([
        ["SKU-001", "Producto", "100.00", "50.00", "25", "", "", "", ""],  # Invalid IVA
    ])
    r = _upload(client, admin_token, content, "/api/productos/import/preview")
    body = r.json()
    assert body["filas_validas"] == 0
    assert body["filas_invalidas"] == 1
    assert "IVA debe ser 0 o 19" in body["errores"][0]["motivo"]


def test_import_precio_negativo(client, admin_token):
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([
        ["SKU-001", "Producto", "-100.00", "50.00", "19", "", "", "", ""],  # Negative price
    ])
    r = _upload(client, admin_token, content, "/api/productos/import/preview")
    body = r.json()
    assert body["filas_validas"] == 0
    assert body["filas_invalidas"] == 1
    assert ">= 0" in body["errores"][0]["motivo"]


def test_import_con_familia_crea_tipo(client, admin_token):
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([
        ["SKU-001", "Producto", "100.00", "50.00", "19", "", "Nueva Categoría", "", ""],
    ])
    r = _upload(client, admin_token, content)
    assert r.status_code == 200
    body = r.json()
    assert body["creadas"] == 1
    # verify tipo was created
    get_r = client.get(
        "/api/productos/?q=SKU-001",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    prods = get_r.json()
    assert len(prods[0]["tipos"]) > 0
    assert prods[0]["tipos"][0]["nombre"] == "Nueva Categoría"


def test_import_sin_familia_no_crea_tipo(client, admin_token):
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([
        ["SKU-001", "Producto", "100.00", "50.00", "19", "", "", "", ""],  # No familia
    ])
    r = _upload(client, admin_token, content)
    assert r.status_code == 200
    body = r.json()
    assert body["creadas"] == 1
    # verify no tipo
    get_r = client.get(
        "/api/productos/?q=SKU-001",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    prods = get_r.json()
    assert len(prods[0]["tipos"]) == 0


def test_import_requiere_admin(client, vendedor_token):
    content = _xlsx([["SKU-001", "Producto", "", "", "", "100.00", "50.00", "19", ""]])
    r = _upload(client, vendedor_token, content)
    assert r.status_code == 403


def test_import_requiere_auth(client):
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([["SKU-001", "Producto", "100.00", "50.00", "19", "", "", "", ""]])
    r = client.post(
        "/api/productos/import",
        files={"archivo": ("p.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 401


def test_import_reporte_download(client, admin_token):
    content = _xlsx([
        ["SKU-001", "Bueno", "", "", "", "100.00", "50.00", "19", ""],
        ["SKU-002", "Mal precio", "", "", "", "abc", "50.00", "19", ""],
    ])
    # Do import
    import_r = _upload(client, admin_token, content)
    body = import_r.json()
    # Download report
    report_r = client.post(
        "/api/productos/import/report",
        json={"detalles": body["detalles"]},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert report_r.status_code == 200
    assert "spreadsheetml" in report_r.headers["content-type"]


def test_import_sku_case_insensitive(client, admin_token):
    # Create with lowercase
    client.post(
        "/api/productos/",
        json={"nombre": "Producto", "sku": "sku-001"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    # Import with uppercase
    # ALL_COLUMNS: sku, nombre, precio_base, costo, iva, descripcion, familia, unidad, afecto
    content = _xlsx([
        ["SKU-001", "Updated", "100.00", "50.00", "19", "", "", "", ""],
    ])
    r = _upload(client, admin_token, content, "/api/productos/import/preview")
    body = r.json()
    # Should recognize as existing
    assert body["a_actualizar"] == 1
    assert body["a_crear"] == 0


def test_import_empty_file(client, admin_token):
    wb = openpyxl.Workbook()
    ws = wb.active
    buf = io.BytesIO()
    wb.save(buf)
    r = _upload(client, admin_token, buf.getvalue(), "/api/productos/import/preview")
    assert r.status_code == 400


def test_import_invalid_xlsx(client, admin_token):
    r = client.post(
        "/api/productos/import/preview",
        files={"archivo": ("p.xlsx", b"not valid xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 400

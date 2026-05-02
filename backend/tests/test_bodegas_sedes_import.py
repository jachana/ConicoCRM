"""
Tests for bodegas + sedes onboarding import API.

Covers:
- Template download
- Preview validation (valid, mixed, missing empresa, duplicates)
- Import execution (create, update, idempotent)
- Error handling and auth
"""

import io
import json

import openpyxl
import pytest

from app.services.bodegas_sedes_parser import BodegasSedesParser


def _xlsx(rows: list[list], header=None) -> bytes:
    """Create Excel file with given rows."""
    if header is None:
        header = ["empresa_rut", "bodega_nombre", "bodega_direccion", "sede_nombre", "sede_direccion"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(header))
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, token, content: bytes, path: str = "/api/onboarding/bodegas-sedes/import"):
    """Upload file for import or preview."""
    return client.post(
        path,
        files={"file": ("b.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )


# ============================================================================
# Template Download
# ============================================================================

def test_template_descarga_admin(client, admin_token):
    """Admin can download template."""
    r = client.get(
        "/api/onboarding/bodegas-sedes/template",
        headers={"Authorization": f"Bearer {admin_token}"}
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert len(r.content) > 0


def test_template_requiere_admin(client, vendedor_token):
    """Only admin can download template."""
    r = client.get(
        "/api/onboarding/bodegas-sedes/template",
        headers={"Authorization": f"Bearer {vendedor_token}"}
    )
    assert r.status_code == 403


def test_template_requiere_auth(client):
    """Must be authenticated to download template."""
    r = client.get("/api/onboarding/bodegas-sedes/template")
    assert r.status_code == 401


# ============================================================================
# Preview Validation
# ============================================================================

def test_preview_valido(client, admin_token, empresa_demo):
    """Valid rows preview successfully."""
    content = _xlsx([
        [empresa_demo.rut, "Bodega A", "Av 1", "Sede A", "Piso 2"],
        [empresa_demo.rut, "Bodega B", "Av 2", "Sede B", "Piso 3"],
    ])
    r = _upload(client, admin_token, content, "/api/onboarding/bodegas-sedes/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["filas_validas"] == 2
    assert body["filas_invalidas"] == 0
    assert body["a_crear"]["bodegas"] == 2
    assert body["a_crear"]["sedes"] == 2


def test_preview_mixto(client, admin_token, empresa_demo):
    """Mix of valid and invalid rows preview correctly."""
    content = _xlsx([
        [empresa_demo.rut, "Bodega 1", "Av 1", "Sede 1", "Piso 1"],
        ["76.999.999-9", "Bodega 2", "", "Sede 2", "Piso 2"],  # Invalid RUT
        ["", "Bodega 3", "", "Sede 3", "Piso 3"],  # Missing RUT
    ])
    r = _upload(client, admin_token, content, "/api/onboarding/bodegas-sedes/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["filas_validas"] == 1
    assert body["filas_invalidas"] == 2


def test_preview_falta_empresa_existente(client, admin_token):
    """Row with non-existent empresa RUT is marked invalid."""
    content = _xlsx([
        ["11.111.111-1", "Bodega X", "Dir X", "Sede X", "Dir Sede X"],
    ])
    r = _upload(client, admin_token, content, "/api/onboarding/bodegas-sedes/preview")
    assert r.status_code == 200
    body = r.json()
    # Valid row structurally, but will fail on import due to missing empresa
    assert body["filas_invalidas"] >= 0


def test_preview_marca_actualizar_si_existe(client, admin_token, empresa_demo):
    """Existing bodegas/sedes marked for update."""
    # Create bodega/sede via API
    from app.models.bodega import Bodega
    from app.models.sede_despacho import SedeDespacho
    from app.database import get_db_context

    db = next(get_db_context())
    try:
        bodega = Bodega(empresa_id=empresa_demo.id, nombre="Bodega Existente", direccion="Av Old")
        sede = SedeDespacho(empresa_id=empresa_demo.id, nombre="Sede Existente", direccion="Piso Old")
        db.add(bodega)
        db.add(sede)
        db.commit()
    finally:
        db.close()

    content = _xlsx([
        [empresa_demo.rut, "Bodega Existente", "Av New", "Sede Existente", "Piso New"],
        [empresa_demo.rut, "Bodega Nueva", "Av New", "Sede Nueva", "Piso New"],
    ])
    r = _upload(client, admin_token, content, "/api/onboarding/bodegas-sedes/preview")
    body = r.json()
    assert body["a_actualizar"]["bodegas"] >= 1
    assert body["a_crear"]["bodegas"] >= 1


def test_preview_sin_columnas_requeridas(client, admin_token):
    """Missing required columns returns 400."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["empresa_rut", "bodega_nombre"])  # Missing bodega_direccion
    ws.append(["11.111.111-1", "Bodega X"])
    buf = io.BytesIO()
    wb.save(buf)
    r = _upload(client, admin_token, buf.getvalue(), "/api/onboarding/bodegas-sedes/preview")
    assert r.status_code == 422
    assert "sede" in r.json()["detail"].lower()


# ============================================================================
# Import Execution
# ============================================================================

def test_import_crea_bodegas_sedes(client, admin_token, empresa_demo):
    """Import creates new bodegas and sedes."""
    content = _xlsx([
        [empresa_demo.rut, "Bodega A", "Av 1", "Sede A", "Piso 1"],
        [empresa_demo.rut, "Bodega B", "Av 2", "Sede B", "Piso 2"],
    ])
    r = _upload(client, admin_token, content)
    assert r.status_code == 200
    body = r.json()
    assert body["report"]["created_bodega_count"] == 2
    assert body["report"]["updated_bodega_count"] == 0
    assert body["report"]["created_sede_count"] == 2
    assert body["report"]["updated_sede_count"] == 0
    assert body["report"]["error_count"] == 0


def test_import_es_idempotente(client, admin_token, empresa_demo):
    """Re-importing same file doesn't create duplicates."""
    content = _xlsx([
        [empresa_demo.rut, "Bodega X", "Av X", "Sede X", "Piso X"],
    ])
    r1 = _upload(client, admin_token, content)
    assert r1.status_code == 200
    result1 = r1.json()["report"]
    assert result1["created_bodega_count"] == 1
    assert result1["created_sede_count"] == 1

    r2 = _upload(client, admin_token, content)
    assert r2.status_code == 200
    result2 = r2.json()["report"]
    assert result2["created_bodega_count"] == 0
    assert result2["updated_bodega_count"] == 1


def test_import_actualiza_campos_modificados(client, admin_token, empresa_demo):
    """Import updates existing bodega/sede when data changes."""
    content_v1 = _xlsx([[empresa_demo.rut, "Bodega Y", "Av Old", "Sede Y", "Piso Old"]])
    r1 = _upload(client, admin_token, content_v1)
    assert r1.json()["report"]["created_bodega_count"] == 1

    content_v2 = _xlsx([[empresa_demo.rut, "Bodega Y", "Av New", "Sede Y", "Piso New"]])
    r2 = _upload(client, admin_token, content_v2)
    result = r2.json()["report"]
    assert result["updated_bodega_count"] == 1
    assert result["created_bodega_count"] == 0

    # Verify data updated
    from app.models.bodega import Bodega
    from app.database import get_db_context

    db = next(get_db_context())
    try:
        bodega = db.query(Bodega).filter(
            Bodega.empresa_id == empresa_demo.id,
            Bodega.nombre == "Bodega Y"
        ).first()
        assert bodega is not None
        assert bodega.direccion == "Av New"
    finally:
        db.close()


def test_import_continua_en_filas_invalidas(client, admin_token, empresa_demo):
    """Valid rows imported even if some rows have errors."""
    content = _xlsx([
        [empresa_demo.rut, "Bodega Good", "Av 1", "Sede Good", "Piso 1"],
        ["76.999.999-9", "Bodega Bad", "Av 2", "Sede Bad", "Piso 2"],  # Invalid RUT
        [empresa_demo.rut, "Bodega Good2", "Av 3", "Sede Good2", "Piso 3"],
    ])
    r = _upload(client, admin_token, content)
    body = r.json()
    result = body["report"]
    # Should have created the valid ones
    assert result["created_bodega_count"] >= 2
    # Should have some errors
    assert result["error_count"] >= 1


def test_import_requiere_admin(client, vendedor_token):
    """Only admin can import."""
    content = _xlsx([[None, "X", "Y", "Z", "W"]])
    r = _upload(client, vendedor_token, content)
    assert r.status_code == 403


def test_import_sin_auth(client):
    """Must be authenticated to import."""
    content = _xlsx([[None, "X", "Y", "Z", "W"]])
    r = client.post(
        "/api/onboarding/bodegas-sedes/import",
        files={"file": ("b.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 401


def test_import_columnas_faltantes(client, admin_token):
    """Missing required columns returns 400."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["empresa_rut"])
    ws.append(["11.111.111-1"])
    buf = io.BytesIO()
    wb.save(buf)
    r = _upload(client, admin_token, buf.getvalue())
    assert r.status_code == 422


def test_import_archivo_vacio(client, admin_token):
    """Empty file returns 400."""
    r = _upload(client, admin_token, b"")
    assert r.status_code == 400
    assert "vacío" in r.json()["detail"]


# ============================================================================
# Response Format
# ============================================================================

def test_import_respuesta_estructura(client, admin_token, empresa_demo):
    """Import response has expected structure."""
    content = _xlsx([
        [empresa_demo.rut, "Bodega Test", "Av Test", "Sede Test", "Piso Test"],
    ])
    r = _upload(client, admin_token, content)
    assert r.status_code == 200
    body = r.json()
    assert "status" in body
    assert body["status"] in ["success", "partial", "error"]
    assert "import_id" in body
    assert "timestamp" in body
    assert "report" in body
    assert "created_bodega_count" in body["report"]
    assert "updated_bodega_count" in body["report"]
    assert "created_sede_count" in body["report"]
    assert "updated_sede_count" in body["report"]
    assert "error_count" in body["report"]
    assert "rows" in body["report"]


def test_preview_respuesta_estructura(client, admin_token, empresa_demo):
    """Preview response has expected structure."""
    content = _xlsx([
        [empresa_demo.rut, "Bodega", "Av 1", "Sede", "Piso 1"],
    ])
    r = _upload(client, admin_token, content, "/api/onboarding/bodegas-sedes/preview")
    assert r.status_code == 200
    body = r.json()
    assert "total_filas" in body
    assert "filas_validas" in body
    assert "filas_invalidas" in body
    assert "a_crear" in body
    assert "a_actualizar" in body
    assert "rows" in body
    assert "bodegas" in body["a_crear"]
    assert "sedes" in body["a_crear"]

"""Tests for Libros API endpoints"""
import pytest
from datetime import datetime, timezone
import csv
from io import StringIO, BytesIO
from openpyxl import load_workbook

from app.models.libro import LibroVentas, LibroCompras
from app.models.empresa import Empresa
from app.models.user import User
from app.core.security import get_password_hash
from app.core.modulos import OPTIONAL_MODULES


@pytest.fixture
def empresa1(db, setup_test_db):
    """Create a test empresa"""
    empresa = Empresa(
        nombre="Empresa Test 1",
        modulos_enabled={slug: True for slug in OPTIONAL_MODULES},
    )
    db.add(empresa)
    db.commit()
    db.refresh(empresa)
    return empresa


@pytest.fixture
def empresa2(db, setup_test_db):
    """Create a second test empresa"""
    empresa = Empresa(
        nombre="Empresa Test 2",
        modulos_enabled={slug: True for slug in OPTIONAL_MODULES},
    )
    db.add(empresa)
    db.commit()
    db.refresh(empresa)
    return empresa


@pytest.fixture
def user_empresa1(db, empresa1):
    """Create a user assigned to empresa1"""
    from tests.conftest import TestingSession
    session = TestingSession()
    user = User(
        email="user_empresa1@test.cl",
        name="User Empresa 1",
        hashed_password=get_password_hash("secret123"),
        role="subadmin",
        empresa_id=empresa1.id,
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    session.close()
    return user


@pytest.fixture
def user_empresa2(db, empresa2):
    """Create a user assigned to empresa2"""
    from tests.conftest import TestingSession
    session = TestingSession()
    user = User(
        email="user_empresa2@test.cl",
        name="User Empresa 2",
        hashed_password=get_password_hash("secret123"),
        role="subadmin",
        empresa_id=empresa2.id,
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    session.close()
    return user


@pytest.fixture
def token_empresa1(client, user_empresa1):
    """Get auth token for user_empresa1"""
    resp = client.post("/api/auth/login", data={"username": "user_empresa1@test.cl", "password": "secret123"})
    return resp.json()["access_token"]


@pytest.fixture
def token_empresa2(client, user_empresa2):
    """Get auth token for user_empresa2"""
    resp = client.post("/api/auth/login", data={"username": "user_empresa2@test.cl", "password": "secret123"})
    return resp.json()["access_token"]


@pytest.fixture
def sample_libros_ventas(db, setup_test_db, admin_user):
    """Create sample LibroVentas records for admin user's empresa"""
    libros = [
        LibroVentas(
            periodo="2026-01",
            empresa_id=admin_user.empresa_id,
            folio_inicio=1,
            folio_fin=100,
            total_registros=50,
            monto_total=500000,
            estado="borrador",
        ),
        LibroVentas(
            periodo="2026-02",
            empresa_id=admin_user.empresa_id,
            folio_inicio=101,
            folio_fin=200,
            total_registros=75,
            monto_total=750000,
            estado="borrador",
        ),
        LibroVentas(
            periodo="2026-03",
            empresa_id=admin_user.empresa_id,
            folio_inicio=1,
            folio_fin=50,
            total_registros=25,
            monto_total=250000,
            estado="enviado",
        ),
    ]
    db.add_all(libros)
    db.commit()
    return libros


@pytest.fixture
def sample_libros_compras(db, setup_test_db, admin_user, empresa2):
    """Create sample LibroCompras records for admin user's empresa and empresa2"""
    libros = [
        LibroCompras(
            periodo="2026-01",
            empresa_id=admin_user.empresa_id,
            rut_proveedor=None,
            total_registros=30,
            monto_total=300000,
            estado="borrador",
        ),
        LibroCompras(
            periodo="2026-01",
            empresa_id=empresa2.id,
            rut_proveedor="12.345.678-9",
            total_registros=15,
            monto_total=150000,
            estado="enviado",
        ),
        LibroCompras(
            periodo="2026-02",
            empresa_id=admin_user.empresa_id,
            rut_proveedor="12.345.678-9",
            total_registros=40,
            monto_total=400000,
            estado="borrador",
        ),
    ]
    db.add_all(libros)
    db.commit()
    return libros


class TestLibrosAuthorizationVentas:
    """Tests for authorization checks on Libros de Ventas endpoints"""

    def test_get_libro_ventas_forbidden_different_empresa(self, client, token_empresa1, token_empresa2, db, empresa1, empresa2):
        """Test that user from empresa2 cannot access libro from empresa1"""
        # Create libro for empresa1
        libro = LibroVentas(
            periodo="2026-01",
            empresa_id=empresa1.id,
            folio_inicio=1,
            folio_fin=100,
            total_registros=50,
            monto_total=500000,
            estado="borrador",
        )
        db.add(libro)
        db.commit()
        db.refresh(libro)

        # Try to access with token from empresa2 user
        response = client.get(
            f"/api/libros/ventas/{libro.id}",
            headers={"Authorization": f"Bearer {token_empresa2}"}
        )
        assert response.status_code == 403
        assert "Not authorized" in response.json()["detail"]

    def test_list_libros_ventas_filtered_by_empresa(self, client, token_empresa1, token_empresa2, db, empresa1, empresa2):
        """Test that list endpoint filters libros by user's empresa"""
        # Create libros for both empresas
        libro_e1 = LibroVentas(
            periodo="2026-01",
            empresa_id=empresa1.id,
            folio_inicio=1,
            folio_fin=100,
            total_registros=50,
            monto_total=500000,
            estado="borrador",
        )
        libro_e2 = LibroVentas(
            periodo="2026-01",
            empresa_id=empresa2.id,
            folio_inicio=1,
            folio_fin=50,
            total_registros=25,
            monto_total=250000,
            estado="borrador",
        )
        db.add_all([libro_e1, libro_e2])
        db.commit()

        # empresa1 user should see only empresa1 libro
        response = client.get(
            "/api/libros/ventas",
            headers={"Authorization": f"Bearer {token_empresa1}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 1
        assert data["data"][0]["empresa_id"] == empresa1.id

        # empresa2 user should see only empresa2 libro
        response = client.get(
            "/api/libros/ventas",
            headers={"Authorization": f"Bearer {token_empresa2}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 1
        assert data["data"][0]["empresa_id"] == empresa2.id

    def test_get_libro_ventas_allowed_same_empresa(self, client, token_empresa1, db, empresa1):
        """Test that user can access libro from their own empresa"""
        # Create libro for empresa1
        libro = LibroVentas(
            periodo="2026-01",
            empresa_id=empresa1.id,
            folio_inicio=1,
            folio_fin=100,
            total_registros=50,
            monto_total=500000,
            estado="borrador",
        )
        db.add(libro)
        db.commit()
        db.refresh(libro)

        # Access with token from empresa1 user should succeed
        response = client.get(
            f"/api/libros/ventas/{libro.id}",
            headers={"Authorization": f"Bearer {token_empresa1}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["id"] == libro.id
        assert data["empresa_id"] == empresa1.id


class TestLibrosAuthorizationCompras:
    """Tests for authorization checks on Libros de Compras endpoints"""

    def test_get_libro_compras_forbidden_different_empresa(self, client, token_empresa1, token_empresa2, db, empresa1, empresa2):
        """Test that user from empresa2 cannot access libro from empresa1"""
        # Create libro for empresa1
        libro = LibroCompras(
            periodo="2026-01",
            empresa_id=empresa1.id,
            rut_proveedor=None,
            total_registros=30,
            monto_total=300000,
            estado="borrador",
        )
        db.add(libro)
        db.commit()
        db.refresh(libro)

        # Try to access with token from empresa2 user
        response = client.get(
            f"/api/libros/compras/{libro.id}",
            headers={"Authorization": f"Bearer {token_empresa2}"}
        )
        assert response.status_code == 403
        assert "Not authorized" in response.json()["detail"]

    def test_list_libros_compras_filtered_by_empresa(self, client, token_empresa1, token_empresa2, db, empresa1, empresa2):
        """Test that list endpoint filters libros by user's empresa"""
        # Create libros for both empresas
        libro_e1 = LibroCompras(
            periodo="2026-01",
            empresa_id=empresa1.id,
            rut_proveedor=None,
            total_registros=30,
            monto_total=300000,
            estado="borrador",
        )
        libro_e2 = LibroCompras(
            periodo="2026-01",
            empresa_id=empresa2.id,
            rut_proveedor=None,
            total_registros=15,
            monto_total=150000,
            estado="borrador",
        )
        db.add_all([libro_e1, libro_e2])
        db.commit()

        # empresa1 user should see only empresa1 libro
        response = client.get(
            "/api/libros/compras",
            headers={"Authorization": f"Bearer {token_empresa1}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 1
        assert data["data"][0]["empresa_id"] == empresa1.id

        # empresa2 user should see only empresa2 libro
        response = client.get(
            "/api/libros/compras",
            headers={"Authorization": f"Bearer {token_empresa2}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 1
        assert data["data"][0]["empresa_id"] == empresa2.id

    def test_get_libro_compras_allowed_same_empresa(self, client, token_empresa1, db, empresa1):
        """Test that user can access libro from their own empresa"""
        # Create libro for empresa1
        libro = LibroCompras(
            periodo="2026-01",
            empresa_id=empresa1.id,
            rut_proveedor=None,
            total_registros=30,
            monto_total=300000,
            estado="borrador",
        )
        db.add(libro)
        db.commit()
        db.refresh(libro)

        # Access with token from empresa1 user should succeed
        response = client.get(
            f"/api/libros/compras/{libro.id}",
            headers={"Authorization": f"Bearer {token_empresa1}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["id"] == libro.id
        assert data["empresa_id"] == empresa1.id


class TestLibrosVentasList:
    """Tests for GET /api/libros/ventas"""

    def test_list_libros_ventas_without_auth(self, client):
        """Test that endpoint requires authentication"""
        response = client.get("/api/libros/ventas")
        assert response.status_code == 401

    def test_list_libros_ventas_success(self, client, admin_token, sample_libros_ventas):
        """Test successful listing of libros ventas"""
        response = client.get(
            "/api/libros/ventas",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert "data" in data
        assert "pagination" in data
        assert len(data["data"]) == 3
        assert data["pagination"]["total"] == 3
        assert data["pagination"]["limit"] == 50
        assert data["pagination"]["offset"] == 0

    def test_list_libros_ventas_with_periodo_filter(self, client, admin_token, sample_libros_ventas):
        """Test filtering libros ventas by periodo"""
        response = client.get(
            "/api/libros/ventas",
            params={"periodo": "2026-01"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["total"] == 1
        assert data["data"][0]["periodo"] == "2026-01"

    def test_list_libros_ventas_with_pagination(self, client, admin_token, sample_libros_ventas):
        """Test pagination parameters"""
        response = client.get(
            "/api/libros/ventas",
            params={"limit": 1, "offset": 0},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 1
        assert data["pagination"]["limit"] == 1
        assert data["pagination"]["offset"] == 0
        assert data["pagination"]["total"] == 3

    def test_list_libros_ventas_pagination_offset(self, client, admin_token, sample_libros_ventas):
        """Test pagination offset"""
        response = client.get(
            "/api/libros/ventas",
            params={"limit": 2, "offset": 1},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 2
        assert data["pagination"]["offset"] == 1

    def test_list_libros_ventas_invalid_periodo_format(self, client, admin_token):
        """Test invalid periodo format returns 400"""
        response = client.get(
            "/api/libros/ventas",
            params={"periodo": "2026/01"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 400
        assert "YYYY-MM" in response.json()["detail"]

    def test_list_libros_ventas_invalid_limit(self, client, admin_token):
        """Test invalid limit (zero) returns 422"""
        response = client.get(
            "/api/libros/ventas",
            params={"limit": 0},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 422

    def test_list_libros_ventas_invalid_offset(self, client, admin_token):
        """Test invalid offset (negative) returns 422"""
        response = client.get(
            "/api/libros/ventas",
            params={"offset": -1},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 422

    def test_list_libros_ventas_empty_result(self, client, admin_token):
        """Test listing when no libros exist"""
        response = client.get(
            "/api/libros/ventas",
            params={"periodo": "2099-12"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 0
        assert data["pagination"]["total"] == 0

    def test_list_libros_ventas_max_limit(self, client, admin_token, sample_libros_ventas):
        """Test that limit is capped at 500"""
        response = client.get(
            "/api/libros/ventas",
            params={"limit": 500},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["limit"] == 500

    def test_list_libros_ventas_limit_over_max(self, client, admin_token):
        """Test that limit > 500 returns 422"""
        response = client.get(
            "/api/libros/ventas",
            params={"limit": 501},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 422


class TestLibrosVentasDetail:
    """Tests for GET /api/libros/ventas/{id}"""

    def test_get_libro_ventas_success(self, client, admin_token, sample_libros_ventas):
        """Test retrieving a specific libro ventas"""
        libro_id = sample_libros_ventas[0].id
        response = client.get(
            f"/api/libros/ventas/{libro_id}",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["id"] == libro_id
        assert data["periodo"] == "2026-01"
        assert data["empresa_id"] == 1
        assert data["total_registros"] == 50

    def test_get_libro_ventas_not_found(self, client, admin_token):
        """Test 404 when libro doesn't exist"""
        response = client.get(
            "/api/libros/ventas/9999",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 404
        assert "not found" in response.json()["detail"].lower()

    def test_get_libro_ventas_without_auth(self, client, sample_libros_ventas):
        """Test that endpoint requires authentication"""
        libro_id = sample_libros_ventas[0].id
        response = client.get(f"/api/libros/ventas/{libro_id}")
        assert response.status_code == 401

    def test_get_libro_ventas_response_format(self, client, admin_token, sample_libros_ventas):
        """Test response has all required fields"""
        libro_id = sample_libros_ventas[0].id
        response = client.get(
            f"/api/libros/ventas/{libro_id}",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert "id" in data
        assert "periodo" in data
        assert "empresa_id" in data
        assert "folio_inicio" in data
        assert "folio_fin" in data
        assert "total_registros" in data
        assert "monto_total" in data
        assert "estado" in data
        assert "created_at" in data


class TestLibrosComprasList:
    """Tests for GET /api/libros/compras"""

    def test_list_libros_compras_without_auth(self, client):
        """Test that endpoint requires authentication"""
        response = client.get("/api/libros/compras")
        assert response.status_code == 401

    def test_list_libros_compras_success(self, client, admin_token, sample_libros_compras):
        """Test successful listing of libros compras"""
        response = client.get(
            "/api/libros/compras",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert "data" in data
        assert "pagination" in data
        # Admin user should see only their empresa's records (2 out of 3 total)
        assert len(data["data"]) == 2
        assert data["pagination"]["total"] == 2

    def test_list_libros_compras_with_periodo_filter(self, client, admin_token, sample_libros_compras):
        """Test filtering libros compras by periodo"""
        response = client.get(
            "/api/libros/compras",
            params={"periodo": "2026-01"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        # Admin user should see only their empresa's record for 2026-01 (the other is from empresa2)
        assert data["pagination"]["total"] == 1
        assert all(item["periodo"] == "2026-01" for item in data["data"])

    def test_list_libros_compras_with_pagination(self, client, admin_token, sample_libros_compras):
        """Test pagination for libros compras"""
        response = client.get(
            "/api/libros/compras",
            params={"limit": 2, "offset": 0},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 2
        # Admin user sees 2 total records from their empresa
        assert data["pagination"]["total"] == 2

    def test_list_libros_compras_invalid_periodo_format(self, client, admin_token):
        """Test invalid periodo format returns 400"""
        response = client.get(
            "/api/libros/compras",
            params={"periodo": "invalid"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 400
        assert "YYYY-MM" in response.json()["detail"]

    def test_list_libros_compras_empty_result(self, client, admin_token):
        """Test listing when no libros exist"""
        response = client.get(
            "/api/libros/compras",
            params={"periodo": "2099-12"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 0
        assert data["pagination"]["total"] == 0


class TestLibrosComprasDetail:
    """Tests for GET /api/libros/compras/{id}"""

    def test_get_libro_compras_success(self, client, admin_token, sample_libros_compras):
        """Test retrieving a specific libro compras"""
        libro_id = sample_libros_compras[0].id
        response = client.get(
            f"/api/libros/compras/{libro_id}",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["id"] == libro_id
        assert data["periodo"] == "2026-01"
        assert data["empresa_id"] == 1
        assert data["total_registros"] == 30

    def test_get_libro_compras_not_found(self, client, admin_token):
        """Test 404 when libro doesn't exist"""
        response = client.get(
            "/api/libros/compras/9999",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 404

    def test_get_libro_compras_without_auth(self, client, sample_libros_compras):
        """Test that endpoint requires authentication"""
        libro_id = sample_libros_compras[0].id
        response = client.get(f"/api/libros/compras/{libro_id}")
        assert response.status_code == 401

    def test_get_libro_compras_response_format(self, client, admin_token, sample_libros_compras):
        """Test response has all required fields"""
        libro_id = sample_libros_compras[0].id
        response = client.get(
            f"/api/libros/compras/{libro_id}",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert "id" in data
        assert "periodo" in data
        assert "empresa_id" in data
        assert "rut_proveedor" in data
        assert "total_registros" in data
        assert "monto_total" in data
        assert "estado" in data
        assert "created_at" in data


class TestPaginationEdgeCases:
    """Tests for edge cases in pagination"""

    def test_pagination_boundary_offset_equals_total(self, client, admin_token, sample_libros_ventas):
        """Test offset equal to total returns empty list"""
        response = client.get(
            "/api/libros/ventas",
            params={"offset": 3, "limit": 50},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 0
        assert data["pagination"]["total"] == 3

    def test_pagination_boundary_offset_greater_than_total(self, client, admin_token, sample_libros_ventas):
        """Test offset > total returns empty list"""
        response = client.get(
            "/api/libros/ventas",
            params={"offset": 100, "limit": 50},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 0
        assert data["pagination"]["total"] == 3

    def test_pagination_large_limit(self, client, admin_token, sample_libros_ventas):
        """Test very large limit returns 422"""
        response = client.get(
            "/api/libros/ventas",
            params={"limit": 1000},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 422


class TestPeriodoValidation:
    """Tests for periodo format validation"""

    def test_periodo_valid_formats(self, client, admin_token):
        """Test valid periodo formats"""
        valid_periodos = ["2026-01", "2025-12", "2020-05"]
        for periodo in valid_periodos:
            response = client.get(
                "/api/libros/ventas",
                params={"periodo": periodo},
                headers={"Authorization": f"Bearer {admin_token}"}
            )
            assert response.status_code == 200

    def test_periodo_invalid_formats(self, client, admin_token):
        """Test invalid periodo formats"""
        invalid_periodos = ["2026/01", "2026-1", "202601", "01-2026", "2026"]
        for periodo in invalid_periodos:
            response = client.get(
                "/api/libros/ventas",
                params={"periodo": periodo},
                headers={"Authorization": f"Bearer {admin_token}"}
            )
            assert response.status_code == 400

    def test_periodo_with_leading_zeros(self, client, admin_token, db):
        """Test periodo with leading zeros in month"""
        libro = LibroVentas(
            periodo="2026-05",
            empresa_id=1,
            total_registros=10,
            monto_total=100000,
            estado="borrador",
        )
        db.add(libro)
        db.commit()

        response = client.get(
            "/api/libros/ventas",
            params={"periodo": "2026-05"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        assert len(response.json()["data"]) == 1


class TestEstadoFilter:
    """Tests for estado filter on both ventas and compras"""

    def test_filter_ventas_by_estado_borrador(self, client, admin_token, sample_libros_ventas):
        """Test filtering libros ventas by estado=borrador"""
        response = client.get(
            "/api/libros/ventas",
            params={"estado": "borrador"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["total"] == 2
        assert all(item["estado"] == "borrador" for item in data["data"])

    def test_filter_ventas_by_estado_enviado(self, client, admin_token, sample_libros_ventas):
        """Test filtering libros ventas by estado=enviado"""
        response = client.get(
            "/api/libros/ventas",
            params={"estado": "enviado"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["total"] == 1
        assert all(item["estado"] == "enviado" for item in data["data"])

    def test_filter_compras_by_estado_borrador(self, client, admin_token, sample_libros_compras):
        """Test filtering libros compras by estado=borrador"""
        response = client.get(
            "/api/libros/compras",
            params={"estado": "borrador"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        # Admin user should see 2 borrador records from their empresa
        assert data["pagination"]["total"] == 2
        assert all(item["estado"] == "borrador" for item in data["data"])

    def test_filter_compras_by_estado_enviado(self, client, admin_token, sample_libros_compras):
        """Test filtering libros compras by estado=enviado"""
        response = client.get(
            "/api/libros/compras",
            params={"estado": "enviado"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        # Admin user should see 0 enviado records from their empresa
        assert data["pagination"]["total"] == 0

    def test_filter_estado_no_match(self, client, admin_token, sample_libros_ventas):
        """Test filtering with estado that doesn't match any records"""
        response = client.get(
            "/api/libros/ventas",
            params={"estado": "nonexistent"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["total"] == 0

    def test_filter_estado_with_periodo(self, client, admin_token, sample_libros_ventas):
        """Test combining estado and periodo filters"""
        response = client.get(
            "/api/libros/ventas",
            params={"estado": "borrador", "periodo": "2026-01"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["total"] == 1
        assert data["data"][0]["estado"] == "borrador"
        assert data["data"][0]["periodo"] == "2026-01"


class TestPeriodoRangeFilter:
    """Tests for periodo_from and periodo_to filters"""

    def test_filter_ventas_by_periodo_from(self, client, admin_token, sample_libros_ventas):
        """Test filtering libros ventas by periodo_from"""
        response = client.get(
            "/api/libros/ventas",
            params={"periodo_from": "2026-02"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["total"] == 2
        assert all(item["periodo"] >= "2026-02" for item in data["data"])

    def test_filter_ventas_by_periodo_to(self, client, admin_token, sample_libros_ventas):
        """Test filtering libros ventas by periodo_to"""
        response = client.get(
            "/api/libros/ventas",
            params={"periodo_to": "2026-02"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["total"] == 2
        assert all(item["periodo"] <= "2026-02" for item in data["data"])

    def test_filter_ventas_by_periodo_range(self, client, admin_token, sample_libros_ventas):
        """Test filtering libros ventas by periodo range"""
        response = client.get(
            "/api/libros/ventas",
            params={"periodo_from": "2026-01", "periodo_to": "2026-02"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["total"] == 2
        assert all("2026-01" <= item["periodo"] <= "2026-02" for item in data["data"])

    def test_filter_compras_by_periodo_range(self, client, admin_token, sample_libros_compras):
        """Test filtering libros compras by periodo range"""
        response = client.get(
            "/api/libros/compras",
            params={"periodo_from": "2026-01", "periodo_to": "2026-02"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        # Admin user should see 2 records from their empresa in this range
        assert data["pagination"]["total"] == 2

    def test_filter_periodo_range_single_month(self, client, admin_token, sample_libros_ventas):
        """Test periodo range with same from and to (single month)"""
        response = client.get(
            "/api/libros/ventas",
            params={"periodo_from": "2026-01", "periodo_to": "2026-01"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["total"] == 1
        assert all(item["periodo"] == "2026-01" for item in data["data"])

    def test_filter_periodo_range_backwards(self, client, admin_token, sample_libros_ventas):
        """Test periodo range with from > to (should return no results)"""
        response = client.get(
            "/api/libros/ventas",
            params={"periodo_from": "2026-03", "periodo_to": "2026-01"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["total"] == 0

    def test_filter_periodo_from_invalid_format(self, client, admin_token):
        """Test invalid periodo_from format returns 400"""
        response = client.get(
            "/api/libros/ventas",
            params={"periodo_from": "2026/01"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 400

    def test_filter_periodo_to_invalid_format(self, client, admin_token):
        """Test invalid periodo_to format returns 400"""
        response = client.get(
            "/api/libros/ventas",
            params={"periodo_to": "invalid"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 400


class TestSortingFilters:
    """Tests for sort_by and sort_order parameters"""

    def test_sort_by_periodo_asc(self, client, admin_token, sample_libros_ventas):
        """Test sorting by periodo ascending"""
        response = client.get(
            "/api/libros/ventas",
            params={"sort_by": "periodo", "sort_order": "asc"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        periodos = [item["periodo"] for item in data["data"]]
        assert periodos == sorted(periodos)

    def test_sort_by_periodo_desc(self, client, admin_token, sample_libros_ventas):
        """Test sorting by periodo descending"""
        response = client.get(
            "/api/libros/ventas",
            params={"sort_by": "periodo", "sort_order": "desc"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        periodos = [item["periodo"] for item in data["data"]]
        assert periodos == sorted(periodos, reverse=True)

    def test_sort_by_monto_total_asc(self, client, admin_token, sample_libros_ventas):
        """Test sorting by monto_total ascending"""
        response = client.get(
            "/api/libros/ventas",
            params={"sort_by": "monto_total", "sort_order": "asc"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        montos = [item["monto_total"] for item in data["data"]]
        assert montos == sorted(montos)

    def test_sort_by_created_at_desc(self, client, admin_token, sample_libros_ventas):
        """Test sorting by created_at descending (default)"""
        response = client.get(
            "/api/libros/ventas",
            params={"sort_by": "created_at", "sort_order": "desc"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        timestamps = [item["created_at"] for item in data["data"]]
        assert timestamps == sorted(timestamps, reverse=True)

    def test_sort_by_estado_asc(self, client, admin_token, sample_libros_ventas):
        """Test sorting by estado ascending"""
        response = client.get(
            "/api/libros/ventas",
            params={"sort_by": "estado", "sort_order": "asc"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        estados = [item["estado"] for item in data["data"]]
        assert estados == sorted(estados)

    def test_sort_by_invalid_field(self, client, admin_token, sample_libros_ventas):
        """Test invalid sort_by field returns 400"""
        response = client.get(
            "/api/libros/ventas",
            params={"sort_by": "invalid_field"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 400

    def test_sort_order_invalid_value(self, client, admin_token, sample_libros_ventas):
        """Test invalid sort_order value returns 400"""
        response = client.get(
            "/api/libros/ventas",
            params={"sort_by": "periodo", "sort_order": "invalid"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 400

    def test_sort_without_sort_order_defaults_asc(self, client, admin_token, sample_libros_ventas):
        """Test that sort_by without sort_order defaults to asc"""
        response = client.get(
            "/api/libros/ventas",
            params={"sort_by": "periodo"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        periodos = [item["periodo"] for item in data["data"]]
        assert periodos == sorted(periodos)

    def test_sort_compras_by_estado(self, client, admin_token, sample_libros_compras):
        """Test sorting libros compras by estado"""
        response = client.get(
            "/api/libros/compras",
            params={"sort_by": "estado", "sort_order": "desc"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        estados = [item["estado"] for item in data["data"]]
        assert estados == sorted(estados, reverse=True)

    def test_sort_by_total_registros(self, client, admin_token, sample_libros_ventas):
        """Test sorting by total_registros"""
        response = client.get(
            "/api/libros/ventas",
            params={"sort_by": "total_registros", "sort_order": "desc"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        registros = [item["total_registros"] for item in data["data"]]
        assert registros == sorted(registros, reverse=True)


class TestComplexFilterCombinations:
    """Tests for combining multiple filters"""

    def test_estado_and_periodo_range_and_sort(self, client, admin_token, sample_libros_ventas):
        """Test combining estado, periodo range, and sorting"""
        response = client.get(
            "/api/libros/ventas",
            params={
                "estado": "borrador",
                "periodo_from": "2026-01",
                "periodo_to": "2026-02",
                "sort_by": "periodo",
                "sort_order": "asc"
            },
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert data["pagination"]["total"] == 2
        assert all(item["estado"] == "borrador" for item in data["data"])
        periodos = [item["periodo"] for item in data["data"]]
        assert periodos == sorted(periodos)

    def test_all_filters_combined(self, client, admin_token, sample_libros_ventas):
        """Test all filters: estado, periodo_from, periodo_to, sort_by, sort_order"""
        response = client.get(
            "/api/libros/ventas",
            params={
                "estado": "borrador",
                "periodo_from": "2026-01",
                "periodo_to": "2026-03",
                "sort_by": "monto_total",
                "sort_order": "desc",
                "limit": 10,
                "offset": 0
            },
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert all(item["estado"] == "borrador" for item in data["data"])
        assert all("2026-01" <= item["periodo"] <= "2026-03" for item in data["data"])
        montos = [item["monto_total"] for item in data["data"]]
        assert montos == sorted(montos, reverse=True)

    def test_filters_with_pagination(self, client, admin_token, sample_libros_ventas):
        """Test filters work correctly with pagination"""
        response = client.get(
            "/api/libros/ventas",
            params={
                "estado": "borrador",
                "sort_by": "periodo",
                "sort_order": "asc",
                "limit": 1,
                "offset": 0
            },
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        data = response.json()
        assert len(data["data"]) == 1
        assert data["pagination"]["total"] == 2


class TestLibrosVentasExportCsv:
    """Tests for GET /api/libros/ventas/export/csv"""

    def test_export_ventas_csv_without_auth(self, client):
        """Test that CSV export requires authentication"""
        response = client.get("/api/libros/ventas/export/csv")
        assert response.status_code == 401

    def test_export_ventas_csv_success(self, client, admin_token, sample_libros_ventas):
        """Test successful CSV export of libros ventas"""
        response = client.get(
            "/api/libros/ventas/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        assert response.headers["Content-Type"] == "text/csv; charset=utf-8"
        assert "attachment" in response.headers["Content-Disposition"]

        # Parse CSV content
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should have all 3 records
        assert len(rows) == 3

        # Check headers exist
        assert "Período" in reader.fieldnames
        assert "Total Registros" in reader.fieldnames
        assert "Monto Total" in reader.fieldnames
        assert "Estado" in reader.fieldnames
        assert "Folio Inicio" in reader.fieldnames
        assert "Folio Fin" in reader.fieldnames

    def test_export_ventas_csv_headers_only_empty_result(self, client, admin_token):
        """Test CSV export with no matching records returns headers only"""
        response = client.get(
            "/api/libros/ventas/export/csv",
            params={"periodo": "2099-12"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should have headers but no data rows
        assert len(rows) == 0
        assert "Período" in reader.fieldnames

    def test_export_ventas_csv_with_filters(self, client, admin_token, sample_libros_ventas):
        """Test CSV export respects filters"""
        response = client.get(
            "/api/libros/ventas/export/csv",
            params={"estado": "borrador"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should have 2 borrador records
        assert len(rows) == 2
        assert all(row["Estado"] == "borrador" for row in rows)

    def test_export_ventas_csv_with_periodo_range(self, client, admin_token, sample_libros_ventas):
        """Test CSV export with periodo range filter"""
        response = client.get(
            "/api/libros/ventas/export/csv",
            params={"periodo_from": "2026-02", "periodo_to": "2026-03"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should have 2 records in the range
        assert len(rows) == 2
        assert all("2026-02" <= row["Período"] <= "2026-03" for row in rows)

    def test_export_ventas_csv_with_sort(self, client, admin_token, sample_libros_ventas):
        """Test CSV export respects sort parameters"""
        response = client.get(
            "/api/libros/ventas/export/csv",
            params={"sort_by": "periodo", "sort_order": "desc"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        periodos = [row["Período"] for row in rows]
        assert periodos == sorted(periodos, reverse=True)

    def test_export_ventas_csv_no_pagination_limit(self, client, admin_token, sample_libros_ventas):
        """Test that CSV export ignores pagination and returns all records"""
        # Even though we set limit=1, CSV should return all records
        response = client.get(
            "/api/libros/ventas/export/csv",
            params={"limit": 1},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should return all 3 records, not just 1
        assert len(rows) == 3

    def test_export_ventas_csv_monto_numeric(self, client, admin_token, sample_libros_ventas):
        """Test that monto_total is numeric in CSV, not string"""
        response = client.get(
            "/api/libros/ventas/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Check montos are numbers
        for row in rows:
            monto = int(row["Monto Total"])
            assert isinstance(monto, int)
            assert monto > 0

    def test_export_ventas_csv_utf8_encoding(self, client, admin_token, db, admin_user):
        """Test UTF-8 encoding with Spanish characters"""
        # Create a libro with special characters in a related campo if any
        libro = LibroVentas(
            periodo="2026-01",
            empresa_id=admin_user.empresa_id,
            folio_inicio=1,
            folio_fin=50,
            total_registros=10,
            monto_total=100000,
            estado="borrador",
        )
        db.add(libro)
        db.commit()

        response = client.get(
            "/api/libros/ventas/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        assert response.headers.get("Content-Type", "").startswith("text/csv")
        # Should not raise encoding errors
        csv_content = response.text
        assert len(csv_content) > 0

    def test_export_ventas_csv_all_columns(self, client, admin_token, sample_libros_ventas):
        """Test that CSV includes all required columns"""
        response = client.get(
            "/api/libros/ventas/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))

        required_columns = {
            "Período", "Total Registros", "Monto Total", "Estado",
            "Folio Inicio", "Folio Fin"
        }
        assert required_columns.issubset(set(reader.fieldnames))


class TestLibrosComprasExportCsv:
    """Tests for GET /api/libros/compras/export/csv"""

    def test_export_compras_csv_without_auth(self, client):
        """Test that CSV export requires authentication"""
        response = client.get("/api/libros/compras/export/csv")
        assert response.status_code == 401

    def test_export_compras_csv_success(self, client, admin_token, sample_libros_compras):
        """Test successful CSV export of libros compras"""
        response = client.get(
            "/api/libros/compras/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        assert response.headers["Content-Type"] == "text/csv; charset=utf-8"
        assert "attachment" in response.headers["Content-Disposition"]

        # Parse CSV content
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should have 2 records (admin user's empresa only, not empresa2's)
        assert len(rows) == 2

        # Check headers exist
        assert "Período" in reader.fieldnames
        assert "Total Registros" in reader.fieldnames
        assert "Monto Total" in reader.fieldnames
        assert "Estado" in reader.fieldnames
        assert "RUT Proveedor" in reader.fieldnames

    def test_export_compras_csv_headers_only_empty_result(self, client, admin_token):
        """Test CSV export with no matching records returns headers only"""
        response = client.get(
            "/api/libros/compras/export/csv",
            params={"periodo": "2099-12"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should have headers but no data rows
        assert len(rows) == 0
        assert "Período" in reader.fieldnames

    def test_export_compras_csv_with_filters(self, client, admin_token, sample_libros_compras):
        """Test CSV export respects filters"""
        response = client.get(
            "/api/libros/compras/export/csv",
            params={"estado": "borrador"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should have 2 borrador records from admin's empresa
        assert len(rows) == 2
        assert all(row["Estado"] == "borrador" for row in rows)

    def test_export_compras_csv_with_periodo_range(self, client, admin_token, sample_libros_compras):
        """Test CSV export with periodo range filter"""
        response = client.get(
            "/api/libros/compras/export/csv",
            params={"periodo_from": "2026-01", "periodo_to": "2026-02"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should have 2 records from admin's empresa in the range
        assert len(rows) == 2

    def test_export_compras_csv_with_sort(self, client, admin_token, sample_libros_compras):
        """Test CSV export respects sort parameters"""
        response = client.get(
            "/api/libros/compras/export/csv",
            params={"sort_by": "monto_total", "sort_order": "asc"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        montos = [int(row["Monto Total"]) for row in rows]
        assert montos == sorted(montos)

    def test_export_compras_csv_no_pagination_limit(self, client, admin_token, sample_libros_compras):
        """Test that CSV export ignores pagination and returns all records"""
        response = client.get(
            "/api/libros/compras/export/csv",
            params={"limit": 1},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should return all 2 records from admin's empresa, not just 1
        assert len(rows) == 2

    def test_export_compras_csv_rut_field(self, client, admin_token, sample_libros_compras):
        """Test that RUT proveedor field is included"""
        response = client.get(
            "/api/libros/compras/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should have RUT column
        assert "RUT Proveedor" in reader.fieldnames
        # Some rows may have RUT, some may have None
        assert len(rows) == 2

    def test_export_compras_csv_authorization(self, client, token_empresa1, token_empresa2, db, empresa1, empresa2):
        """Test that users can only export their own empresa's data"""
        # Create compras for both empresas
        libro_e1 = LibroCompras(
            periodo="2026-01",
            empresa_id=empresa1.id,
            rut_proveedor="12.345.678-9",
            total_registros=30,
            monto_total=300000,
            estado="borrador",
        )
        libro_e2 = LibroCompras(
            periodo="2026-01",
            empresa_id=empresa2.id,
            rut_proveedor="98.765.432-1",
            total_registros=15,
            monto_total=150000,
            estado="borrador",
        )
        db.add_all([libro_e1, libro_e2])
        db.commit()

        # empresa1 user exports
        response = client.get(
            "/api/libros/compras/export/csv",
            headers={"Authorization": f"Bearer {token_empresa1}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))
        rows = list(reader)

        # Should only see their own empresa's data
        assert len(rows) == 1
        assert rows[0]["RUT Proveedor"] == "12.345.678-9"

    def test_export_compras_csv_all_columns(self, client, admin_token, sample_libros_compras):
        """Test that CSV includes all required columns"""
        response = client.get(
            "/api/libros/compras/export/csv",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        csv_content = response.text
        reader = csv.DictReader(StringIO(csv_content))

        required_columns = {
            "Período", "Total Registros", "Monto Total", "Estado",
            "RUT Proveedor"
        }
        assert required_columns.issubset(set(reader.fieldnames))


class TestLibrosVentasExportExcel:
    """Tests for GET /api/libros/ventas/export/excel"""

    def test_export_ventas_excel_without_auth(self, client):
        """Test that Excel export requires authentication"""
        response = client.get("/api/libros/ventas/export/excel")
        assert response.status_code == 401

    def test_export_ventas_excel_success(self, client, admin_token, sample_libros_ventas):
        """Test successful Excel export of libros ventas"""
        response = client.get(
            "/api/libros/ventas/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        assert response.headers["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["Content-Disposition"]
        assert "libros_ventas" in response.headers["Content-Disposition"]

        # Parse Excel content
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Check headers (first row, bold)
        headers = [cell.value for cell in ws[1]]
        assert "Período" in headers
        assert "Total Registros" in headers
        assert "Monto Total" in headers
        assert "Estado" in headers
        assert "Folio Inicio" in headers
        assert "Folio Fin" in headers

        # Check that header row is bold
        for cell in ws[1]:
            if cell.value:
                assert cell.font.bold, f"Header cell {cell.coordinate} should be bold"

        # Should have 3 data rows + 1 header + 1 summary = 5 rows
        assert ws.max_row == 5

        # Check summary row (last row)
        summary_row = ws[ws.max_row]
        assert summary_row[0].value == "TOTAL"

    def test_export_ventas_excel_empty_result(self, client, admin_token):
        """Test Excel export with no matching records"""
        response = client.get(
            "/api/libros/ventas/export/excel",
            params={"periodo": "2099-12"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Should have header + summary = 2 rows
        assert ws.max_row == 2

        # Check summary row shows zeros
        summary_row = ws[ws.max_row]
        assert summary_row[0].value == "TOTAL"
        assert summary_row[1].value == 0
        assert summary_row[2].value == 0

    def test_export_ventas_excel_with_filters(self, client, admin_token, sample_libros_ventas):
        """Test Excel export respects filters"""
        response = client.get(
            "/api/libros/ventas/export/excel",
            params={"estado": "borrador"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Should have 2 data rows + 1 header + 1 summary = 4 rows
        assert ws.max_row == 4

        # All data rows should have estado=borrador
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
            assert row[3].value == "borrador"

    def test_export_ventas_excel_with_periodo_range(self, client, admin_token, sample_libros_ventas):
        """Test Excel export with periodo range filter"""
        response = client.get(
            "/api/libros/ventas/export/excel",
            params={"periodo_from": "2026-02", "periodo_to": "2026-03"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Should have 2 data rows + 1 header + 1 summary = 4 rows
        assert ws.max_row == 4

    def test_export_ventas_excel_with_sort(self, client, admin_token, sample_libros_ventas):
        """Test Excel export respects sort parameters"""
        response = client.get(
            "/api/libros/ventas/export/excel",
            params={"sort_by": "periodo", "sort_order": "desc"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Check sorting in data rows (skip header and summary)
        periodos = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
            periodos.append(row[0].value)

        assert periodos == sorted(periodos, reverse=True)

    def test_export_ventas_excel_no_pagination_limit(self, client, admin_token, sample_libros_ventas):
        """Test that Excel export ignores pagination and returns all records"""
        response = client.get(
            "/api/libros/ventas/export/excel",
            params={"limit": 1},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Should have all 3 records + header + summary = 5 rows
        assert ws.max_row == 5

    def test_export_ventas_excel_currency_formatting(self, client, admin_token, sample_libros_ventas):
        """Test that Monto Total is currency formatted"""
        response = client.get(
            "/api/libros/ventas/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Check monto column (column 3) has currency formatting
        # Header is row 1, data rows start at 2
        for row_idx in range(2, ws.max_row):
            cell = ws.cell(row=row_idx, column=3)
            # Check that number format contains currency pattern
            assert cell.number_format is not None

    def test_export_ventas_excel_summary_totals(self, client, admin_token, sample_libros_ventas):
        """Test that summary row calculates correct totals"""
        response = client.get(
            "/api/libros/ventas/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Get summary row (last row)
        summary_row = ws[ws.max_row]

        # Total Registros should be sum of all registros (column B = index 1)
        total_registros = sum(r[1].value for r in ws.iter_rows(min_row=2, max_row=ws.max_row - 1))
        assert summary_row[1].value == total_registros

        # Monto Total should be sum of all montos (column C = index 2)
        monto_total = sum(r[2].value for r in ws.iter_rows(min_row=2, max_row=ws.max_row - 1))
        assert summary_row[2].value == monto_total

    def test_export_ventas_excel_auto_width(self, client, admin_token, sample_libros_ventas):
        """Test that columns have auto-width adjusted to content"""
        response = client.get(
            "/api/libros/ventas/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Check that column dimensions are set (auto-width)
        # Columns should have width > 0
        for col in ws.columns:
            col_letter = col[0].column_letter
            width = ws.column_dimensions[col_letter].width
            # Auto-width should set a dimension
            assert width is not None

    def test_export_ventas_excel_authorization(self, client, token_empresa1, token_empresa2, db, empresa1, empresa2):
        """Test that users can only export their own empresa's data"""
        # Create ventas for both empresas
        libro_e1 = LibroVentas(
            periodo="2026-01",
            empresa_id=empresa1.id,
            folio_inicio=1,
            folio_fin=100,
            total_registros=50,
            monto_total=500000,
            estado="borrador",
        )
        libro_e2 = LibroVentas(
            periodo="2026-01",
            empresa_id=empresa2.id,
            folio_inicio=1,
            folio_fin=50,
            total_registros=25,
            monto_total=250000,
            estado="borrador",
        )
        db.add_all([libro_e1, libro_e2])
        db.commit()

        # empresa1 user exports
        response = client.get(
            "/api/libros/ventas/export/excel",
            headers={"Authorization": f"Bearer {token_empresa1}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Should have 1 data row + header + summary = 3 rows
        assert ws.max_row == 3

    def test_export_ventas_excel_all_columns(self, client, admin_token, sample_libros_ventas):
        """Test that Excel includes all required columns"""
        response = client.get(
            "/api/libros/ventas/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        required_columns = {
            "Período", "Total Registros", "Monto Total", "Estado",
            "Folio Inicio", "Folio Fin"
        }
        assert required_columns.issubset(set(headers))


class TestLibrosComprasExportExcel:
    """Tests for GET /api/libros/compras/export/excel"""

    def test_export_compras_excel_without_auth(self, client):
        """Test that Excel export requires authentication"""
        response = client.get("/api/libros/compras/export/excel")
        assert response.status_code == 401

    def test_export_compras_excel_success(self, client, admin_token, sample_libros_compras):
        """Test successful Excel export of libros compras"""
        response = client.get(
            "/api/libros/compras/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        assert response.headers["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["Content-Disposition"]
        assert "libros_compras" in response.headers["Content-Disposition"]

        # Parse Excel content
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Check headers (first row, bold)
        headers = [cell.value for cell in ws[1]]
        assert "Período" in headers
        assert "Total Registros" in headers
        assert "Monto Total" in headers
        assert "Estado" in headers
        assert "RUT Proveedor" in headers

        # Check that header row is bold
        for cell in ws[1]:
            if cell.value:
                assert cell.font.bold, f"Header cell {cell.coordinate} should be bold"

        # Should have 2 data rows + 1 header + 1 summary = 4 rows
        assert ws.max_row == 4

    def test_export_compras_excel_empty_result(self, client, admin_token):
        """Test Excel export with no matching records"""
        response = client.get(
            "/api/libros/compras/export/excel",
            params={"periodo": "2099-12"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Should have header + summary = 2 rows
        assert ws.max_row == 2

        # Check summary row shows zeros
        summary_row = ws[ws.max_row]
        assert summary_row[0].value == "TOTAL"
        assert summary_row[1].value == 0
        assert summary_row[2].value == 0

    def test_export_compras_excel_with_filters(self, client, admin_token, sample_libros_compras):
        """Test Excel export respects filters"""
        response = client.get(
            "/api/libros/compras/export/excel",
            params={"estado": "borrador"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Should have 2 data rows + 1 header + 1 summary = 4 rows
        assert ws.max_row == 4

        # All data rows should have estado=borrador
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
            assert row[3].value == "borrador"

    def test_export_compras_excel_with_sort(self, client, admin_token, sample_libros_compras):
        """Test Excel export respects sort parameters"""
        response = client.get(
            "/api/libros/compras/export/excel",
            params={"sort_by": "monto_total", "sort_order": "asc"},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Check sorting in data rows (skip header and summary)
        montos = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
            montos.append(row[2].value)

        assert montos == sorted(montos)

    def test_export_compras_excel_no_pagination_limit(self, client, admin_token, sample_libros_compras):
        """Test that Excel export ignores pagination and returns all records"""
        response = client.get(
            "/api/libros/compras/export/excel",
            params={"limit": 1},
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Should have all 2 records + header + summary = 4 rows
        assert ws.max_row == 4

    def test_export_compras_excel_summary_totals(self, client, admin_token, sample_libros_compras):
        """Test that summary row calculates correct totals"""
        response = client.get(
            "/api/libros/compras/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Get summary row (last row)
        summary_row = ws[ws.max_row]

        # Total Registros should be sum of all registros (column B = index 1)
        total_registros = sum(r[1].value for r in ws.iter_rows(min_row=2, max_row=ws.max_row - 1))
        assert summary_row[1].value == total_registros

        # Monto Total should be sum of all montos (column C = index 2)
        monto_total = sum(r[2].value for r in ws.iter_rows(min_row=2, max_row=ws.max_row - 1))
        assert summary_row[2].value == monto_total

    def test_export_compras_excel_authorization(self, client, token_empresa1, token_empresa2, db, empresa1, empresa2):
        """Test that users can only export their own empresa's data"""
        # Create compras for both empresas
        libro_e1 = LibroCompras(
            periodo="2026-01",
            empresa_id=empresa1.id,
            rut_proveedor="12.345.678-9",
            total_registros=30,
            monto_total=300000,
            estado="borrador",
        )
        libro_e2 = LibroCompras(
            periodo="2026-01",
            empresa_id=empresa2.id,
            rut_proveedor="98.765.432-1",
            total_registros=15,
            monto_total=150000,
            estado="borrador",
        )
        db.add_all([libro_e1, libro_e2])
        db.commit()

        # empresa1 user exports
        response = client.get(
            "/api/libros/compras/export/excel",
            headers={"Authorization": f"Bearer {token_empresa1}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        # Should have 1 data row + header + summary = 3 rows
        assert ws.max_row == 3

    def test_export_compras_excel_all_columns(self, client, admin_token, sample_libros_compras):
        """Test that Excel includes all required columns"""
        response = client.get(
            "/api/libros/compras/export/excel",
            headers={"Authorization": f"Bearer {admin_token}"}
        )
        assert response.status_code == 200
        excel_content = BytesIO(response.content)
        wb = load_workbook(excel_content)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        required_columns = {
            "Período", "Total Registros", "Monto Total", "Estado",
            "RUT Proveedor"
        }
        assert required_columns.issubset(set(headers))

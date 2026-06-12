import datetime

import pytest


def _create_cliente_bulk(client, admin_token, empresa_id=None):
    payload = {"nombre": "Cliente Bulk Test", "rut": "22.222.222-2"}
    if empresa_id:
        payload["empresa_id"] = empresa_id
    r = client.post("/api/clientes/", json=payload, headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 201
    return r.json()["id"]


def _create_factura_bulk(client, admin_token, cliente_id, empresa_id, total_neto=10000,
                          fecha_vencimiento=None, fecha=None):
    payload = {
        "cliente_id": cliente_id,
        "empresa_id": empresa_id,
        "lineas": [{"orden": 0, "descripcion": "Item", "cantidad": 1, "valor_neto": total_neto}],
    }
    if fecha_vencimiento:
        payload["fecha_vencimiento"] = fecha_vencimiento
    if fecha:
        payload["fecha"] = fecha
    r = client.post("/api/facturas/", json=payload, headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 201, r.text
    return r.json()


def test_listar_empresas_sin_auth(client):
    r = client.get("/api/empresas/")
    assert r.status_code == 401


def test_crear_empresa(client, admin_token):
    r = client.post(
        "/api/empresas/",
        json={"nombre": "Empresa A", "rut": "76.123.456-0", "razon_social": "Empresa A Ltda."},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 201
    data = r.json()
    assert data["nombre"] == "Empresa A"
    assert data["rut"] == "76.123.456-0"
    assert data["razon_social"] == "Empresa A Ltda."
    assert "id" in data


def test_crear_empresa_rut_duplicado(client, admin_token):
    client.post("/api/empresas/", json={"nombre": "Emp A", "rut": "76.000.001-9"}, headers={"Authorization": f"Bearer {admin_token}"})
    r = client.post("/api/empresas/", json={"nombre": "Emp B", "rut": "76.000.001-9"}, headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 409


def test_listar_empresas(client, admin_token):
    client.post("/api/empresas/", json={"nombre": "Emp X"}, headers={"Authorization": f"Bearer {admin_token}"})
    r = client.get("/api/empresas/", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    assert len(r.json()) >= 1


def test_buscar_empresa_por_nombre(client, admin_token):
    client.post("/api/empresas/", json={"nombre": "Constructora XYZ"}, headers={"Authorization": f"Bearer {admin_token}"})
    r = client.get("/api/empresas/?q=constructora", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    assert any(e["nombre"] == "Constructora XYZ" for e in r.json())


def test_buscar_empresa_por_rut(client, admin_token):
    client.post("/api/empresas/", json={"nombre": "Emp Z", "rut": "77.777.777-7"}, headers={"Authorization": f"Bearer {admin_token}"})
    r = client.get("/api/empresas/?q=77.777.777", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    assert any(e["rut"] == "77.777.777-7" for e in r.json())


def test_obtener_empresa(client, admin_token):
    r = client.post("/api/empresas/", json={"nombre": "Emp Y"}, headers={"Authorization": f"Bearer {admin_token}"})
    eid = r.json()["id"]
    r2 = client.get(f"/api/empresas/{eid}", headers={"Authorization": f"Bearer {admin_token}"})
    assert r2.status_code == 200
    assert r2.json()["nombre"] == "Emp Y"


def test_obtener_empresa_inexistente(client, admin_token):
    r = client.get("/api/empresas/99999", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 404


def test_actualizar_empresa(client, admin_token):
    r = client.post("/api/empresas/", json={"nombre": "Antigua"}, headers={"Authorization": f"Bearer {admin_token}"})
    eid = r.json()["id"]
    r2 = client.patch(f"/api/empresas/{eid}", json={"nombre": "Nueva", "sector": "Construcción"}, headers={"Authorization": f"Bearer {admin_token}"})
    assert r2.status_code == 200
    assert r2.json()["nombre"] == "Nueva"
    assert r2.json()["sector"] == "Construcción"


def test_eliminar_empresa(client, admin_token):
    r = client.post("/api/empresas/", json={"nombre": "Para Borrar"}, headers={"Authorization": f"Bearer {admin_token}"})
    eid = r.json()["id"]
    r2 = client.delete(f"/api/empresas/{eid}", headers={"Authorization": f"Bearer {admin_token}"})
    assert r2.status_code == 204
    r3 = client.get(f"/api/empresas/{eid}", headers={"Authorization": f"Bearer {admin_token}"})
    assert r3.status_code == 404


def test_eliminar_empresa_con_clientes_falla(client, admin_token):
    emp = client.post("/api/empresas/", json={"nombre": "Emp Con Clientes"}, headers={"Authorization": f"Bearer {admin_token}"}).json()
    client.post("/api/clientes/", json={"nombre": "Cliente X", "empresa_id": emp["id"]}, headers={"Authorization": f"Bearer {admin_token}"})
    r = client.delete(f"/api/empresas/{emp['id']}", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 409
    assert "clientes asociados" in r.json()["detail"]


def test_vendedor_puede_ver_empresas(client, vendedor_token):
    r = client.get("/api/empresas/", headers={"Authorization": f"Bearer {vendedor_token}"})
    assert r.status_code == 200


def test_vendedor_no_puede_crear_empresa(client, vendedor_token):
    r = client.post("/api/empresas/", json={"nombre": "Intento"}, headers={"Authorization": f"Bearer {vendedor_token}"})
    assert r.status_code == 403


def test_actualizar_rut_duplicado(client, admin_token):
    client.post("/api/empresas/", json={"nombre": "Emp 1", "rut": "76.111.111-6"}, headers={"Authorization": f"Bearer {admin_token}"})
    r2 = client.post("/api/empresas/", json={"nombre": "Emp 2", "rut": "76.222.222-1"}, headers={"Authorization": f"Bearer {admin_token}"})
    eid = r2.json()["id"]
    r3 = client.patch(f"/api/empresas/{eid}", json={"rut": "76.111.111-6"}, headers={"Authorization": f"Bearer {admin_token}"})
    assert r3.status_code == 422  # RUT is immutable after creation


def test_exportar_excel(client, admin_token):
    r = client.get("/api/empresas/export/excel", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]


def test_deuda_bulk_lista_vacia(client, admin_token):
    r = client.get("/api/empresas/deuda-bulk", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    # admin_user fixture creates a default empresa (no facturas, no credit terms).
    # The endpoint returns all empresas; verify none have outstanding debt.
    assert all(float(item["deuda_total"]) == 0 for item in r.json())


def test_deuda_bulk_empresa_sin_facturas(client, admin_token):
    emp = client.post(
        "/api/empresas/",
        json={"nombre": "Emp Bulk", "plazo_credito": "30 Dias", "linea_credito": 5000000},
        headers={"Authorization": f"Bearer {admin_token}"},
    ).json()
    r = client.get("/api/empresas/deuda-bulk", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    items = r.json()
    item = next(i for i in items if i["empresa_id"] == emp["id"])
    assert float(item["deuda_total"]) == 0
    assert float(item["deuda_vencida"]) == 0
    assert item["plazo_credito"] == "30 Dias"
    assert float(item["linea_credito"]) == 5000000


def test_deuda_bulk_con_factura_sin_pagar(client, admin_token):
    emp = client.post(
        "/api/empresas/", json={"nombre": "Emp Deudora"}, headers={"Authorization": f"Bearer {admin_token}"}
    ).json()
    cid = _create_cliente_bulk(client, admin_token, emp["id"])
    _create_factura_bulk(client, admin_token, cid, emp["id"], total_neto=10000)
    r = client.get("/api/empresas/deuda-bulk", headers={"Authorization": f"Bearer {admin_token}"})
    item = next(i for i in r.json() if i["empresa_id"] == emp["id"])
    # 10000 neto * 1.19 IVA = 11900
    assert float(item["deuda_total"]) == pytest.approx(11900.0)
    assert float(item["deuda_vencida"]) == 0  # no fecha_vencimiento and plazo is None


def test_deuda_bulk_vencida_por_fecha_vencimiento(client, admin_token):
    emp = client.post(
        "/api/empresas/", json={"nombre": "Emp Vencida"}, headers={"Authorization": f"Bearer {admin_token}"}
    ).json()
    cid = _create_cliente_bulk(client, admin_token, emp["id"])
    past_date = (datetime.date.today() - datetime.timedelta(days=10)).isoformat()
    _create_factura_bulk(client, admin_token, cid, emp["id"], total_neto=5000, fecha_vencimiento=past_date)
    r = client.get("/api/empresas/deuda-bulk", headers={"Authorization": f"Bearer {admin_token}"})
    item = next(i for i in r.json() if i["empresa_id"] == emp["id"])
    # 5000 neto * 1.19 = 5950
    assert float(item["deuda_total"]) == pytest.approx(5950.0)
    assert float(item["deuda_vencida"]) == pytest.approx(5950.0)


def test_deuda_bulk_vencida_por_plazo(client, admin_token):
    emp = client.post(
        "/api/empresas/", json={"nombre": "Emp Plazo Vencido", "plazo_credito": "30 Dias"},
        headers={"Authorization": f"Bearer {admin_token}"}
    ).json()
    cid = _create_cliente_bulk(client, admin_token, emp["id"])
    old_date = (datetime.date.today() - datetime.timedelta(days=60)).isoformat()
    _create_factura_bulk(client, admin_token, cid, emp["id"], total_neto=3000, fecha=old_date)
    r = client.get("/api/empresas/deuda-bulk", headers={"Authorization": f"Bearer {admin_token}"})
    item = next(i for i in r.json() if i["empresa_id"] == emp["id"])
    # 60 days old, 30-day plazo → vencida
    assert float(item["deuda_vencida"]) == pytest.approx(3570.0)  # 3000 * 1.19


def test_deuda_bulk_no_vencida_si_plazo_especial(client, admin_token):
    emp = client.post(
        "/api/empresas/", json={"nombre": "Emp Especial", "plazo_credito": "Especial"},
        headers={"Authorization": f"Bearer {admin_token}"}
    ).json()
    cid = _create_cliente_bulk(client, admin_token, emp["id"])
    _create_factura_bulk(client, admin_token, cid, emp["id"], total_neto=2000)
    r = client.get("/api/empresas/deuda-bulk", headers={"Authorization": f"Bearer {admin_token}"})
    item = next(i for i in r.json() if i["empresa_id"] == emp["id"])
    assert float(item["deuda_total"]) == pytest.approx(2380.0)
    assert float(item["deuda_vencida"]) == 0  # Especial + no fecha_vencimiento → skip


def test_deuda_bulk_factura_anulada_no_cuenta(client, admin_token):
    emp = client.post(
        "/api/empresas/", json={"nombre": "Emp Anulada"}, headers={"Authorization": f"Bearer {admin_token}"}
    ).json()
    cid = _create_cliente_bulk(client, admin_token, emp["id"])
    f = _create_factura_bulk(client, admin_token, cid, emp["id"], total_neto=10000)
    # Anular la factura
    client.patch(
        f"/api/facturas/{f['id']}/estado",
        json={"estado": "anulada"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    r = client.get("/api/empresas/deuda-bulk", headers={"Authorization": f"Bearer {admin_token}"})
    item = next(i for i in r.json() if i["empresa_id"] == emp["id"])
    assert float(item["deuda_total"]) == 0


def test_deuda_bulk_sin_auth(client):
    r = client.get("/api/empresas/deuda-bulk")
    assert r.status_code == 401


# ── Export endpoints ──────────────────────────────────────────────────────────

def _setup_empresa_con_factura(client, admin_token):
    """Create empresa + cliente + factura with one line item. Returns (empresa_id, factura_id)."""
    emp = client.post(
        "/api/empresas/",
        json={"nombre": "Emp Export", "rut": "76.555.555-8"},
        headers={"Authorization": f"Bearer {admin_token}"},
    ).json()
    cid = _create_cliente_bulk(client, admin_token, emp["id"])
    f = _create_factura_bulk(client, admin_token, cid, emp["id"], total_neto=50000)
    return emp["id"], f["id"]


def test_export_facturas_xlsx_200(client, admin_token):
    eid, _ = _setup_empresa_con_factura(client, admin_token)
    r = client.get(
        f"/api/empresas/{eid}/export/facturas?format=xlsx",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")


def test_export_facturas_csv_200(client, admin_token):
    eid, _ = _setup_empresa_con_factura(client, admin_token)
    r = client.get(
        f"/api/empresas/{eid}/export/facturas?format=csv",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    assert "text/csv" in r.headers.get("content-type", "")
    lines = r.content.decode("utf-8-sig").strip().splitlines()
    assert len(lines) >= 2  # header + at least one data row


def test_export_facturas_pdf_200(client, admin_token):
    eid, _ = _setup_empresa_con_factura(client, admin_token)
    r = client.get(
        f"/api/empresas/{eid}/export/facturas?format=pdf",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    assert "pdf" in r.headers.get("content-type", "")


def test_export_facturas_columns_filter(client, admin_token):
    from io import BytesIO
    import openpyxl
    eid, _ = _setup_empresa_con_factura(client, admin_token)
    r = client.get(
        f"/api/empresas/{eid}/export/facturas?format=xlsx&columns=numero&columns=total",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert headers == ["Nº", "Total"]


def test_export_facturas_404_empresa_inexistente(client, admin_token):
    r = client.get(
        "/api/empresas/99999/export/facturas",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 404


def test_export_facturas_400_formato_invalido(client, admin_token):
    eid, _ = _setup_empresa_con_factura(client, admin_token)
    r = client.get(
        f"/api/empresas/{eid}/export/facturas?format=docx",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 400


def test_export_facturas_501_send_to(client, admin_token):
    eid, _ = _setup_empresa_con_factura(client, admin_token)
    r = client.get(
        f"/api/empresas/{eid}/export/facturas?send_to=email",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 501


def test_export_productos_xlsx_200(client, admin_token):
    eid, _ = _setup_empresa_con_factura(client, admin_token)
    r = client.get(
        f"/api/empresas/{eid}/export/productos?format=xlsx",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")


def test_export_productos_csv_200(client, admin_token):
    eid, _ = _setup_empresa_con_factura(client, admin_token)
    r = client.get(
        f"/api/empresas/{eid}/export/productos?format=csv",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    assert "text/csv" in r.headers.get("content-type", "")
    lines = r.content.decode("utf-8-sig").strip().splitlines()
    assert len(lines) >= 2


def test_export_productos_columns_filter(client, admin_token):
    from io import BytesIO
    import openpyxl
    eid, _ = _setup_empresa_con_factura(client, admin_token)
    r = client.get(
        f"/api/empresas/{eid}/export/productos?format=xlsx&columns=sku&columns=descripcion",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert headers == ["SKU", "Descripción"]


def test_export_productos_404_empresa_inexistente(client, admin_token):
    r = client.get(
        "/api/empresas/99999/export/productos",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 404


def test_export_productos_400_formato_invalido(client, admin_token):
    eid, _ = _setup_empresa_con_factura(client, admin_token)
    r = client.get(
        f"/api/empresas/{eid}/export/productos?format=docx",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 400


def test_export_productos_501_send_to(client, admin_token):
    eid, _ = _setup_empresa_con_factura(client, admin_token)
    r = client.get(
        f"/api/empresas/{eid}/export/productos?send_to=whatsapp",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 501


def test_buscar_empresa_insensible_tildes(client, admin_token):
    client.post("/api/empresas/", json={"nombre": "Manufacturas Ñoño"}, headers={"Authorization": f"Bearer {admin_token}"})
    r = client.get("/api/empresas/?q=Nono", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    assert any(e["nombre"] == "Manufacturas Ñoño" for e in r.json())


def test_buscar_empresa_acento_en_query(client, admin_token):
    client.post("/api/empresas/", json={"nombre": "Construcciones Rapidas"}, headers={"Authorization": f"Bearer {admin_token}"})
    r = client.get("/api/empresas/?q=Rápidas", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    assert any(e["nombre"] == "Construcciones Rapidas" for e in r.json())


def test_crear_empresa_rut_invalido_rechazado(client, admin_token):
    r = client.post(
        "/api/empresas/",
        json={"nombre": "Emp Inv", "rut": "12.345.678-9"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 422


# ---------------------------------------------------------------------------
# Sub-recursos ventas — GET /api/empresas/{id}/cotizaciones y /nota-ventas
# ---------------------------------------------------------------------------


def _crear_cotizacion_db(db, cliente_id, vendedor_id, empresa_id=None, numero=1,
                         fecha=None, estado="no_definido", total="1000"):
    from decimal import Decimal
    from app.models.cotizacion import Cotizacion
    cot = Cotizacion(
        numero=numero,
        cliente_id=cliente_id,
        vendedor_id=vendedor_id,
        empresa_id=empresa_id,
        fecha=fecha or datetime.date(2026, 1, 1),
        estado=estado,
        total=Decimal(total),
    )
    db.add(cot)
    db.commit()
    db.refresh(cot)
    return cot


def _crear_nv_db(db, cliente_id, empresa_id=None, numero=None, fecha=None,
                 estado="pendiente", total="1000"):
    from decimal import Decimal
    from app.models.nota_venta import NotaVenta
    nv = NotaVenta(
        numero=numero,
        cliente_id=cliente_id,
        empresa_id=empresa_id,
        fecha=fecha or datetime.date(2026, 1, 1),
        estado=estado,
        total=Decimal(total),
    )
    db.add(nv)
    db.commit()
    db.refresh(nv)
    return nv


def test_cotizaciones_empresa_filtra_y_ordena(client, admin_token, admin_user, db):
    emp_a = client.post("/api/empresas/", json={"nombre": "Emp Ventas A"}, headers={"Authorization": f"Bearer {admin_token}"}).json()
    emp_b = client.post("/api/empresas/", json={"nombre": "Emp Ventas B"}, headers={"Authorization": f"Bearer {admin_token}"}).json()
    cid = _create_cliente_bulk(client, admin_token, empresa_id=emp_a["id"])

    c1 = _crear_cotizacion_db(db, cid, admin_user.id, empresa_id=emp_a["id"], numero=1,
                              fecha=datetime.date(2026, 1, 10), total="1000")
    c2 = _crear_cotizacion_db(db, cid, admin_user.id, empresa_id=emp_a["id"], numero=2,
                              fecha=datetime.date(2026, 2, 10), total="2000")
    _crear_cotizacion_db(db, cid, admin_user.id, empresa_id=emp_b["id"], numero=3,
                         fecha=datetime.date(2026, 3, 10))

    r = client.get(f"/api/empresas/{emp_a['id']}/cotizaciones", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    data = r.json()
    assert [d["id"] for d in data] == [c2.id, c1.id]  # desc por fecha
    assert data[0]["numero"] == 2
    assert float(data[0]["total"]) == 2000.0
    assert data[0]["estado"] == "no_definido"
    assert data[0]["fecha"] == "2026-02-10"


def test_cotizaciones_empresa_filtro_estado(client, admin_token, admin_user, db):
    emp = client.post("/api/empresas/", json={"nombre": "Emp Cot Estado"}, headers={"Authorization": f"Bearer {admin_token}"}).json()
    cid = _create_cliente_bulk(client, admin_token, empresa_id=emp["id"])
    _crear_cotizacion_db(db, cid, admin_user.id, empresa_id=emp["id"], numero=1, estado="aceptada")
    _crear_cotizacion_db(db, cid, admin_user.id, empresa_id=emp["id"], numero=2, estado="rechazada")

    r = client.get(
        f"/api/empresas/{emp['id']}/cotizaciones?estado=aceptada",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    data = r.json()
    assert len(data) == 1
    assert data[0]["estado"] == "aceptada"


def test_nota_ventas_empresa_filtra_y_ordena(client, admin_token, db):
    emp_a = client.post("/api/empresas/", json={"nombre": "Emp NV A"}, headers={"Authorization": f"Bearer {admin_token}"}).json()
    emp_b = client.post("/api/empresas/", json={"nombre": "Emp NV B"}, headers={"Authorization": f"Bearer {admin_token}"}).json()
    cid = _create_cliente_bulk(client, admin_token, empresa_id=emp_a["id"])

    n1 = _crear_nv_db(db, cid, empresa_id=emp_a["id"], numero=10,
                      fecha=datetime.date(2026, 1, 5), total="500")
    n2 = _crear_nv_db(db, cid, empresa_id=emp_a["id"], numero=None,
                      fecha=datetime.date(2026, 2, 5), total="700")
    _crear_nv_db(db, cid, empresa_id=emp_b["id"], numero=30)

    r = client.get(f"/api/empresas/{emp_a['id']}/nota-ventas", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    data = r.json()
    assert [d["id"] for d in data] == [n2.id, n1.id]  # desc por fecha
    assert data[0]["numero"] is None
    assert data[1]["numero"] == 10
    assert float(data[1]["total"]) == 500.0
    assert data[0]["estado"] == "pendiente"


def test_nota_ventas_empresa_filtro_estado(client, admin_token, db):
    emp = client.post("/api/empresas/", json={"nombre": "Emp NV Estado"}, headers={"Authorization": f"Bearer {admin_token}"}).json()
    cid = _create_cliente_bulk(client, admin_token, empresa_id=emp["id"])
    _crear_nv_db(db, cid, empresa_id=emp["id"], numero=1, estado="facturada")
    _crear_nv_db(db, cid, empresa_id=emp["id"], numero=2, estado="pendiente")

    r = client.get(
        f"/api/empresas/{emp['id']}/nota-ventas?estado=facturada",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    data = r.json()
    assert len(data) == 1
    assert data[0]["estado"] == "facturada"


def test_ventas_empresa_404_si_no_existe(client, admin_token):
    r = client.get("/api/empresas/99999/cotizaciones", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 404
    r = client.get("/api/empresas/99999/nota-ventas", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 404


def test_ventas_empresa_vendedor_403_si_no_asignada(client, admin_token, vendedor_token):
    emp = client.post("/api/empresas/", json={"nombre": "Emp Ajena Ventas"}, headers={"Authorization": f"Bearer {admin_token}"}).json()
    r = client.get(f"/api/empresas/{emp['id']}/cotizaciones", headers={"Authorization": f"Bearer {vendedor_token}"})
    assert r.status_code == 403
    r = client.get(f"/api/empresas/{emp['id']}/nota-ventas", headers={"Authorization": f"Bearer {vendedor_token}"})
    assert r.status_code == 403

"""
Tests for categorias (TipoProducto) onboarding import API.

Also includes direct unit tests for CategoriasParser (TestCategoriasParserUnit).

Covers:
1. Template download returns bytes with xlsx content-type
2. Preview with valid file returns correct counts
3. Preview with missing `nombre` column returns 422 ParseError
4. Preview marks existing categories as "omitir"
5. Import creates new TipoProducto entries in DB
6. Import skips (omits) existing categories
7. Import is idempotent (re-upload same file returns cached result)
8. Admin-only: non-admin returns 403
"""

import io

import openpyxl
import pytest

from tests.conftest import TestingSession


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _xlsx(rows: list[list], headers=None) -> bytes:
    """Build a minimal .xlsx with given header + data rows."""
    if headers is None:
        headers = ["nombre"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, token, content: bytes, path: str, filename="cats.xlsx"):
    return client.post(
        path,
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )


BASE = "/api/onboarding/categorias"


# ---------------------------------------------------------------------------
# 1. Template download
# ---------------------------------------------------------------------------

def test_template_returns_xlsx(client, admin_token):
    """Template endpoint returns xlsx bytes with single-column template."""
    r = client.get(f"{BASE}/template", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert len(r.content) > 0
    # Verify it is valid xlsx
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    # Row 1 header — single column only
    assert ws.cell(row=1, column=1).value == "nombre"
    assert ws.cell(row=1, column=2).value is None


# ---------------------------------------------------------------------------
# 2. Preview with valid file returns correct counts
# ---------------------------------------------------------------------------

def test_preview_valid_file(client, admin_token):
    """Preview with new categories returns a_crear > 0, no invalids."""
    content = _xlsx([["Frutas"], ["Verduras"], ["Lácteos"]])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["filas_validas"] == 3
    assert body["filas_invalidas"] == 0
    assert body["a_crear"] == 3
    assert body["a_omitir"] == 0
    statuses = [row["status"] for row in body["rows"]]
    assert all(s == "crear" for s in statuses)


# ---------------------------------------------------------------------------
# 3. Preview with missing 'nombre' column returns 422
# ---------------------------------------------------------------------------

def test_preview_missing_nombre_column(client, admin_token):
    """Missing 'nombre' column causes 422 ParseError."""
    content = _xlsx([["Frutas"]], headers=["familia"])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 422
    assert "nombre" in r.json()["detail"].lower()


# ---------------------------------------------------------------------------
# 4. Preview marks existing categories as "omitir"
# ---------------------------------------------------------------------------

def test_preview_marks_existing_as_omitir(client, admin_token, setup_test_db):
    """Categories already in DB are marked 'omitir' during preview."""
    from app.models.tipo_producto import TipoProducto

    db = TestingSession()
    try:
        db.add(TipoProducto(nombre="Frutas"))
        db.commit()
    finally:
        db.close()

    content = _xlsx([["Frutas"], ["Verduras"]])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["a_omitir"] == 1
    assert body["a_crear"] == 1

    rows_by_nombre = {row["nombre"]: row for row in body["rows"]}
    assert rows_by_nombre["Frutas"]["status"] == "omitir"
    assert rows_by_nombre["Verduras"]["status"] == "crear"


# ---------------------------------------------------------------------------
# 5. Import creates new TipoProducto entries in DB
# ---------------------------------------------------------------------------

def test_import_creates_tipos(client, admin_token):
    """Import endpoint creates new TipoProducto rows."""
    content = _xlsx([["Bebidas"], ["Snacks"]])
    r = _upload(client, admin_token, content, f"{BASE}/import")
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "success"
    assert body["report"]["created_count"] == 2
    assert body["report"]["error_count"] == 0

    from app.models.tipo_producto import TipoProducto

    db = TestingSession()
    try:
        tipos = db.query(TipoProducto).all()
        nombres = {t.nombre for t in tipos}
        assert "Bebidas" in nombres
        assert "Snacks" in nombres
    finally:
        db.close()


# ---------------------------------------------------------------------------
# 6. Import skips (omits) existing categories
# ---------------------------------------------------------------------------

def test_import_skips_existing(client, admin_token, setup_test_db):
    """Import skips categories that already exist (case-insensitive)."""
    from app.models.tipo_producto import TipoProducto

    db = TestingSession()
    try:
        db.add(TipoProducto(nombre="Carnes"))
        db.commit()
    finally:
        db.close()

    content = _xlsx([["Carnes"], ["Pescados"]])
    r = _upload(client, admin_token, content, f"{BASE}/import")
    assert r.status_code == 200
    body = r.json()
    report = body["report"]
    assert report["created_count"] == 1
    assert report["omitted_count"] == 1
    assert report["error_count"] == 0

    db2 = TestingSession()
    try:
        count = db2.query(TipoProducto).filter(TipoProducto.nombre == "Carnes").count()
        assert count == 1  # not duplicated
    finally:
        db2.close()


# ---------------------------------------------------------------------------
# 7. Import is idempotent
# ---------------------------------------------------------------------------

def test_import_idempotent(client, admin_token, setup_test_db):
    """Same idempotency_key returns cached result without duplicate inserts."""
    content = _xlsx([["Panadería"]])
    key = "test-idem-key-001"

    r1 = client.post(
        f"{BASE}/import?idempotency_key={key}",
        files={"file": ("c.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r1.status_code == 200
    body1 = r1.json()
    assert body1["report"]["created_count"] == 1

    r2 = client.post(
        f"{BASE}/import?idempotency_key={key}",
        files={"file": ("c.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r2.status_code == 200
    body2 = r2.json()
    # Cached — same import_id
    assert body2["import_id"] == body1["import_id"]
    assert body2["report"]["created_count"] == 1

    # Only one row in DB
    from app.models.tipo_producto import TipoProducto

    db = TestingSession()
    try:
        count = db.query(TipoProducto).filter(TipoProducto.nombre == "Panadería").count()
        assert count == 1
    finally:
        db.close()


# ---------------------------------------------------------------------------
# 8. Admin-only: non-admin returns 403
# ---------------------------------------------------------------------------

def test_template_requires_admin(client, vendedor_token):
    """Non-admin cannot access template."""
    r = client.get(f"{BASE}/template", headers={"Authorization": f"Bearer {vendedor_token}"})
    assert r.status_code == 403


def test_preview_requires_admin(client, vendedor_token):
    """Non-admin cannot access preview."""
    content = _xlsx([["X"]])
    r = _upload(client, vendedor_token, content, f"{BASE}/preview")
    assert r.status_code == 403


def test_import_requires_admin(client, vendedor_token):
    """Non-admin cannot import."""
    content = _xlsx([["X"]])
    r = _upload(client, vendedor_token, content, f"{BASE}/import")
    assert r.status_code == 403


# ---------------------------------------------------------------------------
# Bonus: within-file duplicate is an error row
# ---------------------------------------------------------------------------

def test_preview_within_file_duplicate(client, admin_token):
    """Same nombre twice in file → second is invalid."""
    content = _xlsx([["Frutas"], ["Frutas"]])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["filas_validas"] == 1
    assert body["filas_invalidas"] == 1
    error_row = next(row for row in body["rows"] if row["status"] == "error")
    assert "duplicado" in error_row["errors"][0].lower()


# ---------------------------------------------------------------------------
# Bonus: empty file returns 400
# ---------------------------------------------------------------------------

def test_preview_empty_file(client, admin_token):
    """Empty file upload returns 400."""
    r = client.post(
        f"{BASE}/preview",
        files={"file": ("empty.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 400


# ---------------------------------------------------------------------------
# Direct parser unit tests
# ---------------------------------------------------------------------------

class TestCategoriasParserUnit:
    """Unit tests for CategoriasParser — no HTTP, no DB."""

    def test_valid_single_row(self, setup_test_db):
        """Single valid row → a_crear=1, a_omitir=0."""
        from app.services.categorias_parser import CategoriasParser

        content = _xlsx([["Frutas"]])
        result = CategoriasParser.parse(content, "test.xlsx")
        assert result.a_crear == 1
        assert result.a_omitir == 0
        assert len(result.valid_rows) == 1
        assert len(result.invalid_rows) == 0
        assert result.valid_rows[0].nombre == "Frutas"
        assert result.valid_rows[0].status == "crear"

    def test_within_file_duplicate(self, setup_test_db):
        """Second occurrence of same nombre → InvalidRow with 'duplicado' in motivo."""
        from app.services.categorias_parser import CategoriasParser

        content = _xlsx([["Verduras"], ["Verduras"]])
        result = CategoriasParser.parse(content, "test.xlsx")
        assert result.a_crear == 1
        assert len(result.valid_rows) == 1
        assert len(result.invalid_rows) == 1
        assert "duplicado" in result.invalid_rows[0].motivo.lower()

    def test_existing_nombres_marks_omitir(self, setup_test_db):
        """Row matching existing_nombres → status 'omitir'."""
        from app.services.categorias_parser import CategoriasParser

        existing = {"lácteos"}
        content = _xlsx([["Lácteos"], ["Bebidas"]])
        result = CategoriasParser.parse(content, "test.xlsx", existing_nombres=existing)
        assert result.a_omitir == 1
        assert result.a_crear == 1
        omit_row = next(r for r in result.valid_rows if r.status == "omitir")
        assert omit_row.nombre == "Lácteos"

"""
Tests for Tareas onboarding import API.

Covers:
1. GET /template returns xlsx with correct headers
2. POST /preview with valid xlsx returns correct counts
3. POST /preview with missing required columns returns 422
4. POST /import creates tareas in DB
5. POST /import idempotency (re-run same file → omitir)
6. POST /import with unknown asignado_email → fallback to admin
7. POST /import with unknown rut_cliente → pendiente flag in report
8. Non-admin returns 403
9. POST /import skips rows with estado=cerrado
"""

import io

import openpyxl
import pytest

from tests.conftest import TestingSession

BASE = "/api/onboarding/tareas"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

DEFAULT_HEADERS = [
    "descripcion",
    "fecha_vencimiento",
    "rut_cliente",
    "tipo",
    "asignado_email",
    "estado",
    "prioridad",
]


def _xlsx(rows: list[list], headers=None, sheet_name="Tareas") -> bytes:
    """Build a minimal .xlsx in the expected Tareas format."""
    if headers is None:
        headers = DEFAULT_HEADERS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, token, content: bytes, path: str, filename="tareas.xlsx"):
    return client.post(
        path,
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def cliente_tarea(setup_test_db):
    from app.models.cliente import Cliente
    db = TestingSession()
    c = Cliente(nombre="Cliente Tarea Test", rut="12.345.678-9")
    db.add(c)
    db.commit()
    db.refresh(c)
    db.close()
    return c


@pytest.fixture
def vendedor_user_tarea(setup_test_db):
    from app.models.user import User
    from app.core.security import get_password_hash
    db = TestingSession()
    user = User(
        email="vendedor_tarea@conico.cl",
        name="Vendedor Tarea",
        hashed_password=get_password_hash("secret123"),
        role="vendedor",
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    db.close()
    return user


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestGetTemplate:
    def test_returns_xlsx(self, client, admin_token):
        resp = client.get(f"{BASE}/template", headers={"Authorization": f"Bearer {admin_token}"})
        assert resp.status_code == 200
        ct = resp.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "application/vnd" in ct

        # Parse the returned bytes as a workbook and verify sheet + headers
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert "Tareas" in wb.sheetnames
        ws = wb["Tareas"]
        header_row = [cell.value for cell in ws[1]]
        assert "descripcion" in header_row
        assert "fecha_vencimiento" in header_row

    def test_non_admin_forbidden(self, client, vendedor_token):
        resp = client.get(f"{BASE}/template", headers={"Authorization": f"Bearer {vendedor_token}"})
        assert resp.status_code == 403


class TestPreview:
    def test_valid_file_returns_correct_counts(self, client, admin_token, cliente_tarea, vendedor_user_tarea):
        content = _xlsx([
            ["Llamar al cliente", "2024-06-15", "12.345.678-9", "llamada", "vendedor_tarea@conico.cl", "pendiente", "Alta"],
            ["Enviar cotización", "2024-06-20", None, "email", None, "pendiente", None],
        ])
        resp = _upload(client, admin_token, content, f"{BASE}/preview")
        assert resp.status_code == 200
        data = resp.json()
        assert data["a_crear"] == 2
        assert data["a_omitir"] == 0
        assert data["invalid_count"] == 0
        assert len(data["valid"]) == 2
        assert data["sin_cliente"] == 0  # 1 has known rut, 1 has no rut

    def test_missing_required_columns_returns_422(self, client, admin_token):
        content = _xlsx(
            [["Llamar al cliente", "2024-06-15"]],
            headers=["descripcion", "fecha"],  # wrong header name
        )
        resp = _upload(client, admin_token, content, f"{BASE}/preview")
        assert resp.status_code == 422
        assert "fecha_vencimiento" in resp.json()["detail"]

    def test_wrong_sheet_name_returns_422(self, client, admin_token):
        content = _xlsx(
            [["Llamar al cliente", "2024-06-15"]],
            headers=["descripcion", "fecha_vencimiento"],
            sheet_name="HojaIncorrecta",
        )
        resp = _upload(client, admin_token, content, f"{BASE}/preview")
        assert resp.status_code == 422
        assert "Tareas" in resp.json()["detail"]

    def test_unknown_asignado_email_flagged(self, client, admin_token):
        content = _xlsx([
            ["Tarea con email desconocido", "2024-07-01", None, "llamada", "desconocido@noemail.cl", "pendiente", None],
        ])
        resp = _upload(client, admin_token, content, f"{BASE}/preview")
        assert resp.status_code == 200
        data = resp.json()
        assert data["asignado_fallback"] == 1

    def test_unknown_rut_cliente_flagged(self, client, admin_token):
        content = _xlsx([
            ["Tarea cliente desconocido", "2024-07-01", "99.999.999-9", "llamada", None, "pendiente", None],
        ])
        resp = _upload(client, admin_token, content, f"{BASE}/preview")
        assert resp.status_code == 200
        data = resp.json()
        assert data["sin_cliente"] == 1

    def test_invalid_tipo_flagged_as_invalid(self, client, admin_token):
        content = _xlsx([
            ["Tarea tipo inválido", "2024-07-01", None, "telepatia", None, "pendiente", None],
        ])
        resp = _upload(client, admin_token, content, f"{BASE}/preview")
        assert resp.status_code == 200
        data = resp.json()
        assert data["invalid_count"] == 1
        assert data["a_crear"] == 0

    def test_estado_cerrado_skipped_silently(self, client, admin_token):
        content = _xlsx([
            ["Tarea pendiente", "2024-07-01", None, "llamada", None, "pendiente", None],
            ["Tarea cerrada", "2024-07-02", None, "email", None, "cerrado", None],
        ])
        resp = _upload(client, admin_token, content, f"{BASE}/preview")
        assert resp.status_code == 200
        data = resp.json()
        # cerrado row is skipped silently (not counted in valid or invalid)
        assert data["a_crear"] == 1
        assert data["invalid_count"] == 0

    def test_non_admin_forbidden(self, client, vendedor_token):
        content = _xlsx([["Tarea", "2024-07-01"]])
        resp = _upload(client, vendedor_token, content, f"{BASE}/preview")
        assert resp.status_code == 403


class TestImport:
    def test_creates_tareas_in_db(self, client, admin_token, cliente_tarea, vendedor_user_tarea, setup_test_db):
        content = _xlsx([
            ["Llamar al cliente", "2024-06-15", "12.345.678-9", "llamada", "vendedor_tarea@conico.cl", "pendiente", "Alta"],
            ["Enviar cotización", "2024-06-20", None, "email", None, "pendiente", None],
        ])
        resp = _upload(client, admin_token, content, f"{BASE}/import")
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "success"
        report = data["report"]
        assert report["created_count"] == 2
        assert report["omitted_count"] == 0
        assert report["error_count"] == 0

        # Verify DB
        db = TestingSession()
        from app.models.tarea import Tarea
        tareas = db.query(Tarea).filter(Tarea.origen == "importado").all()
        assert len(tareas) == 2
        titulos = {t.titulo for t in tareas}
        assert "[Alta] Llamar al cliente" in titulos
        assert "Enviar cotización" in titulos
        db.close()

    def test_idempotency_second_run_omits(self, client, admin_token, setup_test_db):
        content = _xlsx([
            ["Tarea idempotente", "2024-08-01", None, "email", None, "pendiente", None],
        ])
        # First import
        resp1 = _upload(client, admin_token, content, f"{BASE}/import")
        assert resp1.status_code == 200
        assert resp1.json()["report"]["created_count"] == 1

        # Second import — same content, same dedup_key
        resp2 = _upload(client, admin_token, content, f"{BASE}/import")
        assert resp2.status_code == 200
        report2 = resp2.json()["report"]
        assert report2["created_count"] == 0
        assert report2["omitted_count"] == 1

    def test_unknown_asignado_email_falls_back_to_admin(self, client, admin_token, setup_test_db):
        content = _xlsx([
            ["Tarea asignada a desconocido", "2024-09-01", None, "llamada", "nadie@nowhere.cl", "pendiente", None],
        ])
        resp = _upload(client, admin_token, content, f"{BASE}/import")
        assert resp.status_code == 200
        data = resp.json()
        assert data["report"]["created_count"] == 1

        db = TestingSession()
        from app.models.tarea import Tarea
        from app.models.user import User
        tarea = db.query(Tarea).filter(Tarea.titulo == "Tarea asignada a desconocido").first()
        assert tarea is not None
        admin = db.query(User).filter(User.role == "admin").first()
        assert tarea.asignado_id == admin.id
        db.close()

    def test_unknown_rut_cliente_creates_with_no_cliente(self, client, admin_token, setup_test_db):
        content = _xlsx([
            ["Tarea cliente desconocido", "2024-09-15", "00.000.000-0", "visita", None, "pendiente", None],
        ])
        resp = _upload(client, admin_token, content, f"{BASE}/import")
        assert resp.status_code == 200
        data = resp.json()
        assert data["report"]["created_count"] == 1
        row = data["report"]["rows"][0]
        assert row["import_status"] == "creado_sin_cliente"

        db = TestingSession()
        from app.models.tarea import Tarea
        tarea = db.query(Tarea).filter(Tarea.titulo == "Tarea cliente desconocido").first()
        assert tarea is not None
        assert tarea.cliente_id is None
        db.close()

    def test_import_idempotency_key_returns_cached_result(self, client, admin_token, setup_test_db):
        content = _xlsx([
            ["Tarea con clave", "2024-10-01", None, "cobranza", None, "pendiente", None],
        ])
        idem_key = "test-idempotency-key-tareas-001"

        resp1 = client.post(
            f"{BASE}/import?idempotency_key={idem_key}",
            files={"file": ("tareas.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert resp1.status_code == 200
        assert resp1.json()["import_id"] == idem_key

        # Second call with same key should return cached report
        resp2 = client.post(
            f"{BASE}/import?idempotency_key={idem_key}",
            files={"file": ("tareas.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert resp2.status_code == 200
        assert resp2.json()["import_id"] == idem_key
        # Same report returned from DB cache
        assert resp2.json()["report"]["created_count"] == resp1.json()["report"]["created_count"]

    def test_estado_cerrado_skipped_not_imported(self, client, admin_token, setup_test_db):
        content = _xlsx([
            ["Tarea pendiente ok", "2024-11-01", None, "llamada", None, "pendiente", None],
            ["Tarea cerrada skip", "2024-11-02", None, "email", None, "cerrado", None],
        ])
        resp = _upload(client, admin_token, content, f"{BASE}/import")
        assert resp.status_code == 200
        report = resp.json()["report"]
        assert report["created_count"] == 1
        assert report["error_count"] == 0

        db = TestingSession()
        from app.models.tarea import Tarea
        count = db.query(Tarea).filter(Tarea.origen == "importado").count()
        assert count == 1
        db.close()

    def test_prioridad_prefixed_in_titulo(self, client, admin_token, setup_test_db):
        content = _xlsx([
            ["Visitar cliente urgente", "2024-12-01", None, "visita", None, "pendiente", "Alta"],
        ])
        resp = _upload(client, admin_token, content, f"{BASE}/import")
        assert resp.status_code == 200

        db = TestingSession()
        from app.models.tarea import Tarea
        tarea = db.query(Tarea).filter(Tarea.titulo.like("[Alta]%")).first()
        assert tarea is not None
        assert tarea.titulo == "[Alta] Visitar cliente urgente"
        db.close()

    def test_non_admin_forbidden(self, client, vendedor_token, setup_test_db):
        content = _xlsx([["Tarea", "2024-07-01"]])
        resp = _upload(client, vendedor_token, content, f"{BASE}/import")
        assert resp.status_code == 403

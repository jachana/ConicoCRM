"""
Tests for NC/ND Históricas onboarding import API.

Covers:
1. Template download returns xlsx with correct headers
2. Preview with valid file returns correct rows
3. Preview with invalid file (missing required col) returns 422
4. Preview marks existing folios as "omitir"
5. Preview sets pendiente_ref=True when folio_referencia not found
6. Import creates NC and ND records in DB
7. Import omits records that already exist (idempotency on data)
8. Import idempotency (same import_id returns cached result)
9. Non-admin returns 403
10. ND without derivable cliente_id becomes an error
"""

import io
from datetime import date
from decimal import Decimal

import openpyxl
import pytest

from tests.conftest import TestingSession

BASE = "/api/onboarding/nc-nd-historicas"
SHEET_NAME = "NC-ND Históricas"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _xlsx(rows: list[list], headers=None, sheet_name=SHEET_NAME) -> bytes:
    """Build a minimal .xlsx with the required sheet name, header + data rows."""
    if headers is None:
        headers = [
            "tipo", "folio", "fecha", "motivo",
            "neto", "iva", "total",
            "folio_referencia", "tipo_referencia",
        ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, token, content: bytes, path: str, filename="nc_nd.xlsx"):
    return client.post(
        path,
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )


def _nc_row(
    tipo="NC",
    folio=1001,
    fecha="2024-01-15",
    motivo="Corrección precio",
    neto=100000,
    iva=19000,
    total=119000,
    folio_referencia="",
    tipo_referencia="",
):
    return [tipo, folio, fecha, motivo, neto, iva, total, folio_referencia, tipo_referencia]


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def cliente_nc(setup_test_db):
    from app.models.cliente import Cliente
    db = TestingSession()
    c = Cliente(nombre="Cliente NC Test", rut="12.345.678-9")
    db.add(c)
    db.commit()
    db.refresh(c)
    db.close()
    return c


@pytest.fixture
def factura_ref(setup_test_db, cliente_nc):
    from app.models.factura import Factura
    db = TestingSession()
    f = Factura(
        numero=500,
        cliente_id=cliente_nc.id,
        fecha=date(2024, 1, 1),
        total_neto=Decimal("100000"),
        total_iva=Decimal("19000"),
        total=Decimal("119000"),
        estado="emitida",
    )
    db.add(f)
    db.commit()
    db.refresh(f)
    db.close()
    return f


@pytest.fixture
def boleta_ref(setup_test_db, cliente_nc):
    from app.models.boleta import Boleta
    db = TestingSession()
    b = Boleta(
        numero=600,
        tipo_dte="39",
        cliente_id=cliente_nc.id,
        fecha=date(2024, 1, 10),
        total_neto=Decimal("50000"),
        total_iva=Decimal("9500"),
        total=Decimal("59500"),
    )
    db.add(b)
    db.commit()
    db.refresh(b)
    db.close()
    return b


# ---------------------------------------------------------------------------
# 1. Template download returns xlsx with correct headers
# ---------------------------------------------------------------------------

def test_template_returns_xlsx(client, admin_token):
    """Template endpoint returns xlsx bytes with correct column headers."""
    r = client.get(f"{BASE}/template", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert len(r.content) > 0

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb[SHEET_NAME]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
    assert "tipo" in headers
    assert "folio" in headers
    assert "fecha" in headers
    assert "motivo" in headers
    assert "neto" in headers
    assert "iva" in headers
    assert "total" in headers


# ---------------------------------------------------------------------------
# 2. Preview with valid file returns correct rows
# ---------------------------------------------------------------------------

def test_preview_valid_nc_file(client, admin_token):
    """Preview with a valid NC row returns 1 valid row, a_crear=1."""
    row = _nc_row()
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["a_crear"] == 1
    assert body["a_omitir"] == 0
    assert body["invalid_count"] == 0
    assert len(body["valid"]) == 1
    v = body["valid"][0]
    assert v["tipo"] == "NC"
    assert v["folio"] == 1001
    assert v["status"] == "crear"
    assert v["pendiente_ref"] is False


def test_preview_valid_nd_file(client, admin_token):
    """Preview with a valid ND row returns 1 valid row, a_crear=1."""
    row = _nc_row(tipo="ND", folio=2001)
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["a_crear"] == 1
    assert len(body["valid"]) == 1
    assert body["valid"][0]["tipo"] == "ND"


# ---------------------------------------------------------------------------
# 3. Preview with invalid file (missing required col) returns 422
# ---------------------------------------------------------------------------

def test_preview_missing_required_column(client, admin_token):
    """Missing required column causes 422 ParseError."""
    content = _xlsx(
        [["NC", 1001, "2024-01-15", "Motivo"]],
        headers=["tipo", "folio", "fecha", "motivo"],  # missing neto, iva, total
    )
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 422
    detail = r.json()["detail"].lower()
    assert "neto" in detail or "iva" in detail or "total" in detail


def test_preview_wrong_sheet_name(client, admin_token):
    """File with wrong sheet name causes 422 ParseError."""
    content = _xlsx([_nc_row()], sheet_name="Hoja1")
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 422


def test_preview_empty_file(client, admin_token):
    """Empty file upload returns 400."""
    r = client.post(
        f"{BASE}/preview",
        files={"file": ("empty.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 400


# ---------------------------------------------------------------------------
# 4. Preview marks existing folios as "omitir"
# ---------------------------------------------------------------------------

def test_preview_marks_existing_nc_as_omitir(client, admin_token, setup_test_db):
    """NC with folio already in DB is marked 'omitir'."""
    from app.models.nota_credito import NotaCredito
    db = TestingSession()
    nc = NotaCredito(
        numero=999,
        fecha=date(2023, 1, 1),
        razon="Existente",
        monto_neto=Decimal("10000"),
        monto_iva=Decimal("1900"),
        monto_total=Decimal("11900"),
        historico=True,
    )
    db.add(nc)
    db.commit()
    db.close()

    row_existing = _nc_row(folio=999, motivo="Existente")
    row_new = _nc_row(folio=1000, motivo="Nueva")
    content = _xlsx([row_existing, row_new])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["a_omitir"] == 1
    assert body["a_crear"] == 1
    statuses = {v["folio"]: v["status"] for v in body["valid"]}
    assert statuses[999] == "omitir"
    assert statuses[1000] == "crear"


# ---------------------------------------------------------------------------
# 5. Preview sets pendiente_ref=True when folio_referencia not found
# ---------------------------------------------------------------------------

def test_preview_pendiente_ref_when_folio_referencia_not_found(client, admin_token):
    """NC with folio_referencia that doesn't exist in facturas/boletas gets pendiente_ref=True."""
    row = _nc_row(folio_referencia=9999, tipo_referencia=33)
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert len(body["valid"]) == 1
    assert body["valid"][0]["pendiente_ref"] is True
    assert body["pendiente_ref"] == 1


def test_preview_no_pendiente_ref_when_folio_referencia_found(client, admin_token, factura_ref):
    """NC with folio_referencia matching an existing factura gets pendiente_ref=False."""
    row = _nc_row(folio_referencia=500, tipo_referencia=33)
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["valid"][0]["pendiente_ref"] is False
    assert body["pendiente_ref"] == 0


# ---------------------------------------------------------------------------
# 6. Import creates NC and ND records in DB
# ---------------------------------------------------------------------------

def test_import_creates_nc(client, admin_token):
    """Import creates NotaCredito + NotaCreditoLinea in DB (no client required)."""
    row = _nc_row(folio=3001)
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/import")
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "success"
    assert body["report"]["created_count"] == 1
    assert body["report"]["error_count"] == 0

    from app.models.nota_credito import NotaCredito, NotaCreditoLinea
    db = TestingSession()
    try:
        nc = db.query(NotaCredito).filter(NotaCredito.numero == 3001).first()
        assert nc is not None
        assert nc.razon == "Corrección precio"
        assert nc.historico is True
        assert nc.dte_estado == "no_emitida"
        lineas = db.query(NotaCreditoLinea).filter(NotaCreditoLinea.nota_credito_id == nc.id).all()
        assert len(lineas) == 1
        assert lineas[0].orden == 1
        assert lineas[0].cantidad == Decimal("1")
    finally:
        db.close()


def test_import_creates_nd_with_factura_reference(client, admin_token, factura_ref, cliente_nc):
    """Import creates NotaDebito linked to a factura's cliente_id."""
    row = _nc_row(tipo="ND", folio=4001, folio_referencia=500, tipo_referencia=33)
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/import")
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "success"
    assert body["report"]["created_count"] == 1

    from app.models.nota_debito import NotaDebito, NotaDebitoLinea
    db = TestingSession()
    try:
        nd = db.query(NotaDebito).filter(NotaDebito.numero == 4001).first()
        assert nd is not None
        assert nd.cliente_id == cliente_nc.id
        assert nd.historico is True
        lineas = db.query(NotaDebitoLinea).filter(NotaDebitoLinea.nota_debito_id == nd.id).all()
        assert len(lineas) == 1
    finally:
        db.close()


def test_import_creates_nd_with_boleta_reference(client, admin_token, boleta_ref, cliente_nc):
    """Import creates NotaDebito linked to a boleta's cliente_id."""
    row = _nc_row(tipo="ND", folio=4002, folio_referencia=600, tipo_referencia=39)
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/import")
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "success"
    assert body["report"]["created_count"] == 1

    from app.models.nota_debito import NotaDebito
    db = TestingSession()
    try:
        nd = db.query(NotaDebito).filter(NotaDebito.numero == 4002).first()
        assert nd is not None
        assert nd.cliente_id == cliente_nc.id
    finally:
        db.close()


def test_import_nc_with_boleta_reference_sets_boleta_id(client, admin_token, boleta_ref, cliente_nc):
    """Import NC with tipo_referencia=39 sets boleta_id on NotaCredito."""
    row = _nc_row(tipo="NC", folio=5001, folio_referencia=600, tipo_referencia=39)
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/import")
    assert r.status_code == 200

    from app.models.nota_credito import NotaCredito
    db = TestingSession()
    try:
        nc = db.query(NotaCredito).filter(NotaCredito.numero == 5001).first()
        assert nc is not None
        assert nc.boleta_id == boleta_ref.id
        assert nc.cliente_id == cliente_nc.id
    finally:
        db.close()


# ---------------------------------------------------------------------------
# 7. Import omits records that already exist
# ---------------------------------------------------------------------------

def test_import_omits_existing_nc(client, admin_token, setup_test_db):
    """Import skips NC whose folio already exists in DB."""
    from app.models.nota_credito import NotaCredito
    db = TestingSession()
    nc = NotaCredito(
        numero=8001,
        fecha=date(2023, 6, 1),
        razon="Ya existe",
        monto_neto=Decimal("10000"),
        monto_iva=Decimal("1900"),
        monto_total=Decimal("11900"),
        historico=True,
    )
    db.add(nc)
    db.commit()
    db.close()

    row_existing = _nc_row(folio=8001, motivo="Ya existe")
    row_new = _nc_row(folio=8002, motivo="Nueva")
    content = _xlsx([row_existing, row_new])
    r = _upload(client, admin_token, content, f"{BASE}/import")
    assert r.status_code == 200
    body = r.json()
    report = body["report"]
    assert report["created_count"] == 1
    assert report["omitted_count"] == 1
    assert report["error_count"] == 0

    db2 = TestingSession()
    try:
        count = db2.query(NotaCredito).filter(NotaCredito.numero == 8001).count()
        assert count == 1  # not duplicated
        count_new = db2.query(NotaCredito).filter(NotaCredito.numero == 8002).count()
        assert count_new == 1
    finally:
        db2.close()


# ---------------------------------------------------------------------------
# 8. Import idempotency (same import_id returns cached result)
# ---------------------------------------------------------------------------

def test_import_idempotent(client, admin_token):
    """Same idempotency_key returns cached result without duplicate inserts."""
    row = _nc_row(folio=9001)
    content = _xlsx([row])
    key = "test-nc-idem-key-001"

    r1 = client.post(
        f"{BASE}/import?idempotency_key={key}",
        files={"file": ("nc.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r1.status_code == 200
    body1 = r1.json()
    assert body1["report"]["created_count"] == 1

    r2 = client.post(
        f"{BASE}/import?idempotency_key={key}",
        files={"file": ("nc.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r2.status_code == 200
    body2 = r2.json()
    assert body2["import_id"] == body1["import_id"]
    assert body2["report"]["created_count"] == 1

    from app.models.nota_credito import NotaCredito
    db = TestingSession()
    try:
        count = db.query(NotaCredito).filter(NotaCredito.numero == 9001).count()
        assert count == 1
    finally:
        db.close()


# ---------------------------------------------------------------------------
# 9. Non-admin returns 403
# ---------------------------------------------------------------------------

def test_template_requires_admin(client, vendedor_token):
    """Non-admin cannot access template."""
    r = client.get(f"{BASE}/template", headers={"Authorization": f"Bearer {vendedor_token}"})
    assert r.status_code == 403


def test_preview_requires_admin(client, vendedor_token):
    """Non-admin cannot access preview."""
    content = _xlsx([_nc_row()])
    r = _upload(client, vendedor_token, content, f"{BASE}/preview")
    assert r.status_code == 403


def test_import_requires_admin(client, vendedor_token):
    """Non-admin cannot import."""
    content = _xlsx([_nc_row()])
    r = _upload(client, vendedor_token, content, f"{BASE}/import")
    assert r.status_code == 403


# ---------------------------------------------------------------------------
# 10. ND without derivable cliente_id becomes an error
# ---------------------------------------------------------------------------

def test_import_nd_without_cliente_id_becomes_error(client, admin_token):
    """ND with no folio_referencia (can't derive cliente_id) is an error."""
    row = _nc_row(tipo="ND", folio=7001)  # no folio_referencia
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/import")
    assert r.status_code == 200
    body = r.json()
    report = body["report"]
    assert report["error_count"] == 1
    assert report["created_count"] == 0

    from app.models.nota_debito import NotaDebito
    db = TestingSession()
    try:
        count = db.query(NotaDebito).filter(NotaDebito.numero == 7001).count()
        assert count == 0
    finally:
        db.close()


def test_import_nd_with_nonexistent_folio_referencia_is_error(client, admin_token):
    """ND with folio_referencia not in DB (can't derive cliente_id) is an error."""
    row = _nc_row(tipo="ND", folio=7002, folio_referencia=99999, tipo_referencia=33)
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/import")
    assert r.status_code == 200
    body = r.json()
    report = body["report"]
    assert report["error_count"] == 1
    assert report["created_count"] == 0

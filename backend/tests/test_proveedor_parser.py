import io

import openpyxl
import pytest

from app.services.proveedor_parser import (
    ALL_COLUMNS,
    ParseError,
    _normalizar_rut,
    _validar_rut_modulo11,
    build_template_xlsx,
    parse_proveedores_xlsx,
)

# Verified valid RUTs (módulo 11):
VALID_RUTS = ["76.123.456-0", "11.111.111-1", "12.345.678-5", "16.578.619-K"]


def _make_xlsx(rows: list[list], header: tuple = ALL_COLUMNS) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(header))
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_normalizar_rut_strips_dots_and_spaces():
    assert _normalizar_rut("76.123.456-0") == "76123456-0"
    assert _normalizar_rut(" 76 . 123.456-K ") == "76123456-K"
    assert _normalizar_rut("") == ""


def test_validar_rut_modulo11_accepts_valid():
    for r in VALID_RUTS:
        assert _validar_rut_modulo11(_normalizar_rut(r)), f"Expected valid: {r}"


def test_validar_rut_modulo11_rejects_invalid():
    bad = ["76.123.456-7", "11.111.111-2", "abc", "", "1", "76.123.456", "76.123.456-Z"]
    for r in bad:
        assert not _validar_rut_modulo11(_normalizar_rut(r)), f"Expected invalid: {r}"


def test_parse_template_roundtrip_yields_one_valid_row():
    bytes_ = build_template_xlsx()
    res = parse_proveedores_xlsx(bytes_)
    assert len(res.validas) == 1
    assert res.invalidas == []
    p = res.validas[0]
    assert p.rut_normalizado == "76123456-0"
    assert p.razon_social == "Sociedad Ejemplo Ltda."


def test_parse_missing_required_column_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["rut", "giro"])  # missing razon_social
    ws.append(["76.123.456-0", "Comercio"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ParseError) as ei:
        parse_proveedores_xlsx(buf.getvalue())
    assert "razon_social" in str(ei.value)


def test_parse_empty_workbook_raises():
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ParseError):
        parse_proveedores_xlsx(buf.getvalue())


def test_parse_invalid_rut_marks_row_invalid():
    rows = [
        ["76.123.456-0", "Bueno", "", "", "", "", "", "", ""],
        ["76.123.456-7", "Malo", "", "", "", "", "", "", ""],
    ]
    res = parse_proveedores_xlsx(_make_xlsx(rows))
    assert len(res.validas) == 1
    assert len(res.invalidas) == 1
    assert "RUT inválido" in res.invalidas[0].motivo
    assert res.invalidas[0].fila == 3


def test_parse_empty_rut_marks_row_invalid():
    rows = [
        ["", "Sin RUT", "", "", "", "", "", "", ""],
    ]
    res = parse_proveedores_xlsx(_make_xlsx(rows))
    assert len(res.validas) == 0
    assert len(res.invalidas) == 1
    assert "RUT vacío" in res.invalidas[0].motivo


def test_parse_empty_razon_social_marks_row_invalid():
    rows = [
        ["76.123.456-0", "", "", "", "", "", "", "", ""],
    ]
    res = parse_proveedores_xlsx(_make_xlsx(rows))
    assert len(res.invalidas) == 1
    assert "Razón social" in res.invalidas[0].motivo


def test_parse_duplicate_rut_in_file():
    rows = [
        ["76.123.456-0", "A", "", "", "", "", "", "", ""],
        ["76.123.456-0", "B", "", "", "", "", "", "", ""],
    ]
    res = parse_proveedores_xlsx(_make_xlsx(rows))
    assert len(res.validas) == 1
    assert len(res.invalidas) == 1
    assert "duplicado" in res.invalidas[0].motivo.lower()


def test_parse_normalizes_dotted_rut():
    rows = [["76.123.456-0", "X", "", "", "", "", "", "", ""]]
    res = parse_proveedores_xlsx(_make_xlsx(rows))
    assert res.validas[0].rut_normalizado == "76123456-0"


def test_parse_skips_blank_rows():
    rows = [
        ["76.123.456-0", "A", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["11.111.111-1", "B", "", "", "", "", "", "", ""],
    ]
    res = parse_proveedores_xlsx(_make_xlsx(rows))
    assert len(res.validas) == 2
    assert len(res.invalidas) == 0


def test_parse_optional_columns_can_be_missing():
    # only required columns present
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["rut", "razon_social"])
    ws.append(["76.123.456-0", "Solo lo mínimo"])
    buf = io.BytesIO()
    wb.save(buf)
    res = parse_proveedores_xlsx(buf.getvalue())
    assert len(res.validas) == 1
    assert res.validas[0].giro is None
    assert res.validas[0].condicion_pago is None

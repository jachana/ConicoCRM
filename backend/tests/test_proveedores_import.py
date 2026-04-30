import io

import openpyxl

from app.services.proveedor_parser import ALL_COLUMNS


def _xlsx(rows: list[list], header=ALL_COLUMNS) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(header))
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, token, content: bytes, path: str = "/api/proveedores/import"):
    return client.post(
        path,
        files={"archivo": ("p.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )


def test_template_descarga_admin(client, admin_token):
    r = client.get("/api/proveedores/import/template", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert len(r.content) > 0


def test_template_requiere_admin(client, vendedor_token):
    r = client.get("/api/proveedores/import/template", headers={"Authorization": f"Bearer {vendedor_token}"})
    assert r.status_code == 403


def test_template_requiere_auth(client):
    r = client.get("/api/proveedores/import/template")
    assert r.status_code == 401


def test_preview_valido(client, admin_token):
    content = _xlsx([
        ["76.123.456-0", "Empresa A", "Giro A", "Av 1", "Stgo", "Juan", "j@a.cl", "+56 9 1", "30d"],
        ["11.111.111-1", "Empresa B", "", "", "", "", "", "", ""],
    ])
    r = _upload(client, admin_token, content, "/api/proveedores/import/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["filas_validas"] == 2
    assert body["filas_invalidas"] == 0
    assert body["a_crear"] == 2
    assert body["a_actualizar"] == 0


def test_preview_mixto(client, admin_token):
    content = _xlsx([
        ["76.123.456-0", "Buena", "", "", "", "", "", "", ""],
        ["76.123.456-7", "RUT mal DV", "", "", "", "", "", "", ""],
        ["", "Sin RUT", "", "", "", "", "", "", ""],
    ])
    r = _upload(client, admin_token, content, "/api/proveedores/import/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["filas_validas"] == 1
    assert body["filas_invalidas"] == 2


def test_preview_sin_columnas_requeridas(client, admin_token):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["rut", "giro"])
    ws.append(["76.123.456-0", "x"])
    buf = io.BytesIO(); wb.save(buf)
    r = _upload(client, admin_token, buf.getvalue(), "/api/proveedores/import/preview")
    assert r.status_code == 400
    assert "razon_social" in r.json()["detail"]


def test_preview_marca_actualizar_si_existe(client, admin_token):
    # crea uno primero por POST normal
    client.post(
        "/api/proveedores/",
        json={"nombre": "Antiguo", "rut": "76123456-0"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    content = _xlsx([
        ["76.123.456-0", "Nuevo nombre", "", "", "", "", "", "", ""],
        ["11.111.111-1", "Nuevo", "", "", "", "", "", "", ""],
    ])
    r = _upload(client, admin_token, content, "/api/proveedores/import/preview")
    body = r.json()
    assert body["a_crear"] == 1
    assert body["a_actualizar"] == 1


def test_import_crea_proveedores(client, admin_token):
    content = _xlsx([
        ["76.123.456-0", "Empresa A", "Comercio", "Av 1", "Stgo", "Juan", "j@a.cl", "+56 9 1", "30d"],
        ["11.111.111-1", "Empresa B", "", "", "", "", "", "", ""],
    ])
    r = _upload(client, admin_token, content)
    assert r.status_code == 200
    body = r.json()
    assert body["creadas"] == 2
    assert body["actualizadas"] == 0
    assert body["errores"] == 0
    assert len(body["detalles"]) == 2

    listado = client.get("/api/proveedores/", headers={"Authorization": f"Bearer {admin_token}"}).json()
    by_rut = {p["rut"]: p for p in listado}
    assert by_rut["76123456-0"]["razon_social"] == "Empresa A"
    assert by_rut["76123456-0"]["giro"] == "Comercio"
    assert by_rut["76123456-0"]["condicion_pago"] == "30d"


def test_import_es_idempotente(client, admin_token):
    content = _xlsx([
        ["76.123.456-0", "Empresa A", "", "", "", "", "", "", ""],
    ])
    r1 = _upload(client, admin_token, content)
    assert r1.json()["creadas"] == 1
    r2 = _upload(client, admin_token, content)
    body2 = r2.json()
    assert body2["creadas"] == 0
    assert body2["sin_cambio"] == 1


def test_import_actualiza_campos_modificados(client, admin_token):
    content_v1 = _xlsx([["76.123.456-0", "Empresa A", "Giro V1", "", "", "", "", "", "30d"]])
    _upload(client, admin_token, content_v1)
    content_v2 = _xlsx([["76.123.456-0", "Empresa A", "Giro V2", "", "", "", "", "", "60d"]])
    r2 = _upload(client, admin_token, content_v2)
    body = r2.json()
    assert body["actualizadas"] == 1
    assert body["creadas"] == 0
    listado = client.get("/api/proveedores/", headers={"Authorization": f"Bearer {admin_token}"}).json()
    p = next(x for x in listado if x["rut"] == "76123456-0")
    assert p["giro"] == "Giro V2"
    assert p["condicion_pago"] == "60d"


def test_import_reporta_filas_invalidas_sin_abortar_validas(client, admin_token):
    content = _xlsx([
        ["76.123.456-0", "Buena", "", "", "", "", "", "", ""],
        ["76.123.456-7", "Mala DV", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],  # blank — skipped
        ["11.111.111-1", "Otra buena", "", "", "", "", "", "", ""],
    ])
    r = _upload(client, admin_token, content)
    body = r.json()
    assert body["creadas"] == 2
    assert body["errores"] == 1
    estados = {d["estado"] for d in body["detalles"]}
    assert "creada" in estados and "error" in estados


def test_import_requiere_admin(client, vendedor_token):
    content = _xlsx([["76.123.456-0", "X", "", "", "", "", "", "", ""]])
    r = _upload(client, vendedor_token, content)
    assert r.status_code == 403


def test_import_sin_auth(client):
    content = _xlsx([["76.123.456-0", "X", "", "", "", "", "", "", ""]])
    r = client.post(
        "/api/proveedores/import",
        files={"archivo": ("p.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 401


def test_import_columnas_faltantes(client, admin_token):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["rut"])
    ws.append(["76.123.456-0"])
    buf = io.BytesIO(); wb.save(buf)
    r = _upload(client, admin_token, buf.getvalue())
    assert r.status_code == 400


def test_reporte_descargable(client, admin_token):
    content = _xlsx([["76.123.456-0", "X", "", "", "", "", "", "", ""]])
    importado = _upload(client, admin_token, content).json()
    r = client.post(
        "/api/proveedores/import/report",
        json=importado,
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert len(r.content) > 0

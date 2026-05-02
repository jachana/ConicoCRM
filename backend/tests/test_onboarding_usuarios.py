"""
Tests for onboarding usuarios import API and parser.

Covers:
1. Template download returns xlsx bytes
2. Preview validates required columns
3. Preview returns valid/invalid row counts
4. Import creates new users with temp passwords in result
5. Import updates existing users (no temp_password in result)
6. Idempotency: re-running same file returns cached result
7. Admin-only: returns 403 if not admin
"""

import io
import json
import uuid

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.main import app
from app.models.user import User
from app.services.usuarios_parser import UsuariosParser, ParseError, ParseResult


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_xlsx(rows: list[list], add_doc_row: bool = True) -> bytes:
    """Build a minimal xlsx from a header row + optional doc row + data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # rows[0] is assumed to be the header
    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    if add_doc_row and len(rows) == 1:
        # Insert documentation row between header and (empty) data
        ws.insert_rows(2)
        ws.cell(row=2, column=1, value="docs")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def _usuarios_xlsx(data_rows: list[list]) -> bytes:
    """
    Build an xlsx with proper header (row 1), doc row (row 2), then data rows.
    data_rows: each is [email, nombre, rol, rut, activo]
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ["email", "nombre", "rol", "rut", "activo"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    # Documentation row (row 2 — skipped by parser)
    ws.cell(row=2, column=1, value="docs row")

    for r_offset, row in enumerate(data_rows, 3):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_offset, column=c_idx, value=val)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Unit tests — Parser
# ---------------------------------------------------------------------------

class TestUsuariosParserUnit:
    """Direct parser tests (no HTTP)."""

    def test_generate_template_returns_bytes(self):
        tmpl = UsuariosParser.generate_template()
        assert isinstance(tmpl, bytes)
        assert len(tmpl) > 0

    def test_generate_template_is_valid_xlsx(self):
        tmpl = UsuariosParser.generate_template()
        wb = openpyxl.load_workbook(io.BytesIO(tmpl))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert "email" in headers
        assert "nombre" in headers
        assert "rol" in headers

    def test_parse_valid_rows(self):
        content = _usuarios_xlsx([
            ["juan@test.cl", "Juan", "vendedor", "12345678-9", "true"],
            ["maria@test.cl", "María", "admin", None, None],
        ])
        result = UsuariosParser.parse(content, "test.xlsx")
        assert result.valid_count == 2
        assert result.invalid_count == 0

    def test_parse_normalizes_email_to_lowercase(self):
        content = _usuarios_xlsx([
            ["Juan@Test.CL", "Juan", "vendedor", None, None],
        ])
        result = UsuariosParser.parse(content, "test.xlsx")
        assert result.valid_count == 1
        assert result.valid_rows[0].email == "juan@test.cl"

    def test_parse_normalizes_rol_to_lowercase(self):
        content = _usuarios_xlsx([
            ["a@b.cl", "Ana", "ADMIN", None, None],
        ])
        result = UsuariosParser.parse(content, "test.xlsx")
        assert result.valid_count == 1
        assert result.valid_rows[0].rol == "admin"

    def test_parse_invalid_rol(self):
        content = _usuarios_xlsx([
            ["a@b.cl", "Ana", "gerente", None, None],
        ])
        result = UsuariosParser.parse(content, "test.xlsx")
        assert result.valid_count == 0
        assert result.invalid_count == 1
        assert any("rol" in e.lower() for e in result.invalid_rows[0].errors)

    def test_parse_missing_email(self):
        content = _usuarios_xlsx([
            [None, "Sin Email", "vendedor", None, None],
        ])
        result = UsuariosParser.parse(content, "test.xlsx")
        assert result.invalid_count == 1

    def test_parse_invalid_email_format(self):
        content = _usuarios_xlsx([
            ["notemail", "Name", "vendedor", None, None],
        ])
        result = UsuariosParser.parse(content, "test.xlsx")
        assert result.invalid_count == 1

    def test_parse_missing_nombre(self):
        content = _usuarios_xlsx([
            ["a@b.cl", None, "vendedor", None, None],
        ])
        result = UsuariosParser.parse(content, "test.xlsx")
        assert result.invalid_count == 1

    def test_parse_activo_defaults_to_true(self):
        content = _usuarios_xlsx([
            ["a@b.cl", "Ana", "vendedor", None, None],
        ])
        result = UsuariosParser.parse(content, "test.xlsx")
        assert result.valid_rows[0].activo is True

    def test_parse_activo_false_values(self):
        for val in ["false", "0", "no"]:
            content = _usuarios_xlsx([
                ["a@b.cl", "Ana", "vendedor", None, val],
            ])
            result = UsuariosParser.parse(content, "test.xlsx")
            assert result.valid_rows[0].activo is False, f"Expected False for '{val}'"

    def test_parse_a_crear_a_actualizar_counts(self):
        content = _usuarios_xlsx([
            ["nuevo@test.cl", "Nuevo", "vendedor", None, None],
            ["existente@test.cl", "Existente", "admin", None, None],
        ])
        result = UsuariosParser.parse(
            content, "test.xlsx", existing_emails={"existente@test.cl"}
        )
        assert result.a_crear == 1
        assert result.a_actualizar == 1

    def test_parse_missing_required_column_raises(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Only email + nombre, missing rol
        ws.cell(row=1, column=1, value="email")
        ws.cell(row=1, column=2, value="nombre")
        ws.cell(row=2, column=1, value="docs")
        ws.cell(row=3, column=1, value="a@b.cl")
        ws.cell(row=3, column=2, value="A")
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        with pytest.raises(ParseError, match="rol"):
            UsuariosParser.parse(out.getvalue(), "test.xlsx")

    def test_parse_non_xlsx_raises(self):
        with pytest.raises(ParseError, match="xlsx"):
            UsuariosParser.parse(b"not excel", "test.csv")

    def test_parse_empty_file_raises(self):
        wb = openpyxl.Workbook()
        wb.active.title = "empty"
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        with pytest.raises(ParseError):
            UsuariosParser.parse(out.getvalue(), "empty.xlsx")

    def test_hash_key_is_deterministic(self):
        content = _usuarios_xlsx([
            ["a@b.cl", "Ana", "vendedor", None, None],
        ])
        r1 = UsuariosParser.parse(content, "test.xlsx")
        r2 = UsuariosParser.parse(content, "test.xlsx")
        assert r1.valid_rows[0].hash_key == r2.valid_rows[0].hash_key

    def test_hash_key_differs_for_different_email(self):
        c1 = _usuarios_xlsx([["a@b.cl", "Ana", "vendedor", None, None]])
        c2 = _usuarios_xlsx([["x@b.cl", "Ana", "vendedor", None, None]])
        r1 = UsuariosParser.parse(c1, "test.xlsx")
        r2 = UsuariosParser.parse(c2, "test.xlsx")
        assert r1.valid_rows[0].hash_key != r2.valid_rows[0].hash_key


# ---------------------------------------------------------------------------
# Integration tests — HTTP endpoints
# ---------------------------------------------------------------------------

class TestTemplateEndpoint:
    """GET /api/onboarding/usuarios/template"""

    def test_returns_xlsx(self, client, admin_token):
        resp = client.get(
            "/api/onboarding/usuarios/template",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")

    def test_template_contains_expected_headers(self, client, admin_token):
        resp = client.get(
            "/api/onboarding/usuarios/template",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        row1 = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        for col in ["email", "nombre", "rol"]:
            assert col in row1, f"Missing column: {col}"

    def test_admin_only_403_for_vendedor(self, client, vendedor_token):
        resp = client.get(
            "/api/onboarding/usuarios/template",
            headers={"Authorization": f"Bearer {vendedor_token}"},
        )
        assert resp.status_code == 403


class TestPreviewEndpoint:
    """POST /api/onboarding/usuarios/preview"""

    def test_preview_valid_file(self, client, admin_token):
        content = _usuarios_xlsx([
            ["juan@test.cl", "Juan", "vendedor", None, None],
        ])
        resp = client.post(
            "/api/onboarding/usuarios/preview",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["filas_validas"] == 1
        assert data["filas_invalidas"] == 0
        assert data["total_filas"] == 1

    def test_preview_detects_invalid_rows(self, client, admin_token):
        content = _usuarios_xlsx([
            ["invalid-email", "Juan", "vendedor", None, None],
            ["ok@test.cl", "Maria", "admin", None, None],
        ])
        resp = client.post(
            "/api/onboarding/usuarios/preview",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["filas_invalidas"] == 1
        assert data["filas_validas"] == 1

    def test_preview_rejects_missing_required_column(self, client, admin_token):
        # File with only email + nombre, missing rol
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="email")
        ws.cell(row=1, column=2, value="nombre")
        ws.cell(row=2, column=1, value="docs")
        ws.cell(row=3, column=1, value="a@b.cl")
        ws.cell(row=3, column=2, value="A")
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        resp = client.post(
            "/api/onboarding/usuarios/preview",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"file": ("usuarios.xlsx", out.getvalue(), "application/octet-stream")},
        )
        assert resp.status_code == 422

    def test_preview_shows_a_crear_a_actualizar(self, client, admin_token, setup_test_db):
        from tests.conftest import TestingSession
        from app.core.security import get_password_hash as _hash

        # Pre-insert a user
        db = TestingSession()
        existing = User(
            email="existing@test.cl",
            name="Existing",
            hashed_password=_hash("x"),
            role="vendedor",
        )
        db.add(existing)
        db.commit()
        db.close()

        content = _usuarios_xlsx([
            ["existing@test.cl", "Existing Updated", "admin", None, None],
            ["new@test.cl", "New User", "vendedor", None, None],
        ])
        resp = client.post(
            "/api/onboarding/usuarios/preview",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["a_actualizar"] == 1
        assert data["a_crear"] == 1

    def test_preview_admin_only_403(self, client, vendedor_token):
        content = _usuarios_xlsx([["a@b.cl", "A", "vendedor", None, None]])
        resp = client.post(
            "/api/onboarding/usuarios/preview",
            headers={"Authorization": f"Bearer {vendedor_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 403


class TestImportEndpoint:
    """POST /api/onboarding/usuarios/import"""

    def test_import_creates_new_users(self, client, admin_token, setup_test_db):
        from tests.conftest import TestingSession

        content = _usuarios_xlsx([
            ["newuser@test.cl", "New User", "vendedor", "12345678-9", "true"],
        ])
        resp = client.post(
            "/api/onboarding/usuarios/import",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "success"
        assert data["report"]["created_count"] == 1
        assert data["report"]["updated_count"] == 0

        # Verify user was actually created
        db = TestingSession()
        user = db.query(User).filter(User.email == "newuser@test.cl").first()
        db.close()
        assert user is not None
        assert user.role == "vendedor"

    def test_import_new_user_has_temp_password_in_result(self, client, admin_token):
        content = _usuarios_xlsx([
            ["temppass@test.cl", "Temp User", "subadmin", None, None],
        ])
        resp = client.post(
            "/api/onboarding/usuarios/import",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 200
        rows = resp.json()["report"]["rows"]
        created_rows = [r for r in rows if r["status"] == "created"]
        assert len(created_rows) == 1
        assert created_rows[0]["temp_password"] is not None
        assert len(created_rows[0]["temp_password"]) > 0

    def test_import_updates_existing_user(self, client, admin_token, setup_test_db):
        from tests.conftest import TestingSession
        from app.core.security import get_password_hash as _hash

        # Pre-insert a user
        db = TestingSession()
        original_hash = _hash("originalpassword")
        existing = User(
            email="update@test.cl",
            name="Old Name",
            hashed_password=original_hash,
            role="vendedor",
            is_active=True,
        )
        db.add(existing)
        db.commit()
        db.refresh(existing)
        original_id = existing.id
        db.close()

        content = _usuarios_xlsx([
            ["update@test.cl", "New Name", "admin", None, "false"],
        ])
        resp = client.post(
            "/api/onboarding/usuarios/import",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["report"]["updated_count"] == 1
        assert data["report"]["created_count"] == 0

        # Verify updates
        db = TestingSession()
        user = db.query(User).filter(User.id == original_id).first()
        assert user.name == "New Name"
        assert user.role == "admin"
        assert user.is_active is False
        # Password must not change
        assert user.hashed_password == original_hash
        db.close()

    def test_import_updated_user_has_no_temp_password(self, client, admin_token, setup_test_db):
        from tests.conftest import TestingSession
        from app.core.security import get_password_hash as _hash

        db = TestingSession()
        existing = User(
            email="nopw@test.cl",
            name="Name",
            hashed_password=_hash("x"),
            role="vendedor",
        )
        db.add(existing)
        db.commit()
        db.close()

        content = _usuarios_xlsx([["nopw@test.cl", "Name Updated", "subadmin", None, None]])
        resp = client.post(
            "/api/onboarding/usuarios/import",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 200
        rows = resp.json()["report"]["rows"]
        updated_rows = [r for r in rows if r["status"] == "updated"]
        assert len(updated_rows) == 1
        assert updated_rows[0]["temp_password"] is None

    def test_import_idempotency_returns_cached(self, client, admin_token):
        content = _usuarios_xlsx([
            ["idempotent@test.cl", "Idempotent User", "vendedor", None, None],
        ])
        key = str(uuid.uuid4())

        resp1 = client.post(
            f"/api/onboarding/usuarios/import?idempotency_key={key}",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp1.status_code == 200
        import_id_1 = resp1.json()["import_id"]

        # Second call with same key — returns cached
        resp2 = client.post(
            f"/api/onboarding/usuarios/import?idempotency_key={key}",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp2.status_code == 200
        import_id_2 = resp2.json()["import_id"]

        assert import_id_1 == import_id_2

    def test_import_admin_only_403_vendedor(self, client, vendedor_token):
        content = _usuarios_xlsx([["a@b.cl", "A", "vendedor", None, None]])
        resp = client.post(
            "/api/onboarding/usuarios/import",
            headers={"Authorization": f"Bearer {vendedor_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 403

    def test_import_partial_on_invalid_rows(self, client, admin_token):
        content = _usuarios_xlsx([
            ["valid@test.cl", "Valid User", "vendedor", None, None],
            ["bad-email", "Bad User", "vendedor", None, None],  # invalid
        ])
        resp = client.post(
            "/api/onboarding/usuarios/import",
            headers={"Authorization": f"Bearer {admin_token}"},
            files={"file": ("usuarios.xlsx", content, "application/octet-stream")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "partial"
        assert data["report"]["created_count"] == 1
        assert data["report"]["error_count"] == 1

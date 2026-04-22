"""Tests for flat-sheet export endpoints."""
from io import BytesIO
import openpyxl


def _setup_cotizacion(client, admin_token):
    r = client.post("/api/clientes/", json={"nombre": "CLI Export", "rut": "12345678-9"},
                    headers={"Authorization": f"Bearer {admin_token}"})
    cli_id = r.json()["id"]
    r1 = client.post("/api/productos/", json={"nombre": "Tornillo M5", "sku": "T-M5", "precio_venta": 100, "precio_costo": 80},
                     headers={"Authorization": f"Bearer {admin_token}"})
    prod1_id = r1.json()["id"]
    r2 = client.post("/api/productos/", json={"nombre": "Tuerca M5", "sku": "TU-M5", "precio_venta": 50, "precio_costo": 40},
                     headers={"Authorization": f"Bearer {admin_token}"})
    prod2_id = r2.json()["id"]
    r = client.post("/api/cotizaciones/", json={
        "cliente_id": cli_id,
        "lineas": [
            {"orden": 0, "descripcion": "Tornillo M5", "sku": "T-M5",
             "cantidad": 10, "valor_neto": 100, "producto_id": prod1_id},
            {"orden": 1, "descripcion": "Tuerca M5", "sku": "TU-M5",
             "cantidad": 20, "valor_neto": 50, "producto_id": prod2_id},
        ],
    }, headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 201, r.text
    return r.json()


def test_cotizaciones_export_returns_xlsx(client, admin_token):
    _setup_cotizacion(client, admin_token)
    r = client.get("/api/cotizaciones/export/excel",
                   headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    ctype = r.headers.get("content-type", "")
    assert "spreadsheetml" in ctype or "openxmlformats" in ctype


def test_cotizaciones_export_single_flat_sheet(client, admin_token):
    _setup_cotizacion(client, admin_token)
    r = client.get("/api/cotizaciones/export/excel",
                   headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    assert len(wb.sheetnames) == 1
    assert wb.sheetnames[0] == "Cotizaciones"
    ws = wb.active
    # One header row + two line rows for our cotizacion
    assert ws.max_row >= 3


def test_cotizaciones_export_columns_param(client, admin_token):
    _setup_cotizacion(client, admin_token)
    r = client.get(
        "/api/cotizaciones/export/excel?columns=numero&columns=cliente_nombre&columns=descripcion",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    assert headers == ["Nº COT", "Cliente", "Descripción"]


def test_cotizaciones_export_default_columns_when_none_specified(client, admin_token):
    _setup_cotizacion(client, admin_token)
    r = client.get("/api/cotizaciones/export/excel",
                   headers={"Authorization": f"Bearer {admin_token}"})
    wb = openpyxl.load_workbook(BytesIO(r.content))
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    # Default columns must include these
    assert "Nº COT" in headers
    assert "Cliente" in headers
    assert "Total Neto" in headers

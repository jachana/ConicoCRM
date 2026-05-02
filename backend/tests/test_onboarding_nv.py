"""
Tests for NV (Nota de Venta) onboarding import API.

Covers:
1. Template download returns xlsx with correct headers
2. Preview with valid file returns correct NV groups
3. Preview with missing required column returns 422
4. Preview groups rows with same numero_nv
5. Preview marks NVs with existing numero_nv as "omitir"
6. Import creates NotaVenta + NotaVentaLinea in DB
7. Import omits NVs that already exist
8. Import idempotency (same import_id returns cached result)
9. Non-admin returns 403
"""

import io

import openpyxl
import pytest

from tests.conftest import TestingSession

BASE = "/api/onboarding/nv-abiertas"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _xlsx(rows: list[list], headers=None) -> bytes:
    """Build a minimal .xlsx with given header + data rows."""
    if headers is None:
        headers = [
            "numero_nv", "rut_cliente", "rut_empresa", "fecha", "estado",
            "vendedor_email", "nota", "numero_oc_cliente", "sku",
            "descripcion", "formato", "cantidad", "valor_neto_unitario",
        ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, token, content: bytes, path: str, filename="nvs.xlsx"):
    return client.post(
        path,
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )


# ---------------------------------------------------------------------------
# Fixtures: create minimal required data in DB
# ---------------------------------------------------------------------------

@pytest.fixture
def cliente_nv(setup_test_db):
    from app.models.cliente import Cliente
    db = TestingSession()
    c = Cliente(nombre="Cliente NV Test", rut="12.345.678-9")
    db.add(c)
    db.commit()
    db.refresh(c)
    db.close()
    return c


@pytest.fixture
def empresa_nv(setup_test_db):
    from app.models.empresa import Empresa
    db = TestingSession()
    e = Empresa(nombre="Empresa NV Test", rut="76.543.210-K")
    db.add(e)
    db.commit()
    db.refresh(e)
    db.close()
    return e


@pytest.fixture
def producto_nv(setup_test_db):
    from app.models.producto import Producto
    db = TestingSession()
    p = Producto(nombre="Producto Test", sku="SKU-001")
    db.add(p)
    db.commit()
    db.refresh(p)
    db.close()
    return p


def _make_row(
    rut_cliente="12.345.678-9",
    fecha="2024-01-15",
    descripcion="Producto de prueba",
    cantidad=2,
    valor_neto_unitario=10000,
    numero_nv="NV-001",
    rut_empresa="",
    estado="pendiente",
    vendedor_email="",
    nota="",
    numero_oc_cliente="",
    sku="",
    formato="",
):
    return [
        numero_nv, rut_cliente, rut_empresa, fecha, estado,
        vendedor_email, nota, numero_oc_cliente, sku,
        descripcion, formato, cantidad, valor_neto_unitario,
    ]


# ---------------------------------------------------------------------------
# 1. Template download returns xlsx with correct headers
# ---------------------------------------------------------------------------

def test_template_returns_xlsx(client, admin_token):
    """Template endpoint returns xlsx bytes with correct column headers."""
    r = client.get(f"{BASE}/template", headers={"Authorization": f"Bearer {admin_token}"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert len(r.content) > 0

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    # Verify required headers in row 1
    headers = [ws.cell(row=1, column=c).value for c in range(1, 14)]
    assert "rut_cliente" in headers
    assert "descripcion" in headers
    assert "cantidad" in headers
    assert "valor_neto_unitario" in headers
    assert "fecha" in headers


# ---------------------------------------------------------------------------
# 2. Preview with valid file returns correct NV groups
# ---------------------------------------------------------------------------

def test_preview_valid_file(client, admin_token, cliente_nv):
    """Preview with a valid single-row NV returns 1 valid group."""
    row = _make_row()
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["nvs_validas"] == 1
    assert body["nvs_invalidas"] == 0
    assert body["a_crear"] == 1
    assert body["a_omitir"] == 0
    assert len(body["nvs"]) == 1
    nv = body["nvs"][0]
    assert nv["rut_cliente"] == "12.345.678-9"
    assert nv["status"] == "crear"
    assert len(nv["lineas"]) == 1
    assert nv["lineas"][0]["descripcion"] == "Producto de prueba"
    assert nv["lineas"][0]["cantidad"] == 2
    assert nv["total_neto"] == pytest.approx(20000)
    assert nv["total_iva"] == pytest.approx(3800)
    assert nv["total"] == pytest.approx(23800)


# ---------------------------------------------------------------------------
# 3. Preview with missing required column returns 422
# ---------------------------------------------------------------------------

def test_preview_missing_required_column(client, admin_token):
    """Missing required column causes 422 ParseError."""
    content = _xlsx([["Producto A", 2, 10000]], headers=["descripcion", "cantidad", "valor_neto_unitario"])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 422
    detail = r.json()["detail"].lower()
    # Missing rut_cliente and/or fecha
    assert "rut_cliente" in detail or "fecha" in detail


# ---------------------------------------------------------------------------
# 4. Preview groups rows with same numero_nv
# ---------------------------------------------------------------------------

def test_preview_groups_rows_by_numero_nv(client, admin_token, cliente_nv):
    """Two rows with same numero_nv are grouped into one NV with 2 lineas."""
    row1 = _make_row(numero_nv="NV-100", descripcion="Linea 1", cantidad=1, valor_neto_unitario=5000)
    row2 = _make_row(numero_nv="NV-100", descripcion="Linea 2", cantidad=3, valor_neto_unitario=2000)
    content = _xlsx([row1, row2])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["nvs_validas"] == 1  # 2 rows → 1 NV group
    assert len(body["nvs"]) == 1
    nv = body["nvs"][0]
    assert len(nv["lineas"]) == 2
    # total_neto = 1*5000 + 3*2000 = 11000
    assert nv["total_neto"] == pytest.approx(11000)


# ---------------------------------------------------------------------------
# 5. Preview marks NVs with existing numero_nv as "omitir"
# ---------------------------------------------------------------------------

def test_preview_marks_existing_numero_nv_as_omitir(client, admin_token, cliente_nv, setup_test_db):
    """NV with integer numero_nv already in DB is marked 'omitir'."""
    from app.models.nota_venta import NotaVenta
    from decimal import Decimal
    from datetime import date

    db = TestingSession()
    existing_nv = NotaVenta(
        numero=999,
        cliente_id=cliente_nv.id,
        fecha=date(2023, 1, 1),
        estado="pendiente",
        total_neto=Decimal("0"),
        total_iva=Decimal("0"),
        total=Decimal("0"),
    )
    db.add(existing_nv)
    db.commit()
    db.close()

    row_existing = _make_row(numero_nv="999", descripcion="NV existente")
    row_new = _make_row(numero_nv="1000", descripcion="NV nueva")
    content = _xlsx([row_existing, row_new])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["a_omitir"] == 1
    assert body["a_crear"] == 1
    statuses = {nv["numero_nv"]: nv["status"] for nv in body["nvs"]}
    assert statuses["999"] == "omitir"
    assert statuses["1000"] == "crear"


# ---------------------------------------------------------------------------
# 6. Import creates NotaVenta + NotaVentaLinea in DB
# ---------------------------------------------------------------------------

def test_import_creates_nv_and_lineas(client, admin_token, cliente_nv, producto_nv):
    """Import creates NotaVenta and NotaVentaLinea records in DB."""
    row = _make_row(sku="SKU-001", descripcion="Producto Test", cantidad=3, valor_neto_unitario=10000)
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/import")
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "success"
    assert body["report"]["created_count"] == 1
    assert body["report"]["error_count"] == 0

    from app.models.nota_venta import NotaVenta, NotaVentaLinea
    from decimal import Decimal

    db = TestingSession()
    try:
        nvs = db.query(NotaVenta).filter(NotaVenta.cliente_id == cliente_nv.id).all()
        assert len(nvs) == 1
        nv = nvs[0]
        assert nv.total_neto == Decimal("30000")
        lineas = db.query(NotaVentaLinea).filter(NotaVentaLinea.nv_id == nv.id).all()
        assert len(lineas) == 1
        assert lineas[0].descripcion == "Producto Test"
        assert lineas[0].cantidad == 3
        assert lineas[0].sku == "SKU-001"
    finally:
        db.close()


# ---------------------------------------------------------------------------
# 7. Import omits NVs that already exist
# ---------------------------------------------------------------------------

def test_import_omits_existing_nv(client, admin_token, cliente_nv, setup_test_db):
    """Import skips NVs whose integer numero already exists in DB."""
    from app.models.nota_venta import NotaVenta
    from decimal import Decimal
    from datetime import date

    db = TestingSession()
    existing_nv = NotaVenta(
        numero=500,
        cliente_id=cliente_nv.id,
        fecha=date(2023, 6, 1),
        estado="pendiente",
        total_neto=Decimal("0"),
        total_iva=Decimal("0"),
        total=Decimal("0"),
    )
    db.add(existing_nv)
    db.commit()
    db.close()

    row_existing = _make_row(numero_nv="500", descripcion="Ya existe")
    row_new = _make_row(numero_nv="501", descripcion="Nueva NV")
    content = _xlsx([row_existing, row_new])
    r = _upload(client, admin_token, content, f"{BASE}/import")
    assert r.status_code == 200
    body = r.json()
    report = body["report"]
    assert report["created_count"] == 1
    assert report["omitted_count"] == 1
    assert report["error_count"] == 0

    db2 = TestingSession()
    try:
        count = db2.query(NotaVenta).filter(NotaVenta.numero == 500).count()
        assert count == 1  # not duplicated
        count_new = db2.query(NotaVenta).filter(NotaVenta.numero == 501).count()
        assert count_new == 1
    finally:
        db2.close()


# ---------------------------------------------------------------------------
# 8. Import idempotency (same import_id returns cached result)
# ---------------------------------------------------------------------------

def test_import_idempotent(client, admin_token, cliente_nv, setup_test_db):
    """Same idempotency_key returns cached result without duplicate inserts."""
    row = _make_row(numero_nv="777", descripcion="Idempotent NV")
    content = _xlsx([row])
    key = "test-nv-idem-key-001"

    r1 = client.post(
        f"{BASE}/import?idempotency_key={key}",
        files={"file": ("nv.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r1.status_code == 200
    body1 = r1.json()
    assert body1["report"]["created_count"] == 1

    r2 = client.post(
        f"{BASE}/import?idempotency_key={key}",
        files={"file": ("nv.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r2.status_code == 200
    body2 = r2.json()
    assert body2["import_id"] == body1["import_id"]
    assert body2["report"]["created_count"] == 1

    # Only one NV in DB with numero=777
    from app.models.nota_venta import NotaVenta
    db = TestingSession()
    try:
        count = db.query(NotaVenta).filter(NotaVenta.numero == 777).count()
        assert count == 1
    finally:
        db.close()


# ---------------------------------------------------------------------------
# 9. Non-admin returns 403
# ---------------------------------------------------------------------------

def test_template_requires_admin(client, vendedor_token):
    """Non-admin cannot access template."""
    r = client.get(f"{BASE}/template", headers={"Authorization": f"Bearer {vendedor_token}"})
    assert r.status_code == 403


def test_preview_requires_admin(client, vendedor_token):
    """Non-admin cannot access preview."""
    content = _xlsx([_make_row()])
    r = _upload(client, vendedor_token, content, f"{BASE}/preview")
    assert r.status_code == 403


def test_import_requires_admin(client, vendedor_token):
    """Non-admin cannot import."""
    content = _xlsx([_make_row()])
    r = _upload(client, vendedor_token, content, f"{BASE}/import")
    assert r.status_code == 403


# ---------------------------------------------------------------------------
# Extra: Invalid rows (missing rut_cliente) are captured
# ---------------------------------------------------------------------------

def test_preview_invalid_row_missing_rut_cliente(client, admin_token):
    """Row missing rut_cliente is captured as invalid."""
    row = _make_row(rut_cliente="")
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["nvs_invalidas"] == 1
    assert body["nvs_validas"] == 0
    assert len(body["invalid_rows"]) == 1
    assert "rut_cliente" in body["invalid_rows"][0]["motivo"].lower()


# ---------------------------------------------------------------------------
# Extra: Empty file returns 400
# ---------------------------------------------------------------------------

def test_preview_empty_file(client, admin_token):
    """Empty file upload returns 400."""
    r = client.post(
        f"{BASE}/preview",
        files={"file": ("empty.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 400


# ---------------------------------------------------------------------------
# Extra: Non-integer numero_nv is always "crear" regardless of any match
# ---------------------------------------------------------------------------

def test_preview_non_integer_numero_nv_always_crear(client, admin_token, cliente_nv):
    """Non-integer numero_nv like 'NV-001' is always status=crear (no omit check)."""
    row = _make_row(numero_nv="NV-001")
    content = _xlsx([row])
    r = _upload(client, admin_token, content, f"{BASE}/preview")
    assert r.status_code == 200
    body = r.json()
    assert body["a_crear"] == 1
    assert body["a_omitir"] == 0

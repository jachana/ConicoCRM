"""
Onboarding Precios Especiales Import API

Endpoints for bulk-importing customer/empresa-specific pricing during onboarding.
Idempotent by (rut_entidad, sku): re-import updates vigencia/precio of existing record.

Endpoints:
  GET  /template  - Download Excel template
  POST /preview   - Validate and preview file (no DB writes)
  POST /import    - Execute transactional import with idempotency
"""

import io
import json
import uuid
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.config import require_admin
from app.models.cliente import Cliente
from app.models.empresa import Empresa
from app.models.import_report import ImportReport
from app.models.precio_especial_cliente import PrecioEspecialCliente
from app.models.producto import Producto
from app.models.user import User

router = APIRouter()


# ============================================================================
# DB lookup helpers
# ============================================================================

def _get_clientes_by_rut(db: Session) -> dict[str, int]:
    rows = db.query(Cliente.rut, Cliente.id).filter(Cliente.rut.isnot(None)).all()
    return {r[0]: r[1] for r in rows if r[0]}


def _get_empresas_by_rut(db: Session) -> dict[str, int]:
    rows = db.query(Empresa.rut, Empresa.id).filter(Empresa.rut.isnot(None)).all()
    return {r[0]: r[1] for r in rows if r[0]}


def _get_productos_by_sku(db: Session) -> dict[str, int]:
    rows = db.query(Producto.sku, Producto.id).filter(Producto.sku.isnot(None)).all()
    return {r[0]: r[1] for r in rows if r[0]}


def _get_existing_precios(db: Session) -> set[tuple[str, str]]:
    """Return set of (rut_entidad, sku) pairs that already exist."""
    rows = db.query(PrecioEspecialCliente.rut_entidad, PrecioEspecialCliente.sku).all()
    return {(r[0], r[1]) for r in rows}


# ============================================================================
# Parse dataclass
# ============================================================================

@dataclass
class PrecioRow:
    row_num: int
    rut_entidad: str
    sku: str
    precio_especial: Optional[Decimal]
    descuento_pct: Optional[Decimal]
    vigencia_desde: Optional[date]
    vigencia_hasta: Optional[date]
    status: str  # "crear" | "actualizar" | "pendiente" | "error"
    motivo: Optional[str]
    cliente_id: Optional[int]
    empresa_id: Optional[int]
    producto_id: Optional[int]


# ============================================================================
# Parse helpers
# ============================================================================

def _parse_date_cell(value) -> Optional[date]:
    """Parse a cell value that may be a date object or YYYY-MM-DD string."""
    if value is None:
        return None
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def _parse_decimal_cell(value) -> Optional[Decimal]:
    """Parse a cell value to Decimal, or None if empty."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _str_cell(value) -> str:
    """Convert cell value to stripped string."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_rows(
    content: bytes,
    clientes_by_rut: dict[str, int],
    empresas_by_rut: dict[str, int],
    productos_by_sku: dict[str, int],
    existing_precios: set[tuple[str, str]],
) -> tuple[list[PrecioRow], list[dict]]:
    """
    Parse xlsx bytes. Returns (valid_rows, invalid_rows).
    valid_rows includes crear/actualizar/pendiente.
    invalid_rows is a list of dicts with row_num, rut_raw, sku_raw, motivo.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    valid_rows: list[PrecioRow] = []
    invalid_rows: list[dict] = []

    rows_iter = iter(ws.iter_rows(values_only=True))
    # Skip header row
    try:
        next(rows_iter)
    except StopIteration:
        return valid_rows, invalid_rows

    for row_idx, row in enumerate(rows_iter, start=2):
        # Columns: rut_cliente_o_empresa, sku, precio_especial, descuento_pct, vigencia_desde, vigencia_hasta
        rut_raw = _str_cell(row[0] if len(row) > 0 else None)
        sku_raw = _str_cell(row[1] if len(row) > 1 else None)

        # Skip blank rows
        if not rut_raw and not sku_raw:
            continue

        precio_raw = row[2] if len(row) > 2 else None
        desc_raw = row[3] if len(row) > 3 else None
        desde_raw = row[4] if len(row) > 4 else None
        hasta_raw = row[5] if len(row) > 5 else None

        def _invalid(motivo: str) -> None:
            invalid_rows.append({
                "row_num": row_idx,
                "rut_raw": rut_raw,
                "sku_raw": sku_raw,
                "motivo": motivo,
            })

        # Required fields
        if not rut_raw:
            _invalid("rut_cliente_o_empresa es requerido")
            continue
        if not sku_raw:
            _invalid("sku es requerido")
            continue

        precio_especial = _parse_decimal_cell(precio_raw)
        descuento_pct = _parse_decimal_cell(desc_raw)

        # XOR: at least one must be set
        if precio_especial is None and descuento_pct is None:
            _invalid("Se requiere precio_especial o descuento_pct (al menos uno)")
            continue

        # descuento_pct range check
        if descuento_pct is not None and not (Decimal("0") <= descuento_pct <= Decimal("100")):
            _invalid(f"descuento_pct debe estar entre 0 y 100 (recibido: {descuento_pct})")
            continue

        vigencia_desde = _parse_date_cell(desde_raw)
        vigencia_hasta = _parse_date_cell(hasta_raw)

        # Date order check
        if vigencia_desde and vigencia_hasta and vigencia_desde > vigencia_hasta:
            _invalid("vigencia_desde debe ser anterior o igual a vigencia_hasta")
            continue

        # Resolve rut → cliente/empresa
        cliente_id = clientes_by_rut.get(rut_raw)
        empresa_id = empresas_by_rut.get(rut_raw)

        # Resolve sku → producto (not found is OK, still import)
        producto_id = productos_by_sku.get(sku_raw)

        # Determine status
        rut_found = cliente_id is not None or empresa_id is not None
        if not rut_found:
            status = "pendiente"
            motivo: Optional[str] = f"rut '{rut_raw}' no encontrado en clientes ni empresas"
        elif (rut_raw, sku_raw) in existing_precios:
            status = "actualizar"
            motivo = None
        else:
            status = "crear"
            motivo = None

        # If sku not found, mark as pendiente (but still insert)
        if producto_id is None and status not in ("pendiente",):
            status = "pendiente"
            motivo = f"sku '{sku_raw}' no encontrado en productos"

        valid_rows.append(PrecioRow(
            row_num=row_idx,
            rut_entidad=rut_raw,
            sku=sku_raw,
            precio_especial=precio_especial,
            descuento_pct=descuento_pct,
            vigencia_desde=vigencia_desde,
            vigencia_hasta=vigencia_hasta,
            status=status,
            motivo=motivo,
            cliente_id=cliente_id,
            empresa_id=empresa_id,
            producto_id=producto_id,
        ))

    wb.close()
    return valid_rows, invalid_rows


def _row_to_dict(row: PrecioRow) -> dict:
    return {
        "row_num": row.row_num,
        "rut_entidad": row.rut_entidad,
        "sku": row.sku,
        "precio_especial": float(row.precio_especial) if row.precio_especial is not None else None,
        "descuento_pct": float(row.descuento_pct) if row.descuento_pct is not None else None,
        "vigencia_desde": row.vigencia_desde.isoformat() if row.vigencia_desde else None,
        "vigencia_hasta": row.vigencia_hasta.isoformat() if row.vigencia_hasta else None,
        "status": row.status,
        "motivo": row.motivo,
        "cliente_id": row.cliente_id,
        "empresa_id": row.empresa_id,
        "producto_id": row.producto_id,
    }


# ============================================================================
# GET /template
# ============================================================================

@router.get("/template")
def get_template(perms: tuple[User, Session] = Depends(require_admin)):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Precios Especiales"

        headers = ["rut_cliente_o_empresa", "sku", "precio_especial", "descuento_pct", "vigencia_desde", "vigencia_hasta"]
        ws.append(headers)

        # Example row 1: precio_especial set
        ws.append(["76354771-K", "SKU001", 1200, "", "2026-01-01", "2026-12-31"])
        # Example row 2: descuento_pct set
        ws.append(["76354771-K", "SKU002", "", 15.00, "2026-01-01", ""])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=plantilla_precios_especiales.xlsx"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando plantilla: {e}")


# ============================================================================
# POST /preview
# ============================================================================

@router.post("/preview")
async def preview_precios_especiales(
    file: UploadFile,
    perms: tuple[User, Session] = Depends(require_admin),
):
    _, db = perms

    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="El archivo está vacío")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo archivo: {e}")

    clientes_by_rut = _get_clientes_by_rut(db)
    empresas_by_rut = _get_empresas_by_rut(db)
    productos_by_sku = _get_productos_by_sku(db)
    existing_precios = _get_existing_precios(db)

    try:
        valid_rows, invalid_rows = _parse_rows(
            content,
            clientes_by_rut=clientes_by_rut,
            empresas_by_rut=empresas_by_rut,
            productos_by_sku=productos_by_sku,
            existing_precios=existing_precios,
        )
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Error parseando archivo: {e}")

    a_crear = sum(1 for r in valid_rows if r.status == "crear")
    a_actualizar = sum(1 for r in valid_rows if r.status == "actualizar")
    a_pendiente = sum(1 for r in valid_rows if r.status == "pendiente")

    return {
        "total_filas": len(valid_rows) + len(invalid_rows),
        "a_crear": a_crear,
        "a_actualizar": a_actualizar,
        "a_pendiente": a_pendiente,
        "filas_invalidas": len(invalid_rows),
        "rows": [_row_to_dict(r) for r in valid_rows],
        "invalid_rows": invalid_rows,
    }


# ============================================================================
# POST /import
# ============================================================================

@router.post("/import")
async def import_precios_especiales(
    file: UploadFile,
    idempotency_key: Optional[str] = None,
    perms: tuple[User, Session] = Depends(require_admin),
):
    current_user, db = perms

    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="El archivo está vacío")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo archivo: {e}")

    import_id = idempotency_key or str(uuid.uuid4())

    # Idempotency: return cached result if same import_id
    existing_report = (
        db.query(ImportReport).filter(ImportReport.import_id == import_id).first()
    )
    if existing_report:
        report_data = json.loads(existing_report.report_json or "{}")
        return {
            "status": existing_report.status,
            "import_id": existing_report.import_id,
            "timestamp": existing_report.created_at.isoformat(),
            "report": report_data,
        }

    clientes_by_rut = _get_clientes_by_rut(db)
    empresas_by_rut = _get_empresas_by_rut(db)
    productos_by_sku = _get_productos_by_sku(db)
    existing_precios = _get_existing_precios(db)

    try:
        valid_rows, invalid_rows = _parse_rows(
            content,
            clientes_by_rut=clientes_by_rut,
            empresas_by_rut=empresas_by_rut,
            productos_by_sku=productos_by_sku,
            existing_precios=existing_precios,
        )
    except Exception as e:
        err_report = ImportReport(
            import_id=import_id,
            status="error",
            error_message=str(e),
            filename=file.filename,
            total_rows=0,
        )
        db.add(err_report)
        db.commit()
        raise HTTPException(status_code=422, detail=f"Error parseando archivo: {e}")

    try:
        created_count = 0
        updated_count = 0
        pending_count = 0
        error_count = len(invalid_rows)

        for row in valid_rows:
            if row.status == "error":
                error_count += 1
                continue

            if row.status == "actualizar":
                existing = (
                    db.query(PrecioEspecialCliente)
                    .filter(
                        PrecioEspecialCliente.rut_entidad == row.rut_entidad,
                        PrecioEspecialCliente.sku == row.sku,
                    )
                    .first()
                )
                if existing:
                    existing.precio_especial = row.precio_especial
                    existing.descuento_pct = row.descuento_pct
                    existing.vigencia_desde = row.vigencia_desde
                    existing.vigencia_hasta = row.vigencia_hasta
                    existing.cliente_id = row.cliente_id
                    existing.empresa_id = row.empresa_id
                    existing.producto_id = row.producto_id
                    updated_count += 1
                else:
                    # Race condition: insert instead
                    record = PrecioEspecialCliente(
                        rut_entidad=row.rut_entidad,
                        sku=row.sku,
                        precio_especial=row.precio_especial,
                        descuento_pct=row.descuento_pct,
                        vigencia_desde=row.vigencia_desde,
                        vigencia_hasta=row.vigencia_hasta,
                        cliente_id=row.cliente_id,
                        empresa_id=row.empresa_id,
                        producto_id=row.producto_id,
                    )
                    db.add(record)
                    created_count += 1

            else:
                # "crear" or "pendiente" — insert new record
                record = PrecioEspecialCliente(
                    rut_entidad=row.rut_entidad,
                    sku=row.sku,
                    precio_especial=row.precio_especial,
                    descuento_pct=row.descuento_pct,
                    vigencia_desde=row.vigencia_desde,
                    vigencia_hasta=row.vigencia_hasta,
                    cliente_id=row.cliente_id,
                    empresa_id=row.empresa_id,
                    producto_id=row.producto_id,
                )
                db.add(record)
                if row.status == "pendiente":
                    pending_count += 1
                else:
                    created_count += 1

        overall_status = (
            "success" if error_count == 0
            else "partial" if (created_count + updated_count + pending_count) > 0
            else "error"
        )

        report_payload = {
            "created_count": created_count,
            "updated_count": updated_count,
            "pending_count": pending_count,
            "error_count": error_count,
            "total_rows": len(valid_rows) + len(invalid_rows),
        }

        import_report = ImportReport(
            import_id=import_id,
            status=overall_status,
            created_count=created_count,
            updated_count=updated_count,
            pending_count=pending_count,
            error_count=error_count,
            total_rows=len(valid_rows) + len(invalid_rows),
            filename=file.filename,
            report_json=json.dumps(report_payload, ensure_ascii=False, default=str),
        )
        db.add(import_report)
        db.commit()
        db.refresh(import_report)

        return {
            "status": overall_status,
            "import_id": import_id,
            "timestamp": import_report.created_at.isoformat(),
            "report": report_payload,
        }

    except Exception as e:
        db.rollback()
        err_report = ImportReport(
            import_id=import_id,
            status="error",
            error_message=f"Transaction error: {e}",
            filename=file.filename,
            total_rows=0,
        )
        db.add(err_report)
        db.commit()
        raise HTTPException(status_code=500, detail=f"Import error: {e}")

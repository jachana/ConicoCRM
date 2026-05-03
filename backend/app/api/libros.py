"""API endpoints for retrieving Libros (sales and purchase books)"""
import re
import csv
from io import StringIO, BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import asc, desc
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

from app.api.deps import require_permission
from app.database import get_db
from app.models.libro import LibroVentas, LibroCompras
from app.models.user import User
from app.schemas.libro import (
    LibroVentasRead, LibroComprasRead,
    LibroGenerarRequest, LibroEnviarResponse,
)
from app.services.libro_service import LibroService
from app.services.dte_service import get_dte_service

router = APIRouter()

# Valid sort fields for libros endpoints
VALID_SORT_FIELDS = {"id", "periodo", "estado", "monto_total", "total_registros", "created_at"}
VALID_SORT_ORDERS = {"asc", "desc"}


def _validate_pagination(limit: int, offset: int) -> None:
    """Validate pagination parameters.

    Raises:
        HTTPException: If limit or offset are invalid
    """
    if limit < 1:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="limit must be >= 1"
        )
    if offset < 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="offset must be >= 0"
        )


def _validate_periodo(periodo: str | None) -> None:
    """Validate periodo format (YYYY-MM).

    Raises:
        HTTPException: If periodo format is invalid
    """
    if periodo is None:
        return
    if not re.match(r'^\d{4}-\d{2}$', periodo):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="periodo must be in YYYY-MM format"
        )


def _validate_sort_params(sort_by: str | None, sort_order: str | None) -> None:
    """Validate sort parameters.

    Raises:
        HTTPException: If sort_by or sort_order are invalid
    """
    if sort_by is None:
        return

    if sort_by not in VALID_SORT_FIELDS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"sort_by must be one of: {', '.join(sorted(VALID_SORT_FIELDS))}"
        )

    if sort_order and sort_order not in VALID_SORT_ORDERS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"sort_order must be one of: {', '.join(VALID_SORT_ORDERS)}"
        )


def _apply_sort(query, model, sort_by: str | None, sort_order: str | None):
    """Apply sorting to a query.

    Args:
        query: SQLAlchemy query object
        model: The model class (LibroVentas or LibroCompras)
        sort_by: Field name to sort by
        sort_order: 'asc' or 'desc'

    Returns:
        Sorted query
    """
    if sort_by is None:
        # Default: sort by created_at descending
        return query.order_by(desc(model.created_at))

    sort_order = sort_order or "asc"
    sort_column = getattr(model, sort_by)

    if sort_order == "asc":
        return query.order_by(asc(sort_column))
    else:
        return query.order_by(desc(sort_column))


# ── Libros de Ventas ──────────────────────────────────────────────────

@router.get("/ventas", response_model=dict)
def listar_libros_ventas(
    periodo: str | None = Query(None, description="Filter by exact YYYY-MM period"),
    periodo_from: str | None = Query(None, description="Filter by YYYY-MM period from (inclusive)"),
    periodo_to: str | None = Query(None, description="Filter by YYYY-MM period to (inclusive)"),
    estado: str | None = Query(None, description="Filter by estado (borrador, enviado, etc)"),
    sort_by: str | None = Query(None, description=f"Sort by field: {', '.join(sorted(VALID_SORT_FIELDS))}"),
    sort_order: str | None = Query(None, description="Sort order: asc or desc"),
    limit: int = Query(50, ge=1, le=500, description="Pagination limit"),
    offset: int = Query(0, ge=0, description="Pagination offset"),
    perms: tuple[User, Session] = require_permission("libros", "view"),
):
    """Get list of sales books (Libros de Ventas) for current user's empresa.

    Parameters:
    - periodo: Optional filter by exact YYYY-MM period
    - periodo_from: Optional filter by YYYY-MM period from (inclusive)
    - periodo_to: Optional filter by YYYY-MM period to (inclusive)
    - estado: Optional filter by estado (borrador, enviado, etc)
    - sort_by: Optional sort field
    - sort_order: Optional sort order (asc or desc, default asc)
    - limit: Number of records per page (default 50, max 500)
    - offset: Number of records to skip (default 0)

    Returns:
    - data: List of LibroVentasRead
    - pagination: {limit, offset, total}
    """
    current_user, db = perms

    # Validate parameters
    _validate_pagination(limit, offset)
    _validate_periodo(periodo)
    _validate_periodo(periodo_from)
    _validate_periodo(periodo_to)
    _validate_sort_params(sort_by, sort_order)

    # Ensure user has an empresa_id for authorization
    if not current_user.empresa_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User is not assigned to an empresa"
        )

    # Build query filtered by current user's empresa
    query = db.query(LibroVentas).filter(LibroVentas.empresa_id == current_user.empresa_id)

    # Apply filters
    if periodo:
        query = query.filter(LibroVentas.periodo == periodo)
    else:
        # Only apply periodo range if periodo is not specified
        if periodo_from:
            query = query.filter(LibroVentas.periodo >= periodo_from)
        if periodo_to:
            query = query.filter(LibroVentas.periodo <= periodo_to)

    if estado:
        query = query.filter(LibroVentas.estado == estado)

    # Count total before applying pagination and sorting
    total = query.count()

    # Apply sorting
    query = _apply_sort(query, LibroVentas, sort_by, sort_order)

    # Apply pagination
    libros = query.offset(offset).limit(limit).all()

    # Convert to response schema
    data = [LibroVentasRead.model_validate(libro) for libro in libros]

    return {
        "data": data,
        "pagination": {
            "limit": limit,
            "offset": offset,
            "total": total
        }
    }


@router.get("/ventas/{id}", response_model=LibroVentasRead)
def obtener_libro_ventas(
    id: int,
    perms: tuple[User, Session] = require_permission("libros", "view"),
):
    """Get a specific sales book by ID.

    Parameters:
    - id: LibroVentas ID

    Returns:
    - LibroVentasRead
    """
    current_user, db = perms

    libro = db.query(LibroVentas).filter_by(id=id).first()

    if not libro:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Libro de Ventas not found"
        )

    # Authorization check: ensure user can only access libros from their empresa
    if libro.empresa_id != current_user.empresa_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to access this libro"
        )

    return LibroVentasRead.model_validate(libro)


# ── Libros de Compras ─────────────────────────────────────────────────

@router.get("/compras", response_model=dict)
def listar_libros_compras(
    periodo: str | None = Query(None, description="Filter by exact YYYY-MM period"),
    periodo_from: str | None = Query(None, description="Filter by YYYY-MM period from (inclusive)"),
    periodo_to: str | None = Query(None, description="Filter by YYYY-MM period to (inclusive)"),
    estado: str | None = Query(None, description="Filter by estado (borrador, enviado, etc)"),
    sort_by: str | None = Query(None, description=f"Sort by field: {', '.join(sorted(VALID_SORT_FIELDS))}"),
    sort_order: str | None = Query(None, description="Sort order: asc or desc"),
    limit: int = Query(50, ge=1, le=500, description="Pagination limit"),
    offset: int = Query(0, ge=0, description="Pagination offset"),
    perms: tuple[User, Session] = require_permission("libros", "view"),
):
    """Get list of purchase books (Libros de Compras) for current user's empresa.

    Parameters:
    - periodo: Optional filter by exact YYYY-MM period
    - periodo_from: Optional filter by YYYY-MM period from (inclusive)
    - periodo_to: Optional filter by YYYY-MM period to (inclusive)
    - estado: Optional filter by estado (borrador, enviado, etc)
    - sort_by: Optional sort field
    - sort_order: Optional sort order (asc or desc, default asc)
    - limit: Number of records per page (default 50, max 500)
    - offset: Number of records to skip (default 0)

    Returns:
    - data: List of LibroComprasRead
    - pagination: {limit, offset, total}
    """
    current_user, db = perms

    # Validate parameters
    _validate_pagination(limit, offset)
    _validate_periodo(periodo)
    _validate_periodo(periodo_from)
    _validate_periodo(periodo_to)
    _validate_sort_params(sort_by, sort_order)

    # Ensure user has an empresa_id for authorization
    if not current_user.empresa_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User is not assigned to an empresa"
        )

    # Build query filtered by current user's empresa
    query = db.query(LibroCompras).filter(LibroCompras.empresa_id == current_user.empresa_id)

    # Apply filters
    if periodo:
        query = query.filter(LibroCompras.periodo == periodo)
    else:
        # Only apply periodo range if periodo is not specified
        if periodo_from:
            query = query.filter(LibroCompras.periodo >= periodo_from)
        if periodo_to:
            query = query.filter(LibroCompras.periodo <= periodo_to)

    if estado:
        query = query.filter(LibroCompras.estado == estado)

    # Count total before applying pagination and sorting
    total = query.count()

    # Apply sorting
    query = _apply_sort(query, LibroCompras, sort_by, sort_order)

    # Apply pagination
    libros = query.offset(offset).limit(limit).all()

    # Convert to response schema
    data = [LibroComprasRead.model_validate(libro) for libro in libros]

    return {
        "data": data,
        "pagination": {
            "limit": limit,
            "offset": offset,
            "total": total
        }
    }


@router.get("/compras/{id}", response_model=LibroComprasRead)
def obtener_libro_compras(
    id: int,
    perms: tuple[User, Session] = require_permission("libros", "view"),
):
    """Get a specific purchase book by ID.

    Parameters:
    - id: LibroCompras ID

    Returns:
    - LibroComprasRead
    """
    current_user, db = perms

    libro = db.query(LibroCompras).filter_by(id=id).first()

    if not libro:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Libro de Compras not found"
        )

    # Authorization check: ensure user can only access libros from their empresa
    if libro.empresa_id != current_user.empresa_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to access this libro"
        )

    return LibroComprasRead.model_validate(libro)


# ── CSV Export Endpoints ──────────────────────────────────────────────────

@router.get("/ventas/export/csv")
def export_libros_ventas_csv(
    periodo: str | None = Query(None, description="Filter by exact YYYY-MM period"),
    periodo_from: str | None = Query(None, description="Filter by YYYY-MM period from (inclusive)"),
    periodo_to: str | None = Query(None, description="Filter by YYYY-MM period to (inclusive)"),
    estado: str | None = Query(None, description="Filter by estado (borrador, enviado, etc)"),
    sort_by: str | None = Query(None, description=f"Sort by field: {', '.join(sorted(VALID_SORT_FIELDS))}"),
    sort_order: str | None = Query(None, description="Sort order: asc or desc"),
    limit: int = Query(None, ge=1, le=500, description="Ignored for CSV export (returns all records)"),
    offset: int = Query(None, ge=0, description="Ignored for CSV export (returns all records)"),
    perms: tuple[User, Session] = require_permission("libros", "view"),
):
    """Export sales books (Libros de Ventas) as CSV.

    Parameters are the same as list endpoint, but pagination is ignored.
    Returns all matching records.

    Returns:
    - CSV file with columns: Período, Total Registros, Monto Total, Estado, Folio Inicio, Folio Fin
    """
    current_user, db = perms

    # Validate parameters
    _validate_periodo(periodo)
    _validate_periodo(periodo_from)
    _validate_periodo(periodo_to)
    _validate_sort_params(sort_by, sort_order)

    # Ensure user has an empresa_id for authorization
    if not current_user.empresa_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User is not assigned to an empresa"
        )

    # Build query filtered by current user's empresa
    query = db.query(LibroVentas).filter(LibroVentas.empresa_id == current_user.empresa_id)

    # Apply filters
    if periodo:
        query = query.filter(LibroVentas.periodo == periodo)
    else:
        if periodo_from:
            query = query.filter(LibroVentas.periodo >= periodo_from)
        if periodo_to:
            query = query.filter(LibroVentas.periodo <= periodo_to)

    if estado:
        query = query.filter(LibroVentas.estado == estado)

    # Apply sorting
    query = _apply_sort(query, LibroVentas, sort_by, sort_order)

    # Get all records (no pagination)
    libros = query.all()

    # Create CSV in memory
    output = StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=["Período", "Total Registros", "Monto Total", "Estado", "Folio Inicio", "Folio Fin"],
        lineterminator="\n"
    )
    writer.writeheader()

    for libro in libros:
        writer.writerow({
            "Período": libro.periodo,
            "Total Registros": libro.total_registros,
            "Monto Total": libro.monto_total,
            "Estado": libro.estado,
            "Folio Inicio": libro.folio_inicio if libro.folio_inicio is not None else "",
            "Folio Fin": libro.folio_fin if libro.folio_fin is not None else "",
        })

    csv_content = output.getvalue()
    output.close()

    return StreamingResponse(
        iter([csv_content.encode("utf-8")]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=libros_ventas.csv"}
    )


@router.get("/compras/export/csv")
def export_libros_compras_csv(
    periodo: str | None = Query(None, description="Filter by exact YYYY-MM period"),
    periodo_from: str | None = Query(None, description="Filter by YYYY-MM period from (inclusive)"),
    periodo_to: str | None = Query(None, description="Filter by YYYY-MM period to (inclusive)"),
    estado: str | None = Query(None, description="Filter by estado (borrador, enviado, etc)"),
    sort_by: str | None = Query(None, description=f"Sort by field: {', '.join(sorted(VALID_SORT_FIELDS))}"),
    sort_order: str | None = Query(None, description="Sort order: asc or desc"),
    limit: int = Query(None, ge=1, le=500, description="Ignored for CSV export (returns all records)"),
    offset: int = Query(None, ge=0, description="Ignored for CSV export (returns all records)"),
    perms: tuple[User, Session] = require_permission("libros", "view"),
):
    """Export purchase books (Libros de Compras) as CSV.

    Parameters are the same as list endpoint, but pagination is ignored.
    Returns all matching records.

    Returns:
    - CSV file with columns: Período, Total Registros, Monto Total, Estado, RUT Proveedor
    """
    current_user, db = perms

    # Validate parameters
    _validate_periodo(periodo)
    _validate_periodo(periodo_from)
    _validate_periodo(periodo_to)
    _validate_sort_params(sort_by, sort_order)

    # Ensure user has an empresa_id for authorization
    if not current_user.empresa_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User is not assigned to an empresa"
        )

    # Build query filtered by current user's empresa
    query = db.query(LibroCompras).filter(LibroCompras.empresa_id == current_user.empresa_id)

    # Apply filters
    if periodo:
        query = query.filter(LibroCompras.periodo == periodo)
    else:
        if periodo_from:
            query = query.filter(LibroCompras.periodo >= periodo_from)
        if periodo_to:
            query = query.filter(LibroCompras.periodo <= periodo_to)

    if estado:
        query = query.filter(LibroCompras.estado == estado)

    # Apply sorting
    query = _apply_sort(query, LibroCompras, sort_by, sort_order)

    # Get all records (no pagination)
    libros = query.all()

    # Create CSV in memory
    output = StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=["Período", "Total Registros", "Monto Total", "Estado", "RUT Proveedor"],
        lineterminator="\n"
    )
    writer.writeheader()

    for libro in libros:
        writer.writerow({
            "Período": libro.periodo,
            "Total Registros": libro.total_registros,
            "Monto Total": libro.monto_total,
            "Estado": libro.estado,
            "RUT Proveedor": libro.rut_proveedor if libro.rut_proveedor is not None else "",
        })

    csv_content = output.getvalue()
    output.close()

    return StreamingResponse(
        iter([csv_content.encode("utf-8")]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=libros_compras.csv"}
    )


# ── Excel Export Endpoints ────────────────────────────────────────────────────

def _create_excel_workbook_ventas(libros: list) -> BytesIO:
    """Create an Excel workbook for Libros de Ventas with formatting.

    Args:
        libros: List of LibroVentas objects

    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Libros de Ventas"

    # Define headers
    headers = ["Período", "Total Registros", "Monto Total", "Estado", "Folio Inicio", "Folio Fin"]
    ws.append(headers)

    # Format header row (bold)
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    # Add data rows
    for libro in libros:
        ws.append([
            libro.periodo,
            libro.total_registros,
            libro.monto_total,
            libro.estado,
            libro.folio_inicio if libro.folio_inicio is not None else "",
            libro.folio_fin if libro.folio_fin is not None else "",
        ])

    # Format monto_total column (column 3) as currency
    for row_idx in range(2, len(libros) + 2):
        cell = ws.cell(row=row_idx, column=3)
        cell.number_format = '$#,##0'

    # Calculate totals
    total_registros = sum(libro.total_registros for libro in libros)
    total_monto = sum(libro.monto_total for libro in libros)

    # Add summary row (row index for last row after appending all data)
    summary_row_idx = len(libros) + 2
    ws.cell(row=summary_row_idx, column=1).value = "TOTAL"
    ws.cell(row=summary_row_idx, column=2).value = total_registros
    ws.cell(row=summary_row_idx, column=3).value = total_monto

    # Format summary row currency (column 3 = Monto Total)
    summary_cell = ws.cell(row=summary_row_idx, column=3)
    summary_cell.number_format = '$#,##0'

    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, ValueError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Return as BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _create_excel_workbook_compras(libros: list) -> BytesIO:
    """Create an Excel workbook for Libros de Compras with formatting.

    Args:
        libros: List of LibroCompras objects

    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Libros de Compras"

    # Define headers
    headers = ["Período", "Total Registros", "Monto Total", "Estado", "RUT Proveedor"]
    ws.append(headers)

    # Format header row (bold)
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    # Add data rows
    for libro in libros:
        ws.append([
            libro.periodo,
            libro.total_registros,
            libro.monto_total,
            libro.estado,
            libro.rut_proveedor if libro.rut_proveedor is not None else "",
        ])

    # Format monto_total column (column 3) as currency
    for row_idx in range(2, len(libros) + 2):
        cell = ws.cell(row=row_idx, column=3)
        cell.number_format = '$#,##0'

    # Calculate totals
    total_registros = sum(libro.total_registros for libro in libros)
    total_monto = sum(libro.monto_total for libro in libros)

    # Add summary row (row index for last row after appending all data)
    summary_row_idx = len(libros) + 2
    ws.cell(row=summary_row_idx, column=1).value = "TOTAL"
    ws.cell(row=summary_row_idx, column=2).value = total_registros
    ws.cell(row=summary_row_idx, column=3).value = total_monto

    # Format summary row currency (column 3 = Monto Total)
    summary_cell = ws.cell(row=summary_row_idx, column=3)
    summary_cell.number_format = '$#,##0'

    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, ValueError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Return as BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/ventas/export/excel")
def export_libros_ventas_excel(
    periodo: str | None = Query(None, description="Filter by exact YYYY-MM period"),
    periodo_from: str | None = Query(None, description="Filter by YYYY-MM period from (inclusive)"),
    periodo_to: str | None = Query(None, description="Filter by YYYY-MM period to (inclusive)"),
    estado: str | None = Query(None, description="Filter by estado (borrador, enviado, etc)"),
    sort_by: str | None = Query(None, description=f"Sort by field: {', '.join(sorted(VALID_SORT_FIELDS))}"),
    sort_order: str | None = Query(None, description="Sort order: asc or desc"),
    limit: int = Query(None, ge=1, le=500, description="Ignored for Excel export (returns all records)"),
    offset: int = Query(None, ge=0, description="Ignored for Excel export (returns all records)"),
    perms: tuple[User, Session] = require_permission("libros", "view"),
):
    """Export sales books (Libros de Ventas) as Excel.

    Parameters are the same as list endpoint, but pagination is ignored.
    Returns all matching records with formatting.

    Returns:
    - Excel file with columns: Período, Total Registros, Monto Total, Estado, Folio Inicio, Folio Fin
    - Includes summary row with totals
    """
    current_user, db = perms

    # Validate parameters
    _validate_periodo(periodo)
    _validate_periodo(periodo_from)
    _validate_periodo(periodo_to)
    _validate_sort_params(sort_by, sort_order)

    # Ensure user has an empresa_id for authorization
    if not current_user.empresa_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User is not assigned to an empresa"
        )

    # Build query filtered by current user's empresa
    query = db.query(LibroVentas).filter(LibroVentas.empresa_id == current_user.empresa_id)

    # Apply filters
    if periodo:
        query = query.filter(LibroVentas.periodo == periodo)
    else:
        if periodo_from:
            query = query.filter(LibroVentas.periodo >= periodo_from)
        if periodo_to:
            query = query.filter(LibroVentas.periodo <= periodo_to)

    if estado:
        query = query.filter(LibroVentas.estado == estado)

    # Apply sorting
    query = _apply_sort(query, LibroVentas, sort_by, sort_order)

    # Get all records (no pagination)
    libros = query.all()

    # Create Excel workbook
    excel_content = _create_excel_workbook_ventas(libros)

    return StreamingResponse(
        iter([excel_content.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=libros_ventas.xlsx"}
    )


@router.get("/compras/export/excel")
def export_libros_compras_excel(
    periodo: str | None = Query(None, description="Filter by exact YYYY-MM period"),
    periodo_from: str | None = Query(None, description="Filter by YYYY-MM period from (inclusive)"),
    periodo_to: str | None = Query(None, description="Filter by YYYY-MM period to (inclusive)"),
    estado: str | None = Query(None, description="Filter by estado (borrador, enviado, etc)"),
    sort_by: str | None = Query(None, description=f"Sort by field: {', '.join(sorted(VALID_SORT_FIELDS))}"),
    sort_order: str | None = Query(None, description="Sort order: asc or desc"),
    limit: int = Query(None, ge=1, le=500, description="Ignored for Excel export (returns all records)"),
    offset: int = Query(None, ge=0, description="Ignored for Excel export (returns all records)"),
    perms: tuple[User, Session] = require_permission("libros", "view"),
):
    """Export purchase books (Libros de Compras) as Excel.

    Parameters are the same as list endpoint, but pagination is ignored.
    Returns all matching records with formatting.

    Returns:
    - Excel file with columns: Período, Total Registros, Monto Total, Estado, RUT Proveedor
    - Includes summary row with totals
    """
    current_user, db = perms

    # Validate parameters
    _validate_periodo(periodo)
    _validate_periodo(periodo_from)
    _validate_periodo(periodo_to)
    _validate_sort_params(sort_by, sort_order)

    # Ensure user has an empresa_id for authorization
    if not current_user.empresa_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User is not assigned to an empresa"
        )

    # Build query filtered by current user's empresa
    query = db.query(LibroCompras).filter(LibroCompras.empresa_id == current_user.empresa_id)

    # Apply filters
    if periodo:
        query = query.filter(LibroCompras.periodo == periodo)
    else:
        if periodo_from:
            query = query.filter(LibroCompras.periodo >= periodo_from)
        if periodo_to:
            query = query.filter(LibroCompras.periodo <= periodo_to)

    if estado:
        query = query.filter(LibroCompras.estado == estado)

    # Apply sorting
    query = _apply_sort(query, LibroCompras, sort_by, sort_order)

    # Get all records (no pagination)
    libros = query.all()

    # Create Excel workbook
    excel_content = _create_excel_workbook_compras(libros)

    return StreamingResponse(
        iter([excel_content.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=libros_compras.xlsx"}
    )


# ── Generate / Submit endpoints ───────────────────────────────────────────────

@router.post("/ventas/generar", response_model=LibroVentasRead, status_code=status.HTTP_201_CREATED)
def generar_libro_ventas(
    body: LibroGenerarRequest,
    perms: tuple[User, Session] = require_permission("libros", "create"),
):
    """Generate (or retrieve existing) LibroVentas for the given period."""
    current_user, db = perms
    if not current_user.empresa_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="User has no empresa")
    try:
        libro = LibroService.generar_libro_ventas(db, current_user.empresa_id, body.periodo)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    return LibroVentasRead.model_validate(libro)


@router.post("/compras/generar", response_model=LibroComprasRead, status_code=status.HTTP_201_CREATED)
def generar_libro_compras(
    body: LibroGenerarRequest,
    perms: tuple[User, Session] = require_permission("libros", "create"),
):
    """Generate (or retrieve existing) LibroCompras for the given period."""
    current_user, db = perms
    if not current_user.empresa_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="User has no empresa")
    try:
        libro = LibroService.generar_libro_compras(db, current_user.empresa_id, body.periodo)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    return LibroComprasRead.model_validate(libro)


@router.post("/ventas/{id}/enviar", response_model=LibroEnviarResponse)
def enviar_libro_ventas(
    id: int,
    perms: tuple[User, Session] = require_permission("libros", "create"),
):
    """Submit a LibroVentas to SII via Lioren. Idempotent: already-sent books return 409."""
    current_user, db = perms
    libro = db.query(LibroVentas).filter_by(id=id).first()
    if not libro:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Libro de Ventas not found")
    if libro.empresa_id != current_user.empresa_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    if libro.estado == "enviado":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Libro ya enviado a SII")

    rut_emisor = LibroService.get_rut_emisor(db)
    detalles = LibroService.get_detalle_ventas(db, libro.empresa_id, libro.periodo)
    svc = get_dte_service()
    payload = svc.build_libro_ventas_payload(rut_emisor, libro.periodo, detalles)

    try:
        lioren_resp = svc.submit_libro("ventas", payload)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Lioren error: {exc}")

    libro.estado = "enviado"
    db.commit()
    db.refresh(libro)

    return LibroEnviarResponse(id=libro.id, periodo=libro.periodo, estado=libro.estado, lioren_response=lioren_resp)


@router.post("/compras/{id}/enviar", response_model=LibroEnviarResponse)
def enviar_libro_compras(
    id: int,
    perms: tuple[User, Session] = require_permission("libros", "create"),
):
    """Submit a LibroCompras to SII via Lioren. Idempotent: already-sent books return 409."""
    current_user, db = perms
    libro = db.query(LibroCompras).filter_by(id=id).first()
    if not libro:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Libro de Compras not found")
    if libro.empresa_id != current_user.empresa_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
    if libro.estado == "enviado":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Libro ya enviado a SII")

    rut_emisor = LibroService.get_rut_emisor(db)
    detalles = LibroService.get_detalle_compras(db, libro.empresa_id, libro.periodo)
    svc = get_dte_service()
    payload = svc.build_libro_compras_payload(rut_emisor, libro.periodo, detalles)

    try:
        lioren_resp = svc.submit_libro("compras", payload)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Lioren error: {exc}")

    libro.estado = "enviado"
    db.commit()
    db.refresh(libro)

    return LibroEnviarResponse(id=libro.id, periodo=libro.periodo, estado=libro.estado, lioren_response=lioren_resp)

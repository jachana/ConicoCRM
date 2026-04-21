"""Reportes — JSON endpoints for Ventas, Cobranza, and Inventario."""
from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import openpyxl
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload
from weasyprint import HTML

from app.api.deps import require_permission
from app.models.dte_emision import DteEmision
from app.models.factura import Factura, FacturaLinea
from app.models.movimiento_inventario import MovimientoInventario
from app.models.orden_compra import OrdenCompra
from app.models.pago import Pago
from app.models.producto import Producto
from app.models.user import User

router = APIRouter()

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _excel_response(wb: openpyxl.Workbook, tab: str, date_from: date, date_to: date) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{tab}-{date_from}-{date_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type=EXCEL_MIME,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

_ZERO = Decimal("0")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _validate_dates(date_from: date, date_to: date) -> None:
    if date_from > date_to:
        raise HTTPException(422, detail="date_from debe ser anterior a date_to")


def _sum_pagos_for_ids(db: Session, factura_ids: list[int]) -> Decimal:
    if not factura_ids:
        return _ZERO
    result = db.query(func.sum(Pago.monto)).filter(Pago.factura_id.in_(factura_ids)).scalar()
    return Decimal(str(result)) if result is not None else _ZERO


# ---------------------------------------------------------------------------
# GET /ventas
# ---------------------------------------------------------------------------

@router.get("/ventas")
def reporte_ventas(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    current_user, db = perms

    base_q = (
        db.query(Factura)
        .options(joinedload(Factura.cliente), joinedload(Factura.vendedor))
        .filter(
            Factura.fecha >= date_from,
            Factura.fecha <= date_to,
            Factura.estado != "anulada",
        )
    )
    if current_user.role == "vendedor":
        base_q = base_q.filter(Factura.vendedor_id == current_user.id)

    facturas = base_q.all()

    # --- KPIs core ---
    factura_ids = [f.id for f in facturas]
    total_vendido = sum((f.total for f in facturas), _ZERO)
    num_facturas = len(facturas)
    ticket_promedio = (total_vendido / num_facturas) if num_facturas else _ZERO

    total_pagado = _sum_pagos_for_ids(db, factura_ids)
    total_por_cobrar = total_vendido - total_pagado
    if total_por_cobrar < _ZERO:
        total_por_cobrar = _ZERO

    # --- Variación vs período anterior ---
    duration_days = (date_to - date_from).days + 1
    prev_to = date_from - timedelta(days=1)
    prev_from = prev_to - timedelta(days=duration_days - 1)

    prev_q = db.query(func.sum(Factura.total)).filter(
        Factura.fecha >= prev_from,
        Factura.fecha <= prev_to,
        Factura.estado != "anulada",
    )
    if current_user.role == "vendedor":
        prev_q = prev_q.filter(Factura.vendedor_id == current_user.id)
    prev_total_raw = prev_q.scalar()
    prev_total = Decimal(str(prev_total_raw)) if prev_total_raw else _ZERO

    if prev_total == _ZERO:
        variacion = 0.0
    else:
        variacion = float(((total_vendido - prev_total) / prev_total) * 100)

    # --- Ventas diarias ---
    daily_map: dict[date, Decimal] = {}
    for f in facturas:
        daily_map[f.fecha] = daily_map.get(f.fecha, _ZERO) + f.total
    ventas_diarias = [
        {"fecha": str(d), "monto": float(m)}
        for d, m in sorted(daily_map.items())
    ]

    # --- Top clientes ---
    cliente_map: dict[int, dict] = {}
    for f in facturas:
        if f.cliente_id is None:
            continue
        entry = cliente_map.setdefault(
            f.cliente_id,
            {
                "cliente_id": f.cliente_id,
                "nombre": f.cliente.nombre if f.cliente else "",
                "total": _ZERO,
                "num_facturas": 0,
            },
        )
        entry["total"] += f.total
        entry["num_facturas"] += 1
    top_clientes = sorted(cliente_map.values(), key=lambda x: x["total"], reverse=True)[:10]
    for c in top_clientes:
        c["total"] = float(c["total"])

    # --- Por vendedor (admin only; vendedor only sees themselves) ---
    vendedor_map: dict[int, dict] = {}
    for f in facturas:
        if f.vendedor_id is None:
            continue
        entry = vendedor_map.setdefault(
            f.vendedor_id,
            {
                "vendedor_id": f.vendedor_id,
                "nombre": f.vendedor.name if f.vendedor else "",
                "total": _ZERO,
                "num_facturas": 0,
            },
        )
        entry["total"] += f.total
        entry["num_facturas"] += 1
    por_vendedor = sorted(vendedor_map.values(), key=lambda x: x["total"], reverse=True)
    for v in por_vendedor:
        v["total"] = float(v["total"])

    return {
        "kpis": {
            "total_vendido": float(total_vendido),
            "num_facturas": num_facturas,
            "ticket_promedio": float(ticket_promedio),
            "total_por_cobrar": float(total_por_cobrar),
            "variacion_vs_periodo_anterior": round(variacion, 2),
        },
        "ventas_diarias": ventas_diarias,
        "top_clientes": top_clientes,
        "por_vendedor": por_vendedor,
    }


# ---------------------------------------------------------------------------
# GET /cobranza
# ---------------------------------------------------------------------------

@router.get("/cobranza")
def reporte_cobranza(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms

    today = date.today()

    facturas = (
        db.query(Factura)
        .options(joinedload(Factura.empresa))
        .filter(
            Factura.fecha >= date_from,
            Factura.fecha <= date_to,
            Factura.estado != "anulada",
        )
        .all()
    )

    factura_ids = [f.id for f in facturas]
    # Build a map of pago totals per factura
    pago_rows = (
        db.query(Pago.factura_id, func.sum(Pago.monto))
        .filter(Pago.factura_id.in_(factura_ids))
        .group_by(Pago.factura_id)
        .all()
    ) if factura_ids else []
    pago_map: dict[int, Decimal] = {fid: Decimal(str(m)) for fid, m in pago_rows}

    total_por_cobrar = _ZERO
    total_vencido = _ZERO
    proximas_a_vencer_7d = _ZERO

    buckets: dict[str, dict] = {
        "d_0_30": {"count": 0, "monto": _ZERO},
        "d_31_60": {"count": 0, "monto": _ZERO},
        "d_61_90": {"count": 0, "monto": _ZERO},
        "d_90_plus": {"count": 0, "monto": _ZERO},
    }

    empresa_map: dict[int, dict] = {}

    for f in facturas:
        saldo = f.total - pago_map.get(f.id, _ZERO)

        if saldo <= _ZERO:
            continue

        total_por_cobrar += saldo

        venc = f.fecha_vencimiento
        if venc is not None:
            if venc < today:
                total_vencido += saldo
                dias_vencida = (today - venc).days
                # Aging bucket (only for overdue)
                if dias_vencida <= 30:
                    buckets["d_0_30"]["count"] += 1
                    buckets["d_0_30"]["monto"] += saldo
                elif dias_vencida <= 60:
                    buckets["d_31_60"]["count"] += 1
                    buckets["d_31_60"]["monto"] += saldo
                elif dias_vencida <= 90:
                    buckets["d_61_90"]["count"] += 1
                    buckets["d_61_90"]["monto"] += saldo
                else:
                    buckets["d_90_plus"]["count"] += 1
                    buckets["d_90_plus"]["monto"] += saldo
            else:
                dias_vencida = 0
                days_until = (venc - today).days
                if 0 <= days_until <= 7:
                    proximas_a_vencer_7d += saldo
        else:
            dias_vencida = 0

        # Aggregate by empresa
        emp_id = f.empresa_id
        if emp_id is not None:
            if emp_id not in empresa_map:
                empresa_map[emp_id] = {
                    "empresa_id": emp_id,
                    "nombre": f.empresa.nombre if f.empresa else "",
                    "saldo": _ZERO,
                    "dias_vencida": dias_vencida,
                }
            else:
                empresa_map[emp_id]["dias_vencida"] = max(
                    empresa_map[emp_id]["dias_vencida"], dias_vencida
                )
            empresa_map[emp_id]["saldo"] += saldo

    # Serialise buckets
    aging_out = {
        k: {"count": v["count"], "monto": float(v["monto"])}
        for k, v in buckets.items()
    }

    por_empresa = [
        {
            "empresa_id": e["empresa_id"],
            "nombre": e["nombre"],
            "saldo": float(e["saldo"]),
            "dias_vencida": e["dias_vencida"],
        }
        for e in sorted(empresa_map.values(), key=lambda x: x["saldo"], reverse=True)
    ]

    return {
        "kpis": {
            "total_por_cobrar": float(total_por_cobrar),
            "total_vencido": float(total_vencido),
            "proximas_a_vencer_7d": float(proximas_a_vencer_7d),
        },
        "aging": aging_out,
        "por_empresa": por_empresa,
    }


# ---------------------------------------------------------------------------
# GET /inventario
# ---------------------------------------------------------------------------

@router.get("/inventario")
def reporte_inventario(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms

    # --- KPIs: current stock (date range ignored) ---
    productos = db.query(Producto).all()

    valor_total_stock = sum(
        (p.precio_costo * p.stock_actual for p in productos), _ZERO
    )
    bajo_minimo_list = [
        p for p in productos
        if p.stock_actual < p.stock_minimo
    ]
    num_sin_stock = sum(1 for p in productos if p.stock_actual <= 0)

    bajo_minimo_out = [
        {
            "producto_id": p.id,
            "nombre": p.nombre,
            "sku": p.sku,
            "stock_actual": p.stock_actual,
            "stock_minimo": p.stock_minimo,
        }
        for p in bajo_minimo_list
    ]

    # --- Top vendidos: use MovimientoInventario in the date range ---
    date_to_dt_exclusive = date_to + timedelta(days=1)

    movimientos = (
        db.query(MovimientoInventario)
        .options(joinedload(MovimientoInventario.producto))
        .filter(
            MovimientoInventario.tipo == "salida",
            MovimientoInventario.created_at >= date_from,
            MovimientoInventario.created_at < date_to_dt_exclusive,
        )
        .all()
    )

    top_map: dict[int, dict] = {}
    for m in movimientos:
        entry = top_map.setdefault(
            m.producto_id,
            {
                "producto_id": m.producto_id,
                "nombre": m.producto.nombre if m.producto else "",
                "cantidad_vendida": 0,
                "monto_total": _ZERO,
            },
        )
        entry["cantidad_vendida"] += m.cantidad
        precio = m.producto.precio_venta if m.producto else _ZERO
        entry["monto_total"] += Decimal(str(m.cantidad)) * precio

    top_vendidos = sorted(top_map.values(), key=lambda x: x["cantidad_vendida"], reverse=True)[:10]
    for t in top_vendidos:
        t["monto_total"] = float(t["monto_total"])

    return {
        "kpis": {
            "valor_total_stock": float(valor_total_stock),
            "num_bajo_minimo": len(bajo_minimo_list),
            "num_sin_stock": num_sin_stock,
        },
        "bajo_minimo": bajo_minimo_out,
        "top_vendidos": top_vendidos,
    }


# ---------------------------------------------------------------------------
# GET /compras
# ---------------------------------------------------------------------------

_OC_PENDING_ESTADOS = {"borrador", "enviada", "pendiente"}

_DTE_LABELS: dict[str, str] = {
    "033": "Factura",
    "061": "Nota Crédito",
    "056": "Nota Débito",
}


@router.get("/compras")
def reporte_compras(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms

    ocs: list[OrdenCompra] = (
        db.query(OrdenCompra)
        .options(joinedload(OrdenCompra.proveedor))
        .filter(
            OrdenCompra.fecha >= date_from,
            OrdenCompra.fecha <= date_to,
        )
        .all()
    )

    total_comprado = sum((oc.total for oc in ocs), _ZERO)
    num_oc_emitidas = len(ocs)
    num_oc_pendientes = sum(1 for oc in ocs if oc.estado in _OC_PENDING_ESTADOS)

    # --- Por proveedor ---
    prov_map: dict[int, dict] = {}
    for oc in ocs:
        if oc.proveedor_id is None:
            continue
        pid = oc.proveedor_id
        entry = prov_map.setdefault(
            pid,
            {
                "proveedor_id": pid,
                "nombre": oc.proveedor.nombre if oc.proveedor else "",
                "total": _ZERO,
                "num_oc": 0,
            },
        )
        entry["total"] += oc.total
        entry["num_oc"] += 1
    por_proveedor = sorted(prov_map.values(), key=lambda x: x["total"], reverse=True)
    for p in por_proveedor:
        p["total"] = float(p["total"])

    # --- Por estado ---
    estado_map: dict[str, dict] = {}
    for oc in ocs:
        entry = estado_map.setdefault(
            oc.estado,
            {"estado": oc.estado, "count": 0, "total": _ZERO},
        )
        entry["count"] += 1
        entry["total"] += oc.total
    por_estado = sorted(estado_map.values(), key=lambda x: x["count"], reverse=True)
    for e in por_estado:
        e["total"] = float(e["total"])

    return {
        "kpis": {
            "total_comprado": float(total_comprado),
            "num_oc_emitidas": num_oc_emitidas,
            "num_oc_pendientes": num_oc_pendientes,
        },
        "por_proveedor": por_proveedor,
        "por_estado": por_estado,
    }


# ---------------------------------------------------------------------------
# GET /margenes
# ---------------------------------------------------------------------------

@router.get("/margenes")
def reporte_margenes(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    current_user, db = perms

    base_q = (
        db.query(Factura)
        .options(joinedload(Factura.lineas))
        .filter(
            Factura.fecha >= date_from,
            Factura.fecha <= date_to,
            Factura.estado != "anulada",
        )
    )
    if current_user.role == "vendedor":
        base_q = base_q.filter(Factura.vendedor_id == current_user.id)

    facturas: list[Factura] = base_q.all()

    # --- Por factura ---
    por_factura_list: list[dict] = []
    all_factura_margen_pcts: list[float] = []

    # --- Por producto (grouped by producto_id) ---
    prod_map: dict[int, dict] = {}

    for f in facturas:
        lineas_con_margen = [ln for ln in f.lineas if ln.margen is not None]
        total_neto_f = sum((ln.valor_neto for ln in f.lineas), _ZERO)
        margen_total = sum(
            (ln.margen * ln.valor_neto for ln in lineas_con_margen), _ZERO
        )
        if total_neto_f > _ZERO:
            margen_pct_f = float(margen_total / total_neto_f * 100)
        else:
            margen_pct_f = 0.0

        all_factura_margen_pcts.append(margen_pct_f)
        por_factura_list.append(
            {
                "factura_id": f.id,
                "numero": f.numero,
                "total": float(f.total),
                "margen_total": float(margen_total),
                "margen_pct": round(margen_pct_f, 2),
            }
        )

        # Aggregate per producto (producto_id)
        for ln in f.lineas:
            if ln.margen is None:
                continue
            if ln.producto_id is None:
                continue
            key = ln.producto_id
            entry = prod_map.setdefault(
                key,
                {
                    "producto_id": ln.producto_id,
                    "nombre": ln.descripcion,
                    "cantidad_vendida": 0,
                    "precio_venta_sum": _ZERO,
                    "precio_venta_count": 0,
                    "margen_pct_sum": Decimal("0"),
                    "margen_pct_count": 0,
                },
            )
            entry["cantidad_vendida"] += ln.cantidad
            if ln.cantidad > 0:
                entry["precio_venta_sum"] += ln.valor_neto / ln.cantidad
                entry["precio_venta_count"] += 1
            entry["margen_pct_sum"] += ln.margen * 100
            entry["margen_pct_count"] += 1

    # Flatten por_producto
    por_producto: list[dict] = []
    for entry in prod_map.values():
        avg_venta = (
            float(entry["precio_venta_sum"] / entry["precio_venta_count"])
            if entry["precio_venta_count"] > 0
            else 0.0
        )
        avg_margen = (
            float(entry["margen_pct_sum"] / entry["margen_pct_count"])
            if entry["margen_pct_count"] > 0
            else 0.0
        )
        por_producto.append(
            {
                "producto_id": entry["producto_id"],
                "nombre": entry["nombre"],
                "cantidad_vendida": entry["cantidad_vendida"],
                "precio_costo_promedio": 0.0,
                "precio_venta_promedio": round(avg_venta, 2),
                "margen_pct": round(avg_margen, 2),
            }
        )
    por_producto.sort(key=lambda x: x["margen_pct"], reverse=True)

    # --- KPIs ---
    margen_promedio_pct = (
        sum(all_factura_margen_pcts) / len(all_factura_margen_pcts)
        if all_factura_margen_pcts
        else 0.0
    )
    mejor_producto = por_producto[0] if por_producto else None
    peor_producto = por_producto[-1] if por_producto else None

    kpis: dict = {
        "margen_promedio_pct": round(margen_promedio_pct, 2),
        "mejor_producto": (
            {"nombre": mejor_producto["nombre"], "margen_pct": mejor_producto["margen_pct"]}
            if mejor_producto
            else None
        ),
        "peor_producto": (
            {"nombre": peor_producto["nombre"], "margen_pct": peor_producto["margen_pct"]}
            if peor_producto
            else None
        ),
    }

    return {
        "kpis": kpis,
        "por_producto": por_producto,
        "por_factura": por_factura_list,
    }


# ---------------------------------------------------------------------------
# GET /dte
# ---------------------------------------------------------------------------

_DTE_PENDING_ESTADOS = {"pendiente", "procesando"}


@router.get("/dte")
def reporte_dte(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms

    date_to_exclusive = date_to + timedelta(days=1)

    emisiones: list[DteEmision] = (
        db.query(DteEmision)
        .filter(
            DteEmision.created_at >= date_from,
            DteEmision.created_at < date_to_exclusive,
        )
        .all()
    )

    total_emitidos = len(emisiones)
    aceptadas = sum(1 for e in emisiones if e.estado == "aceptada")
    rechazadas = sum(1 for e in emisiones if e.estado == "rechazada")
    pendientes = sum(1 for e in emisiones if e.estado in _DTE_PENDING_ESTADOS)

    # --- Por tipo ---
    tipo_map: dict[str, dict] = {}
    for e in emisiones:
        entry = tipo_map.setdefault(
            e.tipo,
            {
                "tipo": e.tipo,
                "label": _DTE_LABELS.get(e.tipo, e.tipo),
                "count": 0,
                "aceptadas": 0,
            },
        )
        entry["count"] += 1
        if e.estado == "aceptada":
            entry["aceptadas"] += 1
    por_tipo = sorted(tipo_map.values(), key=lambda x: x["count"], reverse=True)

    # --- Emisiones list ---
    emisiones_out: list[dict] = []
    for e in emisiones:
        detalle_rechazo: str | None = None
        if e.estado == "rechazada" and e.respuesta_sii is not None:
            detalle_rechazo = e.respuesta_sii.get("detalle")
        emisiones_out.append(
            {
                "id": e.id,
                "tipo": e.tipo,
                "folio": e.folio,
                "estado": e.estado,
                "monto_total": e.monto_total,
                "created_at": e.created_at.date().isoformat(),
                "detalle_rechazo": detalle_rechazo,
            }
        )

    return {
        "kpis": {
            "total_emitidos": total_emitidos,
            "aceptadas": aceptadas,
            "rechazadas": rechazadas,
            "pendientes": pendientes,
        },
        "por_tipo": por_tipo,
        "emisiones": emisiones_out,
    }


# ---------------------------------------------------------------------------
# Excel export helpers (shared data fetchers)
# ---------------------------------------------------------------------------

def _get_inventario(date_from: date, date_to: date, db: Session) -> dict:
    """Return inventario data dict (same logic as reporte_inventario)."""
    productos = db.query(Producto).all()
    valor_total_stock = sum((p.precio_costo * p.stock_actual for p in productos), _ZERO)
    bajo_minimo_list = [p for p in productos if p.stock_actual < p.stock_minimo]
    num_sin_stock = sum(1 for p in productos if p.stock_actual <= 0)

    date_to_dt_exclusive = date_to + timedelta(days=1)
    movimientos = (
        db.query(MovimientoInventario)
        .options(joinedload(MovimientoInventario.producto))
        .filter(
            MovimientoInventario.tipo == "salida",
            MovimientoInventario.created_at >= date_from,
            MovimientoInventario.created_at < date_to_dt_exclusive,
        )
        .all()
    )

    top_map: dict[int, dict] = {}
    for m in movimientos:
        entry = top_map.setdefault(
            m.producto_id,
            {
                "producto_id": m.producto_id,
                "nombre": m.producto.nombre if m.producto else "",
                "cantidad_vendida": 0,
                "monto_total": _ZERO,
            },
        )
        entry["cantidad_vendida"] += m.cantidad
        precio = m.producto.precio_venta if m.producto else _ZERO
        entry["monto_total"] += Decimal(str(m.cantidad)) * precio

    top_vendidos = sorted(top_map.values(), key=lambda x: x["cantidad_vendida"], reverse=True)[:10]
    for t in top_vendidos:
        t["monto_total"] = float(t["monto_total"])

    return {
        "productos": productos,
        "top_vendidos": top_vendidos,
        "kpis": {
            "valor_total_stock": float(valor_total_stock),
            "num_bajo_minimo": len(bajo_minimo_list),
            "num_sin_stock": num_sin_stock,
        },
    }


def _get_margenes(date_from: date, date_to: date, db: Session, vendedor_id: int | None) -> dict:
    """Return margenes data dict (same logic as reporte_margenes)."""
    base_q = (
        db.query(Factura)
        .options(joinedload(Factura.lineas))
        .filter(
            Factura.fecha >= date_from,
            Factura.fecha <= date_to,
            Factura.estado != "anulada",
        )
    )
    if vendedor_id is not None:
        base_q = base_q.filter(Factura.vendedor_id == vendedor_id)

    facturas: list[Factura] = base_q.all()

    por_factura_list: list[dict] = []
    prod_map: dict[int, dict] = {}

    for f in facturas:
        lineas_con_margen = [ln for ln in f.lineas if ln.margen is not None]
        total_neto_f = sum((ln.valor_neto for ln in f.lineas), _ZERO)
        margen_total = sum(
            (ln.margen * ln.valor_neto for ln in lineas_con_margen), _ZERO
        )
        if total_neto_f > _ZERO:
            margen_pct_f = float(margen_total / total_neto_f * 100)
        else:
            margen_pct_f = 0.0

        por_factura_list.append(
            {
                "factura_id": f.id,
                "numero": f.numero,
                "total": float(f.total),
                "margen_total": float(margen_total),
                "margen_pct": round(margen_pct_f, 2),
            }
        )

        for ln in f.lineas:
            if ln.margen is None or ln.producto_id is None:
                continue
            key = ln.producto_id
            entry = prod_map.setdefault(
                key,
                {
                    "nombre": ln.descripcion,
                    "cantidad_vendida": 0,
                    "precio_venta_sum": _ZERO,
                    "precio_venta_count": 0,
                    "margen_pct_sum": Decimal("0"),
                    "margen_pct_count": 0,
                    "precio_costo_sum": _ZERO,
                    "precio_costo_count": 0,
                },
            )
            entry["cantidad_vendida"] += ln.cantidad
            if ln.cantidad > 0:
                entry["precio_venta_sum"] += ln.valor_neto / ln.cantidad
                entry["precio_venta_count"] += 1
            entry["margen_pct_sum"] += ln.margen * 100
            entry["margen_pct_count"] += 1

    por_producto: list[dict] = []
    for entry in prod_map.values():
        avg_venta = (
            float(entry["precio_venta_sum"] / entry["precio_venta_count"])
            if entry["precio_venta_count"] > 0
            else 0.0
        )
        avg_margen = (
            float(entry["margen_pct_sum"] / entry["margen_pct_count"])
            if entry["margen_pct_count"] > 0
            else 0.0
        )
        por_producto.append(
            {
                "nombre": entry["nombre"],
                "cantidad_vendida": entry["cantidad_vendida"],
                "precio_costo_promedio": 0.0,
                "precio_venta_promedio": round(avg_venta, 2),
                "margen_pct": round(avg_margen, 2),
            }
        )

    return {"por_producto": por_producto, "por_factura": por_factura_list}


# ---------------------------------------------------------------------------
# GET /ventas/export/excel
# ---------------------------------------------------------------------------

@router.get("/ventas/export/excel")
def exportar_ventas_excel(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    current_user, db = perms

    base_q = (
        db.query(Factura)
        .options(joinedload(Factura.cliente), joinedload(Factura.vendedor))
        .filter(
            Factura.fecha >= date_from,
            Factura.fecha <= date_to,
            Factura.estado != "anulada",
        )
    )
    if current_user.role == "vendedor":
        base_q = base_q.filter(Factura.vendedor_id == current_user.id)

    facturas = base_q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(["Fecha", "Cliente", "Vendedor", "Total Neto", "IVA", "Total", "Estado"])
    for f in facturas:
        total_neto = round(float(f.total) / 1.19, 2)
        iva = round(float(f.total) - total_neto, 2)
        ws.append([
            f.fecha.strftime("%d/%m/%Y") if f.fecha else "",
            f.cliente.nombre if f.cliente else "",
            f.vendedor.name if f.vendedor else "",
            total_neto,
            iva,
            float(f.total),
            f.estado,
        ])

    return _excel_response(wb, "ventas", date_from, date_to)


# ---------------------------------------------------------------------------
# GET /cobranza/export/excel
# ---------------------------------------------------------------------------

@router.get("/cobranza/export/excel")
def exportar_cobranza_excel(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms

    today = date.today()

    facturas = (
        db.query(Factura)
        .options(joinedload(Factura.empresa))
        .filter(
            Factura.fecha >= date_from,
            Factura.fecha <= date_to,
            Factura.estado != "anulada",
        )
        .all()
    )

    factura_ids = [f.id for f in facturas]
    pago_rows = (
        db.query(Pago.factura_id, func.sum(Pago.monto))
        .filter(Pago.factura_id.in_(factura_ids))
        .group_by(Pago.factura_id)
        .all()
    ) if factura_ids else []
    pago_map: dict[int, Decimal] = {fid: Decimal(str(m)) for fid, m in pago_rows}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cobranza"
    ws.append(["Empresa", "N° Factura", "Fecha Vencimiento", "Total", "Pagado", "Saldo", "Días Vencida"])

    for f in facturas:
        pagado = pago_map.get(f.id, _ZERO)
        saldo = f.total - pagado
        if saldo <= _ZERO:
            continue

        venc = f.fecha_vencimiento
        if venc is not None and venc < today:
            dias_vencida = (today - venc).days
        else:
            dias_vencida = 0

        ws.append([
            f.empresa.nombre if f.empresa else "",
            f.numero,
            venc.strftime("%d/%m/%Y") if venc else "",
            float(f.total),
            float(pagado),
            float(saldo),
            dias_vencida,
        ])

    return _excel_response(wb, "cobranza", date_from, date_to)


# ---------------------------------------------------------------------------
# GET /inventario/export/excel
# ---------------------------------------------------------------------------

@router.get("/inventario/export/excel")
def exportar_inventario_excel(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms

    data = _get_inventario(date_from, date_to, db)
    productos = data["productos"]
    top_vendidos = data["top_vendidos"]

    wb = openpyxl.Workbook()

    # Sheet 1: Stock Actual
    ws1 = wb.active
    ws1.title = "Stock Actual"
    ws1.append(["SKU", "Nombre", "Stock Actual", "Stock Mínimo", "Precio Costo", "Precio Venta", "Valor Stock"])
    for p in productos:
        ws1.append([
            p.sku or "",
            p.nombre,
            p.stock_actual,
            p.stock_minimo,
            float(p.precio_costo),
            float(p.precio_venta),
            float(p.precio_costo * p.stock_actual),
        ])

    # Sheet 2: Top Vendidos
    ws2 = wb.create_sheet("Top Vendidos")
    ws2.append(["Nombre", "Cantidad Vendida", "Monto Total"])
    for t in top_vendidos:
        ws2.append([
            t["nombre"],
            t["cantidad_vendida"],
            t["monto_total"],
        ])

    return _excel_response(wb, "inventario", date_from, date_to)


# ---------------------------------------------------------------------------
# GET /compras/export/excel
# ---------------------------------------------------------------------------

@router.get("/compras/export/excel")
def exportar_compras_excel(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms

    ocs: list[OrdenCompra] = (
        db.query(OrdenCompra)
        .options(joinedload(OrdenCompra.proveedor))
        .filter(
            OrdenCompra.fecha >= date_from,
            OrdenCompra.fecha <= date_to,
        )
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"
    ws.append(["N° OC", "Fecha", "Proveedor", "Estado", "Total Neto", "IVA", "Total"])

    for oc in ocs:
        total_neto = round(float(oc.total) / 1.19, 2)
        iva = round(float(oc.total) - total_neto, 2)
        ws.append([
            oc.numero,
            oc.fecha.strftime("%d/%m/%Y") if oc.fecha else "",
            oc.proveedor.nombre if oc.proveedor else "",
            oc.estado,
            total_neto,
            iva,
            float(oc.total),
        ])

    return _excel_response(wb, "compras", date_from, date_to)


# ---------------------------------------------------------------------------
# GET /margenes/export/excel
# ---------------------------------------------------------------------------

@router.get("/margenes/export/excel")
def exportar_margenes_excel(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    current_user, db = perms

    vendedor_id = current_user.id if current_user.role == "vendedor" else None
    data = _get_margenes(date_from, date_to, db, vendedor_id)
    por_producto = data["por_producto"]
    por_factura = data["por_factura"]

    wb = openpyxl.Workbook()

    # Sheet 1: Márgenes por Producto
    ws1 = wb.active
    ws1.title = "Márgenes por Producto"
    ws1.append(["Nombre", "Cantidad Vendida", "Precio Costo Prom.", "Precio Venta Prom.", "Margen %"])
    for p in por_producto:
        ws1.append([
            p["nombre"],
            p["cantidad_vendida"],
            p["precio_costo_promedio"],
            p["precio_venta_promedio"],
            p["margen_pct"],
        ])

    # Sheet 2: Márgenes por Factura
    ws2 = wb.create_sheet("Márgenes por Factura")
    ws2.append(["N° Factura", "Total", "Margen Total", "Margen %"])
    for f in por_factura:
        ws2.append([
            f["numero"],
            f["total"],
            f["margen_total"],
            f["margen_pct"],
        ])

    return _excel_response(wb, "margenes", date_from, date_to)


# ---------------------------------------------------------------------------
# GET /dte/export/excel
# ---------------------------------------------------------------------------

@router.get("/dte/export/excel")
def exportar_dte_excel(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms

    date_to_exclusive = date_to + timedelta(days=1)

    emisiones: list[DteEmision] = (
        db.query(DteEmision)
        .filter(
            DteEmision.created_at >= date_from,
            DteEmision.created_at < date_to_exclusive,
        )
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DTE"
    ws.append(["Tipo", "Folio", "TrackID", "Estado", "Monto Total", "Fecha Emisión", "Fecha Aceptación"])

    for e in emisiones:
        tipo_label = _DTE_LABELS.get(e.tipo, e.tipo)
        fecha_emision = e.created_at.date().strftime("%d/%m/%Y") if e.created_at is not None else ""
        fecha_aceptacion = e.aceptado_at.strftime("%d/%m/%Y") if e.aceptado_at is not None else ""
        ws.append([
            tipo_label,
            e.folio,
            e.track_id or "",
            e.estado,
            e.monto_total,
            fecha_emision,
            fecha_aceptacion,
        ])

    return _excel_response(wb, "dte", date_from, date_to)


# ---------------------------------------------------------------------------
# Data helpers for ventas, cobranza, compras, dte (used by PDF + future use)
# ---------------------------------------------------------------------------

def _get_ventas(date_from: date, date_to: date, db: Session, vendedor_id: int | None) -> dict:
    """Return ventas data dict."""
    base_q = (
        db.query(Factura)
        .options(joinedload(Factura.cliente), joinedload(Factura.vendedor))
        .filter(
            Factura.fecha >= date_from,
            Factura.fecha <= date_to,
            Factura.estado != "anulada",
        )
    )
    if vendedor_id is not None:
        base_q = base_q.filter(Factura.vendedor_id == vendedor_id)

    facturas = base_q.all()
    factura_ids = [f.id for f in facturas]
    total_vendido = sum((f.total for f in facturas), _ZERO)
    num_facturas = len(facturas)
    ticket_promedio = (total_vendido / num_facturas) if num_facturas else _ZERO
    total_pagado = _sum_pagos_for_ids(db, factura_ids)
    total_por_cobrar = max(total_vendido - total_pagado, _ZERO)

    duration_days = (date_to - date_from).days + 1
    prev_to = date_from - timedelta(days=1)
    prev_from = prev_to - timedelta(days=duration_days - 1)
    prev_q = db.query(func.sum(Factura.total)).filter(
        Factura.fecha >= prev_from,
        Factura.fecha <= prev_to,
        Factura.estado != "anulada",
    )
    if vendedor_id is not None:
        prev_q = prev_q.filter(Factura.vendedor_id == vendedor_id)
    prev_total_raw = prev_q.scalar()
    prev_total = Decimal(str(prev_total_raw)) if prev_total_raw else _ZERO
    variacion = (
        float(((total_vendido - prev_total) / prev_total) * 100) if prev_total != _ZERO else 0.0
    )

    cliente_map: dict[int, dict] = {}
    for f in facturas:
        if f.cliente_id is None:
            continue
        entry = cliente_map.setdefault(
            f.cliente_id,
            {"nombre": f.cliente.nombre if f.cliente else "", "total": _ZERO, "num_facturas": 0},
        )
        entry["total"] += f.total
        entry["num_facturas"] += 1
    top_clientes = sorted(cliente_map.values(), key=lambda x: x["total"], reverse=True)[:10]
    for c in top_clientes:
        c["total"] = float(c["total"])

    vendedor_map: dict[int, dict] = {}
    for f in facturas:
        if f.vendedor_id is None:
            continue
        entry = vendedor_map.setdefault(
            f.vendedor_id,
            {"nombre": f.vendedor.name if f.vendedor else "", "total": _ZERO, "num_facturas": 0},
        )
        entry["total"] += f.total
        entry["num_facturas"] += 1
    por_vendedor = sorted(vendedor_map.values(), key=lambda x: x["total"], reverse=True)
    for v in por_vendedor:
        v["total"] = float(v["total"])

    return {
        "kpis": {
            "total_vendido": float(total_vendido),
            "num_facturas": num_facturas,
            "ticket_promedio": float(ticket_promedio),
            "total_por_cobrar": float(total_por_cobrar),
            "variacion_vs_periodo_anterior": round(variacion, 2),
        },
        "top_clientes": top_clientes,
        "por_vendedor": por_vendedor,
    }


def _get_cobranza(date_from: date, date_to: date, db: Session) -> dict:
    """Return cobranza data dict."""
    today = date.today()
    facturas = (
        db.query(Factura)
        .options(joinedload(Factura.empresa))
        .filter(
            Factura.fecha >= date_from,
            Factura.fecha <= date_to,
            Factura.estado != "anulada",
        )
        .all()
    )
    factura_ids = [f.id for f in facturas]
    pago_rows = (
        db.query(Pago.factura_id, func.sum(Pago.monto))
        .filter(Pago.factura_id.in_(factura_ids))
        .group_by(Pago.factura_id)
        .all()
    ) if factura_ids else []
    pago_map: dict[int, Decimal] = {fid: Decimal(str(m)) for fid, m in pago_rows}

    total_por_cobrar = _ZERO
    total_vencido = _ZERO
    proximas_a_vencer_7d = _ZERO
    buckets: dict[str, dict] = {
        "d_0_30": {"label": "0-30d", "count": 0, "monto": _ZERO},
        "d_31_60": {"label": "31-60d", "count": 0, "monto": _ZERO},
        "d_61_90": {"label": "61-90d", "count": 0, "monto": _ZERO},
        "d_90_plus": {"label": "90+d", "count": 0, "monto": _ZERO},
    }
    empresa_map: dict[int, dict] = {}

    for f in facturas:
        saldo = f.total - pago_map.get(f.id, _ZERO)
        if saldo <= _ZERO:
            continue
        total_por_cobrar += saldo
        venc = f.fecha_vencimiento
        if venc is not None and venc < today:
            total_vencido += saldo
            dias_vencida = (today - venc).days
            if dias_vencida <= 30:
                buckets["d_0_30"]["count"] += 1
                buckets["d_0_30"]["monto"] += saldo
            elif dias_vencida <= 60:
                buckets["d_31_60"]["count"] += 1
                buckets["d_31_60"]["monto"] += saldo
            elif dias_vencida <= 90:
                buckets["d_61_90"]["count"] += 1
                buckets["d_61_90"]["monto"] += saldo
            else:
                buckets["d_90_plus"]["count"] += 1
                buckets["d_90_plus"]["monto"] += saldo
        else:
            dias_vencida = 0
            if venc is not None and 0 <= (venc - today).days <= 7:
                proximas_a_vencer_7d += saldo

        emp_id = f.empresa_id
        if emp_id is not None:
            if emp_id not in empresa_map:
                empresa_map[emp_id] = {
                    "nombre": f.empresa.nombre if f.empresa else "",
                    "saldo": _ZERO,
                    "dias_vencida": dias_vencida,
                }
            else:
                empresa_map[emp_id]["dias_vencida"] = max(
                    empresa_map[emp_id]["dias_vencida"], dias_vencida
                )
            empresa_map[emp_id]["saldo"] += saldo

    por_empresa = sorted(
        [
            {"nombre": v["nombre"], "saldo": float(v["saldo"]), "dias_vencida": v["dias_vencida"]}
            for v in empresa_map.values()
        ],
        key=lambda x: x["saldo"],
        reverse=True,
    )

    return {
        "kpis": {
            "total_por_cobrar": float(total_por_cobrar),
            "total_vencido": float(total_vencido),
            "proximas_a_vencer_7d": float(proximas_a_vencer_7d),
        },
        "buckets": buckets,
        "por_empresa": por_empresa,
    }


def _get_compras(date_from: date, date_to: date, db: Session) -> dict:
    """Return compras data dict."""
    ocs: list[OrdenCompra] = (
        db.query(OrdenCompra)
        .options(joinedload(OrdenCompra.proveedor))
        .filter(
            OrdenCompra.fecha >= date_from,
            OrdenCompra.fecha <= date_to,
        )
        .all()
    )
    total_comprado = sum((oc.total for oc in ocs), _ZERO)
    num_oc_emitidas = len(ocs)
    num_oc_pendientes = sum(1 for oc in ocs if oc.estado in _OC_PENDING_ESTADOS)

    prov_map: dict[int, dict] = {}
    for oc in ocs:
        if oc.proveedor_id is None:
            continue
        entry = prov_map.setdefault(
            oc.proveedor_id,
            {"nombre": oc.proveedor.nombre if oc.proveedor else "", "total": _ZERO, "num_oc": 0},
        )
        entry["total"] += oc.total
        entry["num_oc"] += 1
    por_proveedor = sorted(prov_map.values(), key=lambda x: x["total"], reverse=True)
    for p in por_proveedor:
        p["total"] = float(p["total"])

    estado_map: dict[str, dict] = {}
    for oc in ocs:
        entry = estado_map.setdefault(
            oc.estado, {"estado": oc.estado, "count": 0, "total": _ZERO}
        )
        entry["count"] += 1
        entry["total"] += oc.total
    por_estado = sorted(estado_map.values(), key=lambda x: x["count"], reverse=True)
    for e in por_estado:
        e["total"] = float(e["total"])

    return {
        "kpis": {
            "total_comprado": float(total_comprado),
            "num_oc_emitidas": num_oc_emitidas,
            "num_oc_pendientes": num_oc_pendientes,
        },
        "por_proveedor": por_proveedor,
        "por_estado": por_estado,
    }


def _get_dte(date_from: date, date_to: date, db: Session) -> dict:
    """Return DTE data dict."""
    date_to_exclusive = date_to + timedelta(days=1)
    emisiones: list[DteEmision] = (
        db.query(DteEmision)
        .filter(
            DteEmision.created_at >= date_from,
            DteEmision.created_at < date_to_exclusive,
        )
        .all()
    )
    total_emitidos = len(emisiones)
    aceptadas = sum(1 for e in emisiones if e.estado == "aceptada")
    rechazadas = sum(1 for e in emisiones if e.estado == "rechazada")
    pendientes = sum(1 for e in emisiones if e.estado in _DTE_PENDING_ESTADOS)

    emisiones_out = [
        {
            "tipo": _DTE_LABELS.get(e.tipo, e.tipo),
            "folio": e.folio,
            "estado": e.estado,
            "monto_total": e.monto_total,
            "fecha": e.created_at.date().isoformat() if e.created_at else "",
        }
        for e in emisiones
    ]

    return {
        "kpis": {
            "total_emitidos": total_emitidos,
            "aceptadas": aceptadas,
            "rechazadas": rechazadas,
            "pendientes": pendientes,
        },
        "emisiones": emisiones_out,
    }


# ---------------------------------------------------------------------------
# PDF export utilities
# ---------------------------------------------------------------------------

_PDF_STYLE = """
<style>
  body { font-family: Arial, sans-serif; font-size: 12px; color: #1f2937; margin: 2cm; }
  h1 { font-size: 18px; margin-bottom: 4px; }
  .period { color: #6b7280; font-size: 11px; margin-bottom: 20px; }
  .kpis { display: flex; gap: 16px; margin-bottom: 24px; flex-wrap: wrap; }
  .kpi { border: 1px solid #e5e7eb; border-radius: 8px; padding: 12px 16px; min-width: 120px; }
  .kpi-label { font-size: 10px; color: #6b7280; text-transform: uppercase; }
  .kpi-value { font-size: 16px; font-weight: 700; margin-top: 2px; }
  h2 { font-size: 13px; margin: 20px 0 8px; border-bottom: 1px solid #e5e7eb; padding-bottom: 4px; }
  table { width: 100%; border-collapse: collapse; font-size: 11px; }
  th { background: #f9fafb; text-align: left; padding: 6px 8px; border-bottom: 2px solid #e5e7eb; }
  td { padding: 5px 8px; border-bottom: 1px solid #f3f4f6; }
</style>
"""


def _fmt(n: float | int) -> str:
    return "$" + f"{int(n):,}".replace(",", ".")


def _pdf_response(html: str, tab: str, date_from: date, date_to: date) -> StreamingResponse:
    pdf_bytes = HTML(string=html).write_pdf()
    filename = f"{tab}-{date_from}-{date_to}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# GET /ventas/export/pdf
# ---------------------------------------------------------------------------

@router.get("/ventas/export/pdf")
def exportar_ventas_pdf(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    current_user, db = perms
    vendedor_id = current_user.id if current_user.role == "vendedor" else None
    data = _get_ventas(date_from, date_to, db, vendedor_id)
    kpis = data["kpis"]

    clientes_rows = "".join(
        f"<tr><td>{c['nombre']}</td><td>{c['num_facturas']}</td><td>{_fmt(c['total'])}</td></tr>"
        for c in data["top_clientes"]
    )
    vendedor_rows = "".join(
        f"<tr><td>{v['nombre']}</td><td>{v['num_facturas']}</td><td>{_fmt(v['total'])}</td></tr>"
        for v in data["por_vendedor"]
    )

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">{_PDF_STYLE}</head><body>
<h1>Reporte de Ventas</h1>
<div class="period">{date_from} — {date_to}</div>
<div class="kpis">
  <div class="kpi"><div class="kpi-label">Total Vendido</div><div class="kpi-value">{_fmt(kpis['total_vendido'])}</div></div>
  <div class="kpi"><div class="kpi-label">N° Facturas</div><div class="kpi-value">{kpis['num_facturas']}</div></div>
  <div class="kpi"><div class="kpi-label">Ticket Promedio</div><div class="kpi-value">{_fmt(kpis['ticket_promedio'])}</div></div>
  <div class="kpi"><div class="kpi-label">Por Cobrar</div><div class="kpi-value">{_fmt(kpis['total_por_cobrar'])}</div></div>
  <div class="kpi"><div class="kpi-label">Variación</div><div class="kpi-value">{kpis['variacion_vs_periodo_anterior']}%</div></div>
</div>
<h2>Top Clientes</h2>
<table><thead><tr><th>Nombre</th><th>Facturas</th><th>Total</th></tr></thead>
<tbody>{clientes_rows}</tbody></table>
<h2>Por Vendedor</h2>
<table><thead><tr><th>Vendedor</th><th>Facturas</th><th>Total</th></tr></thead>
<tbody>{vendedor_rows}</tbody></table>
</body></html>"""

    return _pdf_response(html, "ventas", date_from, date_to)


# ---------------------------------------------------------------------------
# GET /cobranza/export/pdf
# ---------------------------------------------------------------------------

@router.get("/cobranza/export/pdf")
def exportar_cobranza_pdf(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms
    data = _get_cobranza(date_from, date_to, db)
    kpis = data["kpis"]
    buckets = data["buckets"]

    aging_rows = "".join(
        f"<tr><td>{v['label']}</td><td>{v['count']}</td><td>{_fmt(float(v['monto']))}</td></tr>"
        for v in buckets.values()
    )
    empresa_rows = "".join(
        f"<tr><td>{e['nombre']}</td><td>{_fmt(e['saldo'])}</td><td>{e['dias_vencida']}</td></tr>"
        for e in data["por_empresa"]
    )

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">{_PDF_STYLE}</head><body>
<h1>Reporte de Cobranza</h1>
<div class="period">{date_from} — {date_to}</div>
<div class="kpis">
  <div class="kpi"><div class="kpi-label">Por Cobrar</div><div class="kpi-value">{_fmt(kpis['total_por_cobrar'])}</div></div>
  <div class="kpi"><div class="kpi-label">Vencido</div><div class="kpi-value">{_fmt(kpis['total_vencido'])}</div></div>
  <div class="kpi"><div class="kpi-label">Vence en 7d</div><div class="kpi-value">{_fmt(kpis['proximas_a_vencer_7d'])}</div></div>
</div>
<h2>Aging</h2>
<table><thead><tr><th>Tramo</th><th>Facturas</th><th>Monto</th></tr></thead>
<tbody>{aging_rows}</tbody></table>
<h2>Por Empresa</h2>
<table><thead><tr><th>Empresa</th><th>Saldo</th><th>Días vencida</th></tr></thead>
<tbody>{empresa_rows}</tbody></table>
</body></html>"""

    return _pdf_response(html, "cobranza", date_from, date_to)


# ---------------------------------------------------------------------------
# GET /inventario/export/pdf
# ---------------------------------------------------------------------------

@router.get("/inventario/export/pdf")
def exportar_inventario_pdf(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms
    data = _get_inventario(date_from, date_to, db)
    kpis = data["kpis"]

    productos = data["productos"]
    bajo_minimo = [p for p in productos if p.stock_actual < p.stock_minimo]
    bajo_rows = "".join(
        f"<tr><td>{p.sku or ''}</td><td>{p.nombre}</td><td>{p.stock_actual}</td><td>{p.stock_minimo}</td></tr>"
        for p in bajo_minimo
    )
    top_rows = "".join(
        f"<tr><td>{t['nombre']}</td><td>{t['cantidad_vendida']}</td><td>{_fmt(t['monto_total'])}</td></tr>"
        for t in data["top_vendidos"]
    )

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">{_PDF_STYLE}</head><body>
<h1>Reporte de Inventario</h1>
<div class="period">{date_from} — {date_to}</div>
<div class="kpis">
  <div class="kpi"><div class="kpi-label">Valor Stock</div><div class="kpi-value">{_fmt(kpis['valor_total_stock'])}</div></div>
  <div class="kpi"><div class="kpi-label">Bajo Mínimo</div><div class="kpi-value">{kpis['num_bajo_minimo']}</div></div>
  <div class="kpi"><div class="kpi-label">Sin Stock</div><div class="kpi-value">{kpis['num_sin_stock']}</div></div>
</div>
<h2>Bajo Mínimo</h2>
<table><thead><tr><th>SKU</th><th>Nombre</th><th>Stock Actual</th><th>Stock Mínimo</th></tr></thead>
<tbody>{bajo_rows}</tbody></table>
<h2>Top Vendidos</h2>
<table><thead><tr><th>Nombre</th><th>Cantidad</th><th>Monto</th></tr></thead>
<tbody>{top_rows}</tbody></table>
</body></html>"""

    return _pdf_response(html, "inventario", date_from, date_to)


# ---------------------------------------------------------------------------
# GET /compras/export/pdf
# ---------------------------------------------------------------------------

@router.get("/compras/export/pdf")
def exportar_compras_pdf(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms
    data = _get_compras(date_from, date_to, db)
    kpis = data["kpis"]

    prov_rows = "".join(
        f"<tr><td>{p['nombre']}</td><td>{p['num_oc']}</td><td>{_fmt(p['total'])}</td></tr>"
        for p in data["por_proveedor"]
    )
    estado_rows = "".join(
        f"<tr><td>{e['estado']}</td><td>{e['count']}</td><td>{_fmt(e['total'])}</td></tr>"
        for e in data["por_estado"]
    )

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">{_PDF_STYLE}</head><body>
<h1>Reporte de Compras</h1>
<div class="period">{date_from} — {date_to}</div>
<div class="kpis">
  <div class="kpi"><div class="kpi-label">Total Comprado</div><div class="kpi-value">{_fmt(kpis['total_comprado'])}</div></div>
  <div class="kpi"><div class="kpi-label">N° OC Emitidas</div><div class="kpi-value">{kpis['num_oc_emitidas']}</div></div>
  <div class="kpi"><div class="kpi-label">OC Pendientes</div><div class="kpi-value">{kpis['num_oc_pendientes']}</div></div>
</div>
<h2>Por Proveedor</h2>
<table><thead><tr><th>Proveedor</th><th>N° OC</th><th>Total</th></tr></thead>
<tbody>{prov_rows}</tbody></table>
<h2>Por Estado</h2>
<table><thead><tr><th>Estado</th><th>Count</th><th>Total</th></tr></thead>
<tbody>{estado_rows}</tbody></table>
</body></html>"""

    return _pdf_response(html, "compras", date_from, date_to)


# ---------------------------------------------------------------------------
# GET /margenes/export/pdf
# ---------------------------------------------------------------------------

@router.get("/margenes/export/pdf")
def exportar_margenes_pdf(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    current_user, db = perms
    vendedor_id = current_user.id if current_user.role == "vendedor" else None
    data = _get_margenes(date_from, date_to, db, vendedor_id)

    por_producto_full = data["por_producto"]
    por_factura = data["por_factura"]

    all_pcts = [p["margen_pct"] for p in por_producto_full]
    margen_promedio_pct = round(sum(all_pcts) / len(all_pcts), 2) if all_pcts else 0.0
    mejor = por_producto_full[0] if por_producto_full else None
    peor = por_producto_full[-1] if por_producto_full else None

    prod_rows = "".join(
        f"<tr><td>{p['nombre']}</td><td>{p['cantidad_vendida']}</td><td>{p['margen_pct']}%</td></tr>"
        for p in por_producto_full[:20]
    )
    fac_rows = "".join(
        f"<tr><td>{f['numero']}</td><td>{_fmt(f['total'])}</td><td>{_fmt(f['margen_total'])}</td><td>{f['margen_pct']}%</td></tr>"
        for f in por_factura
    )

    mejor_txt = f"{mejor['nombre']} ({mejor['margen_pct']}%)" if mejor else "—"
    peor_txt = f"{peor['nombre']} ({peor['margen_pct']}%)" if peor else "—"

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">{_PDF_STYLE}</head><body>
<h1>Reporte de Márgenes</h1>
<div class="period">{date_from} — {date_to}</div>
<div class="kpis">
  <div class="kpi"><div class="kpi-label">Margen Promedio</div><div class="kpi-value">{margen_promedio_pct}%</div></div>
  <div class="kpi"><div class="kpi-label">Mejor Producto</div><div class="kpi-value">{mejor_txt}</div></div>
  <div class="kpi"><div class="kpi-label">Peor Producto</div><div class="kpi-value">{peor_txt}</div></div>
</div>
<h2>Por Producto (Top 20)</h2>
<table><thead><tr><th>Nombre</th><th>Cantidad</th><th>Margen %</th></tr></thead>
<tbody>{prod_rows}</tbody></table>
<h2>Por Factura</h2>
<table><thead><tr><th>N° Factura</th><th>Total</th><th>Margen</th><th>%</th></tr></thead>
<tbody>{fac_rows}</tbody></table>
</body></html>"""

    return _pdf_response(html, "margenes", date_from, date_to)


# ---------------------------------------------------------------------------
# GET /dte/export/pdf
# ---------------------------------------------------------------------------

@router.get("/dte/export/pdf")
def exportar_dte_pdf(
    date_from: date = Query(...),
    date_to: date = Query(...),
    perms: tuple[User, Session] = require_permission("facturas", "view"),
):
    _validate_dates(date_from, date_to)
    _, db = perms
    data = _get_dte(date_from, date_to, db)
    kpis = data["kpis"]

    emision_rows = "".join(
        f"<tr><td>{e['tipo']}</td><td>{e['folio']}</td><td>{e['estado']}</td>"
        f"<td>{_fmt(e['monto_total'] or 0)}</td><td>{e['fecha']}</td></tr>"
        for e in data["emisiones"]
    )

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">{_PDF_STYLE}</head><body>
<h1>Reporte DTE</h1>
<div class="period">{date_from} — {date_to}</div>
<div class="kpis">
  <div class="kpi"><div class="kpi-label">Total Emitidos</div><div class="kpi-value">{kpis['total_emitidos']}</div></div>
  <div class="kpi"><div class="kpi-label">Aceptadas</div><div class="kpi-value">{kpis['aceptadas']}</div></div>
  <div class="kpi"><div class="kpi-label">Rechazadas</div><div class="kpi-value">{kpis['rechazadas']}</div></div>
  <div class="kpi"><div class="kpi-label">Pendientes</div><div class="kpi-value">{kpis['pendientes']}</div></div>
</div>
<h2>Emisiones</h2>
<table><thead><tr><th>Tipo</th><th>Folio</th><th>Estado</th><th>Monto</th><th>Fecha</th></tr></thead>
<tbody>{emision_rows}</tbody></table>
</body></html>"""

    return _pdf_response(html, "dte", date_from, date_to)

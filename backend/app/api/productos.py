import csv
import io as _io
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.config import require_admin
from app.api.deps import require_permission
from app.models.boleta import Boleta, BoletaLinea
from app.models.cliente import Cliente
from app.models.empresa import Empresa
from app.models.factura import Factura, FacturaLinea
from app.models.lista_precios import ListaPrecios, ListaPreciosItem
from app.models.nota_venta import NotaVenta, NotaVentaLinea
from app.models.producto import Producto
from app.models.system_config import SystemConfig
from app.models.tag import ProductoTag
from app.models.tipo_producto import TipoProducto
from app.models.user import User
from app.models.movimiento_inventario import MovimientoInventario
from app.schemas.lista_precios import HistorialCostoItem, HistorialVentaItem, HistorialVentaPage
from app.schemas.producto import (
    BulkPreciosRequest,
    BulkPreciosResponse,
    ProductoBusquedaOutAdmin,
    ProductoBusquedaOutPublic,
    ProductoCreate,
    ProductoOutAdmin,
    ProductoOutPublic,
    ProductoUpdate,
)
from app.services.producto_parser import (
    ALL_COLUMNS,
    ParseError,
    build_template_xlsx,
    parse_productos_xlsx,
)
from app.utils.search import producto_ids_matching
from app.schemas.movimiento_inventario import MovimientoListOut, MovimientoOut


def _sync_tipos(producto: Producto, tipo_ids: list[int], db: Session) -> None:
    if not tipo_ids:
        producto.tipos = []
        return
    ids = list(dict.fromkeys(tipo_ids))
    tipos = db.query(TipoProducto).filter(TipoProducto.id.in_(ids)).all()
    found = {t.id for t in tipos}
    missing = [i for i in ids if i not in found]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tipo(s) no encontrado(s): {missing}",
        )
    producto.tipos = tipos


def _sync_tags(producto: Producto, tag_nombres: list[str], db: Session) -> None:
    nombres = [n.strip().lower() for n in tag_nombres if n.strip()]
    if not nombres:
        producto.tags = []
        return
    existing = {t.nombre: t for t in db.query(ProductoTag).filter(ProductoTag.nombre.in_(nombres)).all()}
    tags = []
    for nombre in nombres:
        if nombre in existing:
            tags.append(existing[nombre])
        else:
            t = ProductoTag(nombre=nombre)
            db.add(t)
            tags.append(t)
    producto.tags = tags


def _stale_cost(fecha: datetime | None, threshold_days: int, now: datetime) -> bool:
    if fecha is None:
        return True
    if fecha.tzinfo is None:
        fecha = fecha.replace(tzinfo=timezone.utc)
    return (now - fecha).days > threshold_days


def _get_threshold_days(db: Session) -> int:
    """Return the configured stale-cost threshold in days, falling back to 60 on
    missing or malformed SystemConfig rows (e.g., admin saved 'banana')."""
    cfg = db.get(SystemConfig, "dias_alerta_costo_desactualizado")
    if cfg is None:
        return 60
    try:
        return int(cfg.value)
    except (TypeError, ValueError):
        return 60


def _serialize_producto(db: Session, producto: Producto, user: User):
    if user.role != "admin":
        return ProductoOutPublic.model_validate(producto).model_dump(mode="json")
    threshold_days = _get_threshold_days(db)
    stale = _stale_cost(producto.precio_costo_actualizado_en, threshold_days, datetime.now(timezone.utc))
    out = ProductoOutAdmin.model_validate(producto).model_dump(mode="json")
    out["costo_desactualizado"] = stale
    return out


router = APIRouter()


@router.get("/export/excel")
def exportar_excel(
    perms: tuple[User, Session] = require_permission("catalogo", "view"),
):
    _, db = perms
    productos = db.query(Producto).order_by(Producto.nombre).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo"
    ws.append(["ID", "Nombre", "Descripción", "Precio Costo", "Precio Venta", "Stock Mínimo", "Stock Actual"])
    for p in productos:
        ws.append([p.id, p.nombre, p.descripcion or "", float(p.precio_costo), float(p.precio_venta), p.stock_minimo, p.stock_actual])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=catalogo.xlsx"},
    )


@router.get("/buscar")
def buscar_productos(
    q: str = Query("", description="Texto a buscar en nombre, SKU, marca, tipo o tag"),
    perms: tuple[User, Session] = require_permission("catalogo", "view"),
):
    user, db = perms
    q = (q or "").strip()
    if q:
        ids = producto_ids_matching(db, q)
        if not ids:
            return []
        query = db.query(Producto).filter(Producto.id.in_(ids))
    else:
        query = db.query(Producto)
    rows = query.order_by(Producto.nombre).limit(20).all()
    schema = ProductoBusquedaOutAdmin if user.role == "admin" else ProductoBusquedaOutPublic
    return [schema.model_validate(p).model_dump(mode="json") for p in rows]


@router.get("/sugerencias")
def sugerencias_productos(
    cliente_id: int | None = Query(None),
    empresa_id: int | None = Query(None),
    perms: tuple[User, Session] = require_permission("catalogo", "view"),
):
    """Top 20 productos comprados por empresa (prevalece) o cliente en últimos 6 meses
    según facturas no anuladas, ordenados por cantidad total descendente.
    """
    user, db = perms
    if cliente_id is None and empresa_id is None:
        return []

    corte = date.today() - timedelta(days=180)
    q = (
        db.query(
            FacturaLinea.producto_id,
            func.sum(FacturaLinea.cantidad).label("total_qty"),
            func.max(Factura.fecha).label("ultima_fecha"),
        )
        .join(Factura, FacturaLinea.factura_id == Factura.id)
        .filter(
            Factura.estado != "anulada",
            Factura.fecha >= corte,
            FacturaLinea.producto_id.is_not(None),
        )
    )
    if empresa_id is not None:
        q = q.filter(Factura.empresa_id == empresa_id)
    else:
        q = q.filter(Factura.cliente_id == cliente_id)

    ranking = (
        q.group_by(FacturaLinea.producto_id)
        .having(func.sum(FacturaLinea.cantidad) > 0)
        .order_by(func.sum(FacturaLinea.cantidad).desc(), func.max(Factura.fecha).desc())
        .limit(20)
        .all()
    )
    if not ranking:
        return []

    ids = [r.producto_id for r in ranking]
    orden = {pid: i for i, pid in enumerate(ids)}
    productos = db.query(Producto).filter(Producto.id.in_(ids)).all()
    productos.sort(key=lambda p: orden[p.id])

    schema = ProductoBusquedaOutAdmin if user.role == "admin" else ProductoBusquedaOutPublic
    return [schema.model_validate(p).model_dump(mode="json") for p in productos]


@router.patch("/bulk-precios", response_model=BulkPreciosResponse)
def actualizar_precios_bulk(
    body: BulkPreciosRequest,
    perms: tuple[User, Session] = require_permission("catalogo", "edit"),
):
    """Actualiza el precio_venta de múltiples productos en una sola transacción.

    Atómico: si algún ID no existe, se aborta sin tocar nada.
    """
    _, db = perms
    ids = [it.id for it in body.items]
    productos = {p.id: p for p in db.query(Producto).filter(Producto.id.in_(ids)).all()}
    faltantes = [i for i in ids if i not in productos]
    if faltantes:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Producto(s) no encontrado(s): {faltantes}",
        )
    for item in body.items:
        productos[item.id].precio_venta = item.precio_venta
    db.commit()
    return BulkPreciosResponse(actualizados=len(ids), ids=ids)


@router.get("/")
def listar_productos(
    q: str = Query("", description="Filtrar por nombre, SKU, marca, tipo o tag"),
    tipo: list[str] = Query(default_factory=list, description="Filtrar por nombre(s) de tipo (multi-valor)"),
    spec: list[str] = Query(default_factory=list, description="Filtrar por spec exacta (multi-valor, case-insensitive)"),
    perms: tuple[User, Session] = require_permission("catalogo", "view"),
):
    user, db = perms
    q = (q or "").strip()
    if q:
        ids = producto_ids_matching(db, q)
        if not ids:
            return []
        query = db.query(Producto).filter(Producto.id.in_(ids))
    else:
        query = db.query(Producto)
    if tipo:
        nombres = [t.strip().lower() for t in tipo if t.strip()]
        if nombres:
            query = (
                query.join(Producto.tipos)
                .filter(TipoProducto.nombre.in_(nombres))
                .distinct()
            )
    rows = query.order_by(Producto.nombre).all()
    if spec:
        spec_lower = {s.strip().lower() for s in spec if s.strip()}
        rows = [p for p in rows if spec_lower.issubset({s.lower() for s in (p.specs or [])})]

    if user.role != "admin":
        return [ProductoOutPublic.model_validate(p).model_dump(mode="json") for p in rows]

    threshold_days = _get_threshold_days(db)
    now = datetime.now(timezone.utc)

    def serialize(p: Producto):
        stale = _stale_cost(p.precio_costo_actualizado_en, threshold_days, now)
        out = ProductoOutAdmin.model_validate(p).model_dump(mode="json")
        out["costo_desactualizado"] = stale
        return out

    return [serialize(p) for p in rows]


@router.post("/", status_code=status.HTTP_201_CREATED)
def crear_producto(
    body: ProductoCreate,
    perms: tuple[User, Session] = require_permission("catalogo", "create"),
):
    user, db = perms
    data = body.model_dump(exclude={"tags", "tipos"})
    producto = Producto(**data)
    db.add(producto)
    db.flush()
    _sync_tags(producto, body.tags, db)
    _sync_tipos(producto, body.tipos, db)
    db.commit()
    db.refresh(producto)
    return _serialize_producto(db, producto, user)


@router.get("/{producto_id}")
def obtener_producto(
    producto_id: int,
    perms: tuple[User, Session] = require_permission("catalogo", "view"),
):
    user, db = perms
    p = db.get(Producto, producto_id)
    if not p:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Producto no encontrado")
    return _serialize_producto(db, p, user)


@router.patch("/{producto_id}")
def actualizar_producto(
    producto_id: int,
    body: ProductoUpdate,
    perms: tuple[User, Session] = require_permission("catalogo", "edit"),
):
    user, db = perms
    p = db.get(Producto, producto_id)
    if not p:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Producto no encontrado")
    for field, value in body.model_dump(exclude_unset=True, exclude={"tags", "tipos"}).items():
        setattr(p, field, value)
    if body.tags is not None:
        _sync_tags(p, body.tags, db)
    if body.tipos is not None:
        _sync_tipos(p, body.tipos, db)
    db.commit()
    db.refresh(p)
    return _serialize_producto(db, p, user)


@router.delete("/{producto_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_producto(
    producto_id: int,
    perms: tuple[User, Session] = require_permission("catalogo", "delete"),
):
    _, db = perms
    p = db.get(Producto, producto_id)
    if not p:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Producto no encontrado")
    db.delete(p)
    db.commit()


@router.get("/{producto_id}/movimientos/export")
def exportar_movimientos_producto(
    producto_id: int,
    perms: tuple[User, Session] = require_permission("inventario", "view"),
):
    _, db = perms
    if not db.get(Producto, producto_id):
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    movimientos = (
        db.query(MovimientoInventario)
        .filter_by(producto_id=producto_id)
        .order_by(MovimientoInventario.created_at.asc())
        .all()
    )
    output = _io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["fecha", "tipo", "cantidad", "signo", "referencia_tipo", "referencia_id", "motivo", "nota", "usuario_id"])
    for m in movimientos:
        writer.writerow([
            m.created_at.isoformat(), m.tipo, m.cantidad, m.signo,
            m.referencia_tipo or "", m.referencia_id or "",
            m.motivo or "", m.nota or "",
            m.usuario_id or "",
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=movimientos_{producto_id}.csv"},
    )


@router.get("/{producto_id}/movimientos", response_model=MovimientoListOut)
def listar_movimientos_producto(
    producto_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    perms: tuple[User, Session] = require_permission("inventario", "view"),
):
    _, db = perms
    if not db.get(Producto, producto_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Producto no encontrado")
    from sqlalchemy.orm import joinedload

    saldo_expr = func.sum(
        MovimientoInventario.signo * MovimientoInventario.cantidad
    ).over(
        order_by=MovimientoInventario.created_at.asc(),
        partition_by=MovimientoInventario.producto_id,
    ).label("saldo")

    base = (
        db.query(MovimientoInventario, saldo_expr)
        .options(joinedload(MovimientoInventario.producto), joinedload(MovimientoInventario.usuario))
        .filter(MovimientoInventario.producto_id == producto_id)
    )
    total = db.query(MovimientoInventario).filter(MovimientoInventario.producto_id == producto_id).count()
    rows = base.order_by(MovimientoInventario.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    items = []
    for mov, saldo in rows:
        out = MovimientoOut.model_validate(mov)
        out.saldo = int(saldo) if saldo is not None else None
        items.append(out)

    return MovimientoListOut(items=items, total=total, page=page, page_size=page_size)


@router.get("/{producto_id}/historial-costos", response_model=list[HistorialCostoItem])
def historial_costos(
    producto_id: int,
    perms: tuple[User, Session] = Depends(require_admin),
):
    _, db = perms
    p = db.get(Producto, producto_id)
    if not p:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    if not p.sku:
        return []
    rows = (
        db.query(ListaPrecios, ListaPreciosItem)
        .join(ListaPreciosItem, ListaPreciosItem.lista_id == ListaPrecios.id)
        .filter(ListaPreciosItem.sku == p.sku)
        .order_by(ListaPrecios.fecha_subida.desc(), ListaPrecios.id.desc())
        .all()
    )
    return [
        HistorialCostoItem(
            fecha_subida=lp.fecha_subida,
            costo_unitario=item.costo_unitario,
            lista_id=lp.id,
            nombre_archivo=lp.nombre_archivo,
        )
        for lp, item in rows
    ]


def _scope_filter_vendedor(doc_model, user_id: int):
    """Filter expression OR(doc.vendedor_id, cliente.vendedor_id, empresa.vendedor_id, doc.empresa_id ∈ empresas asignadas)."""
    from sqlalchemy import or_, select
    empresas_del_vendedor = select(Empresa.id).where(Empresa.vendedor_id == user_id)
    conds = [
        doc_model.vendedor_id == user_id,
        Cliente.vendedor_id == user_id,
        Empresa.vendedor_id == user_id,
    ]
    if hasattr(doc_model, "empresa_id"):
        conds.append(doc_model.empresa_id.in_(empresas_del_vendedor))
    return or_(*conds)


@router.get("/{producto_id}/historial-ventas", response_model=HistorialVentaPage)
def historial_ventas(
    producto_id: int,
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    perms: tuple[User, Session] = require_permission("catalogo", "view"),
):
    """Sales history for a producto: union of NV/Factura/Boleta lines.

    Vendedor only sees lines from docs whose vendedor_id, cliente.vendedor_id,
    empresa.vendedor_id, or doc.empresa_id matches them.
    """
    from decimal import Decimal as D
    user, db = perms
    producto = db.get(Producto, producto_id)
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    is_vendedor = user.role == "vendedor"

    # --- NV ---
    nv_q = (
        db.query(NotaVentaLinea, NotaVenta, Cliente, Empresa)
        .join(NotaVenta, NotaVenta.id == NotaVentaLinea.nv_id)
        .outerjoin(Cliente, Cliente.id == NotaVenta.cliente_id)
        .outerjoin(Empresa, Empresa.id == Cliente.empresa_id)
        .filter(NotaVentaLinea.producto_id == producto_id)
    )
    if is_vendedor:
        nv_q = nv_q.filter(_scope_filter_vendedor(NotaVenta, user.id))

    # --- Factura ---
    fa_q = (
        db.query(FacturaLinea, Factura, Cliente, Empresa)
        .join(Factura, Factura.id == FacturaLinea.factura_id)
        .outerjoin(Cliente, Cliente.id == Factura.cliente_id)
        .outerjoin(Empresa, Empresa.id == Cliente.empresa_id)
        .filter(FacturaLinea.producto_id == producto_id)
    )
    if is_vendedor:
        fa_q = fa_q.filter(_scope_filter_vendedor(Factura, user.id))

    # --- Boleta ---
    bo_q = (
        db.query(BoletaLinea, Boleta, Cliente, Empresa)
        .join(Boleta, Boleta.id == BoletaLinea.boleta_id)
        .outerjoin(Cliente, Cliente.id == Boleta.cliente_id)
        .outerjoin(Empresa, Empresa.id == Cliente.empresa_id)
        .filter(BoletaLinea.producto_id == producto_id)
    )
    if is_vendedor:
        bo_q = bo_q.filter(_scope_filter_vendedor(Boleta, user.id))

    items: list[HistorialVentaItem] = []

    for linea, nv, cliente, empresa in nv_q.all():
        cantidad = D(linea.cantidad or 0)
        precio = D(linea.valor_neto or 0)
        items.append(HistorialVentaItem(
            fecha=nv.fecha,
            doc_tipo="NV",
            doc_id=nv.id,
            doc_numero=nv.numero,
            cliente_id=cliente.id if cliente else None,
            cliente_nombre=cliente.nombre if cliente else None,
            empresa_id=empresa.id if empresa else None,
            empresa_nombre=empresa.nombre if empresa else None,
            cantidad=cantidad,
            precio_unitario=precio,
            total=D(linea.total_neto or 0),
        ))

    for linea, fa, cliente, empresa in fa_q.all():
        cantidad = D(linea.cantidad or 0)
        precio = D(linea.valor_neto or 0)
        items.append(HistorialVentaItem(
            fecha=fa.fecha,
            doc_tipo="Factura",
            doc_id=fa.id,
            doc_numero=fa.numero,
            cliente_id=cliente.id if cliente else None,
            cliente_nombre=cliente.nombre if cliente else None,
            empresa_id=empresa.id if empresa else None,
            empresa_nombre=empresa.nombre if empresa else None,
            cantidad=cantidad,
            precio_unitario=precio,
            total=D(linea.total_neto or 0),
        ))

    for linea, bo, cliente, empresa in bo_q.all():
        cantidad = D(linea.cantidad or 0)
        precio = D(linea.precio_unitario or 0)
        items.append(HistorialVentaItem(
            fecha=bo.fecha,
            doc_tipo="Boleta",
            doc_id=bo.id,
            doc_numero=bo.numero,
            cliente_id=cliente.id if cliente else None,
            cliente_nombre=cliente.nombre if cliente else None,
            empresa_id=empresa.id if empresa else None,
            empresa_nombre=empresa.nombre if empresa else None,
            cantidad=cantidad,
            precio_unitario=precio,
            total=D(linea.total_neto or 0),
        ))

    items.sort(key=lambda x: (x.fecha, x.doc_id), reverse=True)

    total = len(items)
    total_cantidad = sum((it.cantidad for it in items), D(0))
    total_monto = sum((it.total for it in items), D(0))

    page = items[offset: offset + limit]
    return HistorialVentaPage(
        items=page,
        total=total,
        total_cantidad=total_cantidad,
        total_monto=total_monto,
    )


# ============================================================================
# Import endpoints for product catalog
# ============================================================================


@router.get("/import/template")
def descargar_template_import(
    perms: tuple[User, Session] = Depends(require_admin),
):
    return StreamingResponse(
        BytesIO(build_template_xlsx()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos.xlsx"},
    )


def _resumen_preview(content: bytes, db: Session) -> dict:
    parsed = parse_productos_xlsx(content)
    skus_archivo = [p.sku_normalizado for p in parsed.validas]
    # Case-insensitive SKU lookup: find all productos with matching normalized SKU
    existentes = set()
    if skus_archivo:
        # Query all productos and filter by normalized SKU
        all_produtos = db.query(Producto).all()
        existentes = {
            (p.sku.upper() if p.sku else "")
            for p in all_produtos
            if p.sku and p.sku.upper() in skus_archivo
        }

    filas = []
    for p in parsed.validas:
        accion = "actualizar" if p.sku_normalizado in existentes else "crear"
        filas.append({
            "fila": p.fila,
            "sku": p.sku_normalizado,
            "nombre": p.nombre,
            "accion": accion,
        })
    errores = [
        {"fila": r.fila, "sku": r.sku_raw, "nombre": r.nombre_raw, "motivo": r.motivo}
        for r in parsed.invalidas
    ]
    return {
        "total_filas": len(parsed.validas) + len(parsed.invalidas),
        "filas_validas": len(parsed.validas),
        "filas_invalidas": len(parsed.invalidas),
        "a_crear": sum(1 for f in filas if f["accion"] == "crear"),
        "a_actualizar": sum(1 for f in filas if f["accion"] == "actualizar"),
        "filas": filas,
        "errores": errores,
    }


@router.post("/import/preview")
async def previsualizar_import(
    archivo: UploadFile = File(...),
    perms: tuple[User, Session] = Depends(require_admin),
):
    _, db = perms
    content = await archivo.read()
    try:
        return _resumen_preview(content, db)
    except ParseError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/import")
async def importar_productos(
    archivo: UploadFile = File(...),
    perms: tuple[User, Session] = Depends(require_admin),
):
    _, db = perms
    content = await archivo.read()
    try:
        parsed = parse_productos_xlsx(content)
    except ParseError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))

    skus_archivo = [p.sku_normalizado for p in parsed.validas]
    # Case-insensitive SKU lookup: build a map of normalized SKU -> producto
    existentes_productos = (
        db.query(Producto).filter(func.upper(Producto.sku).in_(skus_archivo)).all()
        if skus_archivo
        else []
    )
    existentes = {
        (p.sku.upper() if p.sku else ""): p
        for p in existentes_productos
    }

    detalles: list[dict] = []
    creadas = actualizadas = sin_cambio = errores = 0

    try:
        for p in parsed.validas:
            existente = existentes.get(p.sku_normalizado)
            if existente is None:
                # Create new producto
                producto = Producto(
                    sku=p.sku_normalizado,
                    nombre=p.nombre,
                    descripcion=p.descripcion,
                    precio_venta=Decimal(str(p.precio_base)),
                    precio_costo=Decimal(str(p.costo)),
                    unidad=p.unidad,
                    iva_porcentaje=p.iva,
                )
                db.add(producto)
                db.flush()
                # Add tipo if familia specified
                if p.familia:
                    tipo = db.query(TipoProducto).filter(TipoProducto.nombre.ilike(p.familia.strip())).first()
                    if not tipo:
                        tipo = TipoProducto(nombre=p.familia.strip())
                        db.add(tipo)
                        db.flush()
                    producto.tipos.append(tipo)
                creadas += 1
                detalles.append({"fila": p.fila, "sku": p.sku_normalizado, "nombre": p.nombre, "estado": "creada", "motivo": None})
            else:
                # Update existing producto
                cambios = False
                if existente.nombre != p.nombre:
                    existente.nombre = p.nombre
                    cambios = True
                if p.descripcion is not None and existente.descripcion != p.descripcion:
                    existente.descripcion = p.descripcion
                    cambios = True
                if existente.precio_venta != Decimal(str(p.precio_base)):
                    existente.precio_venta = Decimal(str(p.precio_base))
                    cambios = True
                if existente.precio_costo != Decimal(str(p.costo)):
                    existente.precio_costo = Decimal(str(p.costo))
                    cambios = True
                if p.unidad is not None and existente.unidad != p.unidad:
                    existente.unidad = p.unidad
                    cambios = True
                if existente.iva_porcentaje != p.iva:
                    existente.iva_porcentaje = p.iva
                    cambios = True
                # Update tipo if familia specified
                if p.familia:
                    tipo = db.query(TipoProducto).filter(TipoProducto.nombre.ilike(p.familia.strip())).first()
                    if not tipo:
                        tipo = TipoProducto(nombre=p.familia.strip())
                        db.add(tipo)
                        db.flush()
                    if tipo not in existente.tipos:
                        existente.tipos = [tipo]
                        cambios = True
                if cambios:
                    actualizadas += 1
                    detalles.append({"fila": p.fila, "sku": p.sku_normalizado, "nombre": p.nombre, "estado": "actualizada", "motivo": None})
                else:
                    sin_cambio += 1
                    detalles.append({"fila": p.fila, "sku": p.sku_normalizado, "nombre": p.nombre, "estado": "sin_cambio", "motivo": None})
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Conflicto al guardar: {e.orig}",
        )

    for r in parsed.invalidas:
        errores += 1
        detalles.append({"fila": r.fila, "sku": r.sku_raw, "nombre": r.nombre_raw, "estado": "error", "motivo": r.motivo})

    detalles.sort(key=lambda d: d["fila"])

    return {
        "creadas": creadas,
        "actualizadas": actualizadas,
        "sin_cambio": sin_cambio,
        "errores": errores,
        "detalles": detalles,
    }


@router.post("/import/report")
def descargar_reporte_import(
    payload: dict,
    perms: tuple[User, Session] = Depends(require_admin),
):
    detalles = payload.get("detalles") or []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.append(["fila", "sku", "nombre", "estado", "motivo"])
    for d in detalles:
        ws.append([
            d.get("fila"),
            d.get("sku") or "",
            d.get("nombre") or "",
            d.get("estado") or "",
            d.get("motivo") or "",
        ])
    for i in range(1, 6):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_import_productos.xlsx"},
    )

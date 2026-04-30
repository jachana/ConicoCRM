from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.config import require_admin
from app.api.deps import require_permission
from app.models.proveedor import Proveedor
from app.models.user import User
from app.schemas.proveedor import ProveedorCreate, ProveedorOut, ProveedorUpdate
from app.services.proveedor_parser import (
    ALL_COLUMNS,
    ParseError,
    build_template_xlsx,
    parse_proveedores_xlsx,
)

router = APIRouter()


@router.get("/export/excel")
def exportar_excel(
    perms: tuple[User, Session] = require_permission("proveedores", "view"),
):
    current_user, db = perms
    proveedores = db.query(Proveedor).order_by(Proveedor.nombre).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    ws.append([
        "ID", "Nombre", "RUT", "Razón social", "Giro", "Dirección", "Comuna",
        "Contacto", "Email", "Teléfono", "Condición pago", "Notas",
    ])
    for p in proveedores:
        ws.append([
            p.id, p.nombre, p.rut or "", p.razon_social or "", p.giro or "",
            p.direccion or "", p.comuna or "", p.contacto or "", p.email or "",
            p.telefono or "", p.condicion_pago or "", p.notas or "",
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=proveedores.xlsx"},
    )


@router.get("/import/template")
def descargar_template_import(
    perms: tuple[User, Session] = Depends(require_admin),
):
    return StreamingResponse(
        BytesIO(build_template_xlsx()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_proveedores.xlsx"},
    )


def _resumen_preview(content: bytes, db: Session) -> dict:
    parsed = parse_proveedores_xlsx(content)
    ruts_archivo = [p.rut_normalizado for p in parsed.validas]
    existentes_rows = (
        db.query(Proveedor.rut)
        .filter(Proveedor.rut.in_(ruts_archivo))
        .all()
        if ruts_archivo
        else []
    )
    existentes = {r[0] for r in existentes_rows}

    filas = []
    for p in parsed.validas:
        accion = "actualizar" if p.rut_normalizado in existentes else "crear"
        filas.append({
            "fila": p.fila,
            "rut": p.rut_normalizado,
            "razon_social": p.razon_social,
            "accion": accion,
        })
    errores = [
        {"fila": r.fila, "rut": r.rut_raw, "razon_social": r.razon_social_raw, "motivo": r.motivo}
        for r in parsed.invalidas
    ]
    return {
        "total_filas": len(parsed.validas) + len(parsed.invalidas),
        "filas_validas": len(parsed.validas),
        "filas_invalidas": len(parsed.invalidas),
        "a_crear": sum(1 for f in filas if f["accion"] == "crear"),
        "a_actualizar": sum(1 for f in filas if f["accion"] == "actualizar"),
        "filas": filas,
        "errores": errores,
    }


@router.post("/import/preview")
async def previsualizar_import(
    archivo: UploadFile = File(...),
    perms: tuple[User, Session] = Depends(require_admin),
):
    _, db = perms
    content = await archivo.read()
    try:
        return _resumen_preview(content, db)
    except ParseError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/import")
async def importar_proveedores(
    archivo: UploadFile = File(...),
    perms: tuple[User, Session] = Depends(require_admin),
):
    _, db = perms
    content = await archivo.read()
    try:
        parsed = parse_proveedores_xlsx(content)
    except ParseError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))

    ruts_archivo = [p.rut_normalizado for p in parsed.validas]
    existentes = {
        p.rut: p
        for p in (db.query(Proveedor).filter(Proveedor.rut.in_(ruts_archivo)).all() if ruts_archivo else [])
    }

    detalles: list[dict] = []
    creadas = actualizadas = sin_cambio = errores = 0

    OPTIONAL_FIELDS = ("giro", "direccion", "comuna", "contacto", "email", "telefono", "condicion_pago")

    try:
        for p in parsed.validas:
            existente = existentes.get(p.rut_normalizado)
            if existente is None:
                proveedor = Proveedor(
                    nombre=p.razon_social,
                    rut=p.rut_normalizado,
                    razon_social=p.razon_social,
                    giro=p.giro,
                    direccion=p.direccion,
                    comuna=p.comuna,
                    contacto=p.contacto,
                    email=p.email,
                    telefono=p.telefono,
                    condicion_pago=p.condicion_pago,
                )
                db.add(proveedor)
                creadas += 1
                detalles.append({"fila": p.fila, "rut": p.rut_normalizado, "razon_social": p.razon_social, "estado": "creada", "motivo": None})
            else:
                cambios = False
                if existente.razon_social != p.razon_social:
                    existente.razon_social = p.razon_social
                    cambios = True
                if not existente.nombre or existente.nombre != p.razon_social:
                    existente.nombre = p.razon_social
                    cambios = True
                for f in OPTIONAL_FIELDS:
                    nuevo = getattr(p, f)
                    if nuevo is not None and getattr(existente, f) != nuevo:
                        setattr(existente, f, nuevo)
                        cambios = True
                if cambios:
                    actualizadas += 1
                    detalles.append({"fila": p.fila, "rut": p.rut_normalizado, "razon_social": p.razon_social, "estado": "actualizada", "motivo": None})
                else:
                    sin_cambio += 1
                    detalles.append({"fila": p.fila, "rut": p.rut_normalizado, "razon_social": p.razon_social, "estado": "sin_cambio", "motivo": None})
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Conflicto al guardar: {e.orig}",
        )

    for r in parsed.invalidas:
        errores += 1
        detalles.append({"fila": r.fila, "rut": r.rut_raw, "razon_social": r.razon_social_raw, "estado": "error", "motivo": r.motivo})

    detalles.sort(key=lambda d: d["fila"])

    return {
        "creadas": creadas,
        "actualizadas": actualizadas,
        "sin_cambio": sin_cambio,
        "errores": errores,
        "detalles": detalles,
    }


@router.post("/import/report")
def descargar_reporte_import(
    payload: dict,
    perms: tuple[User, Session] = Depends(require_admin),
):
    detalles = payload.get("detalles") or []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.append(["fila", "rut", "razon_social", "estado", "motivo"])
    for d in detalles:
        ws.append([
            d.get("fila"),
            d.get("rut") or "",
            d.get("razon_social") or "",
            d.get("estado") or "",
            d.get("motivo") or "",
        ])
    for i in range(1, 6):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_import_proveedores.xlsx"},
    )


@router.get("/", response_model=list[ProveedorOut])
def listar_proveedores(
    perms: tuple[User, Session] = require_permission("proveedores", "view"),
):
    _, db = perms
    return db.query(Proveedor).order_by(Proveedor.nombre).all()


@router.post("/", response_model=ProveedorOut, status_code=status.HTTP_201_CREATED)
def crear_proveedor(
    body: ProveedorCreate,
    perms: tuple[User, Session] = require_permission("proveedores", "create"),
):
    _, db = perms
    proveedor = Proveedor(**body.model_dump())
    db.add(proveedor)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="RUT ya registrado")
    db.refresh(proveedor)
    return proveedor


@router.get("/{proveedor_id}", response_model=ProveedorOut)
def obtener_proveedor(
    proveedor_id: int,
    perms: tuple[User, Session] = require_permission("proveedores", "view"),
):
    _, db = perms
    p = db.get(Proveedor, proveedor_id)
    if not p:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proveedor no encontrado")
    return p


@router.patch("/{proveedor_id}", response_model=ProveedorOut)
def actualizar_proveedor(
    proveedor_id: int,
    body: ProveedorUpdate,
    perms: tuple[User, Session] = require_permission("proveedores", "edit"),
):
    _, db = perms
    p = db.get(Proveedor, proveedor_id)
    if not p:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proveedor no encontrado")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(p, field, value)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="RUT ya registrado")
    db.refresh(p)
    return p


@router.delete("/{proveedor_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_proveedor(
    proveedor_id: int,
    perms: tuple[User, Session] = require_permission("proveedores", "delete"),
):
    _, db = perms
    p = db.get(Proveedor, proveedor_id)
    if not p:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proveedor no encontrado")
    db.delete(p)
    db.commit()

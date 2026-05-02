"""
Usuarios Import Parser

Handles parsing and validation of Excel files containing user data
during onboarding. Provides:
- Excel file parsing with openpyxl
- Row-level validation with detailed error messages
- Email format validation
- Role normalization
"""

from __future__ import annotations

import hashlib
import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl


@dataclass(frozen=True)
class ParsedUser:
    """Single validated user row."""
    row_num: int
    email: str
    nombre: str
    rol: str
    rut: Optional[str]
    activo: bool
    hash_key: str


@dataclass(frozen=True)
class InvalidRow:
    """Single invalid row with errors."""
    row_num: int
    email: Optional[str]  # May be missing/invalid
    errors: list[str]


@dataclass
class ParseResult:
    """Results from parsing a usuarios file."""
    valid_rows: list[ParsedUser] = field(default_factory=list)
    invalid_rows: list[InvalidRow] = field(default_factory=list)
    a_crear: int = 0
    a_actualizar: int = 0

    @property
    def valid_count(self) -> int:
        return len(self.valid_rows)

    @property
    def invalid_count(self) -> int:
        return len(self.invalid_rows)


class ParseError(Exception):
    """Raised when file format is invalid (not recoverable per-row)."""
    pass


VALID_ROLES = {"admin", "subadmin", "vendedor"}

TRUTHY_VALUES = {"true", "1", "si", "sí"}
FALSY_VALUES = {"false", "0", "no"}


class UsuariosParser:
    """Parse and validate usuarios import Excel files."""

    REQUIRED_COLUMNS = {"email", "nombre", "rol"}

    OPTIONAL_COLUMNS = {"rut", "activo"}

    ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS

    @staticmethod
    def parse(content: bytes, filename: str, existing_emails: set[str] = None) -> ParseResult:
        """
        Parse and validate a usuarios import file.

        Args:
            content: File bytes
            filename: Original filename (used to detect format)
            existing_emails: Set of existing emails (lowercase) for counting updates

        Returns:
            ParseResult with valid/invalid rows and creation/update counts

        Raises:
            ParseError: If file format is invalid or headers are missing
        """
        parser = UsuariosParser()
        return parser._parse_file(content, filename, existing_emails or set())

    @staticmethod
    def generate_template() -> bytes:
        """
        Generate Excel template with headers, documentation row, and example row.

        Returns:
            Excel file bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        headers = ["email", "nombre", "rol", "rut", "activo"]

        documentation = [
            "Correo electrónico (requerido, único)",
            "Nombre del usuario (requerido)",
            "Rol: admin, subadmin o vendedor (requerido)",
            "RUT del usuario (opcional)",
            "Activo: true/false (opcional, default: true)",
        ]

        example = [
            "juan.perez@empresa.cl",
            "Juan Pérez",
            "vendedor",
            "12345678-9",
            "true",
        ]

        widths = [30, 25, 12, 15, 10]

        # Row 1: Headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True, size=11)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        # Row 2: Documentation row
        for col_num, doc in enumerate(documentation, 1):
            cell = ws.cell(row=2, column=col_num, value=doc)
            cell.font = openpyxl.styles.Font(italic=True, size=9)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
            )

        # Row 3: Example data
        for col_num, value in enumerate(example, 1):
            ws.cell(row=3, column=col_num, value=value)

        # Adjust column widths
        for col_num, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _parse_file(
        self, content: bytes, filename: str, existing_emails: set[str]
    ) -> ParseResult:
        """Internal parse implementation."""
        ext = Path(filename).suffix.lower()
        if ext != ".xlsx":
            raise ParseError(f"Formato no soportado: {ext}. Use .xlsx")

        raw_rows = self._read_xlsx(content)

        if not raw_rows:
            raise ParseError("El archivo está vacío")

        if len(raw_rows) < 3:
            raise ParseError("El archivo debe tener encabezados y al menos una fila de datos")

        # Row 0: headers, Row 1: documentation, Row 2+: data
        header_row = raw_rows[0]
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in header_row
        ]

        # Find column indices
        col_indices: dict[str, int | None] = {}
        for col_name in self.ALL_COLUMNS:
            try:
                col_indices[col_name] = headers.index(col_name)
            except ValueError:
                if col_name in self.REQUIRED_COLUMNS:
                    raise ParseError(
                        f"Columna requerida faltante: '{col_name}'. "
                        f"Columnas encontradas: {[h for h in headers if h]}"
                    )
                col_indices[col_name] = None

        # Normalise existing_emails to lowercase for case-insensitive comparison
        existing_emails_lower = {e.lower() for e in existing_emails}

        valid_rows: list[ParsedUser] = []
        invalid_rows: list[InvalidRow] = []
        a_crear = 0
        a_actualizar = 0

        for row_num, raw_row in enumerate(raw_rows[2:], start=3):
            parsed, invalid = self._parse_row(
                row_num=row_num,
                raw_row=raw_row,
                col_indices=col_indices,
            )

            if parsed is not None:
                valid_rows.append(parsed)
                if parsed.email.lower() in existing_emails_lower:
                    a_actualizar += 1
                else:
                    a_crear += 1
            else:
                invalid_rows.append(invalid)  # type: ignore[arg-type]

        return ParseResult(
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
            a_crear=a_crear,
            a_actualizar=a_actualizar,
        )

    def _parse_row(
        self,
        row_num: int,
        raw_row: list,
        col_indices: dict[str, int | None],
    ) -> tuple[Optional[ParsedUser], Optional[InvalidRow]]:
        """
        Parse and validate a single data row.

        Returns:
            (valid_row, None) if valid
            (None, invalid_row) if invalid
        """
        errors: list[str] = []
        extracted: dict[str, any] = {}

        for col_name in self.ALL_COLUMNS:
            idx = col_indices.get(col_name)
            if idx is None:
                extracted[col_name] = None
            else:
                value = raw_row[idx] if idx < len(raw_row) else None
                extracted[col_name] = value

        email = self._validate_email(extracted["email"], errors)
        nombre = self._validate_nombre(extracted["nombre"], errors)
        rol = self._validate_rol(extracted["rol"], errors)

        if errors:
            return (
                None,
                InvalidRow(row_num=row_num, email=email, errors=errors),
            )

        rut = self._validate_rut(extracted.get("rut"))
        activo = self._validate_activo(extracted.get("activo"))

        hash_key = self._generate_hash_key(email, nombre, rol)  # type: ignore[arg-type]

        return (
            ParsedUser(
                row_num=row_num,
                email=email.lower(),  # type: ignore[union-attr]
                nombre=nombre,  # type: ignore[arg-type]
                rol=rol,  # type: ignore[arg-type]
                rut=rut,
                activo=activo,
                hash_key=hash_key,
            ),
            None,
        )

    @staticmethod
    def _validate_email(value: any, errors: list[str]) -> Optional[str]:
        """Validate and extract email (required)."""
        if value is None or str(value).strip() == "":
            errors.append("email es requerido")
            return None

        email = str(value).strip()
        if "@" not in email or len(email) < 5:
            errors.append(f"email inválido: {email}")
            return None

        return email

    @staticmethod
    def _validate_nombre(value: any, errors: list[str]) -> Optional[str]:
        """Validate and extract nombre (required)."""
        if value is None or str(value).strip() == "":
            errors.append("nombre es requerido")
            return None

        nombre = str(value).strip()
        if len(nombre) < 1:
            errors.append("nombre no puede estar vacío")
            return None

        return nombre

    @staticmethod
    def _validate_rol(value: any, errors: list[str]) -> Optional[str]:
        """Validate and normalize rol (required)."""
        if value is None or str(value).strip() == "":
            errors.append("rol es requerido")
            return None

        rol = str(value).strip().lower()
        if rol not in VALID_ROLES:
            errors.append(
                f"rol inválido: '{value}'. Debe ser uno de: {', '.join(sorted(VALID_ROLES))}"
            )
            return None

        return rol

    @staticmethod
    def _validate_rut(value: any) -> Optional[str]:
        """Extract rut (optional, no validation)."""
        if value is None or str(value).strip() == "":
            return None
        return str(value).strip()

    @staticmethod
    def _validate_activo(value: any) -> bool:
        """Parse activo flag (optional, default True)."""
        if value is None or str(value).strip() == "":
            return True

        v = str(value).strip().lower()
        if v in TRUTHY_VALUES:
            return True
        if v in FALSY_VALUES:
            return False

        # Default to True for unrecognised values
        return True

    @staticmethod
    def _generate_hash_key(email: str, nombre: str, rol: str) -> str:
        """Generate SHA256 hash key for idempotency detection."""
        key_str = f"{email.lower()}|{nombre}|{rol}"
        return hashlib.sha256(key_str.encode("utf-8")).hexdigest()

    @staticmethod
    def _read_xlsx(content: bytes) -> list[list]:
        """Read Excel file and return rows as lists."""
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            return [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

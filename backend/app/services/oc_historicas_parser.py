from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from typing import Optional

import openpyxl


class ParseError(Exception):
    """Raised when file format is invalid (not recoverable per-row)."""
    pass


@dataclass
class OCLine:
    sku: Optional[str]
    descripcion: str
    cantidad: int
    precio_unitario: Decimal
    descuento: Decimal  # 0-100
    total_neto_linea: Decimal
    row_num: int


@dataclass
class OCGroup:
    numero_oc: int
    rut_proveedor: str
    fecha: date
    fecha_entrega_esperada: Optional[date]
    estado: str
    nota: Optional[str]
    total_neto: Decimal
    iva: Decimal
    total: Decimal
    lines: list[OCLine]
    status: str  # "crear" | "omitir"
    row_num: int  # cabecera row


@dataclass
class InvalidRow:
    row_num: int
    numero_oc_raw: Optional[str]
    sheet: str  # "Cabecera OC" or "Detalle OC"
    motivo: str


@dataclass
class ParseResult:
    valid_groups: list[OCGroup] = field(default_factory=list)
    invalid_rows: list[InvalidRow] = field(default_factory=list)
    a_crear: int = 0
    a_omitir: int = 0
    invalid_group_count: int = 0


class OCHistoricasParser:
    """Parse and validate OC históricas import files (two-sheet Excel)."""

    CABECERA_SHEET = "Cabecera OC"
    DETALLE_SHEET = "Detalle OC"

    CABECERA_REQUIRED = {"numero_oc", "rut_proveedor", "fecha", "total_neto", "iva", "total"}
    CABECERA_OPTIONAL = {"estado", "nota", "fecha_entrega_esperada"}

    DETALLE_REQUIRED = {"numero_oc", "descripcion", "cantidad", "precio_unitario"}
    DETALLE_OPTIONAL = {"sku", "descuento"}

    @staticmethod
    def parse(
        content: bytes,
        filename: str,
        proveedores_by_rut: dict,
        productos_by_sku: dict,
        existing_numeros: set,
    ) -> ParseResult:
        """
        Parse and validate a two-sheet Excel file of historical OCs.

        Args:
            content: Excel file bytes
            filename: Original filename (unused, kept for API symmetry)
            proveedores_by_rut: Dict mapping rut -> Proveedor id
            productos_by_sku: Dict mapping sku -> Producto id
            existing_numeros: Set of numero_oc integers already in DB

        Returns:
            ParseResult with valid_groups, invalid_rows, and counts

        Raises:
            ParseError: If file format is invalid or required columns are missing
        """
        proveedores_by_rut = proveedores_by_rut or {}
        productos_by_sku = productos_by_sku or {}
        existing_numeros = existing_numeros or set()

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ParseError(f"No se pudo abrir el archivo Excel: {exc}")

        try:
            sheet_names = wb.sheetnames
            if OCHistoricasParser.CABECERA_SHEET not in sheet_names:
                raise ParseError(
                    f"No se encontró la hoja '{OCHistoricasParser.CABECERA_SHEET}'. "
                    f"Hojas encontradas: {sheet_names}"
                )
            if OCHistoricasParser.DETALLE_SHEET not in sheet_names:
                raise ParseError(
                    f"No se encontró la hoja '{OCHistoricasParser.DETALLE_SHEET}'. "
                    f"Hojas encontradas: {sheet_names}"
                )

            cab_rows = [list(row) for row in wb[OCHistoricasParser.CABECERA_SHEET].iter_rows(values_only=True)]
            det_rows = [list(row) for row in wb[OCHistoricasParser.DETALLE_SHEET].iter_rows(values_only=True)]
        finally:
            wb.close()

        if not cab_rows:
            raise ParseError(f"La hoja '{OCHistoricasParser.CABECERA_SHEET}' está vacía")
        if not det_rows:
            raise ParseError(f"La hoja '{OCHistoricasParser.DETALLE_SHEET}' está vacía")

        result = ParseResult()

        cab_groups, cab_invalid = OCHistoricasParser._parse_cabecera(
            cab_rows, proveedores_by_rut, existing_numeros
        )
        result.invalid_rows.extend(cab_invalid)

        det_invalid = OCHistoricasParser._parse_detalle(
            det_rows, cab_groups, productos_by_sku
        )
        result.invalid_rows.extend(det_invalid)

        # Cross-validation: compare computed neto sum vs cabecera total_neto
        cross_invalid = OCHistoricasParser._cross_validate(cab_groups)
        result.invalid_rows.extend(cross_invalid)

        for numero_oc, group in cab_groups.items():
            result.valid_groups.append(group)
            if group.status == "omitir":
                result.a_omitir += 1
            else:
                result.a_crear += 1

        result.invalid_group_count = len(cab_invalid)

        return result

    @staticmethod
    def _parse_cabecera(
        raw_rows: list[list],
        proveedores_by_rut: dict,
        existing_numeros: set,
    ) -> tuple[dict[int, OCGroup], list[InvalidRow]]:
        """
        Validate Cabecera OC rows and return (groups_by_numero_oc, invalid_rows).
        groups_by_numero_oc preserves file order via dict insertion order.
        """
        header_row = raw_rows[0]
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in header_row
        ]

        missing = OCHistoricasParser.CABECERA_REQUIRED - set(headers)
        if missing:
            raise ParseError(
                f"Hoja '{OCHistoricasParser.CABECERA_SHEET}' — columna(s) requerida(s) faltante(s): "
                f"{sorted(missing)}. Columnas encontradas: {[h for h in headers if h]}"
            )

        col = {h: i for i, h in enumerate(headers) if h}

        data_rows = OCHistoricasParser._skip_doc_row(raw_rows, col, "numero_oc")
        header_offset = len(raw_rows) - len(data_rows)

        groups: dict[int, OCGroup] = {}
        invalid_rows: list[InvalidRow] = []

        def _get(row: list, col_name: str) -> Optional[str]:
            i = col.get(col_name, -1)
            if i < 0 or i >= len(row):
                return None
            val = row[i]
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        for idx, raw_row in enumerate(data_rows):
            row_num = header_offset + idx + 1
            numero_oc_raw = _get(raw_row, "numero_oc")
            errors = []

            # numero_oc
            parsed_numero_oc = None
            if not numero_oc_raw:
                errors.append("numero_oc es requerido")
            else:
                try:
                    parsed_numero_oc = int(float(numero_oc_raw))
                    if parsed_numero_oc <= 0:
                        errors.append(f"numero_oc debe ser un entero positivo (tiene {parsed_numero_oc})")
                        parsed_numero_oc = None
                except (ValueError, TypeError):
                    errors.append(f"numero_oc '{numero_oc_raw}' no es un entero válido")

            if parsed_numero_oc is not None and parsed_numero_oc in groups:
                errors.append(f"numero_oc {parsed_numero_oc} duplicado en la hoja Cabecera OC")

            # rut_proveedor
            rut_proveedor_raw = _get(raw_row, "rut_proveedor")
            if not rut_proveedor_raw:
                errors.append("rut_proveedor es requerido")
            elif rut_proveedor_raw not in proveedores_by_rut:
                errors.append(f"Proveedor con rut '{rut_proveedor_raw}' no encontrado")

            # fecha
            fecha_raw = _get(raw_row, "fecha")
            parsed_fecha = None
            if not fecha_raw:
                errors.append("fecha es requerida")
            else:
                parsed_fecha = OCHistoricasParser._parse_date(fecha_raw)
                if parsed_fecha is None:
                    errors.append(
                        f"fecha '{fecha_raw}' no tiene formato válido (YYYY-MM-DD o DD-MM-YYYY)"
                    )

            # total_neto
            total_neto_raw = _get(raw_row, "total_neto")
            parsed_total_neto = None
            if total_neto_raw is None:
                errors.append("total_neto es requerido")
            else:
                try:
                    parsed_total_neto = Decimal(str(total_neto_raw))
                    if parsed_total_neto < 0:
                        errors.append(f"total_neto debe ser no-negativo (tiene {parsed_total_neto})")
                        parsed_total_neto = None
                except Exception:
                    errors.append(f"total_neto '{total_neto_raw}' no es un número válido")

            # iva
            iva_raw = _get(raw_row, "iva")
            parsed_iva = None
            if iva_raw is None:
                errors.append("iva es requerido")
            else:
                try:
                    parsed_iva = Decimal(str(iva_raw))
                    if parsed_iva < 0:
                        errors.append(f"iva debe ser no-negativo (tiene {parsed_iva})")
                        parsed_iva = None
                except Exception:
                    errors.append(f"iva '{iva_raw}' no es un número válido")

            # total
            total_raw = _get(raw_row, "total")
            parsed_total = None
            if total_raw is None:
                errors.append("total es requerido")
            else:
                try:
                    parsed_total = Decimal(str(total_raw))
                    if parsed_total < 0:
                        errors.append(f"total debe ser no-negativo (tiene {parsed_total})")
                        parsed_total = None
                except Exception:
                    errors.append(f"total '{total_raw}' no es un número válido")

            # estado (optional, default "recibida")
            estado_raw = _get(raw_row, "estado")
            parsed_estado = estado_raw if estado_raw else "recibida"

            # nota (optional)
            nota_raw = _get(raw_row, "nota")

            # fecha_entrega_esperada (optional)
            fecha_entrega_raw = _get(raw_row, "fecha_entrega_esperada")
            parsed_fecha_entrega = None
            if fecha_entrega_raw:
                parsed_fecha_entrega = OCHistoricasParser._parse_date(fecha_entrega_raw)
                if parsed_fecha_entrega is None:
                    errors.append(
                        f"fecha_entrega_esperada '{fecha_entrega_raw}' no tiene formato válido (YYYY-MM-DD o DD-MM-YYYY)"
                    )

            if errors:
                invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    numero_oc_raw=numero_oc_raw,
                    sheet=OCHistoricasParser.CABECERA_SHEET,
                    motivo="; ".join(errors),
                ))
                continue

            status = "omitir" if parsed_numero_oc in existing_numeros else "crear"

            groups[parsed_numero_oc] = OCGroup(
                numero_oc=parsed_numero_oc,
                rut_proveedor=rut_proveedor_raw,
                fecha=parsed_fecha,
                fecha_entrega_esperada=parsed_fecha_entrega,
                estado=parsed_estado,
                nota=nota_raw,
                total_neto=parsed_total_neto,
                iva=parsed_iva,
                total=parsed_total,
                lines=[],
                status=status,
                row_num=row_num,
            )

        return groups, invalid_rows

    @staticmethod
    def _parse_detalle(
        raw_rows: list[list],
        groups: dict[int, OCGroup],
        productos_by_sku: dict,
    ) -> list[InvalidRow]:
        """
        Validate Detalle OC rows and attach OCLine objects to their groups.
        Returns list of invalid rows.
        """
        header_row = raw_rows[0]
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in header_row
        ]

        missing = OCHistoricasParser.DETALLE_REQUIRED - set(headers)
        if missing:
            raise ParseError(
                f"Hoja '{OCHistoricasParser.DETALLE_SHEET}' — columna(s) requerida(s) faltante(s): "
                f"{sorted(missing)}. Columnas encontradas: {[h for h in headers if h]}"
            )

        col = {h: i for i, h in enumerate(headers) if h}

        data_rows = OCHistoricasParser._skip_doc_row(raw_rows, col, "numero_oc")
        header_offset = len(raw_rows) - len(data_rows)

        invalid_rows: list[InvalidRow] = []

        def _get(row: list, col_name: str) -> Optional[str]:
            i = col.get(col_name, -1)
            if i < 0 or i >= len(row):
                return None
            val = row[i]
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        for idx, raw_row in enumerate(data_rows):
            row_num = header_offset + idx + 1
            numero_oc_raw = _get(raw_row, "numero_oc")
            errors = []

            # numero_oc — links to Cabecera
            parsed_numero_oc = None
            if not numero_oc_raw:
                errors.append("numero_oc es requerido")
            else:
                try:
                    parsed_numero_oc = int(float(numero_oc_raw))
                    if parsed_numero_oc <= 0:
                        errors.append(f"numero_oc debe ser un entero positivo (tiene {parsed_numero_oc})")
                        parsed_numero_oc = None
                except (ValueError, TypeError):
                    errors.append(f"numero_oc '{numero_oc_raw}' no es un entero válido")

            if parsed_numero_oc is not None and parsed_numero_oc not in groups:
                errors.append(
                    f"numero_oc {parsed_numero_oc} no encontrado en la hoja '{OCHistoricasParser.CABECERA_SHEET}'"
                )
                invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    numero_oc_raw=numero_oc_raw,
                    sheet=OCHistoricasParser.DETALLE_SHEET,
                    motivo="; ".join(errors),
                ))
                continue

            # descripcion
            descripcion_raw = _get(raw_row, "descripcion")
            if not descripcion_raw:
                errors.append("descripcion es requerida")

            # cantidad
            cantidad_raw = _get(raw_row, "cantidad")
            parsed_cantidad = None
            if not cantidad_raw:
                errors.append("cantidad es requerida")
            else:
                try:
                    parsed_cantidad = int(float(cantidad_raw))
                    if parsed_cantidad <= 0:
                        errors.append(f"cantidad debe ser un entero positivo (tiene {parsed_cantidad})")
                        parsed_cantidad = None
                except (ValueError, TypeError):
                    errors.append(f"cantidad '{cantidad_raw}' no es un número válido")

            # precio_unitario
            precio_raw = _get(raw_row, "precio_unitario")
            parsed_precio = None
            if precio_raw is None:
                errors.append("precio_unitario es requerido")
            else:
                try:
                    parsed_precio = Decimal(str(precio_raw))
                    if parsed_precio < 0:
                        errors.append(f"precio_unitario debe ser no-negativo (tiene {parsed_precio})")
                        parsed_precio = None
                except Exception:
                    errors.append(f"precio_unitario '{precio_raw}' no es un número válido")

            # descuento (optional, 0-100, default 0)
            descuento_raw = _get(raw_row, "descuento")
            parsed_descuento = Decimal("0")
            if descuento_raw is not None:
                try:
                    parsed_descuento = Decimal(str(descuento_raw))
                    if parsed_descuento < 0 or parsed_descuento > 100:
                        errors.append(f"descuento '{descuento_raw}' debe estar entre 0 y 100")
                        parsed_descuento = Decimal("0")
                except Exception:
                    errors.append(f"descuento '{descuento_raw}' no es un número válido")

            # sku (optional; if provided and not in productos_by_sku → mark as error but continue)
            sku_raw = _get(raw_row, "sku")
            if sku_raw and sku_raw not in productos_by_sku:
                errors.append(f"Producto con sku '{sku_raw}' no encontrado")

            if errors:
                invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    numero_oc_raw=numero_oc_raw,
                    sheet=OCHistoricasParser.DETALLE_SHEET,
                    motivo="; ".join(errors),
                ))
                continue

            total_neto_linea = (
                Decimal(str(parsed_cantidad))
                * parsed_precio
                * (Decimal("1") - parsed_descuento / Decimal("100"))
            ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

            groups[parsed_numero_oc].lines.append(OCLine(
                sku=sku_raw,
                descripcion=descripcion_raw,
                cantidad=parsed_cantidad,
                precio_unitario=parsed_precio,
                descuento=parsed_descuento,
                total_neto_linea=total_neto_linea,
                row_num=row_num,
            ))

        return invalid_rows

    @staticmethod
    def _cross_validate(groups: dict[int, OCGroup]) -> list[InvalidRow]:
        """
        Cross-validate each OC group: compare computed neto sum vs cabecera total_neto.
        Tolerance: $1. If mismatch > $1, mark the group as error.
        Removes the group from groups dict if invalid.
        Returns list of invalid rows for cross-validation failures.
        """
        invalid_rows: list[InvalidRow] = []
        to_remove = []

        for numero_oc, group in groups.items():
            if not group.lines:
                # No lines → nothing to cross-validate
                continue

            computed_neto = sum(line.total_neto_linea for line in group.lines)
            diff = abs(computed_neto - group.total_neto)

            if diff > Decimal("1"):
                motivo = (
                    f"Totales no coinciden (diferencia: ${diff.quantize(Decimal('1'), rounding=ROUND_HALF_UP)})"
                )
                invalid_rows.append(InvalidRow(
                    row_num=group.row_num,
                    numero_oc_raw=str(numero_oc),
                    sheet=OCHistoricasParser.CABECERA_SHEET,
                    motivo=motivo,
                ))
                to_remove.append(numero_oc)

            # Check total_neto + iva ≈ total
            computed_total = group.total_neto + group.iva
            total_diff = abs(computed_total - group.total)
            if total_diff > Decimal("1"):
                invalid_rows.append(InvalidRow(
                    row_num=group.row_num,
                    numero_oc_raw=str(group.numero_oc),
                    sheet="Cabecera OC",
                    motivo=f"total_neto + iva ({computed_total}) no coincide con total ({group.total}), diferencia: ${total_diff:.0f}",
                ))
                if numero_oc not in to_remove:
                    to_remove.append(numero_oc)

        for numero_oc in to_remove:
            del groups[numero_oc]

        return invalid_rows

    @staticmethod
    def _skip_doc_row(raw_rows: list[list], col: dict, anchor_col: str) -> list[list]:
        """
        Skip row index 1 (the row immediately after headers) if it looks like
        a documentation/example-label row rather than real data.
        """
        if len(raw_rows) <= 1:
            return []
        anchor_idx = col.get(anchor_col, -1)
        potential_doc = raw_rows[1]
        anchor_val = (
            potential_doc[anchor_idx]
            if anchor_idx >= 0 and anchor_idx < len(potential_doc)
            else None
        )
        if anchor_val is None or str(anchor_val).strip() == "":
            return raw_rows[2:]
        val_str = str(anchor_val).strip()
        doc_keywords = ("requerido", "ej:", "ejemplo", "opcional", "numero_oc")
        if any(kw in val_str.lower() for kw in doc_keywords) or len(val_str) > 200:
            return raw_rows[2:]
        return raw_rows[1:]

    @staticmethod
    def _parse_date(s: str) -> Optional[date]:
        """Parse date string in YYYY-MM-DD, DD-MM-YYYY, or DD/MM/YYYY format.
        Also handles Excel date serial numbers (int/float)."""
        # Handle Excel date serial numbers
        try:
            serial = float(s)
            # Excel epoch is 1900-01-01 (with Lotus 1-2-3 bug: serial 1 = Jan 1 1900)
            from datetime import timedelta
            epoch = date(1899, 12, 30)
            return epoch + timedelta(days=int(serial))
        except (ValueError, TypeError):
            pass

        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def generate_template() -> bytes:
        """
        Generate a two-sheet Excel template: 'Cabecera OC' and 'Detalle OC'.
        Each sheet has a bold/gray header row, an italic/light documentation row,
        and example rows.
        """
        wb = openpyxl.Workbook()

        header_font = openpyxl.styles.Font(bold=True, size=11)
        header_fill = openpyxl.styles.PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        doc_font = openpyxl.styles.Font(italic=True, size=9)
        doc_fill = openpyxl.styles.PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        )

        # ----------------------------------------------------------------
        # Sheet 1: Cabecera OC
        # ----------------------------------------------------------------
        ws_cab = wb.active
        ws_cab.title = OCHistoricasParser.CABECERA_SHEET

        cab_headers = [
            "numero_oc",
            "rut_proveedor",
            "fecha",
            "estado",
            "total_neto",
            "iva",
            "total",
            "nota",
            "fecha_entrega_esperada",
        ]
        cab_documentation = [
            "Requerido; entero positivo (clave de enlace con Detalle OC)",
            "Requerido; RUT del proveedor (debe existir)",
            "Requerido; YYYY-MM-DD o DD-MM-YYYY",
            "Opcional; 'recibida' (default)",
            "Requerido; monto neto (≥ 0)",
            "Requerido; monto IVA (≥ 0)",
            "Requerido; monto total (≥ 0)",
            "Opcional; texto libre",
            "Opcional; YYYY-MM-DD o DD-MM-YYYY",
        ]
        cab_examples = [
            [1001, "76123456-7", "2025-01-15", "recibida", 1000000, 190000, 1190000, "OC histórica", ""],
            [1002, "76123456-7", "2025-02-01", "recibida", 500000, 95000, 595000, "", "2025-02-20"],
        ]

        for col_num, header in enumerate(cab_headers, 1):
            cell = ws_cab.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for col_num, doc in enumerate(cab_documentation, 1):
            cell = ws_cab.cell(row=2, column=col_num, value=doc)
            cell.font = doc_font
            cell.fill = doc_fill

        for row_idx, example in enumerate(cab_examples, 3):
            for col_idx, val in enumerate(example, 1):
                ws_cab.cell(row=row_idx, column=col_idx, value=val)

        cab_col_widths = [14, 18, 18, 14, 16, 14, 14, 30, 22]
        for i, width in enumerate(cab_col_widths, 1):
            ws_cab.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = width

        # ----------------------------------------------------------------
        # Sheet 2: Detalle OC
        # ----------------------------------------------------------------
        ws_det = wb.create_sheet(title=OCHistoricasParser.DETALLE_SHEET)

        det_headers = [
            "numero_oc",
            "sku",
            "descripcion",
            "cantidad",
            "precio_unitario",
            "descuento",
        ]
        det_documentation = [
            "Requerido; debe coincidir con un numero_oc en 'Cabecera OC'",
            "Opcional; SKU del producto (debe existir si se indica)",
            "Requerido; descripción del ítem",
            "Requerido; entero positivo",
            "Requerido; precio neto unitario (≥ 0)",
            "Opcional; descuento 0-100 (default 0)",
        ]
        det_examples = [
            [1001, "PROD001", "Producto ejemplo", 10, 100000, 0],
            [1001, "PROD002", "Otro producto", 1, 0, 0],
            [1002, "", "Servicio de soporte", 5, 100000, 0],
        ]

        for col_num, header in enumerate(det_headers, 1):
            cell = ws_det.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for col_num, doc in enumerate(det_documentation, 1):
            cell = ws_det.cell(row=2, column=col_num, value=doc)
            cell.font = doc_font
            cell.fill = doc_fill

        for row_idx, example in enumerate(det_examples, 3):
            for col_idx, val in enumerate(example, 1):
                ws_det.cell(row=row_idx, column=col_idx, value=val)

        det_col_widths = [14, 14, 35, 12, 18, 14]
        for i, width in enumerate(det_col_widths, 1):
            ws_det.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

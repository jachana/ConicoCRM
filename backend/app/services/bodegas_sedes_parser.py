"""
Bodegas and SedeDespacho Import Parser

Handles parsing and validation of Excel files containing warehouse (bodega)
and dispatch site (sede_despacho) data during onboarding. Provides:
- Excel file parsing with openpyxl
- Row-level validation with detailed error messages
- RUT format validation using clean_rut utility
- Duplicate detection and update vs create classification
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

import openpyxl

from app.utils.rut import clean_rut, validate_rut


@dataclass(frozen=True)
class ParsedRow:
    """Single validated bodega/sede row."""
    row_num: int
    empresa_rut: str
    bodega_nombre: str
    bodega_direccion: Optional[str]
    sede_nombre: str
    sede_direccion: str
    status: str  # "crear" or "actualizar"
    errors: list[str]


@dataclass(frozen=True)
class InvalidRow:
    """Single invalid row with errors."""
    row_num: int
    empresa_rut: Optional[str]  # May be missing
    bodega_nombre: Optional[str]  # May be missing
    bodega_direccion: Optional[str]  # May be missing
    sede_nombre: Optional[str]  # May be missing
    sede_direccion: Optional[str]  # May be missing
    errors: list[str]


@dataclass
class ParseResult:
    """Results from parsing a bodegas_sedes file."""
    valid_rows: list[ParsedRow]
    invalid_rows: list[InvalidRow]
    a_crear: dict  # { "bodegas": count, "sedes": count }
    a_actualizar: dict  # { "bodegas": count, "sedes": count }

    @property
    def valid_count(self) -> int:
        return len(self.valid_rows)

    @property
    def invalid_count(self) -> int:
        return len(self.invalid_rows)


class ParseError(Exception):
    """Raised when file format is invalid (not recoverable per-row)."""
    pass


class BodegasSedesParser:
    """Parse and validate bodegas + sedes import Excel files."""

    # Expected column headers (case-insensitive)
    REQUIRED_COLUMNS = {
        "empresa_rut",
        "bodega_nombre",
        "sede_nombre",
        "sede_direccion",
    }

    OPTIONAL_COLUMNS = {
        "bodega_direccion",
    }

    ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS

    @staticmethod
    def parse(
        content: bytes,
        filename: str,
        existing_empresas: dict = None,
        existing_bodegas: dict = None,
        existing_sedes: dict = None,
    ) -> ParseResult:
        """
        Parse and validate a bodegas + sedes import file.

        Args:
            content: File bytes
            filename: Original filename (used to detect format)
            existing_empresas: Dict of rut -> Empresa object (for validation)
            existing_bodegas: Dict of (empresa_id, nombre) -> Bodega object (for update detection)
            existing_sedes: Dict of (empresa_id, nombre) -> SedeDespacho object (for update detection)

        Returns:
            ParseResult with valid/invalid rows and creation/update counts

        Raises:
            ParseError: If file format is invalid or headers are missing
        """
        parser = BodegasSedesParser()
        return parser._parse_file(
            content,
            filename,
            existing_empresas or {},
            existing_bodegas or {},
            existing_sedes or {},
        )

    @staticmethod
    def generate_template() -> bytes:
        """
        Generate Excel template with headers and example row.

        Returns:
            Excel file bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bodegas y Sedes"

        # Row 1: Headers
        headers = [
            "empresa_rut",
            "bodega_nombre",
            "bodega_direccion",
            "sede_nombre",
            "sede_direccion",
        ]

        documentation = [
            "RUT de empresa (ej: 76543210-K)",
            "Nombre de bodega",
            "Dirección de bodega (opcional)",
            "Nombre de sede de despacho",
            "Dirección de sede de despacho",
        ]

        # Row 1: Header with formatting
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True, size=11)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        # Row 2: Documentation row
        for col_num, doc in enumerate(documentation, 1):
            cell = ws.cell(row=2, column=col_num, value=doc)
            cell.font = openpyxl.styles.Font(italic=True, size=9)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
            )

        # Row 3: Example data
        example = [
            "76543210-K",
            "Bodega Central",
            "Av. Principal 1000, Santiago",
            "Sede Principal",
            "Calle Principal 123, Santiago",
        ]

        for col_num, value in enumerate(example, 1):
            ws.cell(row=3, column=col_num, value=value)

        # Adjust column widths
        widths = [15, 25, 30, 25, 30]
        for col_num, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _parse_file(
        self,
        content: bytes,
        filename: str,
        existing_empresas: dict,
        existing_bodegas: dict,
        existing_sedes: dict,
    ) -> ParseResult:
        """Internal parse implementation."""
        ext = Path(filename).suffix.lower()
        if ext != ".xlsx":
            raise ParseError(f"Formato no soportado: {ext}. Use .xlsx")

        raw_rows = self._read_xlsx(content)

        if not raw_rows:
            raise ParseError("El archivo está vacío")

        if len(raw_rows) < 3:
            raise ParseError("El archivo debe tener encabezados y al menos una fila de datos")

        # Row 0: headers, Row 1: documentation, Row 2+: data
        header_row = raw_rows[0]
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in header_row
        ]

        # Find column indices
        col_indices = {}
        for col_name in self.ALL_COLUMNS:
            try:
                col_indices[col_name] = headers.index(col_name)
            except ValueError:
                if col_name in self.REQUIRED_COLUMNS:
                    raise ParseError(
                        f"Columna requerida faltante: '{col_name}'. "
                        f"Columnas encontradas: {[h for h in headers if h]}"
                    )
                col_indices[col_name] = None

        # Parse data rows (skip header row 0 and documentation row 1)
        valid_rows = []
        invalid_rows = []
        bodegas_a_crear = 0
        bodegas_a_actualizar = 0
        sedes_a_crear = 0
        sedes_a_actualizar = 0

        # Track seen (empresa_id, bodega_nombre) and (empresa_id, sede_nombre) in this file
        # for duplicate detection
        seen_bodegas = {}  # (empresa_id, bodega_nombre) -> row_num
        seen_sedes = {}    # (empresa_id, sede_nombre) -> row_num

        for row_num, raw_row in enumerate(raw_rows[2:], start=3):
            result = self._parse_row(
                row_num=row_num,
                raw_row=raw_row,
                col_indices=col_indices,
                existing_empresas=existing_empresas,
                existing_bodegas=existing_bodegas,
                existing_sedes=existing_sedes,
                seen_bodegas=seen_bodegas,
                seen_sedes=seen_sedes,
            )

            if result[0] is not None:  # Valid
                parsed = result[0]
                valid_rows.append(parsed)

                # Count new vs updates based on per-row status
                # Status is "crear" only if BOTH bodega and sede are new
                # Otherwise "actualizar" means at least one already exists
                if parsed.status == "crear":
                    bodegas_a_crear += 1
                    sedes_a_crear += 1
                else:  # "actualizar"
                    bodegas_a_actualizar += 1
                    sedes_a_actualizar += 1
            else:  # Invalid
                invalid_rows.append(result[1])

        return ParseResult(
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
            a_crear={"bodegas": bodegas_a_crear, "sedes": sedes_a_crear},
            a_actualizar={"bodegas": bodegas_a_actualizar, "sedes": sedes_a_actualizar},
        )

    def _parse_row(
        self,
        row_num: int,
        raw_row: list,
        col_indices: dict[str, int | None],
        existing_empresas: dict,
        existing_bodegas: dict,
        existing_sedes: dict,
        seen_bodegas: dict,
        seen_sedes: dict,
    ) -> tuple[Optional[ParsedRow], Optional[InvalidRow]]:
        """
        Parse and validate a single data row.

        Returns:
            (valid_row, None) if valid
            (None, invalid_row) if invalid
        """
        errors = []
        extracted = {}

        # Extract all fields
        for col_name in self.ALL_COLUMNS:
            idx = col_indices.get(col_name)
            if idx is None:
                extracted[col_name] = None
            else:
                value = raw_row[idx] if idx < len(raw_row) else None
                extracted[col_name] = value

        # Validate required fields
        empresa_rut = self._validate_empresa_rut(
            extracted["empresa_rut"], errors, existing_empresas
        )
        bodega_nombre = self._validate_bodega_nombre(extracted["bodega_nombre"], errors)
        sede_nombre = self._validate_sede_nombre(extracted["sede_nombre"], errors)
        sede_direccion = self._validate_sede_direccion(extracted["sede_direccion"], errors)

        if errors:
            return (
                None,
                InvalidRow(
                    row_num=row_num,
                    empresa_rut=extracted.get("empresa_rut"),
                    bodega_nombre=extracted.get("bodega_nombre"),
                    bodega_direccion=extracted.get("bodega_direccion"),
                    sede_nombre=extracted.get("sede_nombre"),
                    sede_direccion=extracted.get("sede_direccion"),
                    errors=errors,
                ),
            )

        # Optional fields
        bodega_direccion = self._validate_bodega_direccion(extracted.get("bodega_direccion"), errors)

        if errors:
            return (
                None,
                InvalidRow(
                    row_num=row_num,
                    empresa_rut=empresa_rut,
                    bodega_nombre=bodega_nombre,
                    bodega_direccion=bodega_direccion,
                    sede_nombre=sede_nombre,
                    sede_direccion=sede_direccion,
                    errors=errors,
                ),
            )

        # Get empresa to check bodega/sede existence
        empresa = existing_empresas.get(clean_rut(empresa_rut))
        if not empresa:
            # This shouldn't happen because we validated above, but be safe
            errors.append(f"Empresa con RUT {empresa_rut} no encontrada")
            return (
                None,
                InvalidRow(
                    row_num=row_num,
                    empresa_rut=empresa_rut,
                    bodega_nombre=bodega_nombre,
                    bodega_direccion=bodega_direccion,
                    sede_nombre=sede_nombre,
                    sede_direccion=sede_direccion,
                    errors=errors,
                ),
            )

        empresa_id = empresa.id

        # Determine status: "crear" if both bodega and sede are new, else "actualizar"
        bodega_key = (empresa_id, bodega_nombre)
        sede_key = (empresa_id, sede_nombre)

        bodega_exists = bodega_key in existing_bodegas or bodega_key in seen_bodegas
        sede_exists = sede_key in existing_sedes or sede_key in seen_sedes

        # Status is "crear" only if BOTH are completely new
        status = "crear" if (not bodega_exists and not sede_exists) else "actualizar"

        # Track this row in seen_* dicts for within-file duplicate detection
        if bodega_key not in seen_bodegas:
            seen_bodegas[bodega_key] = row_num
        if sede_key not in seen_sedes:
            seen_sedes[sede_key] = row_num

        return (
            ParsedRow(
                row_num=row_num,
                empresa_rut=empresa_rut,
                bodega_nombre=bodega_nombre,
                bodega_direccion=bodega_direccion,
                sede_nombre=sede_nombre,
                sede_direccion=sede_direccion,
                status=status,
                errors=[],
            ),
            None,
        )

    @staticmethod
    def _validate_empresa_rut(value: Any, errors: list[str], existing_empresas: dict) -> Optional[str]:
        """Validate empresa RUT exists in database."""
        if value is None or str(value).strip() == "":
            errors.append("RUT empresa es requerido")
            return None

        rut = clean_rut(str(value).strip())

        if not validate_rut(rut):
            errors.append(f"RUT empresa inválido: {value}")
            return None

        if rut not in existing_empresas:
            errors.append(f"Empresa con RUT {rut} no encontrada en el sistema")
            return None

        return rut

    @staticmethod
    def _validate_bodega_nombre(value: Any, errors: list[str]) -> Optional[str]:
        """Validate and extract bodega_nombre (required)."""
        if value is None or str(value).strip() == "":
            errors.append("Nombre bodega es requerido")
            return None

        nombre = str(value).strip()
        if len(nombre) < 2:
            errors.append("Nombre bodega debe tener al menos 2 caracteres")
            return None

        return nombre

    @staticmethod
    def _validate_bodega_direccion(value: Any, errors: list[str]) -> Optional[str]:
        """Validate and extract bodega_direccion (optional)."""
        if value is None or str(value).strip() == "":
            return None

        return str(value).strip()

    @staticmethod
    def _validate_sede_nombre(value: Any, errors: list[str]) -> Optional[str]:
        """Validate and extract sede_nombre (required)."""
        if value is None or str(value).strip() == "":
            errors.append("Nombre sede es requerido")
            return None

        nombre = str(value).strip()
        if len(nombre) < 2:
            errors.append("Nombre sede debe tener al menos 2 caracteres")
            return None

        return nombre

    @staticmethod
    def _validate_sede_direccion(value: Any, errors: list[str]) -> Optional[str]:
        """Validate and extract sede_direccion (required)."""
        if value is None or str(value).strip() == "":
            errors.append("Dirección sede es requerida")
            return None

        direccion = str(value).strip()
        if len(direccion) < 2:
            errors.append("Dirección sede debe tener al menos 2 caracteres")
            return None

        return direccion

    @staticmethod
    def _read_xlsx(content: bytes) -> list[list]:
        """Read Excel file and return rows as lists."""
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            return [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

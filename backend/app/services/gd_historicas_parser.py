from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

import openpyxl


class ParseError(Exception):
    """Raised when file format is invalid (not recoverable per-row)."""
    pass


@dataclass
class GDLine:
    sku: Optional[str]
    cantidad: Decimal
    row_num: int


@dataclass
class GDGroup:
    folio_guia: int
    rut_receptor: Optional[str]
    fecha: date
    tipo_traslado: int
    sede_destino: Optional[str]
    folio_factura: Optional[int]
    lines: list[GDLine]
    status: str  # "crear" | "omitir"
    row_num: int  # cabecera row


@dataclass
class InvalidRow:
    row_num: int
    folio_guia_raw: Optional[str]
    sheet: str
    motivo: str


@dataclass
class ParseResult:
    valid_groups: list[GDGroup] = field(default_factory=list)
    invalid_rows: list[InvalidRow] = field(default_factory=list)
    a_crear: int = 0
    a_omitir: int = 0
    invalid_group_count: int = 0


class GDHistoricasParser:
    """Parse and validate GD históricas import files (two-sheet Excel)."""

    CABECERA_SHEET = "Cabecera GD"
    DETALLE_SHEET = "Detalle GD"

    CABECERA_REQUIRED = {"folio_guia", "fecha", "tipo_traslado"}
    CABECERA_OPTIONAL = {"rut_receptor", "sede_destino", "folio_factura"}

    DETALLE_REQUIRED = {"folio_guia", "cantidad"}
    DETALLE_OPTIONAL = {"sku"}

    @staticmethod
    def parse(
        content: bytes,
        filename: str,
        existing_folios: set[int],
    ) -> ParseResult:
        existing_folios = existing_folios or set()

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ParseError(f"No se pudo abrir el archivo Excel: {exc}")

        try:
            sheet_names = wb.sheetnames
            if GDHistoricasParser.CABECERA_SHEET not in sheet_names:
                raise ParseError(
                    f"No se encontró la hoja '{GDHistoricasParser.CABECERA_SHEET}'. "
                    f"Hojas encontradas: {sheet_names}"
                )
            if GDHistoricasParser.DETALLE_SHEET not in sheet_names:
                raise ParseError(
                    f"No se encontró la hoja '{GDHistoricasParser.DETALLE_SHEET}'. "
                    f"Hojas encontradas: {sheet_names}"
                )

            cab_rows = [list(row) for row in wb[GDHistoricasParser.CABECERA_SHEET].iter_rows(values_only=True)]
            det_rows = [list(row) for row in wb[GDHistoricasParser.DETALLE_SHEET].iter_rows(values_only=True)]
        finally:
            wb.close()

        if not cab_rows:
            raise ParseError(f"La hoja '{GDHistoricasParser.CABECERA_SHEET}' está vacía")

        result = ParseResult()

        # --- Parse cabecera headers ---
        cab_headers = [str(h).strip().lower() if h is not None else "" for h in cab_rows[0]]
        missing_cab = GDHistoricasParser.CABECERA_REQUIRED - set(cab_headers)
        if missing_cab:
            raise ParseError(
                f"Hoja '{GDHistoricasParser.CABECERA_SHEET}' — columna(s) requerida(s) faltante(s): "
                f"{sorted(missing_cab)}. Columnas encontradas: {[h for h in cab_headers if h]}"
            )
        cab_col = {h: i for i, h in enumerate(cab_headers) if h}

        # --- Parse detalle headers ---
        if not det_rows:
            raise ParseError(f"La hoja '{GDHistoricasParser.DETALLE_SHEET}' está vacía")
        det_headers = [str(h).strip().lower() if h is not None else "" for h in det_rows[0]]
        missing_det = GDHistoricasParser.DETALLE_REQUIRED - set(det_headers)
        if missing_det:
            raise ParseError(
                f"Hoja '{GDHistoricasParser.DETALLE_SHEET}' — columna(s) requerida(s) faltante(s): "
                f"{sorted(missing_det)}. Columnas encontradas: {[h for h in det_headers if h]}"
            )
        det_col = {h: i for i, h in enumerate(det_headers) if h}

        def _get_cab(row: list, col_name: str) -> Optional[str]:
            i = cab_col.get(col_name, -1)
            if i < 0 or i >= len(row):
                return None
            val = row[i]
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        def _get_det(row: list, col_name: str) -> Optional[str]:
            i = det_col.get(col_name, -1)
            if i < 0 or i >= len(row):
                return None
            val = row[i]
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        # --- Build detalle lines dict {folio_guia: [GDLine]} ---
        det_data = GDHistoricasParser._skip_doc_row(det_rows, det_col, "folio_guia")
        det_header_offset = len(det_rows) - len(det_data)
        lines_by_folio: dict[int, list[GDLine]] = {}
        det_invalid: list[InvalidRow] = []

        for idx, raw_row in enumerate(det_data):
            row_num = det_header_offset + idx + 1
            folio_raw = _get_det(raw_row, "folio_guia")
            errors = []

            parsed_folio = None
            if not folio_raw:
                errors.append("folio_guia es requerido")
            else:
                try:
                    parsed_folio = int(float(folio_raw))
                    if parsed_folio <= 0:
                        errors.append(f"folio_guia debe ser positivo (tiene {parsed_folio})")
                        parsed_folio = None
                except (ValueError, TypeError):
                    errors.append(f"folio_guia '{folio_raw}' no es un entero válido")

            cantidad_raw = _get_det(raw_row, "cantidad")
            parsed_cantidad = None
            if cantidad_raw is None:
                errors.append("cantidad es requerida")
            else:
                try:
                    parsed_cantidad = Decimal(str(cantidad_raw))
                    if parsed_cantidad <= 0:
                        errors.append(f"cantidad debe ser positiva (tiene {parsed_cantidad})")
                        parsed_cantidad = None
                except Exception:
                    errors.append(f"cantidad '{cantidad_raw}' no es un número válido")

            sku_raw = _get_det(raw_row, "sku")

            if errors:
                det_invalid.append(InvalidRow(
                    row_num=row_num,
                    folio_guia_raw=folio_raw,
                    sheet=GDHistoricasParser.DETALLE_SHEET,
                    motivo="; ".join(errors),
                ))
                continue

            line = GDLine(sku=sku_raw, cantidad=parsed_cantidad, row_num=row_num)
            lines_by_folio.setdefault(parsed_folio, []).append(line)

        result.invalid_rows.extend(det_invalid)

        # --- Parse cabecera rows ---
        cab_data = GDHistoricasParser._skip_doc_row(cab_rows, cab_col, "folio_guia")
        cab_header_offset = len(cab_rows) - len(cab_data)

        for idx, raw_row in enumerate(cab_data):
            row_num = cab_header_offset + idx + 1
            folio_raw = _get_cab(raw_row, "folio_guia")
            errors = []

            parsed_folio = None
            if not folio_raw:
                errors.append("folio_guia es requerido")
            else:
                try:
                    parsed_folio = int(float(folio_raw))
                    if parsed_folio <= 0:
                        errors.append(f"folio_guia debe ser positivo (tiene {parsed_folio})")
                        parsed_folio = None
                except (ValueError, TypeError):
                    errors.append(f"folio_guia '{folio_raw}' no es un entero válido")

            fecha_raw = _get_cab(raw_row, "fecha")
            parsed_fecha = None
            if not fecha_raw:
                errors.append("fecha es requerida")
            else:
                parsed_fecha = GDHistoricasParser._parse_date(fecha_raw)
                if parsed_fecha is None:
                    errors.append(f"fecha '{fecha_raw}' no tiene formato válido (YYYY-MM-DD o DD-MM-YYYY)")

            tipo_raw = _get_cab(raw_row, "tipo_traslado")
            parsed_tipo = None
            if not tipo_raw:
                errors.append("tipo_traslado es requerido")
            else:
                try:
                    parsed_tipo = int(float(tipo_raw))
                    if parsed_tipo < 1 or parsed_tipo > 9:
                        errors.append(f"tipo_traslado debe ser entre 1 y 9 (tiene {parsed_tipo})")
                        parsed_tipo = None
                except (ValueError, TypeError):
                    errors.append(f"tipo_traslado '{tipo_raw}' no es un entero válido")

            rut_raw = _get_cab(raw_row, "rut_receptor")
            sede_raw = _get_cab(raw_row, "sede_destino")

            folio_fac_raw = _get_cab(raw_row, "folio_factura")
            parsed_folio_fac = None
            if folio_fac_raw is not None:
                try:
                    parsed_folio_fac = int(float(folio_fac_raw))
                    if parsed_folio_fac <= 0:
                        errors.append(f"folio_factura debe ser positivo (tiene {parsed_folio_fac})")
                        parsed_folio_fac = None
                except (ValueError, TypeError):
                    errors.append(f"folio_factura '{folio_fac_raw}' no es un entero válido")

            if errors:
                result.invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    folio_guia_raw=folio_raw,
                    sheet=GDHistoricasParser.CABECERA_SHEET,
                    motivo="; ".join(errors),
                ))
                result.invalid_group_count += 1
                continue

            status = "omitir" if parsed_folio in existing_folios else "crear"

            group = GDGroup(
                folio_guia=parsed_folio,
                rut_receptor=rut_raw,
                fecha=parsed_fecha,
                tipo_traslado=parsed_tipo,
                sede_destino=sede_raw,
                folio_factura=parsed_folio_fac,
                lines=lines_by_folio.get(parsed_folio, []),
                status=status,
                row_num=row_num,
            )
            result.valid_groups.append(group)

            if status == "omitir":
                result.a_omitir += 1
            else:
                result.a_crear += 1

        return result

    @staticmethod
    def _skip_doc_row(raw_rows: list[list], col: dict, anchor_col: str) -> list[list]:
        if len(raw_rows) <= 1:
            return []
        anchor_idx = col.get(anchor_col, -1)
        potential_doc = raw_rows[1]
        anchor_val = (
            potential_doc[anchor_idx]
            if anchor_idx >= 0 and anchor_idx < len(potential_doc)
            else None
        )
        if anchor_val is None or str(anchor_val).strip() == "":
            return raw_rows[2:]
        val_str = str(anchor_val).strip()
        doc_keywords = ("requerido", "ej:", "ejemplo", "opcional", "folio")
        if any(kw in val_str.lower() for kw in doc_keywords) or len(val_str) > 200:
            return raw_rows[2:]
        return raw_rows[1:]

    @staticmethod
    def _parse_date(s: str) -> Optional[date]:
        try:
            serial = float(s)
            from datetime import timedelta
            epoch = date(1899, 12, 30)
            return epoch + timedelta(days=int(serial))
        except (ValueError, TypeError):
            pass
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def generate_template() -> bytes:
        wb = openpyxl.Workbook()

        header_font = openpyxl.styles.Font(bold=True, size=11)
        header_fill = openpyxl.styles.PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        doc_font = openpyxl.styles.Font(italic=True, size=9)
        doc_fill = openpyxl.styles.PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        )

        # --- Cabecera GD sheet ---
        ws_cab = wb.active
        ws_cab.title = GDHistoricasParser.CABECERA_SHEET

        cab_headers = ["folio_guia", "fecha", "tipo_traslado", "rut_receptor", "sede_destino", "folio_factura"]
        cab_docs = [
            "Requerido: número GD (entero positivo)",
            "Requerido: YYYY-MM-DD o DD-MM-YYYY",
            "Requerido: tipo traslado 1-9 (D-05)",
            "Opcional: RUT receptor (ej: 12345678-9)",
            "Opcional: dirección o nombre de sede destino",
            "Opcional: folio de factura asociada",
        ]
        cab_example = [1001, "2024-01-15", 1, "12345678-9", "Bodega Central", ""]

        for col_num, h in enumerate(cab_headers, 1):
            cell = ws_cab.cell(row=1, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for col_num, d in enumerate(cab_docs, 1):
            cell = ws_cab.cell(row=2, column=col_num, value=d)
            cell.font = doc_font
            cell.fill = doc_fill
        for col_idx, v in enumerate(cab_example, 1):
            ws_cab.cell(row=3, column=col_idx, value=v)

        for i, w in enumerate([14, 18, 18, 20, 30, 18], 1):
            ws_cab.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # --- Detalle GD sheet ---
        ws_det = wb.create_sheet(title=GDHistoricasParser.DETALLE_SHEET)

        det_headers = ["folio_guia", "sku", "cantidad"]
        det_docs = [
            "Requerido: folio de la GD (debe existir en Cabecera GD)",
            "Opcional: SKU del producto",
            "Requerido: cantidad (número positivo)",
        ]
        det_example = [1001, "PROD-001", 5]

        for col_num, h in enumerate(det_headers, 1):
            cell = ws_det.cell(row=1, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for col_num, d in enumerate(det_docs, 1):
            cell = ws_det.cell(row=2, column=col_num, value=d)
            cell.font = doc_font
            cell.fill = doc_fill
        for col_idx, v in enumerate(det_example, 1):
            ws_det.cell(row=3, column=col_idx, value=v)

        for i, w in enumerate([14, 20, 14], 1):
            ws_det.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

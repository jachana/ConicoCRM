from __future__ import annotations

import io
from dataclasses import dataclass, field

import openpyxl


REQUIRED_COLUMNS = ("sku", "nombre", "precio_base", "costo", "iva")
OPTIONAL_COLUMNS = ("descripcion", "familia", "unidad", "afecto")
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS


class ParseError(Exception):
    pass


@dataclass
class ParsedProducto:
    fila: int
    sku_raw: str
    sku_normalizado: str
    nombre: str
    precio_base: str  # Will be converted to Decimal in API
    costo: str  # Will be converted to Decimal in API
    descripcion: str | None = None
    familia: str | None = None
    unidad: str | None = None
    iva: int = 19
    afecto: bool = True  # Default to True


@dataclass
class InvalidRow:
    fila: int
    sku_raw: str | None
    nombre_raw: str | None
    motivo: str


@dataclass
class ParseResult:
    validas: list[ParsedProducto] = field(default_factory=list)
    invalidas: list[InvalidRow] = field(default_factory=list)


def _normalizar_sku(sku_raw: str) -> str:
    """Normalize SKU: uppercase, remove spaces."""
    if not sku_raw:
        return ""
    return sku_raw.replace(" ", "").upper()


def _str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _try_float(val: str) -> float | None:
    """Try to parse a string as float."""
    if not val:
        return None
    try:
        return float(val)
    except ValueError:
        return None


def _try_int(val: str) -> int | None:
    """Try to parse a string as int."""
    if not val:
        return None
    try:
        return int(val)
    except ValueError:
        return None


def parse_productos_xlsx(content: bytes) -> ParseResult:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — surface as ParseError
        raise ParseError(f"No se pudo leer el archivo xlsx: {e}")

    try:
        ws = wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if not rows:
        raise ParseError("El archivo no contiene filas")

    header = [_str(h).lower() for h in rows[0]]
    missing = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing:
        raise ParseError(
            f"Columnas requeridas no encontradas: {missing}. Columnas encontradas: {header}"
        )

    def get(row: list, col: str) -> str:
        if col not in header:
            return ""
        i = header.index(col)
        return _str(row[i]) if i < len(row) else ""

    result = ParseResult()
    seen_skus: dict[str, int] = {}

    for idx, raw in enumerate(rows[1:], start=2):
        sku_raw = get(raw, "sku")
        nombre = get(raw, "nombre")
        precio_base = get(raw, "precio_base")
        costo = get(raw, "costo")
        iva_str = get(raw, "iva")

        # Skip completely blank rows
        if not sku_raw and not nombre and not any(_str(c) for c in raw):
            continue

        # Validate required fields
        if not sku_raw:
            result.invalidas.append(InvalidRow(fila=idx, sku_raw=None, nombre_raw=nombre or None, motivo="SKU vacío"))
            continue
        if not nombre:
            result.invalidas.append(InvalidRow(fila=idx, sku_raw=sku_raw, nombre_raw=None, motivo="Nombre vacío"))
            continue

        # Parse numeric fields
        precio_base_float = _try_float(precio_base)
        if precio_base_float is None:
            result.invalidas.append(
                InvalidRow(fila=idx, sku_raw=sku_raw, nombre_raw=nombre, motivo=f"Precio base inválido: '{precio_base}'")
            )
            continue
        if precio_base_float < 0:
            result.invalidas.append(
                InvalidRow(fila=idx, sku_raw=sku_raw, nombre_raw=nombre, motivo=f"Precio base debe ser >= 0: {precio_base_float}")
            )
            continue

        costo_float = _try_float(costo)
        if costo_float is None:
            result.invalidas.append(
                InvalidRow(fila=idx, sku_raw=sku_raw, nombre_raw=nombre, motivo=f"Costo inválido: '{costo}'")
            )
            continue
        if costo_float < 0:
            result.invalidas.append(
                InvalidRow(fila=idx, sku_raw=sku_raw, nombre_raw=nombre, motivo=f"Costo debe ser >= 0: {costo_float}")
            )
            continue

        iva_int = _try_int(iva_str)
        if iva_int is None:
            result.invalidas.append(
                InvalidRow(fila=idx, sku_raw=sku_raw, nombre_raw=nombre, motivo=f"IVA inválido: '{iva_str}'")
            )
            continue
        if iva_int not in (0, 19):
            result.invalidas.append(
                InvalidRow(fila=idx, sku_raw=sku_raw, nombre_raw=nombre, motivo=f"IVA debe ser 0 o 19: {iva_int}")
            )
            continue

        # Normalize SKU
        sku_norm = _normalizar_sku(sku_raw)
        if not sku_norm:
            result.invalidas.append(
                InvalidRow(fila=idx, sku_raw=sku_raw, nombre_raw=nombre, motivo="SKU vacío después de normalización")
            )
            continue

        # Check for duplicates in the same file (keep last row)
        if sku_norm in seen_skus:
            # Mark the previous occurrence as invalid
            # We'll actually just overwrite the last seen index
            pass
        seen_skus[sku_norm] = idx

        # Get optional fields
        descripcion = get(raw, "descripcion") or None
        familia = get(raw, "familia") or None
        unidad = get(raw, "unidad") or None
        afecto_str = get(raw, "afecto") or "1"
        afecto = afecto_str.lower() not in ("0", "false", "no")

        result.validas.append(
            ParsedProducto(
                fila=idx,
                sku_raw=sku_raw,
                sku_normalizado=sku_norm,
                nombre=nombre,
                descripcion=descripcion,
                familia=familia,
                unidad=unidad,
                precio_base=precio_base,
                costo=costo,
                iva=iva_int,
                afecto=afecto,
            )
        )

    # Filter out duplicates - keep only the last occurrence of each SKU
    if result.validas:
        last_by_sku = {}
        for p in result.validas:
            last_by_sku[p.sku_normalizado] = p

        # Find which SKUs had duplicates
        skus_with_dupes = {sku: filas for sku, filas in {p.sku_normalizado: [] for p in result.validas}.items()}
        for p in result.validas:
            skus_with_dupes[p.sku_normalizado].append(p.fila)

        # Mark earlier occurrences as invalid
        seen_final = set()
        validas_dedupe = []
        invalidas_to_add = []

        for p in result.validas:
            if p.sku_normalizado not in seen_final:
                validas_dedupe.append(p)
                seen_final.add(p.sku_normalizado)
            else:
                invalidas_to_add.append(
                    InvalidRow(
                        fila=p.fila,
                        sku_raw=p.sku_raw,
                        nombre_raw=p.nombre,
                        motivo=f"SKU duplicado en archivo (también en fila {last_by_sku[p.sku_normalizado].fila})",
                    )
                )

        result.validas = validas_dedupe
        result.invalidas.extend(invalidas_to_add)

    return result


def build_template_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(list(ALL_COLUMNS))
    ws.append([
        "SKU-001",
        "Producto Ejemplo",
        "Descripción del producto",
        "Categoría",
        "1.0",
        "100.00",
        "50.00",
        "19",
        "1",
    ])
    for i, col in enumerate(ALL_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

import openpyxl


class ParseError(Exception):
    """Raised when file format is invalid (not recoverable per-row)."""
    pass


@dataclass
class NCNDRow:
    tipo: str  # "NC" or "ND"
    folio: int
    fecha: date
    motivo: str
    neto: Decimal
    iva: Decimal
    total: Decimal
    folio_referencia: Optional[int]
    tipo_referencia: Optional[int]
    status: str  # "crear" | "omitir"
    row_num: int


@dataclass
class InvalidRow:
    row_num: int
    folio_raw: Optional[str]
    tipo_raw: Optional[str]
    motivo: str  # error description


@dataclass
class ParseResult:
    valid_rows: list[NCNDRow] = field(default_factory=list)
    invalid_rows: list[InvalidRow] = field(default_factory=list)
    a_crear: int = 0
    a_omitir: int = 0
    invalid_count: int = 0


class NCHistoricasParser:
    """Parse and validate NC/ND históricas import files (single-sheet Excel)."""

    SHEET_NAME = "NC-ND Históricas"

    REQUIRED_COLS = {"tipo", "folio", "fecha", "motivo", "neto", "iva", "total"}
    OPTIONAL_COLS = {"folio_referencia", "tipo_referencia"}

    @staticmethod
    def parse(
        content: bytes,
        filename: str,
        existing_folios: dict[str, set[int]],  # {"NC": {1,2,3}, "ND": {4,5}}
    ) -> ParseResult:
        """
        Parse and validate a single-sheet Excel file of historical NC/ND.

        Args:
            content: Excel file bytes
            filename: Original filename (unused, kept for API symmetry)
            existing_folios: Dict mapping tipo -> set of folio integers already in DB

        Returns:
            ParseResult with valid_rows, invalid_rows, and counts

        Raises:
            ParseError: If file format is invalid or required columns are missing
        """
        existing_folios = existing_folios or {}

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ParseError(f"No se pudo abrir el archivo Excel: {exc}")

        try:
            sheet_names = wb.sheetnames
            if NCHistoricasParser.SHEET_NAME not in sheet_names:
                raise ParseError(
                    f"No se encontró la hoja '{NCHistoricasParser.SHEET_NAME}'. "
                    f"Hojas encontradas: {sheet_names}"
                )

            raw_rows = [list(row) for row in wb[NCHistoricasParser.SHEET_NAME].iter_rows(values_only=True)]
        finally:
            wb.close()

        if not raw_rows:
            raise ParseError(f"La hoja '{NCHistoricasParser.SHEET_NAME}' está vacía")

        result = ParseResult()

        header_row = raw_rows[0]
        # Case/accent-insensitive header matching — normalize to lowercase stripped
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in header_row
        ]

        missing = NCHistoricasParser.REQUIRED_COLS - set(headers)
        if missing:
            raise ParseError(
                f"Hoja '{NCHistoricasParser.SHEET_NAME}' — columna(s) requerida(s) faltante(s): "
                f"{sorted(missing)}. Columnas encontradas: {[h for h in headers if h]}"
            )

        col = {h: i for i, h in enumerate(headers) if h}

        data_rows = NCHistoricasParser._skip_doc_row(raw_rows, col, "folio")
        header_offset = len(raw_rows) - len(data_rows)

        def _get(row: list, col_name: str) -> Optional[str]:
            i = col.get(col_name, -1)
            if i < 0 or i >= len(row):
                return None
            val = row[i]
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        for idx, raw_row in enumerate(data_rows):
            row_num = header_offset + idx + 1
            folio_raw = _get(raw_row, "folio")
            tipo_raw = _get(raw_row, "tipo")
            errors = []

            # tipo — must be "NC" or "ND"
            parsed_tipo = None
            if not tipo_raw:
                errors.append("tipo es requerido")
            else:
                normalized_tipo = tipo_raw.upper()
                if normalized_tipo not in ("NC", "ND"):
                    errors.append(f"tipo '{tipo_raw}' no es válido (debe ser NC o ND)")
                else:
                    parsed_tipo = normalized_tipo

            # folio — positive integer
            parsed_folio = None
            if not folio_raw:
                errors.append("folio es requerido")
            else:
                try:
                    parsed_folio = int(float(folio_raw))
                    if parsed_folio <= 0:
                        errors.append(f"folio debe ser un entero positivo (tiene {parsed_folio})")
                        parsed_folio = None
                except (ValueError, TypeError):
                    errors.append(f"folio '{folio_raw}' no es un entero válido")

            # fecha
            fecha_raw = _get(raw_row, "fecha")
            parsed_fecha = None
            if not fecha_raw:
                errors.append("fecha es requerida")
            else:
                parsed_fecha = NCHistoricasParser._parse_date(fecha_raw)
                if parsed_fecha is None:
                    errors.append(
                        f"fecha '{fecha_raw}' no tiene formato válido (YYYY-MM-DD o DD-MM-YYYY)"
                    )

            # motivo
            motivo_raw = _get(raw_row, "motivo")
            if not motivo_raw:
                errors.append("motivo es requerido")
            elif len(motivo_raw) > 500:
                errors.append(f"motivo excede 500 caracteres (tiene {len(motivo_raw)})")
                motivo_raw = None

            # neto
            neto_raw = _get(raw_row, "neto")
            parsed_neto = None
            if neto_raw is None:
                errors.append("neto es requerido")
            else:
                try:
                    parsed_neto = Decimal(str(neto_raw))
                    if parsed_neto < 0:
                        errors.append(f"neto debe ser no-negativo (tiene {parsed_neto})")
                        parsed_neto = None
                except Exception:
                    errors.append(f"neto '{neto_raw}' no es un número válido")

            # iva
            iva_raw = _get(raw_row, "iva")
            parsed_iva = None
            if iva_raw is None:
                errors.append("iva es requerido")
            else:
                try:
                    parsed_iva = Decimal(str(iva_raw))
                    if parsed_iva < 0:
                        errors.append(f"iva debe ser no-negativo (tiene {parsed_iva})")
                        parsed_iva = None
                except Exception:
                    errors.append(f"iva '{iva_raw}' no es un número válido")

            # total
            total_raw = _get(raw_row, "total")
            parsed_total = None
            if total_raw is None:
                errors.append("total es requerido")
            else:
                try:
                    parsed_total = Decimal(str(total_raw))
                    if parsed_total < 0:
                        errors.append(f"total debe ser no-negativo (tiene {parsed_total})")
                        parsed_total = None
                except Exception:
                    errors.append(f"total '{total_raw}' no es un número válido")

            # neto + iva ≈ total (tolerance ±1.00)
            if parsed_neto is not None and parsed_iva is not None and parsed_total is not None:
                computed_total = parsed_neto + parsed_iva
                diff = abs(computed_total - parsed_total)
                if diff > Decimal("1"):
                    errors.append(
                        f"neto + iva ({computed_total}) no coincide con total ({parsed_total}), "
                        f"diferencia: ${diff:.0f}"
                    )

            # folio_referencia (optional)
            folio_ref_raw = _get(raw_row, "folio_referencia")
            parsed_folio_ref = None
            if folio_ref_raw is not None:
                try:
                    parsed_folio_ref = int(float(folio_ref_raw))
                    if parsed_folio_ref <= 0:
                        errors.append(
                            f"folio_referencia debe ser un entero positivo (tiene {parsed_folio_ref})"
                        )
                        parsed_folio_ref = None
                except (ValueError, TypeError):
                    errors.append(f"folio_referencia '{folio_ref_raw}' no es un entero válido")

            # tipo_referencia (optional)
            tipo_ref_raw = _get(raw_row, "tipo_referencia")
            parsed_tipo_ref = None
            if tipo_ref_raw is not None:
                try:
                    parsed_tipo_ref = int(float(tipo_ref_raw))
                except (ValueError, TypeError):
                    errors.append(f"tipo_referencia '{tipo_ref_raw}' no es un entero válido")

            if errors:
                result.invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    folio_raw=folio_raw,
                    tipo_raw=tipo_raw,
                    motivo="; ".join(errors),
                ))
                continue

            # Idempotency check
            existing_set = existing_folios.get(parsed_tipo, set())
            status = "omitir" if parsed_folio in existing_set else "crear"

            result.valid_rows.append(NCNDRow(
                tipo=parsed_tipo,
                folio=parsed_folio,
                fecha=parsed_fecha,
                motivo=motivo_raw,
                neto=parsed_neto,
                iva=parsed_iva,
                total=parsed_total,
                folio_referencia=parsed_folio_ref,
                tipo_referencia=parsed_tipo_ref,
                status=status,
                row_num=row_num,
            ))

            if status == "omitir":
                result.a_omitir += 1
            else:
                result.a_crear += 1

        result.invalid_count = len(result.invalid_rows)
        return result

    @staticmethod
    def _skip_doc_row(raw_rows: list[list], col: dict, anchor_col: str) -> list[list]:
        """
        Skip row index 1 (the row immediately after headers) if it looks like
        a documentation/example-label row rather than real data.
        """
        if len(raw_rows) <= 1:
            return []
        anchor_idx = col.get(anchor_col, -1)
        potential_doc = raw_rows[1]
        anchor_val = (
            potential_doc[anchor_idx]
            if anchor_idx >= 0 and anchor_idx < len(potential_doc)
            else None
        )
        if anchor_val is None or str(anchor_val).strip() == "":
            return raw_rows[2:]
        val_str = str(anchor_val).strip()
        doc_keywords = ("requerido", "ej:", "ejemplo", "opcional", "folio")
        if any(kw in val_str.lower() for kw in doc_keywords) or len(val_str) > 200:
            return raw_rows[2:]
        return raw_rows[1:]

    @staticmethod
    def _parse_date(s: str) -> Optional[date]:
        """Parse date string in YYYY-MM-DD, DD-MM-YYYY, or DD/MM/YYYY format.
        Also handles Excel date serial numbers (int/float)."""
        # Handle Excel date serial numbers
        try:
            serial = float(s)
            from datetime import timedelta
            epoch = date(1899, 12, 30)
            return epoch + timedelta(days=int(serial))
        except (ValueError, TypeError):
            pass

        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def generate_template() -> bytes:
        """
        Generate a single-sheet Excel template: 'NC-ND Históricas'.
        Has a bold/gray header row, an italic/light documentation row,
        and an example row.
        """
        wb = openpyxl.Workbook()

        header_font = openpyxl.styles.Font(bold=True, size=11)
        header_fill = openpyxl.styles.PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        doc_font = openpyxl.styles.Font(italic=True, size=9)
        doc_fill = openpyxl.styles.PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        )

        ws = wb.active
        ws.title = NCHistoricasParser.SHEET_NAME

        headers = [
            "tipo",
            "folio",
            "fecha",
            "motivo",
            "neto",
            "iva",
            "total",
            "folio_referencia",
            "tipo_referencia",
        ]
        documentation = [
            "Requerido: NC o ND",
            "Requerido: número DTE (entero positivo)",
            "Requerido: YYYY-MM-DD o DD-MM-YYYY",
            "Requerido: razón de la NC/ND (máx 500 caracteres)",
            "Requerido: monto neto (≥ 0)",
            "Requerido: monto IVA (≥ 0)",
            "Requerido: monto total (≥ 0)",
            "Opcional: folio del documento referenciado",
            "Opcional: tipo DTE referenciado (ej: 33=factura, 39=boleta)",
        ]
        example = [
            "NC", 1001, "2024-01-15", "Corrección precio",
            100000, 19000, 119000, 500, 33,
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for col_num, doc in enumerate(documentation, 1):
            cell = ws.cell(row=2, column=col_num, value=doc)
            cell.font = doc_font
            cell.fill = doc_fill

        for col_idx, val in enumerate(example, 1):
            ws.cell(row=3, column=col_idx, value=val)

        col_widths = [10, 12, 18, 40, 16, 14, 14, 22, 22]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

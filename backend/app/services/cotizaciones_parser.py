"""
Cotizaciones Import Parser

Handles parsing and validation of Excel/CSV files containing open/pending
Cotizacion data during onboarding. Provides:
- Excel (.xlsx) and CSV (.csv) file parsing
- Row-level validation with detailed error messages
- Grouping of rows by numero_cot into Cotizacion groups
- Template generation with openpyxl
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Optional

import openpyxl


class ParseError(Exception):
    """Raised when file format is invalid (not recoverable per-row)."""
    pass


@dataclass
class CotizacionLine:
    """Single validated Cotizacion line item."""
    sku: Optional[str]
    descripcion: str
    formato: Optional[str]
    cantidad: int
    precio: Decimal  # = valor_neto unitario
    descuento: Decimal  # 0-100 percent
    total_neto_linea: Decimal


@dataclass
class CotizacionGroup:
    """A group of rows that form a single Cotizacion."""
    numero_cot: Optional[str]
    rut_cliente: str
    rut_empresa: Optional[str]
    fecha: date
    vigencia_hasta: Optional[date]
    estado: str
    vendedor_email: str
    nota: Optional[str]
    lines: list[CotizacionLine]
    total_neto: Decimal
    total_iva: Decimal
    total: Decimal
    status: str  # "crear" | "omitir"
    row_nums: list[int]  # row numbers in the file that belong to this group


@dataclass
class InvalidRow:
    """Single invalid row with error details."""
    row_num: int
    numero_cot_raw: Optional[str]
    motivo: str


@dataclass
class ParseResult:
    """Results from parsing a Cotizaciones file."""
    valid_groups: list[CotizacionGroup] = field(default_factory=list)
    invalid_rows: list[InvalidRow] = field(default_factory=list)
    a_crear: int = 0
    a_omitir: int = 0
    invalid_group_count: int = 0


class CotizacionesParser:
    """Parse and validate Cotizaciones import files."""

    REQUIRED_COLUMNS = {"rut_cliente", "fecha", "vendedor_email", "descripcion", "cantidad", "precio"}

    OPTIONAL_COLUMNS = {
        "numero_cot", "rut_empresa", "vigencia_hasta", "estado",
        "nota", "sku", "formato", "descuento",
    }

    ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS

    VALID_ESTADOS = {"no_definido", "enviada", "aceptada", "rechazada", "vencida"}

    @staticmethod
    def parse(
        content: bytes,
        filename: str,
        clientes_by_rut: Optional[dict] = None,
        empresas_by_rut: Optional[dict] = None,
        vendedores_by_email: Optional[dict] = None,
        productos_by_sku: Optional[dict] = None,
        existing_numeros: Optional[set] = None,
    ) -> ParseResult:
        """
        Parse and validate a Cotizaciones import file.

        Args:
            content: File bytes
            filename: Original filename (used to detect format: .xlsx or .csv)
            clientes_by_rut: Dict mapping rut -> Cliente id (for lookups)
            empresas_by_rut: Dict mapping rut -> Empresa id (for lookups)
            vendedores_by_email: Dict mapping email -> User id (for lookups)
            productos_by_sku: Dict mapping sku -> Producto id (for lookups)
            existing_numeros: Set of integers already in DB as Cotizacion.numero

        Returns:
            ParseResult with valid_groups, invalid_rows, and counts

        Raises:
            ParseError: If file format is invalid or required column missing
        """
        clientes_by_rut = clientes_by_rut or {}
        empresas_by_rut = empresas_by_rut or {}
        vendedores_by_email = vendedores_by_email or {}
        productos_by_sku = productos_by_sku or {}
        existing_numeros = existing_numeros or set()

        parser = CotizacionesParser()
        ext = Path(filename).suffix.lower()

        if ext == ".xlsx":
            raw_rows = parser._read_xlsx(content)
        elif ext == ".csv":
            raw_rows = parser._read_csv(content)
        else:
            raise ParseError(f"Formato no soportado: '{ext}'. Use .xlsx o .csv")

        if not raw_rows:
            raise ParseError("El archivo está vacío")

        # Row 0: headers
        header_row = raw_rows[0]
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in header_row
        ]

        # Find required columns
        missing = CotizacionesParser.REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ParseError(
                f"Columna(s) requerida(s) faltante(s): {sorted(missing)}. "
                f"Columnas encontradas: {[h for h in headers if h]}"
            )

        # Build column index map
        col = {h: i for i, h in enumerate(headers) if h}

        # Auto-detect documentation row (row 1): skip if descripcion cell looks like a doc row
        data_rows = raw_rows[1:]
        if len(raw_rows) > 1:
            desc_idx = col.get("descripcion", -1)
            potential_doc = raw_rows[1]
            desc_val = (
                potential_doc[desc_idx]
                if desc_idx >= 0 and desc_idx < len(potential_doc)
                else None
            )
            if desc_val is None or str(desc_val).strip() == "":
                data_rows = raw_rows[2:]
            else:
                val_str = str(desc_val).strip()
                doc_keywords = ("descripcion", "requerido", "ej:", "ejemplo", "descripción")
                if any(kw in val_str.lower() for kw in doc_keywords) or len(val_str) > 200:
                    data_rows = raw_rows[2:]

        if not data_rows:
            raise ParseError("El archivo no contiene filas de datos")

        header_offset = len(raw_rows) - len(data_rows)

        # ----------------------------------------------------------------
        # Pass 1: Validate rows individually and collect into raw groups
        # ----------------------------------------------------------------
        # raw_groups[key] -> {"header": ..., "rows": [...], "invalid": [...]}
        # key = numero_cot value or None-with-row-num for ungrouped rows
        raw_groups: dict = {}
        # Track row order for groups so we maintain file order
        group_order: list = []
        invalid_rows_global: list[InvalidRow] = []

        def _get(row: list, col_name: str) -> Optional[str]:
            """Get raw cell value from row as string, or None."""
            i = col.get(col_name, -1)
            if i < 0 or i >= len(row):
                return None
            val = row[i]
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        for idx, raw_row in enumerate(data_rows):
            row_num = header_offset + idx + 1  # 1-based

            # Extract fields
            numero_cot_raw = _get(raw_row, "numero_cot")
            rut_cliente_raw = _get(raw_row, "rut_cliente")
            rut_empresa_raw = _get(raw_row, "rut_empresa")
            fecha_raw = _get(raw_row, "fecha")
            vigencia_hasta_raw = _get(raw_row, "vigencia_hasta")
            estado_raw = _get(raw_row, "estado")
            vendedor_email_raw = _get(raw_row, "vendedor_email")
            nota_raw = _get(raw_row, "nota")
            sku_raw = _get(raw_row, "sku")
            descripcion_raw = _get(raw_row, "descripcion")
            formato_raw = _get(raw_row, "formato")
            cantidad_raw = _get(raw_row, "cantidad")
            precio_raw = _get(raw_row, "precio")
            descuento_raw = _get(raw_row, "descuento")

            # Use numero_cot as group key; rows without numero_cot each form their own group
            group_key = numero_cot_raw if numero_cot_raw is not None else f"__row_{row_num}__"

            # Initialize group record
            if group_key not in raw_groups:
                raw_groups[group_key] = {
                    "numero_cot": numero_cot_raw,
                    "valid_rows": [],
                    "invalid_row_nums": [],
                }
                group_order.append(group_key)

            # ---- Validate row ----
            errors = []

            # rut_cliente
            if not rut_cliente_raw:
                errors.append("rut_cliente es requerido")
            elif rut_cliente_raw not in clientes_by_rut:
                errors.append(f"Cliente con rut '{rut_cliente_raw}' no encontrado")

            # rut_empresa (optional but must exist if provided)
            if rut_empresa_raw and rut_empresa_raw not in empresas_by_rut:
                errors.append(f"Empresa con rut '{rut_empresa_raw}' no encontrada")

            # fecha
            parsed_fecha = None
            if not fecha_raw:
                errors.append("fecha es requerida")
            else:
                parsed_fecha = CotizacionesParser._parse_date(fecha_raw)
                if parsed_fecha is None:
                    errors.append(f"fecha '{fecha_raw}' no tiene formato válido (YYYY-MM-DD o DD/MM/YYYY)")

            # vigencia_hasta (optional)
            parsed_vigencia_hasta = None
            if vigencia_hasta_raw:
                parsed_vigencia_hasta = CotizacionesParser._parse_date(vigencia_hasta_raw)
                if parsed_vigencia_hasta is None:
                    errors.append(f"vigencia_hasta '{vigencia_hasta_raw}' no tiene formato válido (YYYY-MM-DD o DD/MM/YYYY)")

            # vendedor_email (required, case-insensitive lookup)
            vendedor_email_norm = None
            if not vendedor_email_raw:
                errors.append("vendedor_email es requerido")
            else:
                for k in vendedores_by_email:
                    if k.lower() == vendedor_email_raw.lower():
                        vendedor_email_norm = k
                        break
                if vendedor_email_norm is None:
                    errors.append(f"Vendedor con email '{vendedor_email_raw}' no encontrado")

            # estado (optional, defaults to 'no_definido')
            estado = "no_definido"
            if estado_raw:
                if estado_raw.lower() not in CotizacionesParser.VALID_ESTADOS:
                    valid_str = ", ".join(f"'{v}'" for v in sorted(CotizacionesParser.VALID_ESTADOS))
                    errors.append(f"estado '{estado_raw}' no válido (use {valid_str})")
                else:
                    estado = estado_raw.lower()

            # descripcion
            if not descripcion_raw:
                errors.append("descripcion es requerida")
            elif len(descripcion_raw) > 500:
                errors.append(f"descripcion excede 500 caracteres (tiene {len(descripcion_raw)})")

            # formato
            if formato_raw and len(formato_raw) > 50:
                errors.append(f"formato excede 50 caracteres (tiene {len(formato_raw)})")

            # cantidad
            parsed_cantidad = None
            if not cantidad_raw:
                errors.append("cantidad es requerida")
            else:
                try:
                    parsed_cantidad = int(float(cantidad_raw))
                    if parsed_cantidad <= 0:
                        errors.append(f"cantidad debe ser un entero positivo (tiene {parsed_cantidad})")
                        parsed_cantidad = None
                except (ValueError, TypeError):
                    errors.append(f"cantidad '{cantidad_raw}' no es un número válido")

            # precio (= valor_neto unitario)
            parsed_precio = None
            if precio_raw is None:
                errors.append("precio es requerido")
            else:
                try:
                    parsed_precio = Decimal(str(precio_raw))
                    if parsed_precio < 0:
                        errors.append(f"precio debe ser no-negativo (tiene {parsed_precio})")
                        parsed_precio = None
                except Exception:
                    errors.append(f"precio '{precio_raw}' no es un número válido")

            # descuento (optional, 0-100)
            parsed_descuento = Decimal("0")
            if descuento_raw is not None:
                try:
                    parsed_descuento = Decimal(str(descuento_raw))
                    if parsed_descuento < 0 or parsed_descuento > 100:
                        errors.append(f"descuento '{descuento_raw}' debe estar entre 0 y 100")
                        parsed_descuento = Decimal("0")
                except Exception:
                    errors.append(f"descuento '{descuento_raw}' no es un número válido")

            # sku (optional but must exist if provided)
            if sku_raw and sku_raw not in productos_by_sku:
                errors.append(f"Producto con sku '{sku_raw}' no encontrado")

            if errors:
                raw_groups[group_key]["invalid_row_nums"].append(row_num)
                invalid_rows_global.append(InvalidRow(
                    row_num=row_num,
                    numero_cot_raw=numero_cot_raw,
                    motivo="; ".join(errors),
                ))
            else:
                # Calculate line total with discount applied
                total_neto_linea = Decimal("0")
                if parsed_cantidad is not None and parsed_precio is not None:
                    total_neto_linea = (
                        Decimal(str(parsed_cantidad))
                        * parsed_precio
                        * (Decimal("1") - parsed_descuento / Decimal("100"))
                    )

                raw_groups[group_key]["valid_rows"].append({
                    "row_num": row_num,
                    "rut_cliente": rut_cliente_raw,
                    "rut_empresa": rut_empresa_raw,
                    "fecha": parsed_fecha,
                    "vigencia_hasta": parsed_vigencia_hasta,
                    "estado": estado,
                    "vendedor_email": vendedor_email_norm,
                    "nota": nota_raw,
                    "sku": sku_raw,
                    "descripcion": descripcion_raw,
                    "formato": formato_raw,
                    "cantidad": parsed_cantidad,
                    "precio": parsed_precio,
                    "descuento": parsed_descuento,
                    "total_neto_linea": total_neto_linea,
                })

        # ----------------------------------------------------------------
        # Pass 2: Build CotizacionGroups from raw_groups
        # ----------------------------------------------------------------
        result = ParseResult()

        for group_key in group_order:
            gdata = raw_groups[group_key]
            numero_cot = gdata["numero_cot"]
            invalid_row_nums = gdata["invalid_row_nums"]
            valid_rows = gdata["valid_rows"]

            # If any row in the group is invalid, the whole group is invalid
            if invalid_row_nums:
                result.invalid_group_count += 1
                continue

            if not valid_rows:
                continue

            # Use header data from first row
            first = valid_rows[0]

            # Check consistency of header fields across rows in same group
            group_errors = []
            for vr in valid_rows[1:]:
                if vr["rut_cliente"] != first["rut_cliente"]:
                    group_errors.append(
                        f"fila {vr['row_num']}: rut_cliente inconsistente con fila {first['row_num']}"
                    )
                if vr["fecha"] != first["fecha"]:
                    group_errors.append(
                        f"fila {vr['row_num']}: fecha inconsistente con fila {first['row_num']}"
                    )
                if vr["vendedor_email"] != first["vendedor_email"]:
                    group_errors.append(
                        f"fila {vr['row_num']}: vendedor_email inconsistente con fila {first['row_num']}"
                    )

            if group_errors:
                for vr in valid_rows:
                    invalid_rows_global.append(InvalidRow(
                        row_num=vr["row_num"],
                        numero_cot_raw=numero_cot,
                        motivo="Grupo Cotizacion inconsistente: " + "; ".join(group_errors),
                    ))
                result.invalid_group_count += 1
                continue

            # Build lines
            lines = []
            for vr in valid_rows:
                lines.append(CotizacionLine(
                    sku=vr["sku"],
                    descripcion=vr["descripcion"],
                    formato=vr["formato"],
                    cantidad=vr["cantidad"],
                    precio=vr["precio"],
                    descuento=vr["descuento"],
                    total_neto_linea=vr["total_neto_linea"],
                ))

            total_neto = sum(line_item.total_neto_linea for line_item in lines)
            total_iva = (total_neto * Decimal("0.19")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            total = total_neto + total_iva

            # Determine status: omitir if numero_cot is integer and already in DB
            status = "crear"
            if numero_cot is not None:
                try:
                    num_int = int(numero_cot)
                    if num_int in existing_numeros:
                        status = "omitir"
                except (ValueError, TypeError):
                    pass  # non-integer numero_cot → always crear

            group = CotizacionGroup(
                numero_cot=numero_cot,
                rut_cliente=first["rut_cliente"],
                rut_empresa=first["rut_empresa"],
                fecha=first["fecha"],
                vigencia_hasta=first["vigencia_hasta"],
                estado=first["estado"],
                vendedor_email=first["vendedor_email"],
                nota=first["nota"],
                lines=lines,
                total_neto=total_neto,
                total_iva=total_iva,
                total=total,
                status=status,
                row_nums=[vr["row_num"] for vr in valid_rows],
            )
            result.valid_groups.append(group)
            if status == "omitir":
                result.a_omitir += 1
            else:
                result.a_crear += 1

        result.invalid_rows = invalid_rows_global
        return result

    @staticmethod
    def _parse_date(s: str) -> Optional[date]:
        """Parse date string in YYYY-MM-DD or DD/MM/YYYY format."""
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def generate_template() -> bytes:
        """
        Generate Excel template with header row (bold/gray), documentation row
        (italic/light) and 3 example rows (2 lines of COT-001, 1 line of COT-002).

        Returns:
            Excel file bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cotizaciones Abiertas"

        headers = [
            "numero_cot",
            "rut_cliente",
            "rut_empresa",
            "fecha",
            "vigencia_hasta",
            "vendedor_email",
            "estado",
            "nota",
            "sku",
            "descripcion",
            "formato",
            "cantidad",
            "precio",
            "descuento",
        ]
        documentation = [
            "Opcional; filas con mismo valor se agrupan en una Cotizacion",
            "Requerido; RUT del cliente (debe existir)",
            "Opcional; RUT de la empresa (debe existir)",
            "Requerido; YYYY-MM-DD o DD/MM/YYYY",
            "Opcional; YYYY-MM-DD o DD/MM/YYYY",
            "Requerido; email del vendedor (debe existir)",
            "Opcional; 'no_definido' (default), 'enviada', 'aceptada', 'rechazada', 'vencida'",
            "Opcional; texto libre",
            "Opcional; SKU del producto (debe existir)",
            "Requerido; descripción del ítem (máx. 500)",
            "Opcional; formato del ítem (máx. 50)",
            "Requerido; entero positivo",
            "Requerido; precio neto unitario (≥ 0)",
            "Opcional; descuento 0-100 (default 0)",
        ]
        examples = [
            ["COT-001", "12.345.678-9", "", "2024-01-15", "2024-02-15", "vendedor@empresa.cl", "enviada", "Urgente", "SKU-A", "Producto A gran formato", "caja", 2, 15000, 10],
            ["COT-001", "12.345.678-9", "", "2024-01-15", "2024-02-15", "vendedor@empresa.cl", "enviada", "Urgente", "SKU-B", "Producto B estándar", "", 5, 8000, 0],
            ["COT-002", "98.765.432-1", "76.543.210-K", "2024-01-20", "", "otro@empresa.cl", "no_definido", "", "", "Servicio de consultoría", "", 1, 120000, 0],
        ]

        header_font = openpyxl.styles.Font(bold=True, size=11)
        header_fill = openpyxl.styles.PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        doc_font = openpyxl.styles.Font(italic=True, size=9)
        doc_fill = openpyxl.styles.PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        )

        # Row 1: headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Row 2: documentation
        for col_num, doc in enumerate(documentation, 1):
            cell = ws.cell(row=2, column=col_num, value=doc)
            cell.font = doc_font
            cell.fill = doc_fill

        # Rows 3-5: examples
        for row_idx, example in enumerate(examples, 3):
            for col_idx, val in enumerate(example, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Column widths
        col_widths = [15, 18, 18, 15, 15, 25, 15, 20, 12, 35, 12, 10, 15, 12]
        col_letters = "ABCDEFGHIJKLMN"
        for letter, width in zip(col_letters, col_widths):
            ws.column_dimensions[letter].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _read_xlsx(content: bytes) -> list[list]:
        """Read Excel file and return rows as lists."""
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ParseError(f"No se pudo abrir el archivo Excel: {exc}")
        try:
            ws = wb.active
            return [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

    @staticmethod
    def _read_csv(content: bytes) -> list[list]:
        """Read CSV file and return rows as lists."""
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = content.decode("latin-1")
            except Exception as exc:
                raise ParseError(f"No se pudo decodificar el CSV: {exc}")

        reader = csv.reader(io.StringIO(text))
        return [list(row) for row in reader]

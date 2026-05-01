"""
Clientes and Empresas Import Parser

Handles parsing and validation of Excel files containing client and company data
during onboarding. Provides:
- Excel file parsing with openpyxl
- Row-level validation with detailed error messages
- RUT format and check digit validation using module 11
- Spanish column support
"""

from __future__ import annotations

import hashlib
import io
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl

from app.utils.rut import validate_rut, clean_rut


@dataclass(frozen=True)
class ParsedRow:
    """Single validated client/company row."""
    row_num: int
    rut_empresa: str
    nombre_empresa: str
    razon_social: Optional[str]
    sector: Optional[str]
    email_empresa: Optional[str]
    rut_cliente: str
    nombre_cliente: str
    email_cliente: Optional[str]
    telefono_cliente: Optional[str]
    direccion_despacho: Optional[str]
    forma_pago: Optional[str]
    hash_key: str


@dataclass(frozen=True)
class InvalidRow:
    """Single invalid row with errors."""
    row_num: int
    rut_empresa: Optional[str]  # May be missing
    rut_cliente: Optional[str]  # May be missing
    errors: list[str]


@dataclass
class ParseResult:
    """Results from parsing a clientes_empresas file."""
    valid_rows: list[ParsedRow]
    invalid_rows: list[InvalidRow]
    a_crear: dict  # { "empresas": count, "clientes": count }
    a_actualizar: dict  # { "empresas": count, "clientes": count }

    @property
    def valid_count(self) -> int:
        return len(self.valid_rows)

    @property
    def invalid_count(self) -> int:
        return len(self.invalid_rows)


class ParseError(Exception):
    """Raised when file format is invalid (not recoverable per-row)."""
    pass


class ClientesEmpresasParser:
    """Parse and validate clientes + empresas import Excel files."""

    # Expected column headers (case-insensitive)
    REQUIRED_COLUMNS = {
        "rut_empresa",
        "nombre_empresa",
        "rut_cliente",
        "nombre_cliente",
    }

    OPTIONAL_COLUMNS = {
        "razon_social",
        "sector",
        "email",
        "email_empresa",
        "email_cliente",
        "telefono_cliente",
        "direccion_despacho",
        "forma_pago",
    }

    ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS

    @staticmethod
    def parse(content: bytes, filename: str, existing_empresa_ruts: set[str] = None, existing_cliente_ruts: set[str] = None) -> ParseResult:
        """
        Parse and validate a clientes + empresas import file.

        Args:
            content: File bytes
            filename: Original filename (used to detect format)
            existing_empresa_ruts: Set of existing empresa RUTs (for counting updates)
            existing_cliente_ruts: Set of existing cliente RUTs (for counting updates)

        Returns:
            ParseResult with valid/invalid rows and creation/update counts

        Raises:
            ParseError: If file format is invalid or headers are missing
        """
        parser = ClientesEmpresasParser()
        return parser._parse_file(content, filename, existing_empresa_ruts or set(), existing_cliente_ruts or set())

    @staticmethod
    def generate_template() -> bytes:
        """
        Generate Excel template with headers and example row.

        Returns:
            Excel file bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Row 1: Headers
        headers = [
            "rut_empresa",
            "nombre_empresa",
            "razon_social",
            "sector",
            "email_empresa",
            "rut_cliente",
            "nombre_cliente",
            "email_cliente",
            "telefono_cliente",
            "direccion_despacho",
            "forma_pago",
        ]

        documentation = [
            "RUT de empresa (ej: 76543210-K)",
            "Nombre de empresa",
            "Razón social (opcional)",
            "Sector (opcional)",
            "Email de empresa (opcional)",
            "RUT de cliente (ej: 12345678-9)",
            "Nombre del cliente",
            "Email del cliente (opcional)",
            "Teléfono del cliente (opcional)",
            "Dirección de despacho (opcional)",
            "Forma de pago (opcional)",
        ]

        # Row 1: Header with formatting
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True, size=11)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Row 2: Documentation row
        for col_num, doc in enumerate(documentation, 1):
            cell = ws.cell(row=2, column=col_num, value=doc)
            cell.font = openpyxl.styles.Font(italic=True, size=9)
            cell.fill = openpyxl.styles.PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # Row 3: Example data
        example = [
            "76543210-K",
            "Constructora ejemplo",
            "Constructora Ejemplo S.A.",
            "Construcción",
            "contacto@constructora.cl",
            "12345678-9",
            "Juan Pérez",
            "juan@example.cl",
            "+56 9 1234 5678",
            "Calle Principal 123, Santiago",
            "30 días",
        ]

        for col_num, value in enumerate(example, 1):
            ws.cell(row=3, column=col_num, value=value)

        # Adjust column widths
        widths = [15, 25, 25, 15, 20, 15, 25, 20, 15, 30, 15]
        for col_num, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _parse_file(self, content: bytes, filename: str, existing_empresa_ruts: set[str], existing_cliente_ruts: set[str]) -> ParseResult:
        """Internal parse implementation."""
        ext = Path(filename).suffix.lower()
        if ext != ".xlsx":
            raise ParseError(f"Formato no soportado: {ext}. Use .xlsx")

        raw_rows = self._read_xlsx(content)

        if not raw_rows:
            raise ParseError("El archivo está vacío")

        if len(raw_rows) < 3:
            raise ParseError("El archivo debe tener encabezados y al menos una fila de datos")

        # Row 0: headers, Row 1: documentation, Row 2+: data
        header_row = raw_rows[0]
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in header_row
        ]

        # Find column indices
        col_indices = {}
        for col_name in self.ALL_COLUMNS:
            try:
                col_indices[col_name] = headers.index(col_name)
            except ValueError:
                if col_name in self.REQUIRED_COLUMNS:
                    raise ParseError(
                        f"Columna requerida faltante: '{col_name}'. "
                        f"Columnas encontradas: {[h for h in headers if h]}"
                    )
                col_indices[col_name] = None

        # Parse data rows (skip header row 0 and documentation row 1)
        valid_rows = []
        invalid_rows = []
        empresas_a_crear = 0
        empresas_a_actualizar = 0
        clientes_a_crear = 0
        clientes_a_actualizar = 0

        for row_num, raw_row in enumerate(raw_rows[2:], start=3):
            result = self._parse_row(
                row_num=row_num,
                raw_row=raw_row,
                col_indices=col_indices
            )

            if result[0] is not None:  # Valid
                parsed = result[0]
                valid_rows.append(parsed)

                # Count new vs updates
                if clean_rut(parsed.rut_empresa) in existing_empresa_ruts:
                    empresas_a_actualizar += 1
                else:
                    empresas_a_crear += 1

                if clean_rut(parsed.rut_cliente) in existing_cliente_ruts:
                    clientes_a_actualizar += 1
                else:
                    clientes_a_crear += 1
            else:  # Invalid
                invalid_rows.append(result[1])

        return ParseResult(
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
            a_crear={"empresas": empresas_a_crear, "clientes": clientes_a_crear},
            a_actualizar={"empresas": empresas_a_actualizar, "clientes": clientes_a_actualizar},
        )

    def _parse_row(
        self,
        row_num: int,
        raw_row: list,
        col_indices: dict[str, int | None],
    ) -> tuple[Optional[ParsedRow], Optional[InvalidRow]]:
        """
        Parse and validate a single data row.

        Returns:
            (valid_row, None) if valid
            (None, invalid_row) if invalid
        """
        errors = []
        extracted = {}

        # Extract all fields
        for col_name in self.ALL_COLUMNS:
            idx = col_indices.get(col_name)
            if idx is None:
                extracted[col_name] = None
            else:
                value = raw_row[idx] if idx < len(raw_row) else None
                extracted[col_name] = value

        # Validate required fields
        rut_empresa = self._validate_rut_empresa(extracted["rut_empresa"], errors)
        nombre_empresa = self._validate_nombre_empresa(extracted["nombre_empresa"], errors)
        rut_cliente = self._validate_rut_cliente(extracted["rut_cliente"], errors)
        nombre_cliente = self._validate_nombre_cliente(extracted["nombre_cliente"], errors)

        if errors:
            return (
                None,
                InvalidRow(
                    row_num=row_num,
                    rut_empresa=rut_empresa,
                    rut_cliente=rut_cliente,
                    errors=errors,
                )
            )

        # Optional fields (validate if present)
        razon_social = self._validate_razon_social(extracted.get("razon_social"), errors)
        sector = self._validate_sector(extracted.get("sector"), errors)
        email_empresa = self._validate_email(extracted.get("email_empresa"), errors, "email_empresa")
        email_cliente = self._validate_email(extracted.get("email_cliente"), errors, "email_cliente")
        telefono_cliente = self._validate_telefono(extracted.get("telefono_cliente"), errors)
        direccion_despacho = self._validate_direccion(extracted.get("direccion_despacho"), errors)
        forma_pago = self._validate_forma_pago(extracted.get("forma_pago"), errors)

        # Handle "email" column fallback (may be in one generic email column)
        if not email_empresa and extracted.get("email"):
            email_empresa = self._validate_email(extracted.get("email"), errors, "email")

        if errors:
            return (
                None,
                InvalidRow(
                    row_num=row_num,
                    rut_empresa=rut_empresa,
                    rut_cliente=rut_cliente,
                    errors=errors,
                )
            )

        # Generate hash key for idempotency
        hash_key = self._generate_hash_key(rut_empresa, rut_cliente, nombre_cliente)

        return (
            ParsedRow(
                row_num=row_num,
                rut_empresa=rut_empresa,
                nombre_empresa=nombre_empresa,
                razon_social=razon_social,
                sector=sector,
                email_empresa=email_empresa,
                rut_cliente=rut_cliente,
                nombre_cliente=nombre_cliente,
                email_cliente=email_cliente,
                telefono_cliente=telefono_cliente,
                direccion_despacho=direccion_despacho,
                forma_pago=forma_pago,
                hash_key=hash_key,
            ),
            None,
        )

    @staticmethod
    def _validate_rut_empresa(value: any, errors: list[str]) -> Optional[str]:
        """Validate and extract RUT empresa with module 11 check."""
        if value is None or str(value).strip() == "":
            errors.append("RUT empresa es requerido")
            return None

        rut = clean_rut(str(value).strip())

        if not rut or len(rut) < 8:
            errors.append(f"RUT empresa inválido: {value}")
            return None

        if not validate_rut(rut):
            errors.append(f"RUT empresa inválido (validación módulo 11): {value}")
            return None

        return rut

    @staticmethod
    def _validate_rut_cliente(value: any, errors: list[str]) -> Optional[str]:
        """Validate and extract RUT cliente with module 11 check."""
        if value is None or str(value).strip() == "":
            errors.append("RUT cliente es requerido")
            return None

        rut = clean_rut(str(value).strip())

        if not rut or len(rut) < 8:
            errors.append(f"RUT cliente inválido: {value}")
            return None

        if not validate_rut(rut):
            errors.append(f"RUT cliente inválido (validación módulo 11): {value}")
            return None

        return rut

    @staticmethod
    def _validate_nombre_empresa(value: any, errors: list[str]) -> Optional[str]:
        """Validate and extract nombre_empresa."""
        if value is None or str(value).strip() == "":
            errors.append("Nombre empresa es requerido")
            return None

        nombre = str(value).strip()
        if len(nombre) < 2:
            errors.append("Nombre empresa debe tener al menos 2 caracteres")
            return None

        return nombre

    @staticmethod
    def _validate_nombre_cliente(value: any, errors: list[str]) -> Optional[str]:
        """Validate and extract nombre_cliente."""
        if value is None or str(value).strip() == "":
            errors.append("Nombre cliente es requerido")
            return None

        nombre = str(value).strip()
        if len(nombre) < 2:
            errors.append("Nombre cliente debe tener al menos 2 caracteres")
            return None

        return nombre

    @staticmethod
    def _validate_razon_social(value: any, errors: list[str]) -> Optional[str]:
        """Validate and extract razon_social (optional)."""
        if value is None or str(value).strip() == "":
            return None

        return str(value).strip()

    @staticmethod
    def _validate_sector(value: any, errors: list[str]) -> Optional[str]:
        """Validate and extract sector (optional)."""
        if value is None or str(value).strip() == "":
            return None

        return str(value).strip()

    @staticmethod
    def _validate_email(value: any, errors: list[str], field_name: str = "email") -> Optional[str]:
        """Validate and extract email (optional)."""
        if value is None or str(value).strip() == "":
            return None

        email = str(value).strip()

        # Basic email validation
        if "@" not in email or len(email) < 5:
            errors.append(f"{field_name} inválido: {email}")
            return None

        return email

    @staticmethod
    def _validate_telefono(value: any, errors: list[str]) -> Optional[str]:
        """Validate and extract telefono (optional)."""
        if value is None or str(value).strip() == "":
            return None

        return str(value).strip()

    @staticmethod
    def _validate_direccion(value: any, errors: list[str]) -> Optional[str]:
        """Validate and extract direccion_despacho (optional)."""
        if value is None or str(value).strip() == "":
            return None

        return str(value).strip()

    @staticmethod
    def _validate_forma_pago(value: any, errors: list[str]) -> Optional[str]:
        """Validate and extract forma_pago (optional)."""
        if value is None or str(value).strip() == "":
            return None

        return str(value).strip()

    @staticmethod
    def _generate_hash_key(rut_empresa: str, rut_cliente: str, nombre_cliente: str) -> str:
        """Generate SHA256 hash key for idempotency detection."""
        key_str = f"{rut_empresa}|{rut_cliente}|{nombre_cliente}"
        return hashlib.sha256(key_str.encode()).hexdigest()

    @staticmethod
    def _read_xlsx(content: bytes) -> list[list]:
        """Read Excel file and return rows as lists."""
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            return [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

from __future__ import annotations

import io
from dataclasses import dataclass, field

import openpyxl


REQUIRED_COLUMNS = ("rut", "razon_social")
OPTIONAL_COLUMNS = ("giro", "direccion", "comuna", "contacto", "email", "telefono", "condicion_pago")
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS


class ParseError(Exception):
    pass


@dataclass
class ParsedProveedor:
    fila: int
    rut_raw: str
    rut_normalizado: str
    razon_social: str
    giro: str | None = None
    direccion: str | None = None
    comuna: str | None = None
    contacto: str | None = None
    email: str | None = None
    telefono: str | None = None
    condicion_pago: str | None = None


@dataclass
class InvalidRow:
    fila: int
    rut_raw: str | None
    razon_social_raw: str | None
    motivo: str


@dataclass
class ParseResult:
    validas: list[ParsedProveedor] = field(default_factory=list)
    invalidas: list[InvalidRow] = field(default_factory=list)


def _normalizar_rut(rut_raw: str) -> str:
    """Strip dots, uppercase, keep digits and hyphen+DV."""
    if not rut_raw:
        return ""
    return rut_raw.replace(".", "").replace(" ", "").upper()


def _validar_rut_modulo11(rut: str) -> bool:
    """Validate Chilean RUT using módulo 11 algorithm."""
    if not rut:
        return False
    cleaned = rut.replace(".", "").replace(" ", "").upper()
    if "-" not in cleaned:
        return False
    cuerpo, dv = cleaned.rsplit("-", 1)
    if not cuerpo.isdigit() or len(cuerpo) < 7:
        return False
    if dv not in "0123456789K":
        return False
    factores = (2, 3, 4, 5, 6, 7)
    suma = sum(int(d) * factores[i % 6] for i, d in enumerate(reversed(cuerpo)))
    resto = 11 - (suma % 11)
    if resto == 11:
        esperado = "0"
    elif resto == 10:
        esperado = "K"
    else:
        esperado = str(resto)
    return dv == esperado


def _str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def parse_proveedores_xlsx(content: bytes) -> ParseResult:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — surface as ParseError
        raise ParseError(f"No se pudo leer el archivo xlsx: {e}")

    try:
        ws = wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if not rows:
        raise ParseError("El archivo no contiene filas")

    header = [_str(h).lower() for h in rows[0]]
    missing = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing:
        raise ParseError(
            f"Columnas requeridas no encontradas: {missing}. Columnas encontradas: {header}"
        )

    def get(row: list, col: str) -> str:
        if col not in header:
            return ""
        i = header.index(col)
        return _str(row[i]) if i < len(row) else ""

    result = ParseResult()
    seen_ruts: dict[str, int] = {}

    for idx, raw in enumerate(rows[1:], start=2):
        rut_raw = get(raw, "rut")
        razon = get(raw, "razon_social")

        if not rut_raw and not razon and not any(_str(c) for c in raw):
            continue  # blank row

        if not rut_raw:
            result.invalidas.append(InvalidRow(fila=idx, rut_raw=None, razon_social_raw=razon or None, motivo="RUT vacío"))
            continue
        if not razon:
            result.invalidas.append(InvalidRow(fila=idx, rut_raw=rut_raw, razon_social_raw=None, motivo="Razón social vacía"))
            continue

        rut_norm = _normalizar_rut(rut_raw)
        if not _validar_rut_modulo11(rut_norm):
            result.invalidas.append(InvalidRow(fila=idx, rut_raw=rut_raw, razon_social_raw=razon, motivo=f"RUT inválido: '{rut_raw}'"))
            continue

        if rut_norm in seen_ruts:
            result.invalidas.append(
                InvalidRow(
                    fila=idx,
                    rut_raw=rut_raw,
                    razon_social_raw=razon,
                    motivo=f"RUT duplicado en archivo (también en fila {seen_ruts[rut_norm]})",
                )
            )
            continue
        seen_ruts[rut_norm] = idx

        result.validas.append(
            ParsedProveedor(
                fila=idx,
                rut_raw=rut_raw,
                rut_normalizado=rut_norm,
                razon_social=razon,
                giro=get(raw, "giro") or None,
                direccion=get(raw, "direccion") or None,
                comuna=get(raw, "comuna") or None,
                contacto=get(raw, "contacto") or None,
                email=get(raw, "email") or None,
                telefono=get(raw, "telefono") or None,
                condicion_pago=get(raw, "condicion_pago") or None,
            )
        )

    return result


def build_template_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    ws.append(list(ALL_COLUMNS))
    ws.append([
        "76.123.456-0",
        "Sociedad Ejemplo Ltda.",
        "Comercio al por mayor",
        "Av. Providencia 1234",
        "Providencia",
        "Juan Pérez",
        "contacto@ejemplo.cl",
        "+56 9 1234 5678",
        "30 días",
    ])
    for i, col in enumerate(ALL_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

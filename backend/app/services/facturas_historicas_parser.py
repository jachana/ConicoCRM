from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from typing import Optional

import openpyxl


class ParseError(Exception):
    """Raised when file format is invalid (not recoverable per-row)."""
    pass


@dataclass
class FacturaLine:
    sku: Optional[str]
    descripcion: str
    cantidad: int
    precio_unitario: Decimal  # = valor_neto unitario
    descuento: Decimal  # 0-100 percent
    exento: bool
    total_neto_linea: Decimal
    iva_linea: Decimal
    total_linea: Decimal
    row_num: int


@dataclass
class FacturaGroup:
    folio: int
    tipo_dte: str
    rut_receptor: str
    rut_empresa: Optional[str]
    fecha_emision: date
    neto: Decimal
    iva: Decimal
    total: Decimal
    estado_pago: str
    nota: Optional[str]
    lines: list[FacturaLine]
    status: str  # "crear" | "omitir"
    cliente_stub: bool  # True if rut_receptor not found in DB
    row_num: int  # cabecera row number


@dataclass
class InvalidRow:
    row_num: int
    folio_raw: Optional[str]
    sheet: str  # "Cabecera DTE" or "Detalle DTE"
    motivo: str


@dataclass
class ParseResult:
    valid_groups: list[FacturaGroup] = field(default_factory=list)
    invalid_rows: list[InvalidRow] = field(default_factory=list)
    a_crear: int = 0
    a_omitir: int = 0
    invalid_count: int = 0
    stubs_a_crear: int = 0  # count of groups where cliente_stub=True and status="crear"


class FacturasHistoricasParser:
    """Parse and validate Facturas/Boletas históricas import files (two-sheet Excel)."""

    CABECERA_SHEET = "Cabecera DTE"
    DETALLE_SHEET = "Detalle DTE"

    CABECERA_REQUIRED = {"tipo_dte", "folio", "rut_receptor", "fecha_emision", "neto", "iva", "total"}
    CABECERA_OPTIONAL = {"rut_empresa", "estado_pago", "nota"}

    DETALLE_REQUIRED = {"folio", "cantidad", "precio_unitario"}
    DETALLE_OPTIONAL = {"sku", "descripcion", "descuento", "exento"}

    VALID_TIPOS_DTE = {"033", "034", "039", "041"}
    VALID_ESTADOS_PAGO = {"emitida", "parcial", "pagada"}

    @staticmethod
    def parse(
        content: bytes,
        clientes_by_rut: dict,
        empresas_by_rut: dict,
        productos_by_sku: dict,
        existing_folios: set[int],
    ) -> ParseResult:
        """
        Parse and validate a two-sheet Excel file of historical invoices.

        Args:
            content: Excel file bytes
            clientes_by_rut: Dict mapping rut -> Cliente id
            empresas_by_rut: Dict mapping rut -> Empresa id
            productos_by_sku: Dict mapping sku -> Producto id
            existing_folios: Set of folio integers already in DB

        Returns:
            ParseResult with valid_groups, invalid_rows, and counts

        Raises:
            ParseError: If file format is invalid or required columns are missing
        """
        clientes_by_rut = clientes_by_rut or {}
        empresas_by_rut = empresas_by_rut or {}
        productos_by_sku = productos_by_sku or {}
        existing_folios = existing_folios or set()

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ParseError(f"No se pudo abrir el archivo Excel: {exc}")

        try:
            sheet_names = wb.sheetnames
            if FacturasHistoricasParser.CABECERA_SHEET not in sheet_names:
                raise ParseError(
                    f"No se encontró la hoja '{FacturasHistoricasParser.CABECERA_SHEET}'. "
                    f"Hojas encontradas: {sheet_names}"
                )
            if FacturasHistoricasParser.DETALLE_SHEET not in sheet_names:
                raise ParseError(
                    f"No se encontró la hoja '{FacturasHistoricasParser.DETALLE_SHEET}'. "
                    f"Hojas encontradas: {sheet_names}"
                )

            cab_rows = [list(row) for row in wb[FacturasHistoricasParser.CABECERA_SHEET].iter_rows(values_only=True)]
            det_rows = [list(row) for row in wb[FacturasHistoricasParser.DETALLE_SHEET].iter_rows(values_only=True)]
        finally:
            wb.close()

        if not cab_rows:
            raise ParseError(f"La hoja '{FacturasHistoricasParser.CABECERA_SHEET}' está vacía")
        if not det_rows:
            raise ParseError(f"La hoja '{FacturasHistoricasParser.DETALLE_SHEET}' está vacía")

        result = ParseResult()

        cab_groups, cab_invalid = FacturasHistoricasParser._parse_cabecera(
            cab_rows, clientes_by_rut, empresas_by_rut, existing_folios
        )
        result.invalid_rows.extend(cab_invalid)

        det_invalid = FacturasHistoricasParser._parse_detalle(
            det_rows, cab_groups, productos_by_sku
        )
        result.invalid_rows.extend(det_invalid)

        for folio, group in cab_groups.items():
            result.valid_groups.append(group)
            if group.status == "omitir":
                result.a_omitir += 1
            else:
                result.a_crear += 1
                if group.cliente_stub:
                    result.stubs_a_crear += 1

        result.invalid_count = len(cab_invalid)

        return result

    @staticmethod
    def _parse_cabecera(
        raw_rows: list[list],
        clientes_by_rut: dict,
        empresas_by_rut: dict,
        existing_folios: set[int],
    ) -> tuple[dict[int, FacturaGroup], list[InvalidRow]]:
        """
        Validate Cabecera DTE rows and return (groups_by_folio, invalid_rows).
        groups_by_folio preserves file order via dict insertion order.
        """
        header_row = raw_rows[0]
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in header_row
        ]

        missing = FacturasHistoricasParser.CABECERA_REQUIRED - set(headers)
        if missing:
            raise ParseError(
                f"Hoja '{FacturasHistoricasParser.CABECERA_SHEET}' — columna(s) requerida(s) faltante(s): "
                f"{sorted(missing)}. Columnas encontradas: {[h for h in headers if h]}"
            )

        col = {h: i for i, h in enumerate(headers) if h}

        data_rows = FacturasHistoricasParser._skip_doc_row(raw_rows, col, "folio")
        header_offset = len(raw_rows) - len(data_rows)

        groups: dict[int, FacturaGroup] = {}
        invalid_rows: list[InvalidRow] = []

        def _get(row: list, col_name: str) -> Optional[str]:
            i = col.get(col_name, -1)
            if i < 0 or i >= len(row):
                return None
            val = row[i]
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        for idx, raw_row in enumerate(data_rows):
            row_num = header_offset + idx + 1
            folio_raw = _get(raw_row, "folio")
            errors = []

            # folio
            parsed_folio = None
            if not folio_raw:
                errors.append("folio es requerido")
            else:
                try:
                    parsed_folio = int(float(folio_raw))
                    if parsed_folio <= 0:
                        errors.append(f"folio debe ser un entero positivo (tiene {parsed_folio})")
                        parsed_folio = None
                except (ValueError, TypeError):
                    errors.append(f"folio '{folio_raw}' no es un entero válido")

            if parsed_folio is not None and parsed_folio in groups:
                errors.append(f"folio {parsed_folio} duplicado en la hoja Cabecera DTE")

            # tipo_dte
            tipo_dte_raw = _get(raw_row, "tipo_dte")
            parsed_tipo_dte = None
            if not tipo_dte_raw:
                errors.append("tipo_dte es requerido")
            else:
                normalized = tipo_dte_raw.zfill(3)
                if normalized not in FacturasHistoricasParser.VALID_TIPOS_DTE:
                    valid_str = ", ".join(f"'{v}'" for v in sorted(FacturasHistoricasParser.VALID_TIPOS_DTE))
                    errors.append(f"tipo_dte '{tipo_dte_raw}' no válido (use {valid_str})")
                else:
                    parsed_tipo_dte = normalized

            # rut_receptor
            rut_receptor_raw = _get(raw_row, "rut_receptor")
            if not rut_receptor_raw:
                errors.append("rut_receptor es requerido")

            # rut_empresa (optional, must exist if provided)
            rut_empresa_raw = _get(raw_row, "rut_empresa")
            if rut_empresa_raw and rut_empresa_raw not in empresas_by_rut:
                errors.append(f"Empresa con rut '{rut_empresa_raw}' no encontrada")

            # fecha_emision
            fecha_raw = _get(raw_row, "fecha_emision")
            parsed_fecha = None
            if not fecha_raw:
                errors.append("fecha_emision es requerida")
            else:
                parsed_fecha = FacturasHistoricasParser._parse_date(fecha_raw)
                if parsed_fecha is None:
                    errors.append(
                        f"fecha_emision '{fecha_raw}' no tiene formato válido (YYYY-MM-DD o DD/MM/YYYY)"
                    )

            # neto
            neto_raw = _get(raw_row, "neto")
            parsed_neto = None
            if neto_raw is None:
                errors.append("neto es requerido")
            else:
                try:
                    parsed_neto = Decimal(str(neto_raw))
                    if parsed_neto < 0:
                        errors.append(f"neto debe ser no-negativo (tiene {parsed_neto})")
                        parsed_neto = None
                except Exception:
                    errors.append(f"neto '{neto_raw}' no es un número válido")

            # iva
            iva_raw = _get(raw_row, "iva")
            parsed_iva = None
            if iva_raw is None:
                errors.append("iva es requerido")
            else:
                try:
                    parsed_iva = Decimal(str(iva_raw))
                    if parsed_iva < 0:
                        errors.append(f"iva debe ser no-negativo (tiene {parsed_iva})")
                        parsed_iva = None
                except Exception:
                    errors.append(f"iva '{iva_raw}' no es un número válido")

            # total
            total_raw = _get(raw_row, "total")
            parsed_total = None
            if total_raw is None:
                errors.append("total es requerido")
            else:
                try:
                    parsed_total = Decimal(str(total_raw))
                    if parsed_total < 0:
                        errors.append(f"total debe ser no-negativo (tiene {parsed_total})")
                        parsed_total = None
                except Exception:
                    errors.append(f"total '{total_raw}' no es un número válido")

            # estado_pago (optional, default "emitida")
            estado_pago_raw = _get(raw_row, "estado_pago")
            parsed_estado_pago = "emitida"
            if estado_pago_raw:
                if estado_pago_raw.lower() not in FacturasHistoricasParser.VALID_ESTADOS_PAGO:
                    valid_str = ", ".join(f"'{v}'" for v in sorted(FacturasHistoricasParser.VALID_ESTADOS_PAGO))
                    errors.append(f"estado_pago '{estado_pago_raw}' no válido (use {valid_str})")
                else:
                    parsed_estado_pago = estado_pago_raw.lower()

            # nota (optional)
            nota_raw = _get(raw_row, "nota")

            if errors:
                invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    folio_raw=folio_raw,
                    sheet=FacturasHistoricasParser.CABECERA_SHEET,
                    motivo="; ".join(errors),
                ))
                continue

            cliente_stub = rut_receptor_raw not in clientes_by_rut
            status = "omitir" if parsed_folio in existing_folios else "crear"

            groups[parsed_folio] = FacturaGroup(
                folio=parsed_folio,
                tipo_dte=parsed_tipo_dte,
                rut_receptor=rut_receptor_raw,
                rut_empresa=rut_empresa_raw,
                fecha_emision=parsed_fecha,
                neto=parsed_neto,
                iva=parsed_iva,
                total=parsed_total,
                estado_pago=parsed_estado_pago,
                nota=nota_raw,
                lines=[],
                status=status,
                cliente_stub=cliente_stub,
                row_num=row_num,
            )

        return groups, invalid_rows

    @staticmethod
    def _parse_detalle(
        raw_rows: list[list],
        groups: dict[int, FacturaGroup],
        productos_by_sku: dict,
    ) -> list[InvalidRow]:
        """
        Validate Detalle DTE rows and attach FacturaLine objects to their groups.
        Returns list of invalid rows.
        """
        header_row = raw_rows[0]
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in header_row
        ]

        missing = FacturasHistoricasParser.DETALLE_REQUIRED - set(headers)
        if missing:
            raise ParseError(
                f"Hoja '{FacturasHistoricasParser.DETALLE_SHEET}' — columna(s) requerida(s) faltante(s): "
                f"{sorted(missing)}. Columnas encontradas: {[h for h in headers if h]}"
            )

        col = {h: i for i, h in enumerate(headers) if h}

        data_rows = FacturasHistoricasParser._skip_doc_row(raw_rows, col, "folio")
        header_offset = len(raw_rows) - len(data_rows)

        invalid_rows: list[InvalidRow] = []

        def _get(row: list, col_name: str) -> Optional[str]:
            i = col.get(col_name, -1)
            if i < 0 or i >= len(row):
                return None
            val = row[i]
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        for idx, raw_row in enumerate(data_rows):
            row_num = header_offset + idx + 1
            folio_raw = _get(raw_row, "folio")
            errors = []

            # folio — links to Cabecera
            parsed_folio = None
            if not folio_raw:
                errors.append("folio es requerido")
            else:
                try:
                    parsed_folio = int(float(folio_raw))
                    if parsed_folio <= 0:
                        errors.append(f"folio debe ser un entero positivo (tiene {parsed_folio})")
                        parsed_folio = None
                except (ValueError, TypeError):
                    errors.append(f"folio '{folio_raw}' no es un entero válido")

            if parsed_folio is not None and parsed_folio not in groups:
                errors.append(
                    f"folio {parsed_folio} no encontrado en la hoja '{FacturasHistoricasParser.CABECERA_SHEET}'"
                )
                invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    folio_raw=folio_raw,
                    sheet=FacturasHistoricasParser.DETALLE_SHEET,
                    motivo="; ".join(errors),
                ))
                continue

            # sku (optional, must exist if provided)
            sku_raw = _get(raw_row, "sku")
            if sku_raw and sku_raw not in productos_by_sku:
                errors.append(f"Producto con sku '{sku_raw}' no encontrado")

            # descripcion (optional, fallback applied after validation)
            descripcion_raw = _get(raw_row, "descripcion")

            # cantidad
            cantidad_raw = _get(raw_row, "cantidad")
            parsed_cantidad = None
            if not cantidad_raw:
                errors.append("cantidad es requerida")
            else:
                try:
                    parsed_cantidad = int(float(cantidad_raw))
                    if parsed_cantidad <= 0:
                        errors.append(f"cantidad debe ser un entero positivo (tiene {parsed_cantidad})")
                        parsed_cantidad = None
                except (ValueError, TypeError):
                    errors.append(f"cantidad '{cantidad_raw}' no es un número válido")

            # precio_unitario
            precio_raw = _get(raw_row, "precio_unitario")
            parsed_precio = None
            if precio_raw is None:
                errors.append("precio_unitario es requerido")
            else:
                try:
                    parsed_precio = Decimal(str(precio_raw))
                    if parsed_precio < 0:
                        errors.append(f"precio_unitario debe ser no-negativo (tiene {parsed_precio})")
                        parsed_precio = None
                except Exception:
                    errors.append(f"precio_unitario '{precio_raw}' no es un número válido")

            # descuento (optional, 0-100, default 0)
            descuento_raw = _get(raw_row, "descuento")
            parsed_descuento = Decimal("0")
            if descuento_raw is not None:
                try:
                    parsed_descuento = Decimal(str(descuento_raw))
                    if parsed_descuento < 0 or parsed_descuento > 100:
                        errors.append(f"descuento '{descuento_raw}' debe estar entre 0 y 100")
                        parsed_descuento = Decimal("0")
                except Exception:
                    errors.append(f"descuento '{descuento_raw}' no es un número válido")

            # exento (optional, default False)
            exento_raw = _get(raw_row, "exento")
            parsed_exento = False
            if exento_raw is not None:
                parsed_exento = FacturasHistoricasParser._parse_bool(exento_raw)
                if parsed_exento is None:
                    errors.append(
                        f"exento '{exento_raw}' no es un valor válido (use True/False, Si/No, 1/0)"
                    )
                    parsed_exento = False

            if errors:
                invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    folio_raw=folio_raw,
                    sheet=FacturasHistoricasParser.DETALLE_SHEET,
                    motivo="; ".join(errors),
                ))
                continue

            descripcion = descripcion_raw if descripcion_raw else (sku_raw if sku_raw else "Item sin descripción")

            total_neto_linea = (
                Decimal(str(parsed_cantidad))
                * parsed_precio
                * (Decimal("1") - parsed_descuento / Decimal("100"))
            ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

            if parsed_exento:
                iva_linea = Decimal("0")
            else:
                iva_linea = (total_neto_linea * Decimal("0.19")).quantize(
                    Decimal("1"), rounding=ROUND_HALF_UP
                )

            total_linea = total_neto_linea + iva_linea

            groups[parsed_folio].lines.append(FacturaLine(
                sku=sku_raw,
                descripcion=descripcion,
                cantidad=parsed_cantidad,
                precio_unitario=parsed_precio,
                descuento=parsed_descuento,
                exento=parsed_exento,
                total_neto_linea=total_neto_linea,
                iva_linea=iva_linea,
                total_linea=total_linea,
                row_num=row_num,
            ))

        return invalid_rows

    @staticmethod
    def _skip_doc_row(raw_rows: list[list], col: dict, anchor_col: str) -> list[list]:
        """
        Skip row index 1 (the row immediately after headers) if it looks like
        a documentation/example-label row rather than real data. Same detection
        logic as cotizaciones_parser.py, keyed on the anchor column.
        """
        if len(raw_rows) <= 1:
            return []
        anchor_idx = col.get(anchor_col, -1)
        potential_doc = raw_rows[1]
        anchor_val = (
            potential_doc[anchor_idx]
            if anchor_idx >= 0 and anchor_idx < len(potential_doc)
            else None
        )
        if anchor_val is None or str(anchor_val).strip() == "":
            return raw_rows[2:]
        val_str = str(anchor_val).strip()
        doc_keywords = ("requerido", "ej:", "ejemplo", "opcional", "folio")
        if any(kw in val_str.lower() for kw in doc_keywords) or len(val_str) > 200:
            return raw_rows[2:]
        return raw_rows[1:]

    @staticmethod
    def _parse_date(s: str) -> Optional[date]:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_bool(s: str) -> Optional[bool]:
        lower = s.strip().lower()
        if lower in ("true", "si", "sí", "1", "yes"):
            return True
        if lower in ("false", "no", "0"):
            return False
        return None

    @staticmethod
    def generate_template() -> bytes:
        """
        Generate a two-sheet Excel template: 'Cabecera DTE' and 'Detalle DTE'.
        Each sheet has a bold/gray header row, an italic/light documentation row,
        and example rows.
        """
        wb = openpyxl.Workbook()

        header_font = openpyxl.styles.Font(bold=True, size=11)
        header_fill = openpyxl.styles.PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        doc_font = openpyxl.styles.Font(italic=True, size=9)
        doc_fill = openpyxl.styles.PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        )

        # ----------------------------------------------------------------
        # Sheet 1: Cabecera DTE
        # ----------------------------------------------------------------
        ws_cab = wb.active
        ws_cab.title = FacturasHistoricasParser.CABECERA_SHEET

        cab_headers = [
            "tipo_dte",
            "folio",
            "rut_receptor",
            "rut_empresa",
            "fecha_emision",
            "neto",
            "iva",
            "total",
            "estado_pago",
            "nota",
        ]
        cab_documentation = [
            "Requerido; '033' factura afecta, '034' factura exenta, '039' boleta afecta, '041' boleta exenta",
            "Requerido; número DTE entero positivo (clave de enlace con Detalle DTE)",
            "Requerido; RUT del cliente receptor (puede no existir, se importa como stub)",
            "Opcional; RUT de la empresa emisora (debe existir si se indica)",
            "Requerido; YYYY-MM-DD o DD/MM/YYYY",
            "Requerido; monto neto (≥ 0)",
            "Requerido; monto IVA (0 para facturas/boletas exentas)",
            "Requerido; monto total (≥ 0)",
            "Opcional; 'emitida' (default), 'parcial', 'pagada'",
            "Opcional; texto libre",
        ]
        cab_examples = [
            ["033", 1001, "12.345.678-9", "", "2024-01-15", 100000, 19000, 119000, "emitida", ""],
            ["039", 1002, "98.765.432-1", "76.543.210-K", "2024-01-20", 50000, 9500, 59500, "pagada", "Boleta consumidor final"],
            ["034", 1003, "11.111.111-1", "", "2024-01-25", 80000, 0, 80000, "parcial", "Factura exenta servicios"],
        ]

        for col_num, header in enumerate(cab_headers, 1):
            cell = ws_cab.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for col_num, doc in enumerate(cab_documentation, 1):
            cell = ws_cab.cell(row=2, column=col_num, value=doc)
            cell.font = doc_font
            cell.fill = doc_fill

        for row_idx, example in enumerate(cab_examples, 3):
            for col_idx, val in enumerate(example, 1):
                ws_cab.cell(row=row_idx, column=col_idx, value=val)

        cab_col_widths = [12, 10, 18, 18, 18, 14, 14, 14, 14, 30]
        for i, width in enumerate(cab_col_widths, 1):
            ws_cab.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = width

        # ----------------------------------------------------------------
        # Sheet 2: Detalle DTE
        # ----------------------------------------------------------------
        ws_det = wb.create_sheet(title=FacturasHistoricasParser.DETALLE_SHEET)

        det_headers = [
            "folio",
            "sku",
            "descripcion",
            "cantidad",
            "precio_unitario",
            "descuento",
            "exento",
        ]
        det_documentation = [
            "Requerido; debe coincidir con un folio en 'Cabecera DTE'",
            "Opcional; SKU del producto (debe existir si se indica)",
            "Opcional; descripción del ítem (si vacío y sin SKU: 'Item sin descripción')",
            "Requerido; entero positivo",
            "Requerido; precio neto unitario (≥ 0)",
            "Opcional; descuento 0-100 (default 0)",
            "Opcional; True/False/Si/No/1/0 (default False)",
        ]
        det_examples = [
            [1001, "SKU-A", "Producto A", 2, 30000, 10, "No"],
            [1001, "SKU-B", "Producto B", 1, 40000, 0, "No"],
            [1002, "", "Servicio consultoría", 1, 50000, 0, "No"],
            [1003, "", "Servicio exento capacitación", 3, 20000, 0, "Si"],
        ]

        for col_num, header in enumerate(det_headers, 1):
            cell = ws_det.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for col_num, doc in enumerate(det_documentation, 1):
            cell = ws_det.cell(row=2, column=col_num, value=doc)
            cell.font = doc_font
            cell.fill = doc_fill

        for row_idx, example in enumerate(det_examples, 3):
            for col_idx, val in enumerate(example, 1):
                ws_det.cell(row=row_idx, column=col_idx, value=val)

        det_col_widths = [10, 12, 35, 10, 18, 12, 10]
        for i, width in enumerate(det_col_widths, 1):
            ws_det.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

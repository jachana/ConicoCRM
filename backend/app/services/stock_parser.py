"""
Stock Import Parser

Parses Excel files with initial stock data (sku, nombre_bodega, cantidad, costo_unitario).
Validates against existing products and bodegas in DB.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional

import openpyxl


REQUIRED_COLUMNS = {"sku", "nombre_bodega", "cantidad"}
OPTIONAL_COLUMNS = {"costo_unitario"}
ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS


class ParseError(Exception):
    pass


@dataclass(frozen=True)
class ParsedStockRow:
    row_num: int
    sku: str
    nombre_bodega: str
    cantidad: int
    costo_unitario: Optional[Decimal]
    producto_id: int
    bodega_id: int
    has_existing_carga: bool  # True → will update existing carga inicial


@dataclass(frozen=True)
class InvalidStockRow:
    row_num: int
    sku: Optional[str]
    nombre_bodega: Optional[str]
    errors: list[str]


@dataclass
class StockParseResult:
    valid_rows: list[ParsedStockRow]
    invalid_rows: list[InvalidStockRow]
    a_crear: int
    a_actualizar: int

    @property
    def valid_count(self) -> int:
        return len(self.valid_rows)

    @property
    def invalid_count(self) -> int:
        return len(self.invalid_rows)


class StockParser:
    @staticmethod
    def generate_template() -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Inicial"

        headers = ["sku", "nombre_bodega", "cantidad", "costo_unitario"]
        docs = [
            "SKU del producto (exacto, case-insensitive)",
            "Nombre de la bodega (exacto)",
            "Cantidad en stock (entero ≥ 0)",
            "Costo unitario en pesos (opcional)",
        ]
        example = ["SKU-001", "Bodega Central", 50, 12500]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True, size=11)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        for col, d in enumerate(docs, 1):
            cell = ws.cell(row=2, column=col, value=d)
            cell.font = openpyxl.styles.Font(italic=True, size=9)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
            )

        for col, v in enumerate(example, 1):
            ws.cell(row=3, column=col, value=v)

        for col, width in enumerate([15, 25, 12, 18], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    @staticmethod
    def parse(
        content: bytes,
        filename: str,
        productos_by_sku: dict[str, Any],
        bodegas_by_nombre: dict[str, Any],
        existing_cargas: set[tuple[int, int]],
    ) -> StockParseResult:
        ext = Path(filename).suffix.lower()
        if ext != ".xlsx":
            raise ParseError(f"Formato no soportado: {ext}. Use .xlsx")

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            raw_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

        if not raw_rows:
            raise ParseError("El archivo está vacío")
        if len(raw_rows) < 2:
            raise ParseError("El archivo debe tener encabezados y al menos una fila de datos")

        header_row = [str(h).strip().lower() if h is not None else "" for h in raw_rows[0]]

        col_indices: dict[str, Optional[int]] = {}
        for col_name in ALL_COLUMNS:
            try:
                col_indices[col_name] = header_row.index(col_name)
            except ValueError:
                if col_name in REQUIRED_COLUMNS:
                    raise ParseError(
                        f"Columna requerida faltante: '{col_name}'. "
                        f"Columnas encontradas: {[h for h in header_row if h]}"
                    )
                col_indices[col_name] = None

        # Auto-skip documentation row (row 2 from template)
        data_rows = raw_rows[1:]
        sku_idx = col_indices["sku"]
        if sku_idx is not None and len(raw_rows) > 1:
            first_val = raw_rows[1][sku_idx] if sku_idx < len(raw_rows[1]) else None
            if first_val is None or (isinstance(first_val, str) and not any(c.isdigit() or c.isalpha() and len(str(first_val).strip()) > 20 for c in str(first_val))):
                # If the first data row looks like a documentation row (very long text), skip it
                if isinstance(first_val, str) and len(first_val.strip()) > 30:
                    data_rows = raw_rows[2:]

        if not data_rows:
            raise ParseError("El archivo no contiene filas de datos")

        valid_rows: list[ParsedStockRow] = []
        invalid_rows: list[InvalidStockRow] = []
        seen: set[tuple[str, str]] = set()  # (sku, nombre_bodega) for within-file dup detection
        a_crear = 0
        a_actualizar = 0
        header_offset = len(raw_rows) - len(data_rows)

        for i, raw_row in enumerate(data_rows):
            row_num = i + header_offset + 1
            errors: list[str] = []

            def get(col: str) -> Any:
                idx = col_indices.get(col)
                if idx is None:
                    return None
                return raw_row[idx] if idx < len(raw_row) else None

            # Skip fully empty rows
            if all(v is None or str(v).strip() == "" for v in raw_row):
                continue

            sku_raw = get("sku")
            nombre_bodega_raw = get("nombre_bodega")
            cantidad_raw = get("cantidad")
            costo_raw = get("costo_unitario")

            # --- SKU ---
            if sku_raw is None or str(sku_raw).strip() == "":
                errors.append("SKU es requerido")
                sku_clean = None
            else:
                sku_clean = str(sku_raw).strip().upper()

            # --- nombre_bodega ---
            if nombre_bodega_raw is None or str(nombre_bodega_raw).strip() == "":
                errors.append("nombre_bodega es requerido")
                bodega_nombre = None
            else:
                bodega_nombre = str(nombre_bodega_raw).strip()

            # --- cantidad ---
            cantidad = None
            if cantidad_raw is None or str(cantidad_raw).strip() == "":
                errors.append("cantidad es requerida")
            else:
                try:
                    cantidad = int(float(str(cantidad_raw).strip()))
                    if cantidad < 0:
                        errors.append("La cantidad no puede ser negativa")
                        cantidad = None
                except (ValueError, TypeError):
                    errors.append(f"cantidad inválida: '{cantidad_raw}' (debe ser entero)")

            # --- costo_unitario (optional) ---
            costo_unitario = None
            if costo_raw is not None and str(costo_raw).strip() != "":
                try:
                    costo_unitario = Decimal(str(costo_raw).strip())
                    if costo_unitario < 0:
                        errors.append("costo_unitario no puede ser negativo")
                        costo_unitario = None
                except InvalidOperation:
                    errors.append(f"costo_unitario inválido: '{costo_raw}'")

            # --- Lookup in DB (only if basic fields valid) ---
            producto_id = None
            bodega_id = None

            if sku_clean and not any("SKU" in e for e in errors):
                producto = productos_by_sku.get(sku_clean)
                if producto is None:
                    errors.append(f"SKU '{sku_clean}' no encontrado en el sistema")
                else:
                    producto_id = producto.id

            if bodega_nombre and not any("bodega" in e.lower() for e in errors):
                bodega = bodegas_by_nombre.get(bodega_nombre)
                if bodega is None:
                    errors.append(f"Bodega '{bodega_nombre}' no encontrada en el sistema")
                else:
                    bodega_id = bodega.id

            # --- Within-file duplicate check ---
            if sku_clean and bodega_nombre:
                key = (sku_clean, bodega_nombre)
                if key in seen:
                    errors.append(f"Fila duplicada en este archivo: SKU '{sku_clean}' + bodega '{bodega_nombre}' ya aparece antes")
                else:
                    seen.add(key)

            if errors:
                invalid_rows.append(InvalidStockRow(
                    row_num=row_num,
                    sku=sku_raw,
                    nombre_bodega=nombre_bodega_raw,
                    errors=errors,
                ))
                continue

            # All valid
            has_existing = (producto_id, bodega_id) in existing_cargas
            if has_existing:
                a_actualizar += 1
            else:
                a_crear += 1

            valid_rows.append(ParsedStockRow(
                row_num=row_num,
                sku=sku_clean,
                nombre_bodega=bodega_nombre,
                cantidad=cantidad,
                costo_unitario=costo_unitario,
                producto_id=producto_id,
                bodega_id=bodega_id,
                has_existing_carga=has_existing,
            ))

        return StockParseResult(
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
            a_crear=a_crear,
            a_actualizar=a_actualizar,
        )

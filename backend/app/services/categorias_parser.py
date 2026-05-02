"""
Categorias (TipoProducto) Import Parser

Handles parsing and validation of Excel/CSV files containing product
category (TipoProducto) data during onboarding. Provides:
- Excel (.xlsx) and CSV (.csv) file parsing
- Row-level validation with detailed error messages
- Case-insensitive duplicate detection (within file and against DB)
- Template generation with openpyxl
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl


@dataclass(frozen=True)
class ParsedRow:
    """Single validated categoria row."""
    row_num: int
    nombre: str
    status: str  # "crear" | "omitir"


@dataclass(frozen=True)
class InvalidRow:
    """Single invalid row with error details."""
    row_num: int
    nombre_raw: Optional[str]
    motivo: str


@dataclass
class ParseResult:
    """Results from parsing a categorias file."""
    valid_rows: list[ParsedRow] = field(default_factory=list)
    invalid_rows: list[InvalidRow] = field(default_factory=list)
    a_crear: int = 0
    a_omitir: int = 0


class ParseError(Exception):
    """Raised when file format is invalid (not recoverable per-row)."""
    pass


class CategoriasParser:
    """Parse and validate categorias (TipoProducto) import files."""

    REQUIRED_COLUMNS = {"nombre"}

    @staticmethod
    def parse(
        content: bytes,
        filename: str,
        existing_nombres: Optional[set[str]] = None,
    ) -> ParseResult:
        """
        Parse and validate a categorias import file.

        Args:
            content: File bytes
            filename: Original filename (used to detect format: .xlsx or .csv)
            existing_nombres: Set of lowercased nombres already in DB (for omit check)

        Returns:
            ParseResult with valid/invalid rows and create/omit counts

        Raises:
            ParseError: If file format is invalid or required column missing
        """
        existing = existing_nombres or set()
        parser = CategoriasParser()
        ext = Path(filename).suffix.lower()

        if ext == ".xlsx":
            raw_rows = parser._read_xlsx(content)
        elif ext == ".csv":
            raw_rows = parser._read_csv(content)
        else:
            raise ParseError(f"Formato no soportado: '{ext}'. Use .xlsx o .csv")

        if not raw_rows:
            raise ParseError("El archivo está vacío")

        # Row 0: headers
        header_row = raw_rows[0]
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in header_row
        ]

        # Find required column indices
        if "nombre" not in headers:
            raise ParseError(
                f"Columna requerida faltante: 'nombre'. "
                f"Columnas encontradas: {[h for h in headers if h]}"
            )
        nombre_col_idx = headers.index("nombre")

        # Auto-detect documentation row (row 1):
        # Skip row 1 if the nombre cell is empty or non-data-looking
        data_rows = raw_rows[1:]
        if len(raw_rows) > 1:
            potential_doc_row = raw_rows[1]
            nombre_val = (
                potential_doc_row[nombre_col_idx]
                if nombre_col_idx < len(potential_doc_row)
                else None
            )
            if nombre_val is None or str(nombre_val).strip() == "":
                # Empty nombre cell → skip as doc row
                data_rows = raw_rows[2:]
            else:
                val_str = str(nombre_val).strip()
                # Heuristic: if the value contains typical doc words or is very long, skip
                doc_keywords = ("nombre", "ej:", "ejemplo", "requerido", "categoría", "categoria")
                if any(kw in val_str.lower() for kw in doc_keywords) or len(val_str) > 60:
                    data_rows = raw_rows[2:]

        if not data_rows:
            raise ParseError("El archivo no contiene filas de datos")

        # Track within-file duplicates (case-insensitive)
        seen_in_file: dict[str, int] = {}  # lower(nombre) -> first row_num

        result = ParseResult()
        header_offset = len(raw_rows) - len(data_rows)  # rows skipped (header + optional doc)

        for idx, raw_row in enumerate(data_rows):
            row_num = header_offset + idx + 1  # 1-based row number in file

            # Extract nombre value
            nombre_raw = (
                raw_row[nombre_col_idx]
                if nombre_col_idx < len(raw_row)
                else None
            )

            if nombre_raw is None or str(nombre_raw).strip() == "":
                result.invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    nombre_raw=None,
                    motivo="Nombre es requerido",
                ))
                continue

            nombre = str(nombre_raw).strip()

            if len(nombre) > 100:
                result.invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    nombre_raw=nombre_raw,
                    motivo=f"Nombre excede 100 caracteres (tiene {len(nombre)})",
                ))
                continue

            nombre_lower = nombre.lower()

            # Within-file duplicate check
            if nombre_lower in seen_in_file:
                first_row = seen_in_file[nombre_lower]
                result.invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    nombre_raw=nombre_raw,
                    motivo=f"Nombre duplicado en el archivo (fila {first_row})",
                ))
                continue

            seen_in_file[nombre_lower] = row_num

            # DB-level existence check → omit (not an error)
            if nombre_lower in existing:
                parsed = ParsedRow(row_num=row_num, nombre=nombre, status="omitir")
                result.valid_rows.append(parsed)
                result.a_omitir += 1
            else:
                parsed = ParsedRow(row_num=row_num, nombre=nombre, status="crear")
                result.valid_rows.append(parsed)
                result.a_crear += 1

        return result

    @staticmethod
    def generate_template() -> bytes:
        """
        Generate Excel template with header row (bold/gray), documentation row
        (italic/light) and 3 example rows.

        Returns:
            Excel file bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Categorías"

        headers = ["nombre", "familia_padre"]
        documentation = [
            "Nombre de la categoría (requerido, máx. 100 chars)",
            "Familia padre (opcional, solo referencia — no se importa)",
        ]
        examples = [
            ["Frutas", "Alimentos"],
            ["Verduras", "Alimentos"],
            ["Lácteos", "Alimentos"],
        ]

        # Row 1: headers — bold + gray background
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True, size=11)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        # Row 2: documentation row — italic + light background
        for col_num, doc in enumerate(documentation, 1):
            cell = ws.cell(row=2, column=col_num, value=doc)
            cell.font = openpyxl.styles.Font(italic=True, size=9)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
            )

        # Rows 3-5: example data
        for row_idx, example in enumerate(examples, 3):
            for col_idx, value in enumerate(example, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 30

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _read_xlsx(content: bytes) -> list[list]:
        """Read Excel file and return rows as lists."""
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ParseError(f"No se pudo abrir el archivo Excel: {exc}")
        try:
            ws = wb.active
            return [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

    @staticmethod
    def _read_csv(content: bytes) -> list[list]:
        """Read CSV file and return rows as lists."""
        try:
            text = content.decode("utf-8-sig")  # handle BOM
        except UnicodeDecodeError:
            try:
                text = content.decode("latin-1")
            except Exception as exc:
                raise ParseError(f"No se pudo decodificar el CSV: {exc}")

        reader = csv.reader(io.StringIO(text))
        return [list(row) for row in reader]

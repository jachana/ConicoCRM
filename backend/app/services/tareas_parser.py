from __future__ import annotations

import hashlib
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

import openpyxl


class ParseError(Exception):
    """Raised when file format is invalid (not recoverable per-row)."""
    pass


TIPOS_VALIDOS = {"llamada", "email", "reunion", "visita", "cobranza", "otro"}
ESTADOS_SKIP = {"cerrado", "completado", "archivado", "cancelado"}
ESTADOS_VALIDOS = {"pendiente", "en_curso"}


@dataclass
class TareaRow:
    descripcion: str
    titulo: str  # truncated to 255
    due_date: date
    tipo: Optional[str]
    rut_cliente: Optional[str]
    asignado_email: Optional[str]
    estado: str
    prioridad: Optional[str]
    dedup_key: str
    row_num: int
    status: str  # "crear" | "omitir"


@dataclass
class InvalidRow:
    row_num: int
    descripcion_raw: Optional[str]
    motivo: str


@dataclass
class ParseResult:
    valid_rows: list[TareaRow] = field(default_factory=list)
    invalid_rows: list[InvalidRow] = field(default_factory=list)
    a_crear: int = 0
    a_omitir: int = 0


def _compute_dedup_key(rut_cliente: Optional[str], tipo: Optional[str], descripcion: str, due: date) -> str:
    raw = "|".join([
        rut_cliente or "",
        tipo or "",
        descripcion,
        due.isoformat(),
    ])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:80]


def _parse_date(s: str) -> Optional[date]:
    # Try Excel serial first
    try:
        serial = float(s)
        from datetime import timedelta
        epoch = date(1899, 12, 30)
        return epoch + timedelta(days=int(serial))
    except (ValueError, TypeError):
        pass
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _skip_doc_row(raw_rows: list[list], col: dict, anchor_col: str) -> list[list]:
    """Skip the optional documentation row (row 2 after headers)."""
    if len(raw_rows) <= 1:
        return []
    anchor_idx = col.get(anchor_col, -1)
    potential_doc = raw_rows[1]
    anchor_val = (
        potential_doc[anchor_idx]
        if anchor_idx >= 0 and anchor_idx < len(potential_doc)
        else None
    )
    if anchor_val is None or str(anchor_val).strip() == "":
        return raw_rows[2:]
    val_str = str(anchor_val).strip()
    doc_keywords = ("requerido", "ej:", "ejemplo", "opcional", "descripcion", "descripción")
    if any(kw in val_str.lower() for kw in doc_keywords) or len(val_str) > 200:
        return raw_rows[2:]
    return raw_rows[1:]


class TareasParser:
    """Parse and validate Tareas import files (single-sheet Excel)."""

    SHEET_NAME = "Tareas"
    REQUIRED_COLS = {"descripcion", "fecha_vencimiento"}
    OPTIONAL_COLS = {"rut_cliente", "tipo", "asignado_email", "estado", "prioridad"}

    @staticmethod
    def parse(
        content: bytes,
        filename: str,
        existing_dedup_keys: set[str],
    ) -> ParseResult:
        existing_dedup_keys = existing_dedup_keys or set()

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ParseError(f"No se pudo abrir el archivo Excel: {exc}")

        try:
            sheet_names = wb.sheetnames
            if TareasParser.SHEET_NAME not in sheet_names:
                raise ParseError(
                    f"No se encontró la hoja '{TareasParser.SHEET_NAME}'. "
                    f"Hojas encontradas: {sheet_names}"
                )
            raw_rows = [list(row) for row in wb[TareasParser.SHEET_NAME].iter_rows(values_only=True)]
        finally:
            wb.close()

        if not raw_rows:
            raise ParseError(f"La hoja '{TareasParser.SHEET_NAME}' está vacía")

        # --- Parse headers ---
        headers = [str(h).strip().lower() if h is not None else "" for h in raw_rows[0]]
        missing = TareasParser.REQUIRED_COLS - set(headers)
        if missing:
            raise ParseError(
                f"Hoja '{TareasParser.SHEET_NAME}' — columna(s) requerida(s) faltante(s): "
                f"{sorted(missing)}. Columnas encontradas: {[h for h in headers if h]}"
            )
        col = {h: i for i, h in enumerate(headers) if h}

        def _get(row: list, col_name: str) -> Optional[str]:
            i = col.get(col_name, -1)
            if i < 0 or i >= len(row):
                return None
            val = row[i]
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        data_rows = _skip_doc_row(raw_rows, col, "descripcion")
        header_offset = len(raw_rows) - len(data_rows)

        result = ParseResult()
        seen_dedup_keys: set[str] = set(existing_dedup_keys)

        for idx, raw_row in enumerate(data_rows):
            row_num = header_offset + idx + 1
            errors = []

            descripcion_raw = _get(raw_row, "descripcion")
            if not descripcion_raw:
                errors.append("descripcion es requerida")

            fecha_raw = _get(raw_row, "fecha_vencimiento")
            parsed_fecha = None
            if not fecha_raw:
                errors.append("fecha_vencimiento es requerida")
            else:
                parsed_fecha = _parse_date(fecha_raw)
                if parsed_fecha is None:
                    errors.append(
                        f"fecha_vencimiento '{fecha_raw}' no tiene formato válido "
                        f"(YYYY-MM-DD, DD-MM-YYYY o DD/MM/YYYY)"
                    )

            if errors:
                result.invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    descripcion_raw=descripcion_raw,
                    motivo="; ".join(errors),
                ))
                continue

            # Optional fields
            tipo_raw = _get(raw_row, "tipo")
            parsed_tipo = None
            if tipo_raw is not None:
                t = tipo_raw.lower()
                if t in TIPOS_VALIDOS:
                    parsed_tipo = t
                else:
                    errors.append(
                        f"tipo '{tipo_raw}' no válido; debe ser uno de: {sorted(TIPOS_VALIDOS)}"
                    )

            estado_raw = _get(raw_row, "estado")
            parsed_estado = "pendiente"
            if estado_raw is not None:
                e = estado_raw.lower()
                if e in ESTADOS_SKIP:
                    # Skip this row silently — it's closed/completed
                    continue
                if e in ESTADOS_VALIDOS:
                    parsed_estado = e
                else:
                    errors.append(
                        f"estado '{estado_raw}' no válido; debe ser pendiente|en_curso "
                        f"(o cerrado/completado para omitir la fila)"
                    )

            if errors:
                result.invalid_rows.append(InvalidRow(
                    row_num=row_num,
                    descripcion_raw=descripcion_raw,
                    motivo="; ".join(errors),
                ))
                continue

            rut_raw = _get(raw_row, "rut_cliente")
            asignado_email_raw = _get(raw_row, "asignado_email")
            prioridad_raw = _get(raw_row, "prioridad")

            # Build titulo: optionally prefix with [Prioridad]
            titulo_base = descripcion_raw[:255]
            if prioridad_raw:
                prefixed = f"[{prioridad_raw}] {descripcion_raw}"
                titulo_base = prefixed[:255]

            dedup_key = _compute_dedup_key(rut_raw, parsed_tipo, descripcion_raw, parsed_fecha)

            if dedup_key in seen_dedup_keys:
                row_obj = TareaRow(
                    descripcion=descripcion_raw,
                    titulo=titulo_base,
                    due_date=parsed_fecha,
                    tipo=parsed_tipo,
                    rut_cliente=rut_raw,
                    asignado_email=asignado_email_raw,
                    estado=parsed_estado,
                    prioridad=prioridad_raw,
                    dedup_key=dedup_key,
                    row_num=row_num,
                    status="omitir",
                )
                result.valid_rows.append(row_obj)
                result.a_omitir += 1
            else:
                seen_dedup_keys.add(dedup_key)
                row_obj = TareaRow(
                    descripcion=descripcion_raw,
                    titulo=titulo_base,
                    due_date=parsed_fecha,
                    tipo=parsed_tipo,
                    rut_cliente=rut_raw,
                    asignado_email=asignado_email_raw,
                    estado=parsed_estado,
                    prioridad=prioridad_raw,
                    dedup_key=dedup_key,
                    row_num=row_num,
                    status="crear",
                )
                result.valid_rows.append(row_obj)
                result.a_crear += 1

        return result

    @staticmethod
    def generate_template() -> bytes:
        wb = openpyxl.Workbook()

        header_font = openpyxl.styles.Font(bold=True, size=11)
        header_fill = openpyxl.styles.PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        doc_font = openpyxl.styles.Font(italic=True, size=9)
        doc_fill = openpyxl.styles.PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        )

        ws = wb.active
        ws.title = TareasParser.SHEET_NAME

        col_headers = [
            "descripcion",
            "fecha_vencimiento",
            "rut_cliente",
            "tipo",
            "asignado_email",
            "estado",
            "prioridad",
        ]
        col_docs = [
            "Requerido: descripción de la tarea (máx. 255 chars para título)",
            "Requerido: YYYY-MM-DD, DD-MM-YYYY o DD/MM/YYYY",
            "Opcional: RUT cliente (ej: 12345678-9)",
            "Opcional: llamada|email|reunion|visita|cobranza|otro (default: otro)",
            "Opcional: email del usuario asignado (ej: vendedor@empresa.cl)",
            "Opcional: pendiente|en_curso; filas con cerrado/completado se omiten",
            "Opcional: Alta|Media|Baja (se prefija al título: [Alta] ...)",
        ]
        col_example = [
            "Llamar al cliente para confirmar pedido",
            "2024-06-15",
            "12345678-9",
            "llamada",
            "vendedor@empresa.cl",
            "pendiente",
            "Alta",
        ]

        for col_num, h in enumerate(col_headers, 1):
            cell = ws.cell(row=1, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for col_num, d in enumerate(col_docs, 1):
            cell = ws.cell(row=2, column=col_num, value=d)
            cell.font = doc_font
            cell.fill = doc_fill
        for col_idx, v in enumerate(col_example, 1):
            ws.cell(row=3, column=col_idx, value=v)

        col_widths = [50, 22, 20, 14, 30, 14, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

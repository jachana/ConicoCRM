"""
Payment Import Parser

Handles parsing and validation of payment import Excel files with comprehensive
error reporting. Provides:
- Excel file parsing with openpyxl
- Row-level validation with detailed error messages
- Type and format validation
- Spanish column support
"""

from __future__ import annotations

import hashlib
import io
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Optional

import openpyxl


# ── Public constants ──────────────────────────────────────────────────────────

REQUIRED_COLUMNS = (
    "fecha_pago", "rut_cliente", "monto", "medio_pago",
    "referencia", "folio_documento", "tipo_documento",
)

VALID_PAYMENT_METHODS = frozenset(["efectivo", "transferencia", "cheque", "tarjeta", "otro"])

VALID_DOCUMENT_TYPES = frozenset(["factura", "boleta", "nota_credito", "nota_debito", "guia_despacho"])


# ── Data model ────────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class ParsedPayment:
    """Single validated payment row."""
    fila: int
    fecha_pago: date
    rut_cliente: str          # normalized (dots/spaces removed)
    rut_cliente_raw: str      # original input
    monto: Decimal
    medio_pago: str           # lowercased
    referencia: str
    folio_documento: str
    tipo_documento: str       # lowercased
    hash_key: str


@dataclass(frozen=True)
class InvalidRow:
    """Single invalid payment row with errors."""
    fila: int
    motivo: str               # joined error string with "; "
    valores_raw: dict         # {col_name: raw_str_value}


@dataclass
class ParseResult:
    """Results from parsing a payment file."""
    validas: list[ParsedPayment]
    invalidas: list[InvalidRow]

    # Backward compat properties
    @property
    def valid_rows(self) -> list[ParsedPayment]:
        return self.validas

    @property
    def invalid_rows(self) -> list[InvalidRow]:
        return self.invalidas

    @property
    def valid_count(self) -> int:
        return len(self.validas)

    @property
    def invalid_count(self) -> int:
        return len(self.invalidas)


class ParseError(Exception):
    """Raised when file format is invalid (not recoverable per-row)."""
    pass


# ── Module-level function ─────────────────────────────────────────────────────

def parse_pagos_xlsx(content: bytes) -> ParseResult:
    """Module-level convenience wrapper around PaymentParser.parse()."""
    return PaymentParser.parse(content)


# ── RUT helpers ───────────────────────────────────────────────────────────────

def _normalize_rut(raw: str) -> str:
    """Remove dots and spaces from RUT string (keeps dash and DV)."""
    return raw.replace(".", "").replace(" ", "")


def _check_rut_dv(rut_digits: str, dv: str) -> bool:
    """Validate Chilean RUT check digit using modulo 11 algorithm."""
    multipliers = [2, 3, 4, 5, 6, 7]
    total = sum(int(d) * multipliers[i % 6] for i, d in enumerate(reversed(rut_digits)))
    rem = 11 - (total % 11)
    expected = "K" if rem == 10 else "0" if rem == 11 else str(rem)
    return dv.upper() == expected


def _validate_rut_format(rut_normalized: str) -> bool:
    """
    Validate that a normalized RUT (no dots, no spaces) has the form
    XXXXXXXX-D where X are digits and D is a digit or K.
    """
    if "-" not in rut_normalized:
        return False
    parts = rut_normalized.rsplit("-", 1)
    if len(parts) != 2:
        return False
    digits, dv = parts
    if not digits.isdigit() or len(digits) < 7:
        return False
    if not (dv.isdigit() or dv.upper() == "K"):
        return False
    return True


# ── Main class ────────────────────────────────────────────────────────────────

class PaymentParser:
    """Parse and validate payment import Excel files."""

    @staticmethod
    def parse(content: bytes, filename: str = None) -> ParseResult:
        """
        Parse and validate a payment import file.

        Args:
            content: File bytes
            filename: Accepted but ignored (all bytes treated as xlsx)

        Returns:
            ParseResult with valid and invalid rows

        Raises:
            ParseError: If file format is invalid or headers are missing
        """
        parser = PaymentParser()
        return parser._parse_file(content)

    @staticmethod
    def generate_template() -> bytes:
        """
        Generate Excel template with headers and one example row.

        Returns:
            Excel file bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pagos"

        headers = list(REQUIRED_COLUMNS)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True, size=11)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        # Row 2: Example data (valid RUT with correct DV)
        example = [
            "2026-05-01",
            "11111111-1",
            "50000.00",
            "transferencia",
            "TRX20260501001",
            "1001",
            "factura",
        ]
        for col_num, value in enumerate(example, 1):
            ws.cell(row=2, column=col_num, value=value)

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _parse_file(self, content: bytes) -> ParseResult:
        """Internal parse implementation."""
        try:
            raw_rows = self._read_xlsx(content)
        except Exception as exc:
            raise ParseError(f"No se pudo leer el archivo: {exc}") from exc

        if not raw_rows:
            raise ParseError(
                "El archivo no contiene filas de datos (solo encabezado o está vacío)"
            )

        # Row 0: headers, Row 1+: data
        header_row = raw_rows[0]
        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in header_row
        ]

        # Validate required columns FIRST (before checking for data rows)
        missing = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing:
            raise ParseError(
                f"Columnas requeridas faltantes: {', '.join(missing)}. "
                f"Columnas encontradas: {[h for h in headers if h]}"
            )

        # After header validation, check that there is at least one data row
        if len(raw_rows) < 2:
            raise ParseError(
                "El archivo no contiene filas de datos (solo encabezado o está vacío)"
            )

        # Build index map
        col_indices = {col: headers.index(col) for col in REQUIRED_COLUMNS}

        validas: list[ParsedPayment] = []
        invalidas: list[InvalidRow] = []

        for row_offset, raw_row in enumerate(raw_rows[1:], start=0):
            fila = row_offset + 2  # header=1, first data row=2

            # Skip entirely blank rows
            if all(cell is None or str(cell).strip() == "" for cell in raw_row):
                continue

            result = self._parse_row(fila=fila, raw_row=raw_row, col_indices=col_indices)

            if result[0] is not None:
                validas.append(result[0])
            else:
                invalidas.append(result[1])

        return ParseResult(validas=validas, invalidas=invalidas)

    def _parse_row(
        self,
        fila: int,
        raw_row: list,
        col_indices: dict[str, int],
    ) -> tuple[Optional[ParsedPayment], Optional[InvalidRow]]:
        """Parse and validate a single data row."""
        errors = []

        def _raw(col: str) -> str:
            idx = col_indices[col]
            val = raw_row[idx] if idx < len(raw_row) else None
            if val is None:
                return ""
            # openpyxl may return datetime/date objects for date cells
            from datetime import datetime as _dt, date as _date
            if isinstance(val, _dt):
                return val.strftime("%Y-%m-%d")
            if isinstance(val, _date):
                return val.strftime("%Y-%m-%d")
            return str(val).strip()

        # Collect raw values for error reporting
        valores_raw = {col: _raw(col) for col in REQUIRED_COLUMNS}

        # Validate each field
        fecha_pago = self._validate_fecha_pago(valores_raw["fecha_pago"], errors)
        rut_cliente, rut_cliente_raw = self._validate_rut_cliente(valores_raw["rut_cliente"], errors)
        monto = self._validate_monto(valores_raw["monto"], errors)
        medio_pago = self._validate_medio_pago(valores_raw["medio_pago"], errors)
        referencia = self._validate_referencia(valores_raw["referencia"], errors)
        folio_documento = self._validate_folio_documento(valores_raw["folio_documento"], errors)
        tipo_documento = self._validate_tipo_documento(valores_raw["tipo_documento"], errors)

        if errors:
            return (
                None,
                InvalidRow(
                    fila=fila,
                    motivo="; ".join(errors),
                    valores_raw=valores_raw,
                ),
            )

        hash_key = self._generate_hash_key(fecha_pago, rut_cliente, monto, referencia)

        return (
            ParsedPayment(
                fila=fila,
                fecha_pago=fecha_pago,
                rut_cliente=rut_cliente,
                rut_cliente_raw=rut_cliente_raw,
                monto=monto,
                medio_pago=medio_pago,
                referencia=referencia,
                folio_documento=folio_documento,
                tipo_documento=tipo_documento,
                hash_key=hash_key,
            ),
            None,
        )

    # ── Field validators ──────────────────────────────────────────────────────

    @staticmethod
    def _validate_rut_cliente(raw: str, errors: list[str]) -> tuple[Optional[str], str]:
        """
        Validate and normalize RUT cliente.

        Returns (normalized_rut, raw_value). normalized_rut is None on error.
        """
        if not raw:
            errors.append("rut_cliente: requerido")
            return None, raw

        normalized = _normalize_rut(raw)

        if not _validate_rut_format(normalized):
            errors.append(f"RUT inválido: {raw}")
            return None, raw

        parts = normalized.rsplit("-", 1)
        digits, dv = parts
        if not _check_rut_dv(digits, dv):
            errors.append(f"RUT inválido: {raw} (dígito verificador incorrecto)")
            return None, raw

        return normalized, raw

    @staticmethod
    def _validate_monto(raw: str, errors: list[str]) -> Optional[Decimal]:
        """Validate and extract monto."""
        if not raw:
            errors.append("monto: requerido")
            return None

        try:
            monto = Decimal(raw)
        except (InvalidOperation, ValueError):
            errors.append(f"monto inválido: '{raw}'")
            return None

        if monto <= 0:
            errors.append(f"monto debe ser positivo: {raw}")
            return None

        return monto

    @staticmethod
    def _validate_fecha_pago(raw: str, errors: list[str]) -> Optional[date]:
        """Validate and extract fecha_pago."""
        if not raw:
            errors.append("fecha_pago: requerido")
            return None

        DATE_FORMATS = ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y"]
        from datetime import datetime

        # openpyxl may store datetimes; str() gives "YYYY-MM-DD HH:MM:SS" — strip the time part
        clean = raw.split(" ")[0] if " " in raw else raw

        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(clean, fmt).date()
            except ValueError:
                continue

        errors.append(f"fecha_pago: formato de fecha inválido '{raw}'")
        return None

    @staticmethod
    def _validate_medio_pago(raw: str, errors: list[str]) -> Optional[str]:
        """Validate and extract medio_pago."""
        if not raw:
            errors.append("medio_pago: requerido")
            return None

        lower = raw.lower()
        if lower not in VALID_PAYMENT_METHODS:
            errors.append(f"medio_pago '{raw}' no válido. Valores: {sorted(VALID_PAYMENT_METHODS)}")
            return None

        return lower

    @staticmethod
    def _validate_referencia(raw: str, errors: list[str]) -> Optional[str]:
        """Validate referencia (now required)."""
        if not raw:
            errors.append("referencia: requerida")
            return None
        return raw

    @staticmethod
    def _validate_folio_documento(raw: str, errors: list[str]) -> Optional[str]:
        """Validate folio_documento (now required)."""
        if not raw:
            errors.append("folio_documento: requerido")
            return None
        return raw

    @staticmethod
    def _validate_tipo_documento(raw: str, errors: list[str]) -> Optional[str]:
        """Validate tipo_documento (now required)."""
        if not raw:
            errors.append("tipo_documento: requerido")
            return None

        lower = raw.lower()
        if lower not in VALID_DOCUMENT_TYPES:
            errors.append(
                f"tipo_documento '{raw}' no válido. Valores: {sorted(VALID_DOCUMENT_TYPES)}"
            )
            return None

        return lower

    @staticmethod
    def _generate_hash_key(
        fecha_pago: date,
        rut_cliente: str,
        monto: Decimal,
        referencia: Optional[str],
    ) -> str:
        """Generate SHA256 hash key for idempotency detection."""
        key_str = f"{fecha_pago}|{rut_cliente}|{monto}|{referencia or ''}"
        return hashlib.sha256(key_str.encode()).hexdigest()

    @staticmethod
    def _read_xlsx(content: bytes) -> list[list]:
        """Read Excel file and return rows as lists."""
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            return [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
